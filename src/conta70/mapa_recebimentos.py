"""Conta 70 — Mapa de Recebimentos (v10).

Visão consolidada, aprovada pela Débora na prévia em Excel, que cruza a
**Capa da Conta 70** (o que caiu e está numerado ou em aberto) com as
**notas emitidas não baixadas** (Movimentação Financeira), organizando tudo em
três blocos com origem explícita:

  - BLOCO CAPA       : origem Capa da Conta 70 (número, data, banco, histórico,
                       identificação CPF/CNPJ ou nome/origem, valor, R/D).
  - BLOCO FINANCEIRO : origem Movimentação Financeira — só sugere NF quando o
                       VALOR bate (Exato) ou por somatório (Soma — conferir).
  - BLOCO MAPA       : decisão do app (status, parceiro efetivo, justificativa).

Regras invioláveis (mesmas da Débora):
  - Zero falso positivo: NF só é sugerida quando o valor bate ao centavo
    (Exato) ou fecha por soma de notas do mesmo CNPJ (Soma — conferir, deixado
    para a Débora decidir). Se o valor difere, NÃO sugere NF.
  - "Numerado na Capa" aceita numeração ALFANUMÉRICA (ex.: "686a"): qualquer
    valor não-vazio na coluna de numeração = atrelado. Só vazio = em aberto.
  - Deixe em branco quando não houver dado — sem placeholder, sem chute.

Este módulo é ADITIVO e independente: não altera o motor de atrelamento
(`casamento.py`). Ele relê a Capa preservando a numeração como texto.
"""
from __future__ import annotations

import re
from itertools import combinations
from typing import Any

import pandas as pd

from src.conta70.casamento import (
    _ler_planilha,
    _detectar_header,
    _to_data,
    extrair_identidade,
)

# ---------------------------------------------------------------------------
# Identificação e descritor de origem
# ---------------------------------------------------------------------------
_BANCOS_SUFIXO = [
    "SANTANDER", "BRADESCO", "SICREDI", "ITAU KING", "ITAU PISA", "ITAÚ PISA",
    "ITAÚ KING", "ITAU", "ITAÚ", "CAIXA", "KING", "PISA", "TRIO", "NUBANK",
    "INTER", "SICOOB", "BANCO DO BRASIL", "MAESTRO",
]
_PREFIXO_DEP = re.compile(r'^\s*"?\s*DEP\s*(N(?:AO|ÃO)?\s*)?IDENT\.?[\s\-:]*', re.I)


def documento_formatado(historico: Any) -> tuple[str, str]:
    """(dígitos, formatado) do CPF/CNPJ no histórico. ('','') se não houver."""
    m = re.search(r"(\d{11,14})", str(historico or ""))
    if not m:
        return "", ""
    v = m.group(1)
    if len(v) >= 14:
        v = v[:14]
        return v, f"{v[:2]}.{v[2:5]}.{v[5:8]}/{v[8:12]}-{v[12:]}"
    if len(v) >= 11:
        v = v[:11]
        return v, f"{v[:3]}.{v[3:6]}.{v[6:9]}-{v[9:]}"
    return v, v


def descritor_origem(historico: Any) -> str:
    """Nome/origem legível quando não há CPF/CNPJ.

    Remove o prefixo "DEP (N) IDENT" e o sufixo de banco/empresa, devolvendo o
    miolo do histórico (ex.: 'THOMAS PEREIRA ALZEMAN', 'PIX TRANSF SHEILA',
    'SISPAG ENGEARQ'). Vazio se não sobrar nada de útil.
    """
    s = str(historico or "").strip().strip('"').strip()
    s = _PREFIXO_DEP.sub("", s)
    partes = re.split(r"\s*-\s*", s)
    while partes and any(b in partes[-1].upper() for b in _BANCOS_SUFIXO):
        partes.pop()
    s = " - ".join(partes)
    s = re.sub(r"\s+", " ", s).strip(" -:").strip()
    return s[:48]


def _banco_do_historico(historico: Any) -> str:
    h = str(historico or "").upper()
    for chave, rotulo in [
        ("SANTANDER", "Santander"), ("BRADESCO", "Bradesco"), ("SICREDI", "Sicredi"),
        ("ITAU", "Itaú"), ("ITAÚ", "Itaú"), ("CAIXA", "Caixa"), ("NUBANK", "Nubank"),
        ("INTER", "Inter"), ("SICOOB", "Sicoob"), ("BANCO DO BRASIL", "Banco do Brasil"),
    ]:
        if chave in h:
            return rotulo
    return ""


def _fmt_numero(v: Any) -> str:
    """Numeração da Capa como texto. Aceita alfanumérico (686a). '' = em aberto."""
    if pd.isna(v):
        return ""
    sv = str(v).strip()
    if sv.lower() in ("", "nan", "none"):
        return ""
    m = re.match(r"^(\d+)\.0$", sv)   # 839.0 -> 839
    return m.group(1) if m else sv


# ---------------------------------------------------------------------------
# Leitura da Capa preservando a numeração como TEXTO
# ---------------------------------------------------------------------------
_SANKHYA_POS = {
    "dt. conciliação", "dt. conciliacao", "núm. único bancário",
    "num. unico bancario", "pré-data", "pre-data", "usuário", "usuario",
    "dt. alteração", "dt. alteracao", "vlr. troco", "vlr. cheque", "conciliado",
}


def carregar_capa_bruta(arquivo: Any) -> pd.DataFrame:
    """Lê a Capa da Conta 70 preservando a numeração ('x') como texto.

    Saída: num_txt (str, '' = em aberto), valor, rd (Receita/Despesa), data,
    historico, iddig, idfmt (CPF/CNPJ), identificacao (doc ou descritor), banco.
    Somente leitura.
    """
    hdr = _detectar_header(arquivo)
    if hasattr(arquivo, "seek"):
        try:
            arquivo.seek(0)
        except Exception:
            pass
    df = _ler_planilha(arquivo, header=hdr)
    df.columns = [str(c).strip() for c in df.columns]

    def col(*nomes):
        for n in nomes:
            for c in df.columns:
                if c.lower() == n.lower():
                    return c
        return None

    c_tipo = col("Tipo de Movimento")
    c_val = col("Vlr. Lançamento", "Valor", "Vlr Lancamento")
    c_rd = col("Receita/Despesa", "Receita/D")
    c_dt = col("Dt. Lançamento", "Data")
    c_hist = col("Histórico", "Historico")

    # coluna de numeração: EXTRA após o Histórico ("x"/vazia/unnamed) — não é
    # coluna de Sankhya. Mantida como TEXTO para aceitar alfanumérico.
    c_num = None
    if c_hist is not None:
        idx_hist = list(df.columns).index(c_hist)
        for c in list(df.columns)[idx_hist + 1:]:
            hl = c.strip().lower()
            if hl in _SANKHYA_POS:
                continue
            eh_rotulo = hl in ("x", "nan", "") or hl.startswith("unnamed")
            nao_vazio = df[c].astype(str).str.strip().replace(
                {"nan": "", "None": ""}).str.len().gt(0).sum()
            if nao_vazio > 0 and eh_rotulo:
                c_num = c
                break

    out = pd.DataFrame()
    if c_tipo is not None:
        df = df[df[c_tipo].notna()].reset_index(drop=True)
    out["num_txt"] = df[c_num].map(_fmt_numero) if c_num else ""
    out["valor"] = pd.to_numeric(df[c_val], errors="coerce") if c_val else pd.NA
    out["rd"] = (df[c_rd].astype(str).str.strip().str.title() if c_rd else "")
    out["data"] = df[c_dt].map(_to_data) if c_dt else pd.NaT
    out["historico"] = df[c_hist].astype(str) if c_hist else ""
    out = out.dropna(subset=["valor"]).reset_index(drop=True)

    docs = out["historico"].map(documento_formatado)
    out["iddig"] = [d[0] for d in docs]
    out["idfmt"] = [d[1] for d in docs]
    out["identificacao"] = [
        fmt if dig else descritor_origem(h)
        for dig, fmt, h in zip(out["iddig"], out["idfmt"], out["historico"])
    ]
    out["banco"] = out["historico"].map(_banco_do_historico)
    out["aberto"] = out["num_txt"] == ""
    return out


# ---------------------------------------------------------------------------
# Casamento por VALOR (Exato / Soma — conferir / nada)
# ---------------------------------------------------------------------------
def _casar_valor(dep_abs: float, notas: list[tuple[str, str, float]]):
    """(tipo, [notas], nome). tipo in {'Exato','Soma',None}. Nunca chuta."""
    for nt, nm, v in notas:
        if abs(dep_abs - v) < 0.005:
            return "Exato", [nt], nm
    if len(notas) >= 2:
        for r in range(2, min(len(notas), 5) + 1):
            for combo in combinations(range(len(notas)), r):
                if abs(dep_abs - sum(notas[i][2] for i in combo)) < 0.005:
                    return "Soma", [notas[i][0] for i in combo], notas[combo[0]][1]
    return None, [], (notas[0][1] if notas else "")


def construir_mapa(capa: pd.DataFrame, faturamento: pd.DataFrame | None) -> tuple[pd.DataFrame, dict]:
    """Monta o Mapa de Recebimentos e o resumo da rodada.

    `capa` = saída de carregar_capa_bruta. `faturamento` = saída de
    casamento.carregar_faturamento (colunas cnpj, nota, nome, valor) ou None.
    """
    notas_por_cnpj: dict[str, list[tuple[str, str, float]]] = {}
    if faturamento is not None and not faturamento.empty:
        fat = faturamento[faturamento["cnpj"].astype(str) != ""].copy()
        fat["vlr"] = pd.to_numeric(fat["valor"], errors="coerce")
        for cn, g in fat.groupby("cnpj"):
            notas_por_cnpj[str(cn)] = [
                (str(r["nota"]).replace(".0", ""), str(r["nome"]).strip(),
                 float(r["vlr"]) if pd.notna(r["vlr"]) else 0.0)
                for _, r in g.iterrows()
            ]

    nf, pnf, tm, rep, pef, stt, stk, jus = [], [], [], [], [], [], [], []
    for _, r in capa.iterrows():
        aberto = bool(r["aberto"])
        dig = r["iddig"]
        val = abs(float(r["valor"])) if pd.notna(r["valor"]) else 0.0
        if not aberto:
            nf.append(""); pnf.append(""); tm.append(""); rep.append("")
            if "Receita" in str(r["rd"]):
                stt.append("Baixada no Sankhya"); stk.append("BAIXA"); pef.append(r["idfmt"])
                jus.append(f"Atrelada na Capa (Nº {r['num_txt']}) — baixa lançada no Sankhya")
            else:
                stt.append("Saída da Conta 70"); stk.append("SAIDA"); pef.append(r["idfmt"])
                jus.append(f"Saída da Conta 70 (Nº {r['num_txt']}) — contrapartida da baixa")
            continue
        notas = notas_por_cnpj.get(dig) if dig else None
        if notas:
            tipo, ns, nome = _casar_valor(val, notas)
            if tipo == "Exato":
                nf.append(ns[0]); pnf.append(nome); tm.append("Exato"); rep.append("1 de 1")
                pef.append(nome); stt.append("Identificado pelo Mapa"); stk.append("MAPA")
                jus.append(f"Valor bate EXATO com a NF {ns[0]} (parceiro {nome}) — origem: Movimentação Financeira")
            elif tipo == "Soma":
                nf.append(", ".join(ns)); pnf.append(nome); tm.append("Soma — conferir"); rep.append(f"{len(ns)} notas")
                pef.append(nome); stt.append("Pendente de baixa"); stk.append("PENDENTE")
                jus.append(f"{len(ns)} notas do CNPJ somam o valor — CONFERIR e decidir (origem: Movimentação Financeira)")
            else:
                nf.append(""); pnf.append(""); tm.append(""); rep.append("")
                pef.append(nome); stt.append("Pendente de baixa"); stk.append("PENDENTE")
                jus.append(f"CNPJ cadastrado (parceiro {nome}), mas nenhum valor de nota em aberto bate — conferir")
        else:
            nf.append(""); pnf.append(""); tm.append(""); rep.append("")
            if dig:
                pef.append(r["idfmt"]); stt.append("Pendente de baixa"); stk.append("PENDENTE")
                jus.append("CPF/CNPJ identificado, sem nota em aberto correspondente — falta baixar")
            else:
                desc = descritor_origem(r["historico"])
                if len(desc) >= 3:
                    pef.append(""); stt.append("Pendente de baixa"); stk.append("PENDENTE")
                    jus.append(f"Identificado por nome/origem no histórico ({desc}) — sem CPF/CNPJ; conferir e baixar")
                else:
                    pef.append(""); stt.append("Sem identificação no extrato bancário"); stk.append("SEM_ID")
                    jus.append("Sem CPF/CNPJ nem nome no histórico do banco")

    m = capa.copy()
    m["nf_baixar"] = nf; m["parceiro_nf"] = pnf; m["tipo_match"] = tm; m["repeticao"] = rep
    m["parceiro_efetivo"] = pef; m["status"] = stt; m["status_key"] = stk; m["justificativa"] = jus
    m["dias"] = (pd.Timestamp.today().normalize() - pd.to_datetime(m["data"], errors="coerce")).dt.days

    ab = m[m["aberto"]]
    des_ab = ab[ab["rd"].str.contains("Despesa", na=False)]
    rec_ab = ab[ab["rd"].str.contains("Receita", na=False)]
    resumo = {
        "n": int(len(m)),
        "n_com_num": int((m["num_txt"] != "").sum()),
        "em_aberto_qtd": int(len(ab)),
        "em_aberto_valor": float(ab["valor"].sum()),
        "saldo_desp_aberto": float(des_ab["valor"].sum()),
        "qtd_desp_aberto": int(len(des_ab)),
        "saldo_rec_aberto": float(rec_ab["valor"].sum()),
        "qtd_rec_aberto": int(len(rec_ab)),
        "exato": int((m["status_key"] == "MAPA").sum()),
        "soma_conferir": int((m["tipo_match"] == "Soma — conferir").sum()),
        "pendente": int((m["status_key"] == "PENDENTE").sum()),
        "sem_id": int((m["status_key"] == "SEM_ID").sum()),
        "baixada": int((m["status_key"] == "BAIXA").sum()),
        "saida": int((m["status_key"] == "SAIDA").sum()),
        "capa_ate": (str(pd.to_datetime(m["data"], errors="coerce").max().date())
                     if m["data"].notna().any() else ""),
    }
    return m, resumo


# rótulos de exibição dos 15 campos, na ordem dos 3 blocos
COLUNAS_MAPA = [
    ("num_txt", "Nº Capa 70"),
    ("data", "Data lançamento"),
    ("banco", "Banco"),
    ("historico", "Histórico da Capa"),
    ("identificacao", "Identificação (CPF/CNPJ / origem)"),
    ("valor", "Valor"),
    ("rd", "R/D"),
    ("nf_baixar", "NF a baixar (sugestão)"),
    ("parceiro_nf", "Parceiro da NF"),
    ("tipo_match", "Tipo de match"),
    ("repeticao", "Repetição"),
    ("status", "Status"),
    ("parceiro_efetivo", "Parceiro efetivo"),
    ("justificativa", "Justificativa do vínculo"),
]

# cor de fundo por status (identidade Grupo LLE)
CORES_STATUS = {
    "BAIXA": ("#0F8C3B", "#FFFFFF"),
    "SAIDA": ("#EDEFF2", "#56657D"),
    "MAPA": ("#0071FE", "#FFFFFF"),
    "PENDENTE": ("#FAC318", "#041747"),
    "SEM_ID": ("#D63031", "#FFFFFF"),
}


# ---------------------------------------------------------------------------
# Exportação para Excel (mesmo layout aprovado na prévia v10)
# ---------------------------------------------------------------------------
def exportar_mapa_excel(m: pd.DataFrame, resumo: dict) -> bytes:
    """Gera o .xlsx do Mapa (abas Mapa/Alerta/Resumo/Legenda + Listas oculta),
    na identidade Grupo LLE, idêntico à prévia aprovada. Retorna os bytes."""
    import io
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter as GL

    A = "Arial"
    def F(sz=10, b=False, c="1A1A1A"): return Font(name=A, size=sz, bold=b, color=c)
    def P(c): return PatternFill("solid", fgColor=c)
    thin = Side(style="thin", color="E4E4E4"); BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment("center", "center", wrap_text=True); L = Alignment("left", "center", wrap_text=True)
    Lw = Alignment("left", "top", wrap_text=True); Rg = Alignment("right", "center")
    NAVY="041747"; AZUL="0071FE"; AMAR="FAC318"; VERDE="0F8C3B"; VERM="D63031"; LARANJA="E67E22"; RX="6A4C93"
    HD_AZ="E6F0FF"; HD_VD="E7F4EC"; HD_RX="F1ECF7"
    ST={"BAIXA":(VERDE,"FFFFFF"),"SAIDA":("EDEFF2","56657D"),"MAPA":(AZUL,"FFFFFF"),
        "PENDENTE":(AMAR,NAVY),"SEM_ID":(VERM,"FFFFFF")}
    def brl(v):
        s=f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X","."); return ("-" if v<0 else "")+"R$ "+s
    def titlebar(ws, cols, txt):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c=ws.cell(1,1,"  "+txt); c.font=F(13,True,"FFFFFF"); c.fill=P(NAVY); c.alignment=L; ws.row_dimensions[1].height=26
        for j in range(1,cols+1): ws.cell(2,j).fill=P(AMAR)
        ws.row_dimensions[2].height=4; ws.sheet_view.showGridLines=False

    wb=Workbook()
    wl=wb.active; wl.title="Listas"; wl["A1"]="Sim"; wl["A2"]="Não"; wl.sheet_state="hidden"

    # ---- MAPA ----
    ws=wb.create_sheet("Mapa")
    cols=[("Nº Capa 70",12),("Data lançamento",13),("Banco",11),("Histórico da Capa",50),
          ("Identificação (CPF/CNPJ / origem)",24),("Valor",13),("R/D",9),
          ("NF a baixar (sugestão)",20),("Parceiro da NF",26),("Tipo de match",16),("Repetição",11),("Confirmar baixa?",14),
          ("Status",30),("Parceiro efetivo",26),("Justificativa do vínculo",54)]
    n=len(cols); titlebar(ws,n,"MAPA DE RECEBIMENTOS — CONTA 70   ·   v10")
    def band(a,b,txt,bg):
        ws.merge_cells(start_row=3,start_column=a,end_row=3,end_column=b)
        cc=ws.cell(3,a,txt); cc.font=F(9,True,"FFFFFF"); cc.fill=P(bg); cc.alignment=C
    band(1,7,"BLOCO CAPA  ·  origem: Capa da Conta 70",AZUL)
    band(8,12,"BLOCO FINANCEIRO  ·  origem: Movimentação Financeira (notas em aberto)",VERDE)
    band(13,15,"BLOCO MAPA  ·  decisão do app",RX)
    ws.row_dimensions[3].height=19
    hdbg=[HD_AZ]*7+[HD_VD]*5+[HD_RX]*3; hdfg=[AZUL]*7+[VERDE]*5+[RX]*3
    for j,((nome,w),bg,fg) in enumerate(zip(cols,hdbg,hdfg),1):
        cc=ws.cell(4,j,nome); cc.font=F(9,True,fg); cc.fill=P(bg); cc.alignment=C; cc.border=BRD; ws.column_dimensions[GL(j)].width=w
    ws.row_dimensions[4].height=30
    sf={k:P(v[0]) for k,v in ST.items()}; sfo={k:F(9,True,v[1]) for k,v in ST.items()}
    r0=5
    for i,(_,r) in enumerate(m.iterrows()):
        rr=r0+i
        ws.cell(rr,1,r["num_txt"]).alignment=C
        cd=ws.cell(rr,2, pd.to_datetime(r["data"]).to_pydatetime() if pd.notna(r["data"]) else None); cd.number_format="DD/MM/YYYY"; cd.alignment=C
        ws.cell(rr,3,r["banco"]).alignment=C
        ws.cell(rr,4,str(r["historico"])).alignment=Lw
        ws.cell(rr,5,r["identificacao"]).alignment=C
        cv=ws.cell(rr,6, float(r["valor"]) if pd.notna(r["valor"]) else None); cv.number_format="#,##0.00;[RED]-#,##0.00"; cv.alignment=Rg
        ws.cell(rr,7,r["rd"]).alignment=C
        ws.cell(rr,8,r["nf_baixar"]).alignment=C
        ws.cell(rr,9,r["parceiro_nf"]).alignment=L
        ws.cell(rr,10,r["tipo_match"]).alignment=C
        ws.cell(rr,11,r["repeticao"]).alignment=C
        ws.cell(rr,12,"")
        sk=r["status_key"]
        cs=ws.cell(rr,13,r["status"]); cs.fill=sf[sk]; cs.font=sfo[sk]; cs.alignment=L
        ws.cell(rr,14,r["parceiro_efetivo"]).alignment=L
        ws.cell(rr,15,r["justificativa"]).alignment=Lw
        ws.cell(rr,17,sk)
    last=r0+len(m)-1
    ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:{GL(n)}{last}"
    dv=DataValidation(type="list", formula1="Listas!$A$1:$A$2", allow_blank=True); ws.add_data_validation(dv); dv.add(f"L{r0}:L{last}")
    ws.column_dimensions[GL(16)].width=2; ws.column_dimensions[GL(17)].hidden=True
    LAST=str(last)

    # ---- ALERTA ----
    wa=wb.create_sheet("Alerta")
    al=m[m["aberto"]].copy(); al=al[al["dias"]>15].copy()
    def fa(x):
        if x<=30: return ("15-30 dias",AMAR,NAVY,"FEF7DE",0)
        if x<=60: return ("31-60 dias",LARANJA,"FFFFFF","FBEAD9",1)
        if x<=90: return ("61-90 dias",VERM,"FFFFFF","FADEDD",2)
        return ("90+ dias","5A0E0E","FFFFFF","E7C9C9",3)
    if not al.empty:
        z=al["dias"].map(fa)
        al["fl"]=[x[0] for x in z]; al["strip"]=[x[1] for x in z]; al["stfg"]=[x[2] for x in z]; al["tint"]=[x[3] for x in z]; al["ord"]=[x[4] for x in z]
        al=al.sort_values(["ord","dias"],ascending=[False,False])
    titlebar(wa,7,"ALERTA — RECEBIMENTOS PARADOS EM ABERTO (aging)")
    wa.merge_cells("A3:G3"); c=wa.cell(3,1,"Só o que está EM ABERTO na Capa (sem número). Dias = hoje − data. Prioridade crescente: amarelo → vinho."); c.font=F(9,False,"777777"); c.alignment=L
    col=1
    for lbl,strip,fg,tint in [("15-30 dias",AMAR,NAVY,"FEF7DE"),("31-60 dias",LARANJA,"FFFFFF","FBEAD9"),("61-90 dias",VERM,"FFFFFF","FADEDD"),("90+ dias","5A0E0E","FFFFFF","E7C9C9")]:
        sub=al[al["fl"]==lbl] if not al.empty else al
        qtd=len(sub); tot=float(sub["valor"].sum()) if not sub.empty else 0.0
        for rr in (5,6,7): wa.merge_cells(start_row=rr,start_column=col,end_row=rr,end_column=col+1)
        t=wa.cell(5,col,lbl); t.fill=P(strip); t.font=F(11,True,fg); t.alignment=C
        q=wa.cell(6,col,f"{qtd} recebimentos"); q.fill=P(tint); q.font=F(10,False,NAVY); q.alignment=C
        v=wa.cell(7,col,brl(tot)); v.fill=P(tint); v.font=F(12,True,strip if strip!=AMAR else "9A6B00"); v.alignment=C
        col+=2
    for rr in (5,6,7): wa.row_dimensions[rr].height=22
    heads=["Faixa","Nº Capa 70","Data","Banco","Identificação (CPF/CNPJ / origem)","Valor","Dias"]; wsw=[13,12,12,11,24,14,8]; hr=9
    for j,(h,w) in enumerate(zip(heads,wsw),1):
        cc=wa.cell(hr,j,h); cc.font=F(9,True,"FFFFFF"); cc.fill=P(NAVY); cc.alignment=C; wa.column_dimensions[GL(j)].width=w
    rr=hr+1
    for _,row in al.iterrows():
        cf=wa.cell(rr,1,row["fl"]); cf.fill=P(row["strip"]); cf.font=F(9,True,row["stfg"]); cf.alignment=C
        wa.cell(rr,2,row["num_txt"]).alignment=C
        cd=wa.cell(rr,3, pd.to_datetime(row["data"]).to_pydatetime() if pd.notna(row["data"]) else None); cd.number_format="DD/MM/YYYY"; cd.alignment=C
        wa.cell(rr,4,row["banco"]).alignment=C; wa.cell(rr,5,row["identificacao"]).alignment=C
        cv=wa.cell(rr,6,float(row["valor"])); cv.number_format="#,##0.00;[RED]-#,##0.00"; cv.alignment=Rg
        wa.cell(rr,7,int(row["dias"]) if pd.notna(row["dias"]) else None).alignment=C; rr+=1
    wa.freeze_panes=f"A{hr+1}"; wa.auto_filter.ref=f"A{hr}:G{max(rr-1,hr)}"

    # ---- RESUMO ----
    wr=wb.create_sheet("Resumo"); titlebar(wr,8,"RESUMO DA RODADA — CONTA 70")
    wr.merge_cells("A3:H3")
    _ate = (pd.to_datetime(resumo["capa_ate"]).strftime("%d/%m/%Y") if resumo.get("capa_ate") else "—")
    c=wr.cell(3,1,f"{resumo['n']:,} lançamentos  ·  Capa atualizada até {_ate}  ·  prévia gerada em {_date.today().strftime('%d/%m/%Y')}".replace(",",".")); c.font=F(9,False,"777777"); c.alignment=L
    wr.merge_cells("A5:F5"); c=wr.cell(5,1,"  SALDO DE DESPESAS EM ABERTO NA CONTA 70"); c.fill=P(NAVY); c.font=F(10,True,"FFFFFF"); c.alignment=L
    wr.merge_cells("A6:C7"); c=wr.cell(6,1,brl(resumo["saldo_desp_aberto"])); c.font=F(24,True,VERM); c.alignment=C
    wr.merge_cells("D6:F6"); c=wr.cell(6,4,f"{resumo['qtd_desp_aberto']} despesas sem número na Capa"); c.font=F(10,False,"555555"); c.alignment=L
    wr.merge_cells("D7:F7"); c=wr.cell(7,4,"Igual ao total EM ABERTO — não há receitas sem número na Capa"); c.font=F(9,False,"888888"); c.alignment=L
    for rr in (5,6,7):
        for cc2 in range(1,7):
            e=wr.cell(rr,cc2); e.border=Border(left=Side(style="thin",color="DDDDDD"),right=Side(style="thin",color="DDDDDD"),top=Side(style="thin",color="DDDDDD"),bottom=Side(style="thin",color="DDDDDD"))
        wr.row_dimensions[rr].height=22
    for rr in (5,6,7): wr.cell(rr,1).border=Border(left=Side(style="medium",color=AMAR),top=wr.cell(rr,1).border.top,bottom=wr.cell(rr,1).border.bottom,right=wr.cell(rr,1).border.right)
    wr.merge_cells("A8:H8"); c=wr.cell(8,1,f"Atreladas na Capa: {resumo['n_com_num']:,} de {resumo['n']:,}  ·  NF sugerida por valor exato: {resumo['exato']}  ·  soma p/ conferir: {resumo['soma_conferir']}".replace(",",".")); c.font=F(9,False,"444444"); c.alignment=L
    wr.cell(10,1,"SITUAÇÃO POR STATUS").font=F(11,True,NAVY)
    for j,h in enumerate(["Status","Qtd","Valor (R$)"],1):
        cc=wr.cell(11,j,h); cc.font=F(9,True,"FFFFFF"); cc.fill=P(NAVY); cc.alignment=C
    STL=[("Baixada no Sankhya","BAIXA",VERDE,"FFFFFF"),("Saída da Conta 70","SAIDA","EDEFF2",NAVY),
         ("Identificado pelo Mapa","MAPA",AZUL,"FFFFFF"),("Pendente de baixa","PENDENTE",AMAR,NAVY),
         ("Sem identificação no extrato bancário","SEM_ID",VERM,"FFFFFF")]
    rr=12
    for lbl,key,bg,fg in STL:
        a=wr.cell(rr,1,lbl); a.fill=P(bg); a.font=F(9,True,fg); a.alignment=L
        wr.cell(rr,2,f'=COUNTIF(Mapa!$Q$5:$Q${LAST},"{key}")').alignment=C
        v=wr.cell(rr,3,f'=SUMIF(Mapa!$Q$5:$Q${LAST},"{key}",Mapa!$F$5:$F${LAST})'); v.number_format="#,##0.00"; v.alignment=Rg
        rr+=1
    a=wr.cell(rr,1,"EM ABERTO (não atrelado)"); a.font=F(9,True,NAVY); a.fill=P("EAF1F9")
    b=wr.cell(rr,2,"=B14+B15+B16"); b.font=F(9,True,NAVY); b.fill=P("EAF1F9"); b.alignment=C
    cc=wr.cell(rr,3,"=C14+C15+C16"); cc.number_format="#,##0.00"; cc.font=F(9,True,NAVY); cc.fill=P("EAF1F9"); cc.alignment=Rg
    wr.merge_cells(start_row=rr+1,start_column=1,end_row=rr+2,end_column=3)
    c=wr.cell(rr+1,1,"Baixada e Saída são lançamentos já atrelados (baixados) e praticamente se anulam. O foco é o EM ABERTO — que fecha com POR BANCO."); c.font=F(8,False,"888888"); c.alignment=Lw
    rr+=2
    wr.column_dimensions["A"].width=42; wr.column_dimensions["B"].width=12; wr.column_dimensions["C"].width=18
    ab_all=m[m["aberto"] & (m["iddig"]!="")].copy()
    if not ab_all.empty:
        top=(ab_all.groupby("idfmt").agg(qtd=("valor","size"),total=("valor","sum"),nome=("parceiro_efetivo","first")).reset_index().sort_values("total").head(5))
    else:
        top=pd.DataFrame(columns=["idfmt","qtd","total","nome"])
    base=rr+3
    wr.cell(base-1,1,"TOP 5 PARCEIROS COM VALOR TRAVADO (em aberto)").font=F(11,True,NAVY)
    for j,h in enumerate(["CPF/CNPJ","Parceiro (se identificado)","Qtd","Valor travado"],1):
        cc=wr.cell(base,j,h); cc.font=F(9,True,"FFFFFF"); cc.fill=P(NAVY); cc.alignment=C
    wr.column_dimensions["B"].width=30
    for i,(_,row) in enumerate(top.iterrows()):
        wr.cell(base+1+i,1,row["idfmt"]).alignment=L
        nm=row["nome"] if (isinstance(row["nome"],str) and "/" not in row["nome"] and "." not in str(row["nome"])[:4]) else ""
        wr.cell(base+1+i,2,nm).alignment=L; wr.cell(base+1+i,3,int(row["qtd"])).alignment=C
        cv=wr.cell(base+1+i,4,float(row["total"])); cv.number_format="#,##0.00;[RED]-#,##0.00"; cv.alignment=Rg
    bcol=6
    _abx=m[m["aberto"]].copy(); _abx["bk"]=_abx["banco"].replace("","(sem banco identificado)")
    dist=(_abx.groupby("bk").agg(qtd=("valor","size"),total=("valor","sum")).reset_index().sort_values("total")) if not _abx.empty else pd.DataFrame(columns=["bk","qtd","total"])
    wr.cell(base-1,bcol,"POR BANCO (em aberto)").font=F(11,True,AZUL)
    for j,h in enumerate(["Banco","Qtd","Valor"],bcol):
        cc=wr.cell(base,j,h); cc.font=F(9,True,"FFFFFF"); cc.fill=P(AZUL); cc.alignment=C
    wr.column_dimensions[GL(bcol)].width=24; wr.column_dimensions[GL(bcol+2)].width=16
    _ri=0
    for _,row in dist.iterrows():
        wr.cell(base+1+_ri,bcol,row["bk"]).alignment=L; wr.cell(base+1+_ri,bcol+1,int(row["qtd"])).alignment=C
        cv=wr.cell(base+1+_ri,bcol+2,float(row["total"])); cv.number_format="#,##0.00;[RED]-#,##0.00"; cv.alignment=Rg
        _ri+=1
    tr=base+1+_ri
    a=wr.cell(tr,bcol,"TOTAL (em aberto)"); a.font=F(9,True,NAVY)
    b=wr.cell(tr,bcol+1,int(dist["qtd"].sum()) if not dist.empty else 0); b.font=F(9,True,NAVY); b.alignment=C
    cv=wr.cell(tr,bcol+2,float(dist["total"].sum()) if not dist.empty else 0.0); cv.number_format="#,##0.00;[RED]-#,##0.00"; cv.font=F(9,True,NAVY); cv.alignment=Rg

    # ---- LEGENDA ----
    wg=wb.create_sheet("Legenda"); wg.column_dimensions["A"].width=3; wg.column_dimensions["B"].width=42; wg.column_dimensions["C"].width=98
    titlebar(wg,3,"LEGENDA — MAPA DE RECEBIMENTOS")
    row=[4]
    def sec(t):
        wg.merge_cells(start_row=row[0],start_column=1,end_row=row[0],end_column=3); c=wg.cell(row[0],1,"  "+t); c.font=F(11,True,"FFFFFF"); c.fill=P(NAVY); c.alignment=L; row[0]+=1
    def it(k,v,bg=None,fg="1A1A1A"):
        a=wg.cell(row[0],2,k); a.font=F(10,True,fg); a.alignment=L
        if bg: a.fill=P(bg)
        b=wg.cell(row[0],3,v); b.font=F(10); b.alignment=Lw; wg.row_dimensions[row[0]].height=32; row[0]+=1
    sec("REGRA DA SUGESTÃO DE NF (bloco verde)")
    it("Exato","O valor do recebimento bate ao centavo com uma nota em aberto → sugere a NF. Status Identificado pelo Mapa.","E7F4EC")
    it("Soma — conferir","O valor bate com a SOMA de 2+ notas do mesmo CNPJ → mostra as notas para você DECIDIR. Não confirma sozinho.","FEF7DE")
    it("Valor difere","O CNPJ tem nota em aberto, mas nenhum valor bate → NÃO sugere NF (fica Pendente; a justificativa avisa para conferir).","FADEDD")
    row[0]+=1
    sec("STATUS")
    it("Baixada no Sankhya","Já atrelada na Capa (tem Nº Capa 70, inclusive alfanumérico como 686a).",VERDE,"FFFFFF")
    it("Identificado pelo Mapa","Em aberto; valor bateu EXATO com uma NF em aberto.",AZUL,"FFFFFF")
    it("Pendente de baixa","Em aberto; identificado, mas sem NF de valor exato (ou soma a conferir).",AMAR,NAVY)
    it("Sem identificação no extrato bancário","Em aberto e sem CPF/CNPJ E sem nome/origem no histórico.",VERM,"FFFFFF")
    it("Saída da Conta 70","Débito de fechamento, já atrelado.","EDEFF2",NAVY)

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
