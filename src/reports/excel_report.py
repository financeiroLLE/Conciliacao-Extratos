"""Geração dos relatórios de saída em Excel e CSV.

Excel: workbook único com 13+ abas (Resumo, Por Banco, Conciliadas, Pendentes,
Divergências, Não Pertence, Boletos, Pix, Tarifas, Pagamentos, Recebimentos,
Sugestões, Auditoria, Pendências Consolidadas).

CSV: cada aba como um .csv separado, todos empacotados em um zip.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..pipeline import ResultadoConciliacao


# Cores do manual de marca Grupo LLE
COR_AZUL_ESCURO = "041747"
COR_AMARELO = "FAC318"
COR_AZUL = "0071FE"
COR_VERDE = "0F8C3B"
COR_BRANCO = "FFFFFF"
COR_CINZA_CLARO = "F4F6FA"

FONTE_HEADER = Font(name="Arial", bold=True, color=COR_BRANCO, size=11)
FONTE_BODY = Font(name="Arial", size=10)
FILL_HEADER = PatternFill("solid", fgColor=COR_AZUL_ESCURO)
FILL_DESTAQUE = PatternFill("solid", fgColor=COR_AMARELO)
BORDA_FINA = Side(border_style="thin", color="D0D0D0")
BORDA = Border(left=BORDA_FINA, right=BORDA_FINA, top=BORDA_FINA, bottom=BORDA_FINA)

FORMATO_BRL = 'R$ #,##0.00;[Red]-R$ #,##0.00;"-"'
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_PCT = '0.0%'


def _aplicar_estilo_header(ws, n_colunas: int, linha: int = 1):
    for col in range(1, n_colunas + 1):
        cell = ws.cell(row=linha, column=col)
        cell.font = FONTE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDA
    ws.row_dimensions[linha].height = 28


def _ajustar_larguras(ws, df: pd.DataFrame, max_width: int = 40):
    for i, col in enumerate(df.columns, start=1):
        # tamanho baseado no header e nos primeiros valores
        valores = [str(col)] + [str(v) for v in df[col].head(50).tolist()]
        largura = min(max(len(v) for v in valores) + 2, max_width)
        ws.column_dimensions[get_column_letter(i)].width = max(largura, 12)


def _escrever_dataframe(ws, df: pd.DataFrame, inicio_linha: int = 1):
    """Escreve um DataFrame em uma worksheet, com estilo de header e formato BRL/data."""
    if df.empty:
        ws.cell(row=inicio_linha, column=1, value="(nenhum registro)").font = Font(
            italic=True, color="808080"
        )
        return inicio_linha + 1

    # Header
    for i, col in enumerate(df.columns, start=1):
        ws.cell(row=inicio_linha, column=i, value=str(col))
    _aplicar_estilo_header(ws, len(df.columns), linha=inicio_linha)

    # Body
    COLS_MOEDA_SUFIXOS = ("valor", "diferenca", "total banco", "total sistema",
                          "total conciliado", "falta conciliar", "falta lancar",
                          "conciliado c/ divergência", "conciliado com divergência")
    COLS_PCT_SUFIXOS = ("percentual conciliado", "percentual_conciliado")

    def _is_col_moeda(nome: str) -> bool:
        n = nome.lower()
        # Bate quando "valor" aparece como token, não dentro de outras palavras
        return any(n == s or n.startswith(s + " ") or n.endswith(" " + s) or s in n.split(" ")
                   for s in COLS_MOEDA_SUFIXOS)

    def _is_col_pct(nome: str) -> bool:
        return nome.lower() in COLS_PCT_SUFIXOS

    for row_idx, (_, row) in enumerate(df.iterrows(), start=inicio_linha + 1):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, pd.Timestamp) or isinstance(val, datetime):
                cell.value = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
                cell.number_format = FORMATO_DATA
            elif _is_col_moeda(str(col_name)) and isinstance(val, (int, float)):
                cell.value = float(val)
                cell.number_format = FORMATO_BRL
            elif _is_col_pct(str(col_name)) and isinstance(val, (int, float)):
                cell.value = float(val) / 100.0
                cell.number_format = FORMATO_PCT
            else:
                cell.value = val
            cell.font = FONTE_BODY
            cell.border = BORDA
            # zebra
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COR_CINZA_CLARO)

    _ajustar_larguras(ws, df)
    ws.freeze_panes = ws.cell(row=inicio_linha + 1, column=1)
    return row_idx + 1


def _aba_resumo_geral(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Resumo Geral")
    kpis = resultado.kpis_globais()

    ws.cell(row=1, column=1, value="CONCILIAÇÃO BANCÁRIA — RESUMO GERAL").font = Font(
        name="Arial", bold=True, size=16, color=COR_AZUL_ESCURO
    )
    ws.cell(row=2, column=1, value=f"Data de referência: "
            f"{resultado.data_referencia.strftime('%d/%m/%Y')}").font = FONTE_BODY
    ws.cell(row=3, column=1, value=f"Tolerância de data: ±{resultado.tolerancia_dias} dia(s)").font = FONTE_BODY
    ws.cell(row=4, column=1, value=f"Contas processadas: {', '.join(resultado.contas_processadas)}").font = FONTE_BODY

    linhas = [
        ("INDICADOR", "VALOR", "header"),
        ("Total Extrato Bancário", kpis["total_extrato_bancario"], "moeda"),
        ("Total Extrato Sistema/Sankhya", kpis["total_extrato_sistema"], "moeda"),
        ("Total Conciliado", kpis["total_conciliado"], "moeda"),
        ("Falta Conciliar (no Sistema)", kpis["falta_conciliar"], "moeda"),
        ("Falta Lançar (no Banco)", kpis["falta_lancar"], "moeda"),
        ("Valor Conciliado com Divergência", kpis["valor_divergencia"], "moeda"),
        ("Percentual Conciliado", kpis["percentual_conciliado"] / 100.0, "pct"),
        ("", "", "vazio"),
        ("Qtde Registros Banco", kpis["qtd_registros_banco"], "int"),
        ("Qtde Registros Sistema", kpis["qtd_registros_sistema"], "int"),
        ("Qtde Conciliados", kpis["qtd_conciliados"], "int"),
        ("Qtde Pendentes Banco", kpis["qtd_pendentes_banco"], "int"),
        ("Qtde Pendentes Sistema", kpis["qtd_pendentes_sistema"], "int"),
        ("Qtde Divergências", kpis["qtd_divergencias"], "int"),
    ]
    linha_atual = 6
    for label, valor, kind in linhas:
        c1 = ws.cell(row=linha_atual, column=1, value=label)
        c2 = ws.cell(row=linha_atual, column=2, value=valor)
        if kind == "header":
            c1.font = FONTE_HEADER
            c1.fill = FILL_HEADER
            c2.font = FONTE_HEADER
            c2.fill = FILL_HEADER
        elif kind == "moeda":
            c1.font = Font(name="Arial", bold=True, size=11, color=COR_AZUL_ESCURO)
            c2.number_format = FORMATO_BRL
            c2.font = FONTE_BODY
        elif kind == "pct":
            c1.font = Font(name="Arial", bold=True, size=11, color=COR_AZUL_ESCURO)
            c2.number_format = FORMATO_PCT
            c2.font = FONTE_BODY
            c2.fill = FILL_DESTAQUE
        elif kind == "int":
            c1.font = FONTE_BODY
            c2.number_format = "#,##0"
            c2.font = FONTE_BODY
        linha_atual += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22


def _aba_resumo_por_banco(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Resumo por Banco")
    kpis_pb = resultado.kpis_por_banco()
    if not kpis_pb:
        ws.cell(row=1, column=1, value="(nenhuma conta processada)").font = Font(italic=True)
        return

    rows = []
    for conta, k in kpis_pb.items():
        rows.append({
            "Conta": conta,
            "Total Banco": k["total_extrato_bancario"],
            "Total Sistema": k["total_extrato_sistema"],
            "Total Conciliado": k["total_conciliado"],
            "Falta Conciliar": k["falta_conciliar"],
            "Falta Lançar": k["falta_lancar"],
            "Conciliado c/ Divergência": k["valor_divergencia"],
            "Percentual Conciliado": k["percentual_conciliado"],
            "Qtde Banco": k["qtd_registros_banco"],
            "Qtde Sistema": k["qtd_registros_sistema"],
            "Qtde Conciliados": k["qtd_conciliados"],
            "Qtde Pendentes Banco": k["qtd_pendentes_banco"],
            "Qtde Pendentes Sistema": k["qtd_pendentes_sistema"],
        })
    df = pd.DataFrame(rows)
    _escrever_dataframe(ws, df)


def _aba_conciliadas(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Conciliadas")
    if resultado.conciliados.empty:
        _escrever_dataframe(ws, pd.DataFrame())
        return
    cols = [
        "banco_data", "banco_conta", "banco_historico", "banco_documento", "banco_valor",
        "sistema_data", "sistema_historico", "sistema_documento", "sistema_valor",
        "dias_diferenca", "status", "motivo",
    ]
    cols_existentes = [c for c in cols if c in resultado.conciliados.columns]
    df = resultado.conciliados[cols_existentes].copy()
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_pendentes(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Pendentes de Conciliação")
    pb = resultado.pendentes_banco.copy()
    ps = resultado.pendentes_sistema.copy()
    if not pb.empty:
        pb["origem"] = "Banco (falta lançar no Sistema)"
    if not ps.empty:
        ps["origem"] = "Sistema (falta no Banco)"
    df = pd.concat([pb, ps], ignore_index=True)
    if df.empty:
        _escrever_dataframe(ws, df)
        return
    cols = ["origem", "data", "conta", "historico", "documento", "valor", "tipo", "natureza"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    df.columns = [c.title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_divergencias(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Conciliadas com Divergência")
    if resultado.divergencias.empty:
        _escrever_dataframe(ws, pd.DataFrame())
        return
    df = resultado.divergencias.copy()
    cols = [
        "data", "conta", "historico_banco", "valor_banco", "historico_sistema",
        "valor_sistema", "diferenca", "status", "motivo",
    ]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_nao_pertence(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Não Pertence à Conta")
    df = resultado.nao_pertence.copy()
    if df.empty:
        _escrever_dataframe(ws, df)
        return
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_por_tipo(wb: Workbook, resultado: "ResultadoConciliacao", tipo: str, nome_aba: str):
    """Cria uma aba com os lançamentos de um tipo específico (Boleto, Pix, etc)."""
    ws = wb.create_sheet(nome_aba)
    frames = []
    if not resultado.banco_completo.empty:
        df_b = resultado.banco_completo[resultado.banco_completo["tipo"] == tipo].copy()
        if not df_b.empty:
            df_b["origem"] = "Banco"
            frames.append(df_b)
    if not resultado.sistema_completo.empty:
        df_s = resultado.sistema_completo[resultado.sistema_completo["tipo"] == tipo].copy()
        if not df_s.empty:
            df_s["origem"] = "Sistema"
            frames.append(df_s)
    if not frames:
        _escrever_dataframe(ws, pd.DataFrame())
        return
    df = pd.concat(frames, ignore_index=True)
    cols = ["origem", "data", "conta", "historico", "documento", "valor", "natureza"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    df.columns = [c.title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_natureza(wb: Workbook, resultado: "ResultadoConciliacao", natureza: str, nome_aba: str):
    ws = wb.create_sheet(nome_aba)
    frames = []
    if not resultado.banco_completo.empty:
        df_b = resultado.banco_completo[resultado.banco_completo["natureza"] == natureza].copy()
        if not df_b.empty:
            df_b["origem"] = "Banco"
            frames.append(df_b)
    if not resultado.sistema_completo.empty:
        df_s = resultado.sistema_completo[resultado.sistema_completo["natureza"] == natureza].copy()
        if not df_s.empty:
            df_s["origem"] = "Sistema"
            frames.append(df_s)
    if not frames:
        _escrever_dataframe(ws, pd.DataFrame())
        return
    df = pd.concat(frames, ignore_index=True)
    cols = ["origem", "data", "conta", "historico", "documento", "valor", "tipo"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    df.columns = [c.title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_duplicidades(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Duplicidades")
    df = resultado.duplicidades.copy()
    if df.empty:
        _escrever_dataframe(ws, df)
        return
    df.columns = [c.title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_excesso_sankhya(wb: Workbook, resultado: "ResultadoConciliacao"):
    """v3: lançamentos excedentes no Sankhya em relação ao banco."""
    ws = wb.create_sheet("Excesso no Sankhya")
    df = getattr(resultado, "excesso_sankhya", pd.DataFrame())
    if df.empty:
        _escrever_dataframe(ws, df)
        return
    df = df.copy()
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_sugestoes(wb: Workbook, resultado: "ResultadoConciliacao"):
    ws = wb.create_sheet("Sugestões Fuzzy")
    df = resultado.sugestoes_fuzzy.copy()
    if df.empty:
        _escrever_dataframe(ws, df)
        return
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_pendencias_consolidadas(wb: Workbook, resultado: "ResultadoConciliacao",
                                   pendencias_anteriores: pd.DataFrame):
    """Aba que serve como INPUT do próximo dia."""
    ws = wb.create_sheet("Pendências Consolidadas")
    # Pendências atuais + anteriores
    frames = []
    if not resultado.pendentes_banco.empty:
        df_b = resultado.pendentes_banco.copy()
        df_b["origem"] = "Banco"
        df_b["dias_pendente"] = (resultado.data_referencia - df_b["data"]).dt.days
        frames.append(df_b)
    if not resultado.pendentes_sistema.empty:
        df_s = resultado.pendentes_sistema.copy()
        df_s["origem"] = "Sistema"
        df_s["dias_pendente"] = (resultado.data_referencia - df_s["data"]).dt.days
        frames.append(df_s)
    if pendencias_anteriores is not None and not pendencias_anteriores.empty:
        df_a = pendencias_anteriores.copy()
        if "origem" not in df_a.columns:
            df_a["origem"] = "Anterior"
        frames.append(df_a)
    if not frames:
        _escrever_dataframe(ws, pd.DataFrame())
        return
    df = pd.concat(frames, ignore_index=True)
    cols_preferidas = [
        "origem", "data", "conta", "historico", "documento", "valor",
        "tipo", "natureza", "dias_pendente",
    ]
    cols_existentes = [c for c in cols_preferidas if c in df.columns]
    df = df[cols_existentes]
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    _escrever_dataframe(ws, df)


def _aba_auditoria(wb: Workbook, resultado: "ResultadoConciliacao", execucao: dict | None):
    ws = wb.create_sheet("Auditoria")
    rows = [
        {"Campo": "Data de execução", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
        {"Campo": "Data de referência", "Valor": resultado.data_referencia.strftime("%d/%m/%Y")},
        {"Campo": "Contas processadas", "Valor": ", ".join(resultado.contas_processadas)},
        {"Campo": "Qtde contas", "Valor": len(resultado.contas_processadas)},
        {"Campo": "Tolerância de data (dias)", "Valor": resultado.tolerancia_dias},
        {"Campo": "Linhas banco", "Valor": len(resultado.banco_completo)},
        {"Campo": "Linhas sistema", "Valor": len(resultado.sistema_completo)},
        {"Campo": "Conciliados", "Valor": len(resultado.conciliados)},
        {"Campo": "Pendentes banco", "Valor": len(resultado.pendentes_banco)},
        {"Campo": "Pendentes sistema", "Valor": len(resultado.pendentes_sistema)},
        {"Campo": "Divergências", "Valor": len(resultado.divergencias)},
        {"Campo": "Duplicidades (grupos)", "Valor": len(resultado.duplicidades)},
        {"Campo": "Não pertence à conta", "Valor": len(resultado.nao_pertence)},
    ]
    if execucao:
        rows.append({"Campo": "ID Execução", "Valor": execucao.get("id", "")})
        rows.append({"Campo": "Versão", "Valor": execucao.get("versao", "")})
        rows.append({"Campo": "Status", "Valor": execucao.get("status", "processado")})
    df = pd.DataFrame(rows)
    _escrever_dataframe(ws, df)


def gerar_relatorio_excel(
    resultado: "ResultadoConciliacao",
    pendencias_anteriores: pd.DataFrame | None = None,
    execucao: dict | None = None,
) -> bytes:
    """Gera o relatório Excel multi-aba e retorna os bytes do arquivo."""
    wb = Workbook()
    # Remove a aba default
    wb.remove(wb.active)

    _aba_resumo_geral(wb, resultado)
    _aba_resumo_por_banco(wb, resultado)
    _aba_conciliadas(wb, resultado)
    _aba_pendentes(wb, resultado)
    _aba_divergencias(wb, resultado)
    _aba_nao_pertence(wb, resultado)
    _aba_por_tipo(wb, resultado, "Boleto", "Boletos")
    _aba_por_tipo(wb, resultado, "Pix", "Pix")
    _aba_por_tipo(wb, resultado, "Tarifa", "Tarifas")
    _aba_natureza(wb, resultado, "Pagamento", "Pagamentos")
    _aba_natureza(wb, resultado, "Recebimento", "Recebimentos")
    _aba_duplicidades(wb, resultado)
    _aba_excesso_sankhya(wb, resultado)
    _aba_sugestoes(wb, resultado)
    if pendencias_anteriores is None:
        pendencias_anteriores = pd.DataFrame()
    _aba_pendencias_consolidadas(wb, resultado, pendencias_anteriores)
    _aba_auditoria(wb, resultado, execucao)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def gerar_relatorio_excel_de_conta(
    resultado: "ResultadoConciliacao",
    conta: str,
) -> bytes:
    """Versão filtrada por uma única conta (mesmas abas, dados só dessa conta)."""
    # Cria um sub-resultado filtrado
    from copy import copy
    sub = copy(resultado)
    sub.banco_completo = (
        resultado.banco_completo[resultado.banco_completo["conta"] == conta].copy()
        if not resultado.banco_completo.empty else resultado.banco_completo
    )
    sub.sistema_completo = (
        resultado.sistema_completo[resultado.sistema_completo["conta"] == conta].copy()
        if not resultado.sistema_completo.empty else resultado.sistema_completo
    )
    sub.conciliados = resultado.conciliados_da_conta(conta)
    sub.pendentes_banco = (
        resultado.pendentes_banco[resultado.pendentes_banco["conta"] == conta].copy()
        if not resultado.pendentes_banco.empty else resultado.pendentes_banco
    )
    sub.pendentes_sistema = (
        resultado.pendentes_sistema[resultado.pendentes_sistema["conta"] == conta].copy()
        if not resultado.pendentes_sistema.empty else resultado.pendentes_sistema
    )
    sub.divergencias = resultado.divergencias_da_conta(conta)
    sub.nao_pertence = resultado.nao_pertence_da_conta(conta)
    sub.contas_processadas = [conta]
    return gerar_relatorio_excel(sub)


def gerar_csvs_zip(resultado: "ResultadoConciliacao") -> bytes:
    """Gera um zip com todos os DataFrames principais em CSV separado."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        def _add(df: pd.DataFrame, nome: str):
            if df is None or df.empty:
                zf.writestr(f"{nome}.csv", "")
                return
            csv_str = df.to_csv(index=False, sep=";", encoding="utf-8-sig", decimal=",")
            zf.writestr(f"{nome}.csv", csv_str)

        _add(resultado.conciliados, "conciliadas")
        _add(resultado.pendentes_banco, "pendentes_banco")
        _add(resultado.pendentes_sistema, "pendentes_sistema")
        _add(resultado.divergencias, "divergencias")
        _add(resultado.nao_pertence, "nao_pertence_a_conta")
        _add(resultado.duplicidades, "duplicidades")
        _add(getattr(resultado, "excesso_sankhya", pd.DataFrame()), "excesso_sankhya")
        _add(getattr(resultado, "possiveis_duplicidades", pd.DataFrame()), "possiveis_duplicidades")
        _add(resultado.sugestoes_fuzzy, "sugestoes_fuzzy")
        _add(resultado.banco_completo, "banco_completo")
        _add(resultado.sistema_completo, "sistema_completo")

        # KPIs como CSV
        kpis_rows = []
        for conta, k in resultado.kpis_por_banco().items():
            row = {"conta": conta, **k}
            kpis_rows.append(row)
        if kpis_rows:
            _add(pd.DataFrame(kpis_rows), "kpis_por_banco")

    buf.seek(0)
    return buf.getvalue()
