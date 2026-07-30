# -*- coding: utf-8 -*-
"""
Exportação Excel da Conciliação de Vendas — Fase 7 (antecipada).

Gera um workbook .xlsx com 8 abas para auditoria mensal:

    1. Resumo            — KPIs da rodada, balanço, contadores por grupo
    2. Grupo 1 Auto      — auto-conciliadas pelo motor (uma linha por parcela)
    3. Grupo 1 Confirm   — confirmadas manualmente entre candidatas ambíguas
    4. Grupo 2 TOP 1722  — vendas já baixadas por cartão (auditoria)
    5. Grupo 3 Aguard    — títulos Sankhya aguardando captura
    6. Grupo 4 Devoluc   — cancelamentos e devoluções
    7. A analisar        — vendas pendentes (ambíguas + sem par)
    8. Histórico         — trilha de auditoria de ações da rodada

API pública:
    gerar_excel(...) -> bytes

Todas as abas usam formatação leve para leitura contábil:
    - cabeçalho em azul navy + amarelo (identidade LLE)
    - colunas de valor formatadas como moeda R$
    - datas em formato dd/mm/yyyy
    - larguras ajustadas para o conteúdo
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ==============================================================================
# ESTILO
# ==============================================================================

_AZUL_NAVY = "0A1730"
_AMARELO = "FFCC00"
_CREME = "FFF6C8"
_CINZA_CLARO = "F2F2F2"

_FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color=_AMARELO)
_FILL_HEADER = PatternFill("solid", fgColor=_AZUL_NAVY)
_ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)

_FONTE_TITULO = Font(name="Calibri", size=14, bold=True, color=_AZUL_NAVY)
_FONTE_SECAO = Font(name="Calibri", size=11, bold=True, color=_AZUL_NAVY)
_FILL_TITULO = PatternFill("solid", fgColor=_AMARELO)
_FILL_SECAO = PatternFill("solid", fgColor=_CREME)

_FMT_MOEDA = 'R$ #,##0.00;-R$ #,##0.00'
_FMT_DATA = "dd/mm/yyyy"


# ==============================================================================
# HELPERS
# ==============================================================================

def _to_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    try:
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        return pd.to_datetime(v).date()
    except Exception:
        return None


def _fmt_moeda_str(v: Any) -> str:
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


def _num(v: Any) -> float:
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _str(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v)


def _label_adquirente(a: Any) -> str:
    s = _str(a).strip().lower()
    return {"getnet": "Getnet", "cielo": "Cielo", "pagseguro": "PagSeguro"}.get(s, s.capitalize() if s else "—")


def _label_bandeira(b: Any) -> str:
    s = _str(b).strip().lower()
    if not s:
        return "—"
    return {
        "visa": "Visa", "master": "Master", "elo": "Elo",
        "vis_mas": "Vis/Mas", "mas_elo": "Mas/Elo",
        "hipercard": "Hipercard", "amex": "Amex",
    }.get(s, s.upper())


def _label_modalidade(m: Any, parcelas: Any = None) -> str:
    s = _str(m).strip().lower()
    if not s:
        return "—"
    if s == "debito":
        return "Débito"
    if s == "credito_avista":
        return "Crédito à vista"
    if s == "credito_parcelado":
        try:
            n = int(parcelas) if parcelas is not None else None
            if n and n > 1:
                return f"Crédito parc {n}×"
        except (ValueError, TypeError):
            pass
        return "Crédito parcelado"
    return s.capitalize()


def _label_classe_sk(c: Any) -> str:
    s = _str(c).strip().lower()
    return {"adiantamento": "Adiantamento", "nota_fiscal": "Nota Fiscal", "outro": "Outro"}.get(s, "—")


def _aplicar_header(ws, colunas: List[str], linha: int = 1):
    """Aplica estilo de cabeçalho em uma linha específica."""
    for col_idx, col_name in enumerate(colunas, start=1):
        cell = ws.cell(row=linha, column=col_idx, value=col_name)
        cell.font = _FONTE_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_HEADER
    ws.row_dimensions[linha].height = 24


def _ajustar_larguras(ws, colunas: List[str], larguras: Optional[Dict[str, int]] = None):
    """Ajusta larguras baseado no nome da coluna (heurística) ou dict explícito."""
    larguras = larguras or {}
    for col_idx, col_name in enumerate(colunas, start=1):
        letra = get_column_letter(col_idx)
        if col_name in larguras:
            ws.column_dimensions[letra].width = larguras[col_name]
        else:
            # heurística: valor -> 14, data -> 12, texto -> 18, ID -> 12
            nlow = col_name.lower()
            if "valor" in nlow or "líquido" in nlow or "bruto" in nlow or "taxa" in nlow:
                ws.column_dimensions[letra].width = 15
            elif "data" in nlow or "venc" in nlow:
                ws.column_dimensions[letra].width = 12
            elif "parceiro" in nlow or "empresa" in nlow or "histórico" in nlow or "descrição" in nlow:
                ws.column_dimensions[letra].width = 32
            elif "nf" in nlow or "nro" in nlow or "id" in nlow or "nsu" in nlow:
                ws.column_dimensions[letra].width = 14
            else:
                ws.column_dimensions[letra].width = 18


def _preencher_linhas(ws, linhas: List[List[Any]], inicio: int = 2, formatos_col: Optional[Dict[int, str]] = None):
    """
    Preenche linhas na worksheet. `formatos_col` mapeia índice-1-based -> number_format.
    """
    formatos_col = formatos_col or {}
    for i, linha in enumerate(linhas, start=inicio):
        for j, val in enumerate(linha, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j in formatos_col:
                cell.number_format = formatos_col[j]


def _chave_venda_original(row: pd.Series) -> Tuple[str, str, str]:
    """Mesma lógica do agrupamento na página — replicada aqui pra ficar autônomo."""
    adq = str(row.get("adquirente") or "")
    nsu = str(row.get("nsu") or "").strip()
    auth = str(row.get("autorizacao") or "").strip()

    if nsu and auth:
        return (adq, nsu, auth)
    if nsu:
        return (adq, nsu, "")
    if auth:
        return (adq, "", auth)

    parc = str(row.get("sk_nome_parceiro") or row.get("nome_parceiro") or "")
    data = row.get("data_prev_pagamento")
    try:
        data_str = pd.to_datetime(data).strftime("%Y-%m-%d") if data is not None else ""
    except Exception:
        data_str = ""
    return (adq, parc, data_str)


# ==============================================================================
# ABA 1 · RESUMO
# ==============================================================================

def _aba_resumo(
    wb: Workbook,
    resultado,
    contadores: Dict[str, int],
    ligacoes_desfeitas: Set,
    confirmadas_manual: Dict,
    historico: List,
    df_cielo,
    df_getnet,
    df_sankhya,
    tolerancia_dias: int,
):
    ws = wb.create_sheet("Resumo")

    ws["A1"] = "Conciliação de Vendas — Resumo da Rodada"
    ws["A1"].font = _FONTE_TITULO
    ws["A1"].fill = _FILL_TITULO
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    linha = 4

    # -------- Seção Balanço --------
    ws.cell(row=linha, column=1, value="BALANÇO FINANCEIRO").font = _FONTE_SECAO
    ws.cell(row=linha, column=1).fill = _FILL_SECAO
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    linha += 1

    tot_cielo = float(df_cielo["valor_bruto"].sum()) if (df_cielo is not None and not df_cielo.empty) else 0.0
    n_cielo = len(df_cielo) if (df_cielo is not None and not df_cielo.empty) else 0

    col_getnet = "valor_parcela_bruto" if (df_getnet is not None and "valor_parcela_bruto" in df_getnet.columns) else "valor_bruto"
    tot_getnet = float(df_getnet[col_getnet].sum()) if (df_getnet is not None and not df_getnet.empty and col_getnet in df_getnet.columns) else 0.0
    n_getnet = len(df_getnet) if (df_getnet is not None and not df_getnet.empty) else 0

    tot_adq = tot_cielo + tot_getnet

    tot_sk = 0.0
    n_sk_elegivel = 0
    if df_sankhya is not None and not df_sankhya.empty:
        try:
            from src.motor_vendas import classificador_sankhya
            df_c = classificador_sankhya.classificar(df_sankhya)
            df_el = classificador_sankhya.filtrar_elegiveis_para_match(df_c)
            if df_el is not None and not df_el.empty:
                tot_sk = float(df_el["vlr_desdobramento"].sum())
                n_sk_elegivel = len(df_el)
        except Exception:
            pass

    linhas_balanco = [
        ["Lado", "Valor (R$)", "Nº itens", "Observação"],
        ["Adquirente (bruto)", tot_adq, n_cielo + n_getnet, f"Cielo {n_cielo} · Getnet {n_getnet}"],
        ["Sankhya elegível", tot_sk, n_sk_elegivel, "Nota fiscal + Adiantamento (em aberto ou TOP 1722)"],
        ["Diferença", tot_adq - tot_sk, "", "Adq − Sankhya"],
    ]

    _aplicar_header(ws, linhas_balanco[0], linha=linha)
    _preencher_linhas(ws, linhas_balanco[1:], inicio=linha + 1, formatos_col={2: _FMT_MOEDA})
    linha += len(linhas_balanco) + 1

    # -------- Seção Contadores por Grupo --------
    ws.cell(row=linha, column=1, value="CONTADORES POR GRUPO").font = _FONTE_SECAO
    ws.cell(row=linha, column=1).fill = _FILL_SECAO
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    linha += 1

    linhas_grupos = [
        ["Grupo", "Qtd", "Descrição", ""],
        ["A analisar", contadores.get("a_analisar", 0), "Ambíguos + venda sem título + desfeitas", ""],
        ["Auto-conciliadas", contadores.get("auto_conciliadas", 0), "Grupo 1 · matches únicos do motor", ""],
        ["Compensadas (TOP 1722)", contadores.get("compensadas", 0), "Grupo 2 · já baixadas por cartão", ""],
        ["Aguardando captura", contadores.get("aguardando", 0), "Grupo 3 · títulos em aberto sem venda casando", ""],
        ["Devoluções", contadores.get("devolucoes", 0), "Grupo 4 · cancelamentos", ""],
        ["", "", "", ""],
        ["Confirmadas manualmente", len(confirmadas_manual), "Débora escolheu entre candidatas ambíguas", ""],
        ["Ligações desfeitas", len(ligacoes_desfeitas), "Débora desfez auto-conciliações do motor", ""],
        ["Ações no histórico", len(historico), "Auditoria da rodada", ""],
    ]

    _aplicar_header(ws, linhas_grupos[0], linha=linha)
    _preencher_linhas(ws, linhas_grupos[1:], inicio=linha + 1)
    linha += len(linhas_grupos) + 1

    # -------- Seção Configuração --------
    ws.cell(row=linha, column=1, value="CONFIGURAÇÃO DA RODADA").font = _FONTE_SECAO
    ws.cell(row=linha, column=1).fill = _FILL_SECAO
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    linha += 1

    linhas_cfg = [
        ["Parâmetro", "Valor", "", ""],
        ["Tolerância de data (dias)", tolerancia_dias, "", ""],
        ["Data da rodada", date.today().strftime("%d/%m/%Y"), "", ""],
    ]
    _aplicar_header(ws, linhas_cfg[0], linha=linha)
    _preencher_linhas(ws, linhas_cfg[1:], inicio=linha + 1)
    linha += len(linhas_cfg) + 1

    # Larguras
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22


# ==============================================================================
# ABAS 2 e 4 · GRUPO 1 AUTO e GRUPO 2 TOP 1722 (mesmo layout)
# ==============================================================================

_COLS_CONCILIADAS = [
    "Adquirente", "Bandeira", "Modalidade",
    "Parcela", "Data Venda", "Nro Único (Adq)",
    "Empresa", "Parceiro", "Tipo Sankhya", "NF Sankhya",
    "Data Vencimento", "Valor (R$)", "Ref NF (adiant.)", "Situação Sankhya",
    "Origem venda (idx)", "Chave venda", "Status",
]


def _aba_conciliadas(wb: Workbook, nome_aba: str, df_grupo: pd.DataFrame, ligacoes_desfeitas: Set):
    ws = wb.create_sheet(nome_aba)

    if df_grupo is None or df_grupo.empty:
        _aplicar_header(ws, _COLS_CONCILIADAS)
        ws.cell(row=2, column=1, value="(sem linhas neste grupo)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, _COLS_CONCILIADAS)
        return

    _aplicar_header(ws, _COLS_CONCILIADAS)
    linhas = []

    for _, row in df_grupo.iterrows():
        chave = _chave_venda_original(row)
        desfeita = chave in ligacoes_desfeitas
        status = "DESFEITA MANUALMENTE" if desfeita else "OK"

        parc_atual = row.get("parcela_atual")
        parc_total = row.get("parcelas_total")
        parc_txt = f"{parc_atual}/{parc_total}" if parc_atual and parc_total and parc_total > 1 else "1/1"

        linhas.append([
            _label_adquirente(row.get("adquirente")),
            _label_bandeira(row.get("bandeira")),
            _label_modalidade(row.get("modalidade"), row.get("parcelas_total")),
            parc_txt,
            _to_date(row.get("data_prev_pagamento")),
            _str(row.get("nsu")),
            _str(row.get("sk_empresa_nome")),
            _str(row.get("sk_nome_parceiro")),
            _label_classe_sk(row.get("sk_classe")),
            _str(row.get("sk_nro_nota")),
            _to_date(row.get("sk_dt_vencimento")),
            _num(row.get("valor_match")),
            _str(row.get("sk_ref_nf")),
            _str(row.get("sk_situacao")),
            _str(row.get("origem_venda")),
            "|".join(str(x) for x in chave),
            status,
        ])

    formatos = {5: _FMT_DATA, 11: _FMT_DATA, 12: _FMT_MOEDA}
    _preencher_linhas(ws, linhas, inicio=2, formatos_col=formatos)
    _ajustar_larguras(ws, _COLS_CONCILIADAS, larguras={"Chave venda": 30, "Ref NF (adiant.)": 14})

    # Freeze cabeçalho
    ws.freeze_panes = "A2"


# ==============================================================================
# ABA 3 · GRUPO 1 · CONFIRMADAS MANUALMENTE
# ==============================================================================

def _aba_confirmadas(wb: Workbook, confirmadas: Dict[str, Any]):
    ws = wb.create_sheet("Grupo 1 Confirm")

    cols = [
        "Chave venda", "NF escolhida", "Tipo Sankhya",
        "Parceiro (Sankhya)", "Valor (R$)", "Vencimento",
    ]
    _aplicar_header(ws, cols)

    if not confirmadas:
        ws.cell(row=2, column=1, value="(nenhuma confirmação manual nesta rodada)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, cols)
        return

    linhas = []
    for chave_str, dados in confirmadas.items():
        linhas.append([
            chave_str,
            _str(dados.get("sk_nro_nota")),
            _label_classe_sk(dados.get("sk_classe")),
            _str(dados.get("sk_nome_parceiro")),
            _num(dados.get("sk_vlr_desdobramento")),
            _to_date(dados.get("sk_dt_vencimento")),
        ])

    _preencher_linhas(ws, linhas, inicio=2, formatos_col={5: _FMT_MOEDA, 6: _FMT_DATA})
    _ajustar_larguras(ws, cols, larguras={"Chave venda": 30})
    ws.freeze_panes = "A2"


# ==============================================================================
# ABA 5 · GRUPO 3 · AGUARDANDO CAPTURA
# ==============================================================================

def _aba_aguardando(wb: Workbook, df_g3: pd.DataFrame):
    ws = wb.create_sheet("Grupo 3 Aguard")

    cols = [
        "Tipo Sankhya", "NF / Referência", "Adquirente inferida",
        "Empresa", "Parceiro", "Data Vencimento", "Valor (R$)",
        "Bandeira Sankhya", "Modalidade Sankhya", "Histórico",
    ]
    _aplicar_header(ws, cols)

    if df_g3 is None or df_g3.empty:
        ws.cell(row=2, column=1, value="(nenhum título aguardando captura)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, cols)
        return

    linhas = []
    for _, row in df_g3.iterrows():
        classe = row.get("classe")
        if classe == "adiantamento":
            ref = f"REF NF {row.get('nro_nota_referenciada')}" if row.get("nro_nota_referenciada") else "sem REF"
        else:
            ref = _str(row.get("nro_nota"))

        linhas.append([
            _label_classe_sk(classe),
            ref,
            _label_adquirente(row.get("adquirente_sankhya")),
            _str(row.get("empresa_nome")),
            _str(row.get("nome_parceiro")),
            _to_date(row.get("dt_vencimento")),
            _num(row.get("vlr_desdobramento")),
            _label_bandeira(row.get("bandeira_sankhya")),
            _label_modalidade(row.get("modalidade_sankhya"), row.get("parcelas_sankhya")),
            _str(row.get("historico"))[:200],
        ])

    _preencher_linhas(ws, linhas, inicio=2, formatos_col={6: _FMT_DATA, 7: _FMT_MOEDA})
    _ajustar_larguras(ws, cols, larguras={"Histórico": 45})
    ws.freeze_panes = "A2"


# ==============================================================================
# ABA 6 · GRUPO 4 · DEVOLUÇÕES
# ==============================================================================

def _aba_devolucoes(wb: Workbook, df_g4: pd.DataFrame):
    ws = wb.create_sheet("Grupo 4 Devoluc")

    cols = [
        "Adquirente", "Bandeira", "Modalidade",
        "Data Venda", "Valor (R$)", "Nro Único", "Autorização",
    ]
    _aplicar_header(ws, cols)

    if df_g4 is None or df_g4.empty:
        ws.cell(row=2, column=1, value="(nenhuma devolução nesta rodada)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, cols)
        return

    linhas = []
    for _, row in df_g4.iterrows():
        linhas.append([
            _label_adquirente(row.get("adquirente")),
            _label_bandeira(row.get("bandeira")),
            _label_modalidade(row.get("modalidade"), row.get("parcelas_total")),
            _to_date(row.get("data_prev_pagamento")),
            _num(row.get("valor_match")),
            _str(row.get("nsu")),
            _str(row.get("autorizacao")),
        ])

    _preencher_linhas(ws, linhas, inicio=2, formatos_col={4: _FMT_DATA, 5: _FMT_MOEDA})
    _ajustar_larguras(ws, cols)
    ws.freeze_panes = "A2"


# ==============================================================================
# ABA 7 · A ANALISAR (ambíguos + venda sem título + desfeitas)
# ==============================================================================

def _aba_a_analisar(wb: Workbook, resultado, ligacoes_desfeitas: Set):
    ws = wb.create_sheet("A analisar")

    cols = [
        "Situação", "Adquirente", "Bandeira", "Modalidade",
        "Parcela", "Data Venda", "Valor (R$)",
        "Nro Único", "Autorização",
        "N candidatos (Sankhya)", "Detalhe",
    ]
    _aplicar_header(ws, cols)

    linhas = []

    # 1. Ambíguos
    if resultado.a_analisar_ambiguos is not None and not resultado.a_analisar_ambiguos.empty:
        for _, row in resultado.a_analisar_ambiguos.iterrows():
            candidatas = row.get("candidatos") or []
            detalhe = " | ".join(
                f"{'ADI' if c.get('classe') == 'adiantamento' else 'NF'} {c.get('nro_nota') or c.get('nro_nota_referenciada') or '?'} · {_fmt_moeda_str(c.get('vlr_desdobramento'))}"
                for c in candidatas[:3]
            )
            if len(candidatas) > 3:
                detalhe += f" ... (+{len(candidatas) - 3})"

            parc_atual = row.get("parcela_atual")
            parc_total = row.get("parcelas_total")
            parc_txt = f"{parc_atual}/{parc_total}" if parc_atual and parc_total and parc_total > 1 else "1/1"

            linhas.append([
                "Múltiplas candidatas",
                _label_adquirente(row.get("adquirente")),
                _label_bandeira(row.get("bandeira")),
                _label_modalidade(row.get("modalidade"), row.get("parcelas_total")),
                parc_txt,
                _to_date(row.get("data_prev_pagamento")),
                _num(row.get("valor_match")),
                _str(row.get("nsu")),
                _str(row.get("autorizacao")),
                len(candidatas),
                detalhe,
            ])

    # 2. Venda sem título (subdivide por urgência)
    hoje = date.today()
    if resultado.a_analisar_venda_sem_titulo is not None and not resultado.a_analisar_venda_sem_titulo.empty:
        for _, row in resultado.a_analisar_venda_sem_titulo.iterrows():
            data_venda = _to_date(row.get("data_prev_pagamento"))
            if data_venda is None:
                dias_txt = ""
                situacao = "Venda sem título · sem data"
            else:
                dias = (hoje - data_venda).days
                if dias < 3:
                    situacao = f"Aguardando faturamento · {dias} dia(s)"
                else:
                    situacao = f"Divergência real · {dias} dias sem par"

            parc_atual = row.get("parcela_atual")
            parc_total = row.get("parcelas_total")
            parc_txt = f"{parc_atual}/{parc_total}" if parc_atual and parc_total and parc_total > 1 else "1/1"

            linhas.append([
                situacao,
                _label_adquirente(row.get("adquirente")),
                _label_bandeira(row.get("bandeira")),
                _label_modalidade(row.get("modalidade"), row.get("parcelas_total")),
                parc_txt,
                data_venda,
                _num(row.get("valor_match")),
                _str(row.get("nsu")),
                _str(row.get("autorizacao")),
                0,
                "Sem par no Sankhya",
            ])

    # 3. Desfeitas manualmente — precisamos localizar cada uma no grupo_1 pra reportar
    if ligacoes_desfeitas and resultado.grupo_1_conciliadas is not None and not resultado.grupo_1_conciliadas.empty:
        # Deduplica por chave — só uma linha por venda desfeita
        vistos = set()
        for _, row in resultado.grupo_1_conciliadas.iterrows():
            chave = _chave_venda_original(row)
            if chave in ligacoes_desfeitas and chave not in vistos:
                vistos.add(chave)
                linhas.append([
                    "Desfeita manualmente",
                    _label_adquirente(row.get("adquirente")),
                    _label_bandeira(row.get("bandeira")),
                    _label_modalidade(row.get("modalidade"), row.get("parcelas_total")),
                    f"1/{row.get('parcelas_total') or 1}",
                    _to_date(row.get("data_prev_pagamento")),
                    _num(row.get("valor_match")),
                    _str(row.get("nsu")),
                    _str(row.get("autorizacao")),
                    0,
                    f"Ligação com {row.get('sk_nome_parceiro')} · NF {row.get('sk_nro_nota')} desfeita",
                ])

    if not linhas:
        ws.cell(row=2, column=1, value="(nada a analisar nesta rodada — motor casou tudo)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, cols)
        return

    _preencher_linhas(ws, linhas, inicio=2, formatos_col={6: _FMT_DATA, 7: _FMT_MOEDA})
    _ajustar_larguras(ws, cols, larguras={"Situação": 26, "Detalhe": 55})
    ws.freeze_panes = "A2"


# ==============================================================================
# ABA 8 · HISTÓRICO
# ==============================================================================

def _aba_historico(wb: Workbook, historico: List[Dict[str, Any]]):
    ws = wb.create_sheet("Histórico")

    cols = ["Quando", "Ação", "Chave venda / NF", "Parceiro", "Adquirente", "Valor (R$)", "Detalhes"]
    _aplicar_header(ws, cols)

    if not historico:
        ws.cell(row=2, column=1, value="(nenhuma ação registrada nesta rodada)").font = Font(italic=True, color="888888")
        _ajustar_larguras(ws, cols)
        return

    linhas = []
    for ev in historico:
        acao = ev.get("acao")
        if acao == "desfazer_ligacao":
            linhas.append([
                ev.get("quando"),
                "Desfazer ligação",
                ev.get("chave_venda"),
                ev.get("nome_parceiro"),
                _label_adquirente(ev.get("adquirente")),
                _num(ev.get("valor_total")),
                f"{ev.get('n_parcelas', 1)} parcela(s) voltaram a 'A analisar'",
            ])
        elif acao == "escolher_candidata":
            linhas.append([
                ev.get("quando"),
                "Escolher candidata",
                f"{ev.get('chave_venda')} → NF {ev.get('sk_nro_nota')}",
                "",
                "",
                "",
                f"Classe escolhida: {_label_classe_sk(ev.get('sk_classe'))}",
            ])
        else:
            linhas.append([ev.get("quando"), acao, "", "", "", "", str(ev)])

    _preencher_linhas(ws, linhas, inicio=2, formatos_col={6: _FMT_MOEDA})
    _ajustar_larguras(ws, cols, larguras={"Quando": 20, "Chave venda / NF": 35, "Detalhes": 55})
    ws.freeze_panes = "A2"


# ==============================================================================
# API PÚBLICA
# ==============================================================================

def gerar_excel(
    resultado,
    confirmadas_manual: Dict,
    ligacoes_desfeitas: Set,
    historico: List,
    contadores: Dict[str, int],
    df_cielo=None,
    df_getnet=None,
    df_sankhya=None,
    tolerancia_dias: int = 2,
) -> bytes:
    """
    Gera o Excel de auditoria com 8 abas.

    Args:
        resultado: ResultadoMotor devolvido por motor_vendas.rodar()
        confirmadas_manual: dict {chave_venda_str: dados_do_candidato_escolhido}
        ligacoes_desfeitas: set de chaves (tupla) que foram desfeitas manualmente
        historico: lista de eventos da rodada
        contadores: dict retornado por _calcular_contadores_pills
        df_cielo, df_getnet, df_sankhya: DataFrames originais (pra Resumo)
        tolerancia_dias: usada na Resumo

    Returns:
        bytes: workbook .xlsx pronto para download
    """
    wb = Workbook()
    # Remove a sheet default
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Abas na ordem
    _aba_resumo(
        wb, resultado, contadores,
        ligacoes_desfeitas, confirmadas_manual, historico,
        df_cielo, df_getnet, df_sankhya, tolerancia_dias,
    )
    _aba_conciliadas(wb, "Grupo 1 Auto", resultado.grupo_1_conciliadas, ligacoes_desfeitas)
    _aba_confirmadas(wb, confirmadas_manual)
    _aba_conciliadas(wb, "Grupo 2 TOP 1722", resultado.grupo_2_ja_baixadas, set())
    _aba_aguardando(wb, resultado.grupo_3_aguardando)
    _aba_devolucoes(wb, resultado.grupo_4_devolucoes)
    _aba_a_analisar(wb, resultado, ligacoes_desfeitas)
    _aba_historico(wb, historico)

    # Salva em bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
