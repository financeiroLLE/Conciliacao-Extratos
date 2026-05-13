"""Gera arquivos de exemplo sintéticos (dados FICTÍCIOS) para o repositório.

Os arquivos gerados são seguros para commitar no GitHub público — não
contém nenhuma informação real de empresa, cliente ou banco.
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

SAMPLES_DIR = Path(__file__).parent.parent / "data" / "samples"
SAMPLES_DIR.mkdir(parents=True, exist_ok=True)

# Dados sintéticos — fornecedores e descrições fictícias
FORNECEDORES_FICT = [
    "FORNECEDOR ALPHA LTDA",
    "BETA INDUSTRIA SA",
    "GAMMA COMERCIO ME",
    "DELTA SERVICOS LTDA",
    "EPSILON DISTRIB SA",
]

CLIENTES_FICT = [
    "CLIENTE NORTE LTDA",
    "CLIENTE SUL ME",
    "CLIENTE LESTE SA",
    "CLIENTE OESTE LTDA",
]


def gerar_extrato_banco_exemplo():
    """Cria um extrato bancário sintético padronizado."""
    base_date = datetime(2026, 5, 4)
    linhas_por_aba = {
        "04.05": [],
        "05.05": [],
        "06.05": [],
        "07.05": [],
    }

    rng = list(range(100))  # determinístico para o exemplo

    # 04.05 — saídas grandes (pagamentos a fornecedores)
    for i, forn in enumerate(FORNECEDORES_FICT):
        linhas_por_aba["04.05"].append({
            "Data": "04/05/2026",
            "Histórico": f"PGTO TITULO OUTRO BANCO - API {forn}",
            "Documento": None,
            "Valor (R$)": -1000.00 * (i + 1),
        })

    # 04.05 — algumas tarifas
    for i, tarifa in enumerate(["TAR LIQ COB COM REG COMPE", "TAR LIQ COB CORBAN DIGITAL"]):
        linhas_por_aba["04.05"].append({
            "Data": "04/05/2026",
            "Histórico": tarifa,
            "Documento": "193956",
            "Valor (R$)": -3.00,
        })

    # 05.05 — entradas (PIX recebidos)
    for i, cli in enumerate(CLIENTES_FICT):
        linhas_por_aba["05.05"].append({
            "Data": "05/05/2026",
            "Histórico": f"PIX RECEBIDO  {cli}",
            "Documento": None,
            "Valor (R$)": 250.00 + i * 100,
        })

    # 06.05 — mistura
    linhas_por_aba["06.05"].extend([
        {"Data": "06/05/2026", "Histórico": "APLICACAO CONTAMAX", "Documento": None, "Valor (R$)": -50000.00},
        {"Data": "06/05/2026", "Histórico": f"PIX ENVIADO  {FORNECEDORES_FICT[0]}", "Documento": None, "Valor (R$)": -1500.00},
        {"Data": "06/05/2026", "Histórico": "PAGAMENTO DE TITULO  9999.0000", "Documento": "310506", "Valor (R$)": -800.50},
    ])

    # 07.05 — incluir um lançamento que NÃO terá contrapartida no sistema
    # (vai virar pendência de banco)
    linhas_por_aba["07.05"].extend([
        {"Data": "07/05/2026", "Histórico": "TAR LIQ COB COM REG COMPE", "Documento": "193956", "Valor (R$)": -3.00},
        {"Data": "07/05/2026", "Histórico": "PIX ENVIADO MISTERIOSO", "Documento": None, "Valor (R$)": -777.77},
        # Esse vai duplicar de propósito (auditoria de duplicidade)
        {"Data": "07/05/2026", "Histórico": "TAR LIQ COB COM REG COMPE", "Documento": "193956", "Valor (R$)": -3.00},
    ])

    # Escreve no Excel
    caminho = SAMPLES_DIR / "extrato_banco_exemplo.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for aba, linhas in linhas_por_aba.items():
        ws = wb.create_sheet(aba)
        df = pd.DataFrame(linhas)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
    wb.save(caminho)
    print(f"✅ {caminho.name} — {sum(len(v) for v in linhas_por_aba.values())} linhas")


def gerar_relatorio_sistema_exemplo():
    """Cria um relatório de Conciliação Bancária sintético."""
    linhas = []

    # Replica a maioria dos lançamentos do banco no sistema (vai conciliar)
    for i, forn in enumerate(FORNECEDORES_FICT):
        linhas.append({
            "Tipo de Movimento": "Financeiro",
            "Núm. Documento": 100 + i,
            "Conciliado": "Sim",
            "Vlr. Lançamento": 1000.00 * (i + 1),
            "Receita/Despesa": "Despesa",
            "Dt. Lançamento": "04/05/2026",
            "Histórico": forn,
            "Núm. Único Bancário": 800000 + i,
            "Usuário": "DEBORA.SILVA",
            "Conta Bancária": "BCO-EXEMPLO-001",
        })

    # Tarifas
    for hist in ["TAR LIQ COB COM REG COMPE", "TAR LIQ COB CORBAN DIGITAL"]:
        linhas.append({
            "Tipo de Movimento": "Financeiro",
            "Núm. Documento": 0,
            "Conciliado": "Sim",
            "Vlr. Lançamento": 3.00,
            "Receita/Despesa": "Despesa",
            "Dt. Lançamento": "04/05/2026",
            "Histórico": hist,
            "Núm. Único Bancário": 900001,
            "Usuário": "DEBORA.SILVA",
            "Conta Bancária": "BCO-EXEMPLO-001",
        })

    # Entradas
    for i, cli in enumerate(CLIENTES_FICT):
        linhas.append({
            "Tipo de Movimento": "Financeiro",
            "Núm. Documento": 200 + i,
            "Conciliado": "Sim",
            "Vlr. Lançamento": 250.00 + i * 100,
            "Receita/Despesa": "Receita",
            "Dt. Lançamento": "05/05/2026",
            "Histórico": cli,
            "Núm. Único Bancário": 700000 + i,
            "Usuário": "DEBORA.SILVA",
            "Conta Bancária": "BCO-EXEMPLO-001",
        })

    # Mistura do dia 06
    linhas.extend([
        {"Tipo de Movimento": "Transferência", "Núm. Documento": 0, "Conciliado": "Sim",
         "Vlr. Lançamento": 50000.00, "Receita/Despesa": "Despesa", "Dt. Lançamento": "06/05/2026",
         "Histórico": "APLICACAO CONTAMAX", "Núm. Único Bancário": 950001,
         "Usuário": "DEBORA.SILVA", "Conta Bancária": "BCO-EXEMPLO-001"},
        {"Tipo de Movimento": "Financeiro", "Núm. Documento": 301, "Conciliado": "Sim",
         "Vlr. Lançamento": 1500.00, "Receita/Despesa": "Despesa", "Dt. Lançamento": "06/05/2026",
         "Histórico": FORNECEDORES_FICT[0], "Núm. Único Bancário": 950002,
         "Usuário": "DEBORA.SILVA", "Conta Bancária": "BCO-EXEMPLO-001"},
        {"Tipo de Movimento": "Financeiro", "Núm. Documento": 310506, "Conciliado": "Sim",
         "Vlr. Lançamento": 800.50, "Receita/Despesa": "Despesa", "Dt. Lançamento": "06/05/2026",
         "Histórico": "PAGAMENTO DE TITULO", "Núm. Único Bancário": 950003,
         "Usuário": "DEBORA.SILVA", "Conta Bancária": "BCO-EXEMPLO-001"},
    ])

    # Dia 07 — lançamento que está no SISTEMA mas NÃO no banco (indevido)
    linhas.append({
        "Tipo de Movimento": "Financeiro", "Núm. Documento": 999, "Conciliado": "Não",
        "Vlr. Lançamento": 555.55, "Receita/Despesa": "Despesa", "Dt. Lançamento": "07/05/2026",
        "Histórico": "LANCAMENTO INDEVIDO TESTE", "Núm. Único Bancário": 0,
        "Usuário": "DEBORA.SILVA", "Conta Bancária": "BCO-EXEMPLO-001",
    })

    # Divergência de valor — mesma data + histórico parecido, valor diferente
    linhas.append({
        "Tipo de Movimento": "Financeiro", "Núm. Documento": 105, "Conciliado": "Sim",
        "Vlr. Lançamento": 1050.00,  # banco tem 1000, sistema tem 1050 → divergência
        "Receita/Despesa": "Despesa", "Dt. Lançamento": "04/05/2026",
        "Histórico": FORNECEDORES_FICT[0] + " (divergente)",  # similar mas diferente
        "Núm. Único Bancário": 999001,
        "Usuário": "DEBORA.SILVA", "Conta Bancária": "BCO-EXEMPLO-001",
    })

    # Escreve mantendo a estrutura: linha 1 título, linha 2 emissão, linha 3 cabeçalho
    caminho = SAMPLES_DIR / "relatorio_sistema_exemplo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"

    ws["A1"] = "Conciliação Bancária"
    ws["A2"] = f"Emissão:{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["B2"] = f"Total de registros:{len(linhas)}"
    ws["C2"] = "Usuário: EXEMPLO"

    cols = list(linhas[0].keys())
    for c_idx, col in enumerate(cols, start=1):
        ws.cell(row=3, column=c_idx, value=col)

    for r_idx, linha in enumerate(linhas, start=4):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=linha[col])

    wb.save(caminho)
    print(f"✅ {caminho.name} — {len(linhas)} linhas")


if __name__ == "__main__":
    gerar_extrato_banco_exemplo()
    gerar_relatorio_sistema_exemplo()
    print("\n🎉 Arquivos de exemplo criados em data/samples/")
