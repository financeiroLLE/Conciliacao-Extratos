"""Testes do módulo CARTÃO (Sprint 1)."""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
import openpyxl

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.cartao import (
    MODALIDADES_VALIDAS,
    carregar_cadastro_taxas,
    carregar_relatorio_adquirente,
    auditar_taxas,
    consolidar_historico,
    encontrar_taxa_vigente,
)


def _criar_taxas_xlsx(linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "adquirente", "modalidade", "parcelas", "taxa_mdr",
        "taxa_antecipacao", "prazo_dias", "vigencia_inicio", "vigencia_fim",
    ])
    for ln in linhas:
        ws.append(list(ln))
    buf = BytesIO(); wb.save(buf); buf.seek(0); buf.name = "taxas.xlsx"
    return buf


def _criar_relatorio_xlsx(linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "data_venda", "adquirente", "modalidade", "parcelas",
        "valor_bruto", "taxa_aplicada", "valor_liquido", "data_prevista_recebimento",
    ])
    for ln in linhas:
        ws.append(list(ln))
    buf = BytesIO(); wb.save(buf); buf.seek(0); buf.name = "adq.xlsx"
    return buf


def teste_modalidades_validas():
    assert "Débito" in MODALIDADES_VALIDAS
    assert "Crédito à vista" in MODALIDADES_VALIDAS
    assert "Crédito parcelado" in MODALIDADES_VALIDAS
    assert "Pix QR Code" in MODALIDADES_VALIDAS
    print("✓ teste_modalidades_validas")


def teste_parse_taxa_com_percentual_e_fracao():
    from src.cartao.cadastro_taxas import _parse_taxa
    assert abs(_parse_taxa("1,39%") - 0.0139) < 1e-9
    assert abs(_parse_taxa("0.0139") - 0.0139) < 1e-9
    assert abs(_parse_taxa("1.39") - 0.0139) < 1e-9
    assert abs(_parse_taxa("0,49%") - 0.0049) < 1e-9
    assert _parse_taxa(None) == 0.0
    assert _parse_taxa("") == 0.0
    print("✓ teste_parse_taxa_com_percentual_e_fracao")


def teste_normalizacao_modalidade():
    from src.cartao.cadastro_taxas import _normalizar_modalidade
    assert _normalizar_modalidade("debito") == "Débito"
    assert _normalizar_modalidade("Débito") == "Débito"
    assert _normalizar_modalidade("CREDITO A VISTA") == "Crédito à vista"
    assert _normalizar_modalidade("Crédito à vista") == "Crédito à vista"
    assert _normalizar_modalidade("Parcelado") == "Crédito parcelado"
    assert _normalizar_modalidade("Pix QR Code") == "Pix QR Code"
    assert _normalizar_modalidade("pix") == "Pix QR Code"
    print("✓ teste_normalizacao_modalidade")


def teste_cadastro_basico():
    buf = _criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
        ("Stone", "Crédito à vista", 1, "2,49%", "1,99%", 30, "01/01/2026", ""),
        ("Cielo", "Pix QR Code", 1, "0,49%", "0%", 0, "01/01/2026", ""),
    ])
    cad = carregar_cadastro_taxas(buf)
    assert len(cad) == 3
    assert abs(cad.iloc[0]["taxa_mdr"] - 0.0139) < 1e-9
    assert cad.iloc[2]["modalidade"] == "Pix QR Code"
    print("✓ teste_cadastro_basico")


def teste_cadastro_sem_colunas_obrigatorias():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["adquirente", "modalidade"])
    ws.append(["Stone", "Débito"])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    try:
        carregar_cadastro_taxas(buf)
        assert False, "deveria ter lançado erro"
    except ValueError as e:
        assert "obrigatórias" in str(e).lower() or "obrigatorias" in str(e).lower()
    print("✓ teste_cadastro_sem_colunas_obrigatorias")


def teste_encontrar_taxa_vigente():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    cfg = encontrar_taxa_vigente(cad, "Stone", "Débito", 1, pd.Timestamp("2026-05-14"))
    assert cfg is not None
    assert abs(cfg["taxa_mdr"] - 0.0139) < 1e-9
    assert encontrar_taxa_vigente(cad, "Rede", "Débito", 1, pd.Timestamp("2026-05-14")) is None
    assert encontrar_taxa_vigente(cad, "Stone", "Pix QR Code", 1, pd.Timestamp("2026-05-14")) is None
    print("✓ teste_encontrar_taxa_vigente")


def teste_auditoria_jeito_b_ok_com_arredondamento():
    """JEITO B: venda R$ 73,42 × 1,389% efetiva = R$ 1,02 = R$ 1,02 esperado → OK."""
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    rel = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("14/05/2026", "Stone", "Débito", 1, "73,42", "1,389267%", "72,40", "15/05/2026"),
    ]))
    res = auditar_taxas(rel, cad)
    linha = res.detalhado.iloc[0]
    assert linha["status"] == "OK", (
        f"esperava OK (centavos batem) mas veio '{linha['status']}'. "
        f"esperado_rs={linha['esperado_rs']} cobrado_rs={linha['cobrado_rs']}"
    )
    print("✓ teste_auditoria_jeito_b_ok_com_arredondamento")


def teste_auditoria_divergencia_real():
    """Stone cobrou 2% no débito quando contrato é 1,39% → R$ 0,61 de divergência."""
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    rel = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("14/05/2026", "Stone", "Débito", 1, "100,00", "2,00%", "98,00", "15/05/2026"),
    ]))
    res = auditar_taxas(rel, cad)
    linha = res.detalhado.iloc[0]
    assert linha["status"] == "Divergente"
    assert abs(linha["diferenca_rs"] - 0.61) < 0.01
    assert res.kpis["impacto_acumulado"] > 0
    print("✓ teste_auditoria_divergencia_real")


def teste_auditoria_sem_contrato():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    rel = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("14/05/2026", "Getnet", "Crédito à vista", 1, "100,00", "3,00%", "97,00", "13/06/2026"),
    ]))
    res = auditar_taxas(rel, cad)
    assert res.detalhado.iloc[0]["status"] == "Sem contrato"
    assert res.kpis["qtd_sem_contrato"] == 1
    print("✓ teste_auditoria_sem_contrato")


def teste_auditoria_pix_qrcode():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Cielo", "Pix QR Code", 1, "0,49%", "0%", 0, "01/01/2026", ""),
    ]))
    rel = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("14/05/2026", "Cielo", "Pix QR Code", 1, "500,00", "0,49%", "497,55", "14/05/2026"),
    ]))
    res = auditar_taxas(rel, cad)
    assert res.detalhado.iloc[0]["status"] == "OK"
    print("✓ teste_auditoria_pix_qrcode")


def teste_historico_acumulado():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    rel_dia_1 = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("12/05/2026", "Stone", "Débito", 1, "100,00", "2,00%", "98,00", "13/05/2026"),
    ]))
    res_dia_1 = auditar_taxas(rel_dia_1, cad)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        res_dia_1.detalhado.to_excel(w, sheet_name="Auditoria", index=False)
    buf.seek(0); buf.name = "audit_dia1.xlsx"

    hist_df, avisos = consolidar_historico([buf])
    assert len(hist_df) == 1
    assert avisos == []

    rel_dia_2 = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("13/05/2026", "Stone", "Débito", 1, "200,00", "2,00%", "196,00", "14/05/2026"),
    ]))
    res_consolidado = auditar_taxas(rel_dia_2, cad, historico=hist_df)
    assert len(res_consolidado.detalhado) == 2
    assert abs(res_consolidado.kpis["impacto_acumulado"] - 1.83) < 0.05
    print("✓ teste_historico_acumulado")


def teste_historico_detecta_duplicatas():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    rel = carregar_relatorio_adquirente(_criar_relatorio_xlsx([
        ("12/05/2026", "Stone", "Débito", 1, "100,00", "2,00%", "98,00", "13/05/2026"),
    ]))
    res = auditar_taxas(rel, cad)
    bufs = []
    for nome in ("a.xlsx", "b.xlsx"):
        b = BytesIO()
        with pd.ExcelWriter(b, engine="openpyxl") as w:
            res.detalhado.to_excel(w, sheet_name="Auditoria", index=False)
        b.seek(0); b.name = nome
        bufs.append(b)
    hist, avisos = consolidar_historico(bufs)
    assert len(avisos) > 0
    assert "duplicado" in avisos[0].lower()
    print("✓ teste_historico_detecta_duplicatas")


def teste_kpis_vazios():
    cad = carregar_cadastro_taxas(_criar_taxas_xlsx([
        ("Stone", "Débito", 1, "1,39%", "0%", 1, "01/01/2026", ""),
    ]))
    res = auditar_taxas(pd.DataFrame(), cad, historico=None)
    assert res.detalhado.empty
    assert res.divergentes.empty
    assert res.kpis["volume_bruto"] == 0.0
    assert res.kpis["impacto_acumulado"] == 0.0
    print("✓ teste_kpis_vazios")


def main():
    testes = [v for k, v in globals().items() if k.startswith("teste_")]
    falhas = []
    for t in testes:
        try:
            t()
        except AssertionError as e:
            print(f"✗ {t.__name__}: {e}")
            falhas.append(t.__name__)
        except Exception as e:
            print(f"✗ {t.__name__}: {type(e).__name__}: {e}")
            falhas.append(t.__name__)
    print()
    print(f"{len(testes) - len(falhas)}/{len(testes)} passaram")
    return 0 if not falhas else 1


if __name__ == "__main__":
    sys.exit(main())
