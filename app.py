"""App Streamlit — Conciliação Bancária Grupo LLE.

Layout: fundo azul institucional, sidebar amarela, cards executivos.
Fluxo: Upload → tela única de Resultado com cards, painel de bancos
e abas internas por status / subabas por tipo.
"""

from __future__ import annotations

import base64
import io
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st

from src.auditoria import (
    listar_execucoes,
    novo_id_execucao,
    registrar_execucao,
    salvar_snapshot,
)
from src.classificacao import TIPOS_PRINCIPAIS
from src.parsers import (
    carregar_extrato_banco,
    carregar_pendencias_anteriores,
    carregar_relatorio_sistema,
)
from src.parsers.adquirente import (
    carregar_extrato_adquirente,
    resumo_por_categoria,
)
from src.pipeline import ResultadoConciliacao, executar_pipeline
from src.reports import (
    gerar_csvs_zip,
    gerar_relatorio_excel,
    gerar_relatorio_excel_de_conta,
)


# ============================================================
# Identidade visual — Grupo LLE
# ============================================================
CORES = {
    "azul_escuro": "#041747",
    "azul_escuro_2": "#0A1F4D",
    "azul": "#0071FE",
    "amarelo": "#FAC318",
    "amarelo_2": "#E5AD0A",
    "verde": "#0F8C3B",
    "vermelho": "#D63031",
    "branco": "#FFFFFF",
    "texto_muted": "#A8B3CC",
    "card_bg": "#0E2456",
    "card_borda": "#1B3266",
}

ASSETS = Path(__file__).parent / "assets"


@st.cache_data
def _logo_data_uri() -> str:
    """Retorna a logo PNG (com fundo transparente) como data URI."""
    # Preferir a versão transparente; cair para a com fundo se a transparente não existir
    for nome in ("logo-grupo-lle-transparente.png", "logo-grupo-lle-branco.png"):
        arq = ASSETS / nome
        if arq.exists():
            b64 = base64.b64encode(arq.read_bytes()).decode("ascii")
            return f"data:image/png;base64,{b64}"
    return ""


# ============================================================
# v5.35: Identificador da conta a partir do Sankhya
# ------------------------------------------------------------
# O match banco × Sankhya casa pela STRING exata do nome da conta. Para a
# usuária não precisar digitar (e errar espaço/underline/maiúscula), lemos os
# nomes de conta de dentro do próprio Sankhya e oferecemos numa lista — assim a
# string é sempre idêntica à que o matcher usa. A pré-seleção usa o nome do
# arquivo/banco só para já marcar o item certo; quem casa é o nome do Sankhya.
# ============================================================
_OPCAO_DIGITAR = "✏️  Digitar outro nome…"


def _norm_conta(s: str) -> str:
    """Normaliza nome de conta para comparar (espaço/underline/maiúscula iguais)."""
    s = str(s).lower().strip()
    s = re.sub(r"[_\-./]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


_STOP_CONTA = {
    "banco", "extrato", "mensal", "conta", "corrente", "cc", "c", "ag", "agencia",
    "do", "da", "de", "dos", "das", "e", "no", "na", "movimento", "conciliacao",
    "relatorio",
}


def _tokens_conta(s: str) -> set[str]:
    """Tokens distintivos de um nome de conta (sem palavras genéricas/números)."""
    return {
        t for t in _norm_conta(s).split()
        if t not in _STOP_CONTA and not t.isdigit() and len(t) >= 3
    }


def _melhor_match_conta(candidato: str, opcoes: list[str]) -> int | None:
    """Índice da conta do Sankhya que melhor casa com o nome do arquivo/banco.

    Retorna None quando não há um match claro (uma conta nova, p.ex.) — aí a UI
    pré-seleciona 'Digitar outro nome…'. Nunca chuta com ambiguidade.
    """
    nc = _norm_conta(candidato)
    if not nc:
        return None
    # 1) match exato (normalizado)
    for i, o in enumerate(opcoes):
        if _norm_conta(o) == nc:
            return i
    # 2) um contém o outro, e só uma conta bate
    contidos = [
        i for i, o in enumerate(opcoes)
        if nc in _norm_conta(o) or _norm_conta(o) in nc
    ]
    if len(contidos) == 1:
        return contidos[0]
    # 3) token distintivo (KING, INOV, PISA...) — casa quando uma conta tem o
    #    MAIOR overlap de tokens e esse máximo é único (sem empate = sem chute).
    tk = _tokens_conta(candidato)
    if tk:
        scores = sorted(
            ((len(_tokens_conta(o) & tk), i) for i, o in enumerate(opcoes)),
            reverse=True,
        )
        if scores and scores[0][0] > 0 and (len(scores) == 1 or scores[0][0] > scores[1][0]):
            return scores[0][1]
    return None


@st.cache_data(show_spinner=False)
def _contas_no_sankhya(payload: tuple, coluna_conta: str) -> list[str]:
    """Nomes de conta presentes no(s) relatório(s) do Sankhya.

    `payload` é uma tupla de (nome_arquivo, bytes) — hashável, então o resultado
    fica em cache e o Sankhya (que pode ter 69k linhas) só é lido uma vez por
    arquivo. As strings retornadas são IDÊNTICAS às que o pipeline usa na conta.
    """
    nomes: set[str] = set()
    for _nome_arq, conteudo in payload:
        try:
            # v5.39: sem o nome do arquivo, o leitor não sabe que é .xls e cai no
            # engine errado (openpyxl → BadZipFile), o except engolia e a lista de
            # contas voltava vazia — por isso o dropdown não aparecia pra .xls.
            bio = io.BytesIO(conteudo)
            bio.name = _nome_arq
            df = carregar_relatorio_sistema(
                bio, coluna_conta=coluna_conta or None
            )
            nomes.update(
                c for c in df["conta"].astype(str).unique() if c and c != "—"
            )
        except Exception:
            continue
    return sorted(nomes)


def _ler_contas_sankhya_da_sessao() -> list[str]:
    """Lê os nomes de conta do Sankhya já enviado (via session_state).

    A coluna 1 (banco) é renderizada ANTES da coluna 2 (Sankhya) no mesmo run,
    então pegamos o arquivo do Sankhya do run anterior pelo session_state.
    """
    arquivos = st.session_state.get("sistema") or []
    if arquivos and not isinstance(arquivos, list):
        arquivos = [arquivos]
    if not arquivos:
        return []
    coluna = st.session_state.get("coluna_conta_sistema", "") or ""
    try:
        payload = tuple((f.name, f.getvalue()) for f in arquivos)
    except Exception:
        return []
    return _contas_no_sankhya(payload, coluna)


@st.cache_data(show_spinner=False)
def _detecta_conta_bytes(nome: str, conteudo: bytes):
    """v5.64: lê o cabeçalho do extrato (banco/agência/conta/empresa) a partir
    dos bytes. Cacheado — no modo 'Várias contas de uma vez' cada rerun da tela
    repassaria todos os arquivos; com cache, cada arquivo é lido uma vez só."""
    import io as _io
    from src.parsers.deteccao_conta import detectar_conta_extrato
    try:
        b = _io.BytesIO(conteudo)
        b.name = nome
        return detectar_conta_extrato(b)
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def _parse_adquirentes(payload: tuple) -> pd.DataFrame:
    """Lê todos os arquivos de adquirente (GetNet/PagBank) e junta num só df,
    no esquema comum de src.parsers.adquirente. payload = ((nome, bytes), ...)."""
    import io as _io

    from src.parsers.adquirente import COLUNAS_SAIDA

    frames = []
    _vistos = set()
    for nome, conteudo in payload:
        try:
            import hashlib as _hl

            _h = _hl.md5(conteudo).hexdigest()
            if _h in _vistos:
                continue  # v5.42: ignora arquivo idêntico subido em duplicata
            _vistos.add(_h)
            bio = _io.BytesIO(conteudo)
            bio.name = nome
            df = carregar_extrato_adquirente(bio)
            if df is not None and not df.empty:
                df = df.copy()
                df["arquivo"] = nome
                frames.append(df)
        except Exception:
            continue
    if not frames:
        return pd.DataFrame(columns=COLUNAS_SAIDA + ["arquivo"])
    return pd.concat(frames, ignore_index=True)


def _adquirentes_da_sessao() -> pd.DataFrame:
    """DataFrame combinado dos extratos de adquirente enviados (ou vazio)."""
    from src.parsers.adquirente import COLUNAS_SAIDA

    arquivos = st.session_state.get("adquirente") or []
    if arquivos and not isinstance(arquivos, list):
        arquivos = [arquivos]
    if not arquivos:
        return pd.DataFrame(columns=COLUNAS_SAIDA + ["arquivo"])
    try:
        payload = tuple((f.name, f.getvalue()) for f in arquivos)
    except Exception:
        return pd.DataFrame(columns=COLUNAS_SAIDA + ["arquivo"])
    return _parse_adquirentes(payload)


def _relatorio_auditoria_da_sessao() -> tuple:
    """Monta o relatório da Auditoria de Cartões a partir dos extratos de
    adquirente JÁ enviados na conciliação (guardados em 'adquirente_bytes').

    Só o extrato CRU da GETNET tem venda + valor bruto + taxa aplicada, então
    é o único auditável por venda. PagBank é recebimento (não traz taxa por
    venda) e é ignorado aqui — informamos quais foram ignorados, sem inventar.
    Retorna (relatorio_df, nomes_ignorados).
    """
    import io as _io

    from src.cartao import eh_extrato_getnet_cru, carregar_extrato_getnet_cru

    itens = st.session_state.get("adquirente_bytes") or []
    frames, ignorados = [], []
    for nome, conteudo in itens:
        try:
            if eh_extrato_getnet_cru(_io.BytesIO(conteudo)):
                rel = carregar_extrato_getnet_cru(_io.BytesIO(conteudo))
                if rel is not None and not rel.empty:
                    frames.append(rel)
                else:
                    ignorados.append(nome)
            else:
                ignorados.append(nome)
        except Exception:
            ignorados.append(nome)
    relatorio = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return relatorio, ignorados


# ============================================================
# Configuração da página
# ============================================================
st.set_page_config(
    page_title="Conciliação Bancária · Grupo LLE",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# Estado de sessão
# ============================================================
if "pagina" not in st.session_state:
    st.session_state.pagina = "Conciliação"
if "resultado" not in st.session_state:
    st.session_state.resultado = None
if "pendencias_anteriores" not in st.session_state:
    st.session_state.pendencias_anteriores = pd.DataFrame()
if "id_execucao_atual" not in st.session_state:
    st.session_state.id_execucao_atual = None
if "banco_conta_selecionada" not in st.session_state:
    st.session_state.banco_conta_selecionada = None
if "fluxo_etapa" not in st.session_state:
    st.session_state.fluxo_etapa = "upload"  # upload | resultado
if "subtab_conciliacao" not in st.session_state:
    st.session_state.subtab_conciliacao = "Todos"


def ir_para(pagina: str):
    st.session_state.pagina = pagina


def selecionar_banco(conta: str):
    st.session_state.banco_conta_selecionada = conta


def voltar_upload():
    st.session_state.fluxo_etapa = "upload"
    st.session_state.banco_conta_selecionada = None


# ============================================================
# CSS global
# ============================================================
LOGO_URI = _logo_data_uri()

st.html(
    f"""
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/@tabler/icons-webfont@3.19.0/dist/tabler-icons.min.css" rel="stylesheet">
<style>
/* ===== Tipografia e fundo ===== */
html, body, [class*="css"], .stMarkdown, .stText, button, input, select, textarea {{
    font-family: 'Montserrat', sans-serif !important;
}}
.stApp {{
    background: linear-gradient(180deg, {CORES["azul_escuro"]} 0%, #061B57 100%) !important;
    color: {CORES["branco"]};
}}

/* ===== v3: ESCONDER faixa branca do topo do Streamlit =====
   v5.8: mantemos o header vis\u00edvel (mas transparente) porque \u00e9 onde fica
   o bot\u00e3o de reabrir sidebar. Com 'display: none' o bot\u00e3o sumia. */
header[data-testid="stHeader"] {{
    background-color: transparent !important;
}}
[data-testid="stToolbar"] {{
    background-color: transparent !important;
}}
#MainMenu {{ visibility: hidden; }}
.stDeployButton,
.stAppDeployButton,
[data-testid="stDeployButton"],
[data-testid="stAppDeployButton"],
[data-testid="stToolbarActions"],
[data-testid="stStatusWidget"] {{
    display: none !important;
    visibility: hidden !important;
}}
footer {{ visibility: hidden; }}

.block-container {{
    padding-top: 0.8rem !important;
    padding-bottom: 4rem;
    max-width: 1500px;
}}
h1, h2, h3, h4, h5, h6, p, span, div, label {{ color: {CORES["branco"]}; }}

/* ===== Sidebar AMARELA — v3 com bordas arredondadas e tom uniforme ===== */
[data-testid="stSidebar"] {{
    background-color: {CORES["amarelo"]} !important;
    border-right: none;
}}
[data-testid="stSidebar"] * {{ color: {CORES["azul_escuro"]} !important; }}
[data-testid="stSidebar"] hr {{ border-color: rgba(4,23,71,0.18) !important; }}

/* Logo: bloco azul com bordas inferiores arredondadas e transição suave */
.lle-sidebar-logo {{
    background: linear-gradient(180deg, {CORES["azul_escuro"]} 0%, #061B57 100%);
    padding: 24px 14px 22px 14px;
    margin: 8px 4px 22px 4px;
    border-radius: 18px;
    text-align: center;
    box-shadow: 0 6px 18px rgba(4,23,71,0.22);
}}
.lle-sidebar-logo img {{
    height: 78px;
    width: auto;
    display: inline-block;
}}
.lle-sidebar-tagline {{
    text-align: center;
    color: {CORES["amarelo"]} !important;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 2px;
    margin-top: 10px;
}}

/* v5.8: bot\u00e3o de reabrir sidebar quando ela est\u00e1 fechada.
   Por padr\u00e3o o Streamlit deixa quase invis\u00edvel no canto. */
[data-testid="stExpandSidebarButton"] {{
    background-color: {CORES["amarelo"]} !important;
    border-radius: 8px !important;
    padding: 6px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.35) !important;
    transition: all 0.18s ease !important;
}}
[data-testid="stExpandSidebarButton"]:hover {{
    box-shadow: 0 4px 12px rgba(0,0,0,0.5) !important;
    transform: scale(1.05);
    background-color: {CORES["amarelo"]} !important;
}}
[data-testid="stExpandSidebarButton"] svg,
[data-testid="stExpandSidebarButton"] *,
[data-testid="stExpandSidebarButton"] button {{
    color: {CORES["azul_escuro"]} !important;
    fill: {CORES["azul_escuro"]} !important;
}}

/* Sidebar buttons — TODOS com mesmo tom, ícones verdes ao lado */
[data-testid="stSidebar"] .stButton > button {{
    background-color: rgba(4,23,71,0.06) !important;
    color: {CORES["azul_escuro"]} !important;
    border: 1px solid rgba(4,23,71,0.22) !important;
    border-radius: 10px !important;
    padding: 12px 16px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    text-align: left !important;
    width: 100% !important;
    transition: all 0.18s ease !important;
}}
[data-testid="stSidebar"] .stButton > button:hover {{
    background-color: {CORES["azul_escuro"]} !important;
    color: {CORES["amarelo"]} !important;
    border-color: {CORES["azul_escuro"]} !important;
    transform: translateX(2px);
}}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {{
    background-color: rgba(4,23,71,0.92) !important;
    border-color: {CORES["azul_escuro"]} !important;
}}
/* Texto do botão ativo amarelo */
[data-testid="stSidebar"] .stButton > button[kind="primary"],
[data-testid="stSidebar"] .stButton > button[kind="primary"] * {{
    color: {CORES["amarelo"]} !important;
}}

/* v5.7: força texto amarelo no bot\u00e3o ativo da sidebar com especificidade alta.
   Sem isso, a regra geral '[kind="primary"] *' (texto azul) sobrescrevia,
   ficando azul sobre azul (invis\u00edvel). */
[data-testid="stSidebar"] .stButton > button[kind="primary"] p,
[data-testid="stSidebar"] .stButton > button[kind="primary"] span,
[data-testid="stSidebar"] .stButton > button[kind="primary"] div,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] p,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] span,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] div,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] {{
    color: {CORES["amarelo"]} !important;
    -webkit-text-fill-color: {CORES["amarelo"]} !important;
}}

/* ===== Header escuro com logo ===== */
.lle-header {{
    background: linear-gradient(135deg, {CORES["azul_escuro"]} 0%, {CORES["azul_escuro_2"]} 100%);
    border-radius: 16px;
    padding: 22px 30px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 24px;
    border: 1px solid {CORES["card_borda"]};
}}
.lle-header img {{ height: 56px; width: auto; }}
.lle-header .lle-title {{
    font-size: 26px;
    font-weight: 800;
    color: {CORES["branco"]};
    line-height: 1.1;
    margin: 0;
    letter-spacing: -0.5px;
}}
.lle-header .lle-subtitle {{
    font-size: 13px;
    color: {CORES["amarelo"]};
    margin-top: 4px;
    font-weight: 600;
    letter-spacing: 0.5px;
}}

/* ===== Cards executivos ===== */
.lle-kpi-row {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 14px;
    margin-bottom: 18px;
}}
.lle-kpi {{
    background: linear-gradient(145deg, {CORES["card_bg"]} 0%, #0B1E48 100%);
    border: 1px solid {CORES["card_borda"]};
    border-radius: 16px;
    padding: 20px 22px;
    position: relative;
    overflow: hidden;
    transition: all 0.22s ease;
    box-shadow: 0 4px 14px rgba(0,0,0,0.18);
}}
.lle-kpi:hover {{
    border-color: {CORES["amarelo"]};
    transform: translateY(-2px);
    box-shadow: 0 8px 22px rgba(0,0,0,0.28);
}}
.lle-kpi::before {{
    content: "";
    position: absolute;
    top: 0; left: 0;
    width: 4px; height: 100%;
    background: linear-gradient(180deg, {CORES["amarelo"]} 0%, {CORES["amarelo_2"]} 100%);
}}
.lle-kpi-label {{
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: {CORES["texto_muted"]} !important;
    margin-bottom: 8px;
}}
.lle-kpi-value {{
    font-size: 26px;
    font-weight: 800;
    color: {CORES["branco"]} !important;
    line-height: 1.1;
}}
.lle-kpi-value.destaque-amarelo {{ color: {CORES["amarelo"]} !important; }}
.lle-kpi-value.destaque-verde {{ color: {CORES["verde"]} !important; }}
.lle-kpi-value.destaque-vermelho {{ color: {CORES["vermelho"]} !important; }}
.lle-kpi-suffix {{
    font-size: 12px;
    font-weight: 500;
    color: {CORES["texto_muted"]} !important;
    margin-top: 4px;
}}

/* v3: sub-stack para Falta Conciliar (vertical) e Investimentos */
.lle-kpi-sub-stack {{
    margin-top: 10px;
    display: flex;
    flex-direction: column;
    gap: 2px;
}}
.lle-kpi-sub-label {{
    font-size: 11px;
    font-weight: 600;
    color: {CORES["branco"]} !important;
    letter-spacing: 0.2px;
    margin-top: 4px;
}}
.lle-kpi-sub-valor {{
    font-size: 14px;
    font-weight: 700;
    color: {CORES["branco"]} !important;
    line-height: 1.2;
}}
.lle-kpi-sub-valor.vermelho {{ color: {CORES["vermelho"]} !important; }}
.lle-kpi-sub-valor.verde {{ color: {CORES["verde"]} !important; }}

/* ===== Painel de bancos (botões) ===== */
.lle-banco-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(260px, 1fr));
    gap: 12px;
}}
.stButton > button.lle-banco-btn {{
    background-color: {CORES["azul_escuro_2"]};
    color: {CORES["branco"]};
    border: 2px solid {CORES["card_borda"]};
    border-radius: 12px;
    padding: 18px 16px;
    font-weight: 700;
    font-size: 14px;
    width: 100%;
    text-align: left;
    transition: all 0.18s ease;
}}

/* Botões primários e gerais — cobre stButton E stDownloadButton */
.stButton > button,
[data-testid="stDownloadButton"] > button,
[data-testid="stBaseButton-secondary"],
[data-testid="stBaseButton-primary"] {{
    background-color: {CORES["azul"]} !important;
    color: {CORES["branco"]} !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    padding: 8px 14px !important;
    min-height: 38px !important;
    font-size: 13.5px !important;
    transition: all 0.18s ease !important;
    box-shadow: 0 2px 6px rgba(0,0,0,0.18) !important;
}}
.stButton > button *,
[data-testid="stDownloadButton"] > button *,
[data-testid="stBaseButton-secondary"] *,
[data-testid="stBaseButton-primary"] * {{
    color: {CORES["branco"]} !important;
}}
.stButton > button:hover,
[data-testid="stDownloadButton"] > button:hover {{
    background-color: {CORES["amarelo"]} !important;
    box-shadow: 0 4px 10px rgba(0,0,0,0.24) !important;
}}
.stButton > button:hover *,
[data-testid="stDownloadButton"] > button:hover * {{
    color: {CORES["azul_escuro"]} !important;
}}
/* Primary = amarelo institucional com texto azul escuro */
.stButton > button[kind="primary"],
[data-testid="stDownloadButton"] > button[kind="primary"],
[data-testid="stBaseButton-primary"] {{
    background-color: {CORES["amarelo"]} !important;
}}
.stButton > button[kind="primary"] *,
[data-testid="stDownloadButton"] > button[kind="primary"] *,
[data-testid="stBaseButton-primary"] * {{
    color: {CORES["azul_escuro"]} !important;
}}
/* Primary HOVER = AZUL com texto branco (item 17) */
.stButton > button[kind="primary"]:hover,
[data-testid="stDownloadButton"] > button[kind="primary"]:hover {{
    background-color: {CORES["azul"]} !important;
}}
.stButton > button[kind="primary"]:hover *,
[data-testid="stDownloadButton"] > button[kind="primary"]:hover * {{
    color: {CORES["branco"]} !important;
}}

/* ===== Inputs e file uploader ===== */
input, textarea, select, [data-baseweb="select"] > div {{
    background-color: {CORES["azul_escuro_2"]} !important;
    color: {CORES["branco"]} !important;
    border-color: {CORES["card_borda"]} !important;
}}
/* v5.35: menu ABERTO do selectbox (popover do BaseWeb). Antes herdava o texto
   branco do tema, mas o popover tem fundo claro → texto branco-no-branco,
   invisível. Força fundo escuro + texto branco, com destaque no hover/selecionado.
   Corrige também o antigo submenu de Tipo (mesmo problema). */
[data-baseweb="popover"] [role="listbox"],
[data-baseweb="popover"] [data-baseweb="menu"],
[data-baseweb="popover"] ul {{
    background-color: {CORES["azul_escuro_2"]} !important;
}}
[data-baseweb="popover"] [role="option"],
[data-baseweb="popover"] li {{
    background-color: {CORES["azul_escuro_2"]} !important;
    color: {CORES["branco"]} !important;
}}
[data-baseweb="popover"] [role="option"] * {{
    color: {CORES["branco"]} !important;
}}
[data-baseweb="popover"] [role="option"]:hover,
[data-baseweb="popover"] [role="option"][aria-selected="true"] {{
    background-color: {CORES["card_bg"]} !important;
    color: {CORES["amarelo"]} !important;
}}
[data-baseweb="popover"] [role="option"]:hover *,
[data-baseweb="popover"] [role="option"][aria-selected="true"] * {{
    color: {CORES["amarelo"]} !important;
}}
.stTextInput input, .stDateInput input {{
    background-color: {CORES["azul_escuro_2"]} !important;
    color: {CORES["branco"]} !important;
    border: 1px solid {CORES["card_borda"]} !important;
}}
[data-testid="stFileUploaderDropzone"] {{
    background-color: {CORES["azul_escuro_2"]} !important;
    border: 2px dashed {CORES["card_borda"]} !important;
}}
[data-testid="stFileUploaderDropzone"]:hover {{
    border-color: {CORES["amarelo"]} !important;
}}
/* v5.37: ANTES pintava TUDO de branco (inclusive o nome do arquivo, que fica
   numa caixa branca → sumia). Agora só as INSTRUÇÕES ("arraste", "200MB") ficam
   brancas; o nome do arquivo mantém a cor escura padrão e aparece na caixa. */
[data-testid="stFileUploaderDropzoneInstructions"],
[data-testid="stFileUploaderDropzoneInstructions"] * {{ color: {CORES["branco"]} !important; }}

/* v5.37/v5.38: menu da coluna do st.dataframe (Sort/filtro) e menus em geral —
   vinham claro-no-claro (texto branco em fundo branco). Cobre vários tipos. */
[data-testid="stDataFrameColumnMenu"],
[data-baseweb="menu"],
[data-baseweb="popover"] [role="menu"],
[role="menu"] {{
    background-color: {CORES["azul_escuro_2"]} !important;
    border: 1px solid {CORES["card_borda"]} !important;
}}
[data-testid="stDataFrameColumnMenu"] *,
[data-baseweb="menu"] *,
[data-baseweb="menu"] [role="option"],
[role="menu"] *,
[role="menuitem"],
[role="menuitem"] * {{
    color: {CORES["branco"]} !important;
}}
[data-testid="stDataFrameColumnMenu"] input,
[data-baseweb="menu"] input,
[role="menu"] input {{
    color: {CORES["branco"]} !important;
    background-color: {CORES["azul_escuro_2"]} !important;
    border: 1px solid {CORES["card_borda"]} !important;
}}

/* v3.4: tooltip do help (?) — texto PRETO em qualquer lugar (sidebar amarela ou main) */
[data-testid="stTooltipContent"],
[data-testid="stTooltipContent"] *,
[role="tooltip"],
[role="tooltip"] * {{
    color: {CORES["azul_escuro"]} !important;
    background-color: #FFFFFF !important;
}}
[data-testid="stTooltipContent"] {{
    border: 1px solid {CORES["azul_escuro"]} !important;
    border-radius: 8px !important;
}}

/* v5.6: o input do number_input estava com texto invisível na sidebar (texto
   escuro sobre fundo escuro). Força cor branca com seletores específicos. */
[data-testid="stNumberInput"] input,
[data-testid="stNumberInputContainer"] input,
input[type="number"] {{
    background-color: {CORES["azul_escuro_2"]} !important;
    color: {CORES["branco"]} !important;
    -webkit-text-fill-color: {CORES["branco"]} !important;
}}

/* v5.6: checkbox da sidebar — o "check" estava invisível */
[data-testid="stCheckbox"] svg {{
    color: {CORES["branco"]} !important;
    fill: {CORES["branco"]} !important;
}}

/* v3.4: botões + / - do number_input — texto BRANCO sobre fundo azul.
   No sidebar amarela, esses botões herdavam azul-escuro do '*' geral. */
[data-testid="stNumberInputContainer"] button,
[data-testid="stNumberInput"] button {{
    background-color: {CORES["azul"]} !important;
    color: {CORES["branco"]} !important;
    border-color: {CORES["azul"]} !important;
}}
[data-testid="stNumberInputContainer"] button *,
[data-testid="stNumberInput"] button *,
[data-testid="stNumberInputContainer"] button svg,
[data-testid="stNumberInput"] button svg {{
    color: {CORES["branco"]} !important;
    fill: {CORES["branco"]} !important;
}}
[data-testid="stNumberInputContainer"] button:hover,
[data-testid="stNumberInput"] button:hover {{
    background-color: {CORES["azul_escuro"]} !important;
}}

/* Radio horizontal */
[role="radiogroup"] label {{ color: {CORES["branco"]} !important; }}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {{
    gap: 4px;
    background-color: {CORES["azul_escuro_2"]};
    padding: 6px;
    border-radius: 12px;
    border: 1px solid {CORES["card_borda"]};
}}
.stTabs [data-baseweb="tab"] {{
    font-weight: 600;
    color: {CORES["texto_muted"]} !important;
    background-color: transparent;
    border-radius: 8px;
    padding: 8px 16px;
}}
.stTabs [aria-selected="true"] {{
    color: {CORES["azul_escuro"]} !important;
    background-color: {CORES["amarelo"]} !important;
}}

/* Dataframes */
[data-testid="stDataFrame"] {{
    background-color: {CORES["card_bg"]};
    border: 1px solid {CORES["card_borda"]};
    border-radius: 10px;
    padding: 4px;
}}

/* Alerts */
.stAlert {{ background-color: {CORES["azul_escuro_2"]} !important; border: 1px solid {CORES["card_borda"]}; }}
.stAlert * {{ color: {CORES["branco"]} !important; }}

/* Divider */
hr {{ border-color: {CORES["card_borda"]} !important; }}

/* Expander */
.streamlit-expanderHeader {{
    background-color: {CORES["azul_escuro_2"]} !important;
    color: {CORES["branco"]} !important;
    border: 1px solid {CORES["card_borda"]} !important;
}}

/* Caption */
.stCaption, [data-testid="stCaptionContainer"] {{ color: {CORES["texto_muted"]} !important; }}

/* Subheader chips */
.lle-section-title {{
    display: inline-block;
    background-color: {CORES["amarelo"]};
    color: {CORES["azul_escuro"]} !important;
    padding: 6px 14px;
    border-radius: 999px;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.5px;
    margin-bottom: 14px;
}}

/* Footer */
.lle-footer {{
    margin-top: 56px;
    padding: 18px 24px;
    text-align: center;
    font-size: 12px;
    color: {CORES["texto_muted"]} !important;
    border-top: 1px solid {CORES["card_borda"]};
}}
.lle-footer a {{ color: {CORES["amarelo"]} !important; text-decoration: none; }}

/* Tabela de detalhamento — cabeçalho mais legível */
table {{ color: {CORES["branco"]}; }}

/* Status badges */
.lle-badge {{
    display: inline-block;
    padding: 3px 10px;
    border-radius: 999px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.4px;
}}
.lle-badge.verde {{ background: {CORES["verde"]}; color: {CORES["branco"]}; }}
.lle-badge.amarelo {{ background: {CORES["amarelo"]}; color: {CORES["azul_escuro"]}; }}
.lle-badge.vermelho {{ background: {CORES["vermelho"]}; color: {CORES["branco"]}; }}
.lle-badge.azul {{ background: {CORES["azul"]}; color: {CORES["branco"]}; }}

/* ============================================================
   v5.12 — Estilo EDITORIAL (Fase 1: Resumo Executivo)
   Componentes novos com prefixo .editorial-*. Convivem com o CSS antigo.
   Princípios: tipografia leve (300/400/500), sem cards, divisores finos,
   cor cirúrgica.
   ============================================================ */

/* Header do resumo (kicker + título + data) */
.editorial-header {{
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    margin-bottom: 8px;
}}
.editorial-kicker {{
    color: {CORES["amarelo"]};
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 3px;
    text-transform: uppercase;
}}
.editorial-title {{
    color: {CORES["branco"]};
    font-size: 22px;
    font-weight: 300;
    margin-top: 6px;
    letter-spacing: -0.3px;
}}
.editorial-meta {{
    color: #6B88B5;
    font-size: 11px;
    font-weight: 400;
    letter-spacing: 0.5px;
    text-align: right;
}}
.editorial-meta + .editorial-meta {{ margin-top: 2px; }}

/* Divisor temático com fade dourado */
.editorial-divisor-fade {{
    height: 1px;
    background: linear-gradient(90deg,
        rgba(250,195,24,0.4) 0%,
        rgba(250,195,24,0.05) 40%,
        transparent 100%);
    margin: 20px 0 36px;
}}

/* Bloco do KPI âncora (% + valor) */
.editorial-destaque-row {{
    display: flex;
    align-items: flex-end;
    gap: 28px;
    margin-bottom: 8px;
}}
.editorial-destaque-col {{ flex: 1; }}
.editorial-destaque-col.bigger {{
    flex: 1.4;
    padding-left: 28px;
    border-left: 1px solid rgba(255,255,255,0.08);
}}
.editorial-label {{
    color: #6B88B5;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 2.5px;
    text-transform: uppercase;
}}
.editorial-valor-hero {{
    color: {CORES["branco"]};
    font-size: 52px;
    font-weight: 300;
    letter-spacing: -1.5px;
    line-height: 1;
    margin-top: 10px;
}}
.editorial-valor-hero .unit {{
    color: #6B88B5;
    font-size: 28px;
    font-weight: 300;
}}
.editorial-valor-grande {{
    color: {CORES["branco"]};
    font-size: 38px;
    font-weight: 300;
    letter-spacing: -1px;
    margin-top: 10px;
    line-height: 1;
}}
.editorial-valor-grande .destaque {{ font-weight: 400; }}
.editorial-valor-grande .centavos {{ color: #6B88B5; }}
.editorial-sub-text {{
    color: #6B88B5;
    font-size: 12px;
    margin-top: 8px;
}}
.editorial-status-pill {{
    font-size: 12px;
    font-weight: 500;
    padding-bottom: 6px;
    display: inline-flex;
    align-items: center;
    gap: 6px;
}}
.editorial-status-pill.verde {{ color: #7DD87D; }}
.editorial-status-pill.amarelo {{ color: {CORES["amarelo"]}; }}
.editorial-status-pill.vermelho {{ color: #FF8A8A; }}

/* Barra de progresso editorial */
.editorial-progresso-wrap {{ margin: 36px 0; }}
.editorial-progresso-track {{
    height: 6px;
    background: rgba(255,255,255,0.04);
    border-radius: 3px;
    position: relative;
    overflow: hidden;
}}
.editorial-progresso-fill {{
    position: absolute;
    left: 0;
    top: 0;
    height: 100%;
    background: linear-gradient(90deg, {CORES["amarelo"]} 0%, {CORES["amarelo_2"]} 100%);
    border-radius: 3px;
}}
.editorial-progresso-legend {{
    display: flex;
    justify-content: space-between;
    margin-top: 10px;
    font-size: 10px;
    color: #6B88B5;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}

/* Linha de métricas em colunas com divisor */
.editorial-cols {{
    display: grid;
    grid-template-columns: repeat(var(--cols, 4), 1fr);
    gap: 0;
    margin: 48px 0 36px;
}}
.editorial-cols .col {{
    padding: 0 20px;
    border-right: 1px solid rgba(255,255,255,0.08);
}}
.editorial-cols .col:first-child {{ padding-left: 0; }}
.editorial-cols .col:last-child {{
    padding-right: 0;
    border-right: none;
}}
.editorial-cols .label {{
    color: #6B88B5;
    font-size: 9px;
    font-weight: 500;
    letter-spacing: 2px;
    text-transform: uppercase;
}}
.editorial-cols .valor {{
    color: {CORES["branco"]};
    font-size: 24px;
    font-weight: 300;
    margin-top: 8px;
    letter-spacing: -0.5px;
}}
.editorial-cols .valor.verde {{ color: #7DD87D; }}
.editorial-cols .valor.amarelo {{ color: {CORES["amarelo"]}; }}
.editorial-cols .valor.vermelho {{ color: #FF8A8A; }}
.editorial-cols .valor.atencao {{ color: #FFD6A0; }}
.editorial-cols .valor .centavos {{ color: #6B88B5; font-size: 14px; }}
.editorial-cols .sub {{
    color: #6B88B5;
    font-size: 10px;
    margin-top: 10px;
}}
.editorial-cols .sub .verde {{ color: #7DD87D; }}
.editorial-cols .sub .vermelho {{ color: #FF8A8A; }}
.editorial-cols .sub .inline {{ margin-right: 10px; }}

/* Faixa horizontal de contagens (entre divisores horizontais) */
.editorial-faixa {{
    display: flex;
    align-items: center;
    gap: 20px;
    padding: 18px 0;
    border-top: 1px solid rgba(255,255,255,0.06);
    border-bottom: 1px solid rgba(255,255,255,0.06);
    margin-bottom: 36px;
}}
.editorial-faixa .item {{
    flex: 1;
    text-align: center;
}}
.editorial-faixa .item .valor {{
    color: {CORES["branco"]};
    font-size: 18px;
    font-weight: 300;
}}
.editorial-faixa .item .valor.verde {{ color: #7DD87D; }}
.editorial-faixa .item .valor.amarelo {{ color: {CORES["amarelo"]}; }}
.editorial-faixa .item .label {{
    color: #6B88B5;
    font-size: 9px;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-top: 2px;
}}
.editorial-faixa .sep {{
    width: 1px;
    height: 24px;
    background: rgba(255,255,255,0.08);
}}

/* Cabeçalho de seção editorial (com linha que se esvai) */
.editorial-secao-head {{
    display: flex;
    align-items: center;
    gap: 14px;
    margin-bottom: 18px;
}}
.editorial-secao-head .titulo {{
    color: {CORES["amarelo"]};
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 2.5px;
    text-transform: uppercase;
}}
.editorial-secao-head .linha {{
    flex: 1;
    height: 1px;
    background: linear-gradient(90deg, rgba(250,195,24,0.2) 0%, transparent 100%);
}}
.editorial-secao-head .contagem {{
    color: #6B88B5;
    font-size: 10px;
}}

/* Lista de exceções (sem cards) */
.editorial-lista {{
    display: flex;
    flex-direction: column;
    gap: 14px;
}}
.editorial-lista-item {{
    display: flex;
    align-items: center;
    gap: 18px;
    padding: 6px 0;
}}
.editorial-lista-item.muted {{ opacity: 0.5; }}
.editorial-lista-divisor {{
    height: 1px;
    background: rgba(255,255,255,0.04);
}}
.editorial-lista-icone {{
    width: 32px;
    height: 32px;
    border-radius: 50%;
    border: 1px solid rgba(107,136,181,0.3);
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
    color: #6B88B5;
}}
.editorial-lista-icone.verde {{
    border-color: rgba(125,216,125,0.3);
    color: #7DD87D;
}}
.editorial-lista-icone.amarelo {{
    border-color: rgba(250,195,24,0.3);
    color: {CORES["amarelo"]};
}}
.editorial-lista-corpo {{ flex: 1; }}
.editorial-lista-titulo {{
    color: {CORES["branco"]};
    font-size: 13px;
    font-weight: 400;
}}
.editorial-lista-titulo .light {{ color: #6B88B5; font-weight: 300; }}
.editorial-lista-sub {{
    color: #6B88B5;
    font-size: 11px;
    margin-top: 2px;
}}
.editorial-lista-direita {{ text-align: right; }}
.editorial-lista-valor {{
    color: {CORES["branco"]};
    font-size: 15px;
    font-weight: 300;
}}
.editorial-lista-valor.muted {{ color: #6B88B5; }}
.editorial-lista-status {{
    font-size: 9px;
    font-weight: 500;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-top: 2px;
}}
.editorial-lista-status.verde {{ color: #7DD87D; }}
.editorial-lista-status.amarelo {{ color: {CORES["amarelo"]}; }}

/* ============================================================
   v5.13 — Estilo OPÇÃO A (cards generosos)
   Cards modernos com hierarquia visual, gradientes sutis,
   bordas superiores semânticas, donut SVG pro %.
   ============================================================ */

/* Grid principal: 3 colunas com proporção 1.4 / 1 / 1 (card âncora maior) */
.opa-grid-hero {{
    display: grid;
    grid-template-columns: 1.4fr 1fr 1fr;
    gap: 16px;
    margin-bottom: 16px;
}}
.opa-grid-secundario {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 16px;
}}
@media (max-width: 900px) {{
    .opa-grid-hero, .opa-grid-secundario {{
        grid-template-columns: 1fr;
    }}
}}

/* Card âncora (gradiente + glow sutil no canto) */
.opa-card-ancora {{
    background: linear-gradient(135deg, #16335F 0%, #0F2548 100%);
    padding: 24px 22px;
    border-radius: 18px;
    position: relative;
    overflow: hidden;
}}
.opa-card-ancora::before {{
    content: "";
    position: absolute;
    top: 0;
    right: 0;
    width: 140px;
    height: 140px;
    background: radial-gradient(circle, rgba(250,195,24,0.08) 0%, transparent 70%);
    pointer-events: none;
}}
.opa-card-ancora-label {{
    color: #8BA3C7;
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}
.opa-card-ancora-valor {{
    color: {CORES["branco"]};
    font-size: 30px;
    font-weight: 500;
    margin-top: 14px;
    letter-spacing: -0.5px;
}}
.opa-rec-desp-row {{
    display: flex;
    gap: 28px;
    margin-top: 18px;
    padding-top: 14px;
    border-top: 1px solid rgba(255,255,255,0.08);
}}
.opa-rec-desp-item .marker {{
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 11px;
    font-weight: 500;
}}
.opa-rec-desp-item .marker.verde {{ color: #7DD87D; }}
.opa-rec-desp-item .marker.vermelho {{ color: #FF8A8A; }}
.opa-rec-desp-item .marker .dot {{
    width: 6px;
    height: 6px;
    border-radius: 50%;
}}
.opa-rec-desp-item .marker.verde .dot {{ background: #7DD87D; }}
.opa-rec-desp-item .marker.vermelho .dot {{ background: #FF8A8A; }}
.opa-rec-desp-item .valor {{
    color: {CORES["branco"]};
    font-size: 15px;
    font-weight: 500;
    margin-top: 4px;
}}

/* Card verde translúcido (Conciliado) */
.opa-card-conciliado {{
    background: linear-gradient(135deg, rgba(125,216,125,0.12) 0%, rgba(125,216,125,0.04) 100%);
    padding: 24px 22px;
    border-radius: 18px;
    border: 1px solid rgba(125,216,125,0.18);
}}
.opa-card-conciliado-label {{
    color: #A8D8A8;
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}
.opa-card-conciliado-valor {{
    color: #7DD87D;
    font-size: 30px;
    font-weight: 500;
    margin-top: 14px;
    letter-spacing: -0.5px;
}}
.opa-card-conciliado-sub {{
    color: #8BA3C7;
    font-size: 12px;
    margin-top: 14px;
    padding-top: 14px;
    border-top: 1px solid rgba(255,255,255,0.06);
}}
.opa-card-conciliado-sub i {{
    color: #7DD87D;
    vertical-align: -2px;
    margin-right: 4px;
}}

/* Card donut (percentual) */
.opa-card-donut {{
    background: #0F2548;
    padding: 24px 22px;
    border-radius: 18px;
    border: 1px solid rgba(250,195,24,0.2);
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    position: relative;
    min-height: 160px;
}}
.opa-card-donut svg {{
    position: absolute;
    top: 50%;
    left: 50%;
    transform: translate(-50%, -50%);
}}
.opa-card-donut-text {{
    position: relative;
    z-index: 1;
    text-align: center;
}}
.opa-card-donut-pct {{
    color: {CORES["amarelo"]};
    font-size: 30px;
    font-weight: 500;
    letter-spacing: -0.5px;
}}
.opa-card-donut-label {{
    color: #8BA3C7;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    margin-top: 2px;
}}

/* Cards secundários (Falta, Divergência, Contas) */
.opa-card-secundario {{
    background: #0F2548;
    padding: 18px 20px;
    border-radius: 14px;
    border-top: 2px solid #888;
}}
.opa-card-secundario.vermelho {{ border-top-color: #FF6B6B; }}
.opa-card-secundario.amarelo {{ border-top-color: {CORES["amarelo"]}; }}
.opa-card-secundario.verde {{ border-top-color: #7DD87D; }}

.opa-card-sec-head {{
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
}}
.opa-card-sec-label {{
    color: #8BA3C7;
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.6px;
    text-transform: uppercase;
}}
.opa-card-sec-valor {{
    color: {CORES["branco"]};
    font-size: 24px;
    font-weight: 500;
    margin-top: 8px;
    letter-spacing: -0.3px;
}}
.opa-card-sec-valor.vermelho {{ color: #FF8A8A; }}
.opa-card-sec-valor.amarelo {{ color: {CORES["amarelo"]}; }}
.opa-card-sec-valor.verde {{ color: #7DD87D; }}

.opa-card-sec-icon {{
    padding: 6px;
    border-radius: 8px;
}}
.opa-card-sec-icon.vermelho {{ background: rgba(255,107,107,0.12); color: #FF8A8A; }}
.opa-card-sec-icon.amarelo {{ background: rgba(250,195,24,0.12); color: {CORES["amarelo"]}; }}
.opa-card-sec-icon.verde {{ background: rgba(125,216,125,0.12); color: #7DD87D; }}
.opa-card-sec-icon i {{ font-size: 14px; }}

.opa-card-sec-sub {{
    font-size: 11px;
    color: #8BA3C7;
    margin-top: 14px;
}}
.opa-card-sec-sub .label {{ color: #8BA3C7; }}
.opa-card-sec-sub .verde-text {{ color: #7DD87D; }}
.opa-card-sec-sub .vermelho-text {{ color: #FF8A8A; }}
.opa-card-sec-sub .item {{ display: inline-block; margin-right: 14px; }}

/* Seção Exceções (chips horizontais leves) */
.opa-excecoes-head {{
    display: flex;
    gap: 10px;
    align-items: center;
    margin: 28px 0 14px;
    padding-top: 20px;
    border-top: 1px dashed rgba(255,255,255,0.08);
}}
.opa-excecoes-head-label {{
    color: #8BA3C7;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}
.opa-excecoes-head-line {{
    flex: 1;
    height: 1px;
    background: rgba(255,255,255,0.06);
}}

.opa-excecoes-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 12px;
}}

.opa-chip {{
    display: flex;
    gap: 12px;
    align-items: center;
    padding: 12px 14px;
    border-radius: 12px;
    background: rgba(139,163,199,0.06);
}}
.opa-chip.verde {{ background: rgba(125,216,125,0.06); }}
.opa-chip.amarelo {{ background: rgba(250,195,24,0.06); }}
.opa-chip-icon {{
    width: 34px;
    height: 34px;
    border-radius: 50%;
    background: rgba(139,163,199,0.15);
    color: #8BA3C7;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
}}
.opa-chip-icon.verde {{ background: rgba(125,216,125,0.15); color: #7DD87D; }}
.opa-chip-icon.amarelo {{ background: rgba(250,195,24,0.15); color: {CORES["amarelo"]}; }}
.opa-chip-icon i {{ font-size: 16px; }}
.opa-chip-label {{
    color: #8BA3C7;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.6px;
    text-transform: uppercase;
}}
.opa-chip-valor {{
    color: {CORES["branco"]};
    font-size: 16px;
    font-weight: 500;
    margin-top: 2px;
}}
.opa-chip-valor .light {{
    color: #8BA3C7;
    font-size: 11px;
    font-weight: 400;
}}

</style>
"""
)


# ============================================================
# Sidebar — navegação
# ============================================================
with st.sidebar:
    if LOGO_URI:
        st.html(
            f"""
            <div class="lle-sidebar-logo">
                <img src="{LOGO_URI}" alt="Grupo LLE" />
                <div class="lle-sidebar-tagline">CONCILIAÇÃO BANCÁRIA</div>
            </div>
            """
        )

    # Páginas principais (lista simples)
    paginas_top = [
        ("📊 Dashboard", "Dashboard"),
        ("✅ Conciliação", "Conciliação"),
    ]
    for label, key in paginas_top:
        is_atual = st.session_state.pagina == key
        st.button(
            label,
            key=f"nav_{key}",
            on_click=ir_para,
            args=(key,),
            type="primary" if is_atual else "secondary",
            use_container_width=True,
        )

    # v5.4: Conta 70 logo após Conciliação (antes do CARTÃO)
    conta70_atual = st.session_state.pagina == "Conta 70"
    st.button(
        "📒 Conta 70",
        key="nav_Conta 70",
        on_click=ir_para,
        args=("Conta 70",),
        type="primary" if conta70_atual else "secondary",
        use_container_width=True,
    )

    # v5.5: Auditoria volta pra dentro do submenu CARTÃO
    cartao_atual = st.session_state.pagina in ("Cadastro de Taxas", "Auditoria de Taxas")
    with st.expander("💳 CARTÃO", expanded=cartao_atual):
        submenus_cartao = [
            ("🏦 Cadastro de Taxas", "Cadastro de Taxas"),
            ("💳 Auditoria de Cartões", "Auditoria de Taxas"),
        ]
        for label, key in submenus_cartao:
            is_atual = st.session_state.pagina == key
            st.button(
                label,
                key=f"nav_{key}",
                on_click=ir_para,
                args=(key,),
                type="primary" if is_atual else "secondary",
                use_container_width=True,
            )

    # Restante das páginas principais
    paginas_bot = [
        ("📂 Histórico", "Histórico"),
        ("🟢 Sobre", "Sobre"),
    ]
    for label, key in paginas_bot:
        is_atual = st.session_state.pagina == key
        st.button(
            label,
            key=f"nav_{key}",
            on_click=ir_para,
            args=(key,),
            type="primary" if is_atual else "secondary",
            use_container_width=True,
        )

    st.divider()

    with st.expander("⚙️ Configurações"):
        rodar_fuzzy = st.checkbox(
            "Gerar sugestões fuzzy",
            value=True,
            help="Aba complementar de revisão manual. Não entra na conciliação automática.",
        )
        tolerancia = st.number_input(
            "Tolerância de data (dias)",
            min_value=0, max_value=10, value=2, step=1,
            help="Aceita diferença de até N dias entre o lançamento no banco e no sistema (fim de semana / feriado).",
        )


# ============================================================
# Header (sempre presente)
# ============================================================
PAGE_INFO = {
    "Dashboard": ("Dashboard", "Visão executiva da última conciliação"),
    "Conciliação": ("Conciliação", "Upload, processamento e detalhamento"),
    "Cadastro de Taxas": ("Cadastro de Taxas", "Contratos com adquirentes de cartão"),
    "Auditoria de Taxas": ("Auditoria de Cartões", "Comparação entre taxa contratada e aplicada"),
    "Conta 70": ("Conta 70", "Controle provisório de créditos bancários não identificados"),
    "Histórico": ("Histórico", "Execuções e reprocessamentos"),
    "Sobre": ("Sobre o sistema", "Como funciona e regras de negócio"),
}
titulo, subtitulo = PAGE_INFO.get(st.session_state.pagina, ("Conciliação", ""))

st.html(
    f"""
    <div class="lle-header">
        {'<img src="' + LOGO_URI + '" alt="Grupo LLE"/>' if LOGO_URI else ''}
        <div>
            <div class="lle-title">{titulo}</div>
            <div class="lle-subtitle">{subtitulo}</div>
        </div>
    </div>
    """
)


# ============================================================
# Helpers de formatação
# ============================================================
def fmt_brl(v: float) -> str:
    if v is None or pd.isna(v):
        return "R$ 0,00"
    sinal = "-" if v < 0 else ""
    return f"{sinal}R$ {abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_int(v) -> str:
    if v is None or pd.isna(v):
        return "0"
    return f"{int(v):,}".replace(",", ".")


def fmt_pct(v: float) -> str:
    if v is None or pd.isna(v):
        return "0,0%"
    return f"{v:.1f}%".replace(".", ",")


def card_kpi(label: str, valor: str, suffix: str = "", classe: str = "") -> str:
    return f"""
    <div class="lle-kpi">
        <div class="lle-kpi-label">{label}</div>
        <div class="lle-kpi-value {classe}">{valor}</div>
        {f'<div class="lle-kpi-suffix">{suffix}</div>' if suffix else ''}
    </div>
    """


def card_kpi_html(label: str, valor: str, suffix_html: str = "", classe: str = "") -> str:
    """Versão de card_kpi onde o suffix é HTML literal (pra blocos verticais)."""
    return f"""
    <div class="lle-kpi">
        <div class="lle-kpi-label">{label}</div>
        <div class="lle-kpi-value {classe}">{valor}</div>
        {suffix_html}
    </div>
    """


def _card_falta_conciliar_vertical(receitas: float, despesas: float) -> str:
    """v3: Falta Conciliar com receitas e despesas EMPILHADAS, ambos em vermelho."""
    return f"""
    <div class="lle-kpi-sub-stack">
        <div class="lle-kpi-sub-label">Receitas:</div>
        <div class="lle-kpi-sub-valor vermelho">{fmt_brl(receitas)}</div>
        <div class="lle-kpi-sub-label">Despesas:</div>
        <div class="lle-kpi-sub-valor vermelho">{fmt_brl(despesas)}</div>
    </div>
    """


def _card_total_com_rec_desp(receitas: float, despesas: float) -> str:
    """v3.1: bloco com Receitas (verde) e Despesas (vermelho) empilhadas embaixo do valor total.
    Usado em 'Total Movimentado no Banco' e 'Total Extrato Sankhya'."""
    return f"""
    <div class="lle-kpi-sub-stack">
        <div class="lle-kpi-sub-label">Receitas:</div>
        <div class="lle-kpi-sub-valor verde">{fmt_brl(receitas)}</div>
        <div class="lle-kpi-sub-label">Despesas:</div>
        <div class="lle-kpi-sub-valor vermelho">{fmt_brl(despesas)}</div>
    </div>
    """


def _card_investimentos(resultado: ResultadoConciliacao) -> str:
    """Card único de Investimentos (Aplicações + Resgates) — substitui 2 cards separados."""
    return _card_investimentos_de_df(resultado.aplicacoes_resgates)


def _card_investimentos_da_conta(resultado: ResultadoConciliacao, conta: str) -> str:
    """Versão filtrada por conta."""
    return _card_investimentos_de_df(resultado.aplicacoes_resgates_da_conta(conta))


def _card_investimentos_de_df(df: pd.DataFrame) -> str:
    """v5.27: card Investimentos simplificado, mesmo tamanho dos outros cards.

    Mudanças desta versão:
    - Card mostra saldo líquido + 2 linhas (Aplicações e Resgates).
    - Rendimentos e alerta de descasamento vão pra dentro da aba 'Aplicações e
      Resgates' — não poluem o card principal.

    DEDUP banco × sankhya: prefere Sankhya quando ambos existem.
    """
    if df.empty:
        return card_kpi("Posição de Investimentos", "—", "sem aplicações/resgates")

    df = df.copy()

    # Filtro defensivo: SALDO não é aplicação nem resgate
    if "historico" in df.columns:
        mask_saldo = df["historico"].astype(str).str.upper().str.contains("SALDO", na=False)
        df = df[~mask_saldo]

    if df.empty:
        return card_kpi("Posição de Investimentos", "—", "sem aplicações/resgates")

    # DEDUP banco × sankhya: preferir Sankhya quando ambos existem
    cols_chave = [c for c in ["data", "valor", "conta", "tipo_aplicacao"] if c in df.columns]
    if cols_chave and "origem" in df.columns:
        df["_ord_origem"] = df["origem"].apply(lambda x: 0 if "Sankhya" in str(x) else 1)
        df = df.sort_values("_ord_origem").drop_duplicates(subset=cols_chave, keep="first")
        df = df.drop(columns=["_ord_origem"])

    aplic = df[df["tipo_aplicacao"] == "Aplicação"] if "tipo_aplicacao" in df.columns else pd.DataFrame()
    resg = df[df["tipo_aplicacao"] == "Resgate"] if "tipo_aplicacao" in df.columns else pd.DataFrame()

    qtd_a = len(aplic)
    val_a = float(aplic["valor"].abs().sum()) if not aplic.empty else 0.0
    qtd_r = len(resg)
    val_r = float(resg["valor"].abs().sum()) if not resg.empty else 0.0

    # Saldo líquido = Resgates - Aplicações (faz sentido contábil)
    saldo_liquido = val_r - val_a

    # v5.45: card enxuto — o "por que não é prejuízo" já está em "Entenda seus
    # cards"; aqui fica só o fato curto + um lembrete apontando pro detalhamento.
    if saldo_liquido < -0.005:
        _status = "Aplicou " + fmt_brl(abs(saldo_liquido)) + " a mais que resgatou"
    elif saldo_liquido > 0.005:
        _status = "Resgatou " + fmt_brl(saldo_liquido) + " a mais que aplicou no mês"
    else:
        _status = "Aplicações e resgates se equilibraram no mês"

    sub = f"""
    <div class="lle-kpi-sub-stack">
        <div class="lle-kpi-sub-label">Aplicações:</div>
        <div class="lle-kpi-sub-valor">{fmt_int(qtd_a)} mov. · {fmt_brl(val_a)}</div>
        <div class="lle-kpi-sub-label">Resgates:</div>
        <div class="lle-kpi-sub-valor">{fmt_int(qtd_r)} mov. · {fmt_brl(val_r)}</div>
        <div class="lle-kpi-sub-label" style="margin-top:6px;">{_status}</div>
        <div class="lle-kpi-sub-label" style="margin-top:2px; opacity:.65;">ℹ️ O que significa? veja em “Entenda seus cards”.</div>
    </div>
    """
    return card_kpi_html("Posição de Investimentos", fmt_brl(saldo_liquido), sub, classe="destaque-amarelo")


def render_cards(cards: list[str]):
    html = '<div class="lle-kpi-row">' + "".join(cards) + "</div>"
    st.html(html)


def section_title(texto: str):
    st.html(f'<div class="lle-section-title">{texto}</div>')


def render_secao_excecoes_regras(resultado, kpis: dict, conta: str | None = None):
    """v5.12: Seção 'EXCEÇÕES E REGRAS APLICADAS' — agrupa os cards de tratamento
    especial (Estornos, Cartão TOP 1722, Investimentos) que antes poluíam o Resumo
    Executivo. Só renderiza se houver pelo menos uma exceção com dados.

    Args:
        resultado: ResultadoConciliacao
        kpis: dict de KPIs (globais ou da conta)
        conta: se informado, filtra Investimentos pela conta
    """
    qtd_est_anu = kpis.get("qtd_estornos_anulados", 0)
    qtd_est_par = kpis.get("qtd_estornos_parciais", 0)
    qtd_top1722 = kpis.get("qtd_top1722_grupos", 0)
    qtd_top1702 = kpis.get("qtd_top1702_grupos", 0)

    # Investimentos: pega global ou por conta
    if conta:
        df_inv = resultado.aplicacoes_resgates_da_conta(conta)
    else:
        df_inv = resultado.aplicacoes_resgates
    tem_invest = not df_inv.empty

    # Só renderiza a seção se houver pelo menos uma exceção
    if not (qtd_est_anu > 0 or qtd_est_par > 0 or qtd_top1722 > 0 or qtd_top1702 > 0 or tem_invest):
        return

    section_title("EXCEÇÕES E REGRAS APLICADAS")
    st.caption(
        "Tratamentos especiais aplicados pelo sistema durante a conciliação — "
        "estes valores já foram considerados nos totais do Resumo Executivo."
    )

    cards = []
    if qtd_est_anu > 0:
        cards.append(card_kpi(
            "♻️ Anulados por Estorno", fmt_int(qtd_est_anu),
            f"valor bruto: {fmt_brl(kpis.get('valor_estornos_anulados', 0.0))}",
            classe="destaque-verde",
        ))
    if qtd_est_par > 0:
        cards.append(card_kpi(
            "⚖️ Estornos Parciais", fmt_int(qtd_est_par),
            f"saldo restante: {fmt_brl(kpis.get('saldo_estornos_parciais', 0.0))}",
            classe="destaque-amarelo",
        ))
    if qtd_top1722 > 0:
        cards.append(card_kpi(
            "🃏 Cartão TOP 1722", fmt_int(qtd_top1722),
            f"valor: {fmt_brl(kpis.get('valor_top1722_conciliado', 0.0))}",
            classe="destaque-verde",
        ))
    if qtd_top1702 > 0:
        _dif1702 = getattr(resultado, "top1702_diferencas", pd.DataFrame())
        _difv = float(_dif1702["diferenca"].sum()) if (not _dif1702.empty and "diferenca" in _dif1702.columns) else 0.0
        _sub = f"valor: {fmt_brl(kpis.get('valor_top1702_conciliado', 0.0))}"
        if abs(_difv) >= 0.005:
            _sub += f" · dif {fmt_brl(_difv)}"
        cards.append(card_kpi(
            "🎫 Boleto TOP 1702", fmt_int(qtd_top1702),
            _sub,
            classe="destaque-amarelo" if abs(_difv) >= 0.005 else "destaque-verde",
        ))
    if tem_invest:
        cards.append(_card_investimentos_de_df(df_inv))

    # Completa com cards vazios pra manter grid de 4
    while len(cards) % 4 != 0:
        cards.append(card_kpi("", "", ""))

    render_cards(cards)
    st.divider()


# ============================================================
# v5.12 — Componentes EDITORIAIS (Fase 1: Resumo Executivo)
# Convivem com os helpers antigos. Outras telas continuam usando os antigos.
# ============================================================

def _split_centavos(valor: float) -> tuple[str, str]:
    """Quebra 874388.20 em ('R$ 874.388', ',20'). Usado pros valores com centavos
    em cinza no estilo editorial."""
    s = fmt_brl(valor)  # 'R$ 874.388,20'
    if "," in s:
        inteiro, cents = s.rsplit(",", 1)
        return inteiro, f",{cents}"
    return s, ""


def editorial_header(kicker: str, titulo: str, meta_esquerda: str = "", meta_direita: str = ""):
    """Cabeçalho editorial: kicker em amarelo + título grande + meta à direita."""
    meta_html = ""
    if meta_esquerda or meta_direita:
        m1 = f'<div class="editorial-meta">{meta_esquerda}</div>' if meta_esquerda else ""
        m2 = f'<div class="editorial-meta">{meta_direita}</div>' if meta_direita else ""
        meta_html = f'<div>{m1}{m2}</div>'
    st.html(
        f"""
        <div class="editorial-header">
            <div>
                <div class="editorial-kicker">{kicker}</div>
                <div class="editorial-title">{titulo}</div>
            </div>
            {meta_html}
        </div>
        <div class="editorial-divisor-fade"></div>
        """
    )


def editorial_kpi_destaque(
    percentual: float,
    valor_conciliado: float,
    valor_total: float,
    label_pct: str = "Conciliado",
    label_valor: str = "Valor conciliado",
    status_text: str | None = None,
    status_cor: str = "verde",  # 'verde' | 'amarelo' | 'vermelho'
):
    """KPI âncora gigante: percentual (52px) + valor (38px) lado a lado."""
    pct_str = f"{percentual:.1f}".replace(".", ",")
    inteiro, cents = _split_centavos(valor_conciliado)

    status_html = ""
    if status_text:
        status_html = (
            f'<div class="editorial-status-pill {status_cor}">'
            f'<i class="ti ti-trending-up"></i> {status_text}'
            f'</div>'
        )

    st.html(
        f"""
        <div class="editorial-destaque-row">
            <div class="editorial-destaque-col">
                <div class="editorial-label">{label_pct}</div>
                <div style="display:flex; align-items:baseline; gap:14px; margin-top:10px;">
                    <div class="editorial-valor-hero">{pct_str}<span class="unit">%</span></div>
                    {status_html}
                </div>
            </div>
            <div class="editorial-destaque-col bigger">
                <div class="editorial-label">{label_valor}</div>
                <div class="editorial-valor-grande">
                    R$ <span class="destaque">{inteiro.replace('R$ ', '')}</span><span class="centavos">{cents}</span>
                </div>
                <div class="editorial-sub-text">de {fmt_brl(valor_total)} movimentados</div>
            </div>
        </div>
        """
    )


def editorial_barra_progresso(percentual: float, label_esq: str = "", label_dir: str = ""):
    """Barra de progresso fina, dourada, com legendas opcionais embaixo."""
    pct_clamped = max(0.0, min(100.0, percentual))
    st.html(
        f"""
        <div class="editorial-progresso-wrap">
            <div class="editorial-progresso-track">
                <div class="editorial-progresso-fill" style="width:{pct_clamped}%;"></div>
            </div>
            <div class="editorial-progresso-legend">
                <span>{label_esq}</span>
                <span>{label_dir}</span>
            </div>
        </div>
        """
    )


def editorial_linha_metricas(metricas: list[dict]):
    """Linha de colunas separadas por divisores verticais.
    Cada métrica: {'label': str, 'valor': str, 'centavos': str?, 'cor': str?, 'sub_html': str?}
    cor: '' (branco) | 'verde' | 'amarelo' | 'vermelho' | 'atencao'
    """
    cols_html = []
    for m in metricas:
        cor_class = m.get("cor", "")
        centavos = m.get("centavos", "")
        sub = m.get("sub_html", "")
        sub_html = f'<div class="sub">{sub}</div>' if sub else ""
        cols_html.append(
            f"""
            <div class="col">
                <div class="label">{m['label']}</div>
                <div class="valor {cor_class}">{m['valor']}<span class="centavos">{centavos}</span></div>
                {sub_html}
            </div>
            """
        )
    cols_count = len(metricas)
    st.html(
        f'<div class="editorial-cols" style="--cols:{cols_count};">{"".join(cols_html)}</div>'
    )


def editorial_faixa_contagens(itens: list[dict]):
    """Faixa horizontal entre dois divisores. Cada item: {'label', 'valor', 'cor'?}"""
    parts = []
    for i, item in enumerate(itens):
        if i > 0:
            parts.append('<div class="sep"></div>')
        cor = item.get("cor", "")
        parts.append(
            f"""
            <div class="item">
                <div class="valor {cor}">{item['valor']}</div>
                <div class="label">{item['label']}</div>
            </div>
            """
        )
    st.html(f'<div class="editorial-faixa">{"".join(parts)}</div>')


def editorial_secao_head(titulo: str, contagem_text: str = ""):
    """Cabeçalho de seção: título amarelo + linha que esmaece + contagem opcional à direita."""
    contagem_html = f'<span class="contagem">{contagem_text}</span>' if contagem_text else ""
    st.html(
        f"""
        <div class="editorial-secao-head">
            <span class="titulo">{titulo}</span>
            <span class="linha"></span>
            {contagem_html}
        </div>
        """
    )


def editorial_lista_excecoes(itens: list[dict]):
    """Lista de exceções (não cards). Cada item:
        {'icone': 'ti-credit-card', 'titulo': str, 'sub': str,
         'valor': str, 'status': str?, 'cor': 'verde'|'amarelo'|'',
         'muted': bool?}
    """
    parts = []
    for i, item in enumerate(itens):
        if i > 0:
            parts.append('<div class="editorial-lista-divisor"></div>')
        cor = item.get("cor", "")
        muted = "muted" if item.get("muted") else ""
        status = item.get("status", "")
        status_html = (
            f'<div class="editorial-lista-status {cor}">{status}</div>' if status else ""
        )
        valor_class = "muted" if item.get("muted") else ""
        # Parte "light" (cinza) no título: separar por travessão
        titulo = item["titulo"]
        if " — " in titulo:
            principal, light = titulo.split(" — ", 1)
            titulo_html = f'{principal} <span class="light">— {light}</span>'
        else:
            titulo_html = titulo
        parts.append(
            f"""
            <div class="editorial-lista-item {muted}">
                <div class="editorial-lista-icone {cor}">
                    <i class="ti {item['icone']}"></i>
                </div>
                <div class="editorial-lista-corpo">
                    <div class="editorial-lista-titulo">{titulo_html}</div>
                    <div class="editorial-lista-sub">{item['sub']}</div>
                </div>
                <div class="editorial-lista-direita">
                    <div class="editorial-lista-valor {valor_class}">{item['valor']}</div>
                    {status_html}
                </div>
            </div>
            """
        )
    st.html(f'<div class="editorial-lista">{"".join(parts)}</div>')


def render_resumo_executivo_editorial(resultado: ResultadoConciliacao, kpis: dict):
    """v5.12 Fase 1: Resumo Executivo no estilo editorial.
    Substitui os 8-11 cards retangulares por um layout em camadas:
    1) Header (kicker + título + data)
    2) KPI âncora (% + valor conciliado, lado a lado)
    3) Barra de progresso
    4) Linha de 4 colunas com divisores (Banco | Sankhya | Falta | Divergência)
    5) Faixa de contagens (Banco · Sankhya · Pares · Conta)
    6) Seção de exceções (lista, não cards)
    """
    # 1) Header
    data_ref = resultado.data_referencia.strftime("%d de %B · %Y")
    # tradução curta dos meses pra PT-BR (cobre EN-US default e variações de locale)
    meses_pt = {
        "January": "janeiro", "February": "fevereiro", "March": "março",
        "April": "abril", "May": "maio", "June": "junho",
        "July": "julho", "August": "agosto", "September": "setembro",
        "October": "outubro", "November": "novembro", "December": "dezembro",
    }
    for en, pt in meses_pt.items():
        data_ref = data_ref.replace(en, pt).replace(en.lower(), pt)

    n_contas = len(resultado.contas_processadas)
    titulo_subtitulo = (
        "Conciliação consolidada"
        if n_contas != 1
        else f"Conciliação · {resultado.contas_processadas[0]}"
    )

    editorial_header(
        kicker="Resumo executivo",
        titulo=titulo_subtitulo,
        meta_esquerda=data_ref,
        meta_direita=f"{n_contas} conta{'s' if n_contas != 1 else ''} processada{'s' if n_contas != 1 else ''}",
    )

    # 2) KPI âncora
    pct = float(kpis["percentual_conciliado"])
    if pct >= 95:
        status_text, status_cor = "dentro da meta", "verde"
    elif pct >= 80:
        status_text, status_cor = "próximo da meta", "amarelo"
    else:
        status_text, status_cor = "abaixo da meta", "vermelho"

    editorial_kpi_destaque(
        percentual=pct,
        valor_conciliado=float(kpis["total_conciliado"]),
        valor_total=float(kpis["total_movimentado_banco"]),
        status_text=status_text,
        status_cor=status_cor,
    )

    # 3) Barra de progresso
    falta = float(kpis["falta_conciliar"])
    diverg = float(kpis["divergencia_sankhya_banco"])
    label_dir = []
    if falta > 0:
        label_dir.append(f"{fmt_brl(falta)} falta")
    if diverg > 0:
        label_dir.append(f"{fmt_brl(diverg)} divergência")
    editorial_barra_progresso(
        percentual=pct,
        label_esq=f"{fmt_brl(float(kpis['total_conciliado']))} conciliado",
        label_dir=" · ".join(label_dir) if label_dir else "—",
    )

    # 4) Linha de 4 colunas: Banco | Sankhya | Falta | Divergência
    inteiro_b, cents_b = _split_centavos(float(kpis["total_movimentado_banco"]))
    inteiro_s, cents_s = _split_centavos(float(kpis["total_extrato_sistema"]))
    inteiro_f, cents_f = _split_centavos(falta)
    inteiro_d, cents_d = _split_centavos(diverg)

    # Subs: receitas/despesas resumidos
    rec_b = float(kpis["receitas_banco"])
    desp_b = float(kpis["despesas_banco"])
    rec_s = float(kpis["receitas_sistema"])
    desp_s = float(kpis["despesas_sistema"])

    sub_banco = (
        f'<span class="inline"><span class="verde">↑ {fmt_brl(rec_b).replace("R$ ", "")}</span></span>'
        f'<span class="vermelho">↓ {fmt_brl(desp_b).replace("R$ ", "")}</span>'
    )
    sub_sankhya = (
        f'<span class="inline"><span class="verde">↑ {fmt_brl(rec_s).replace("R$ ", "")}</span></span>'
        f'<span class="vermelho">↓ {fmt_brl(desp_s).replace("R$ ", "")}</span>'
    )

    # Sub de "Falta": indicar se é só despesas, só receitas ou ambos
    fr = float(kpis["falta_conciliar_receitas"])
    fd = float(kpis["falta_conciliar_despesas"])
    if fr > 0 and fd > 0:
        sub_falta = "receitas e despesas"
    elif fd > 0:
        sub_falta = "apenas em despesas"
    elif fr > 0:
        sub_falta = "apenas em receitas"
    else:
        sub_falta = "—"

    qtd_div = int(kpis.get("qtd_divergencia_sankhya_banco", 0))
    sub_div = f"{qtd_div} lançamento{'s' if qtd_div != 1 else ''}" if qtd_div else "—"

    # Remover "R$ " dos inteiros pra mostrar só o número (R$ já está embutido no contexto)
    editorial_linha_metricas([
        {
            "label": "Banco",
            "valor": inteiro_b.replace("R$ ", ""),
            "centavos": cents_b,
            "sub_html": sub_banco,
        },
        {
            "label": "Sankhya",
            "valor": inteiro_s.replace("R$ ", ""),
            "centavos": cents_s,
            "sub_html": sub_sankhya,
        },
        {
            "label": "Falta conciliar",
            "valor": inteiro_f.replace("R$ ", ""),
            "centavos": cents_f,
            "cor": "vermelho" if falta > 0 else "",
            "sub_html": sub_falta,
        },
        {
            "label": "Divergência",
            "valor": inteiro_d.replace("R$ ", ""),
            "centavos": cents_d,
            "cor": "atencao" if diverg > 0 else "",
            "sub_html": sub_div,
        },
    ])

    # 5) Faixa de contagens
    editorial_faixa_contagens([
        {"label": "Banco", "valor": fmt_int(kpis["qtd_registros_banco"])},
        {"label": "Sankhya", "valor": fmt_int(kpis["qtd_registros_sistema"])},
        {"label": "Pares", "valor": fmt_int(kpis["qtd_conciliados"]), "cor": "verde"},
        {"label": "Conta" if n_contas == 1 else "Contas", "valor": fmt_int(n_contas), "cor": "amarelo"},
    ])

    # 6) Seção de exceções (lista editorial)
    _render_excecoes_editorial(resultado, kpis)


def _render_excecoes_editorial(resultado: ResultadoConciliacao, kpis: dict):
    """Lista de exceções no estilo editorial — substitui a versão de cards."""
    qtd_est_anu = kpis.get("qtd_estornos_anulados", 0)
    qtd_est_par = kpis.get("qtd_estornos_parciais", 0)
    qtd_top1722 = kpis.get("qtd_top1722_grupos", 0)
    qtd_top1702 = kpis.get("qtd_top1702_grupos", 0)
    df_inv = resultado.aplicacoes_resgates
    tem_invest = not df_inv.empty

    # Se nenhuma das regras se aplica, mostra seção minimalista com "nenhuma exceção"
    if not (qtd_est_anu > 0 or qtd_est_par > 0 or qtd_top1722 > 0 or qtd_top1702 > 0 or tem_invest):
        editorial_secao_head("Exceções aplicadas", "nenhuma neste período")
        return

    total_regras = sum(1 for x in [qtd_est_anu > 0, qtd_est_par > 0, qtd_top1722 > 0, qtd_top1702 > 0, tem_invest] if x)
    editorial_secao_head("Exceções aplicadas", f"{total_regras} regra{'s' if total_regras != 1 else ''} ativa{'s' if total_regras != 1 else ''}")

    itens = []

    # TOP 1722
    if qtd_top1722 > 0:
        valor_top = float(kpis.get("valor_top1722_conciliado", 0.0))
        # Detalhamento de qtd banco × sankhya, se possível
        linhas_b = getattr(resultado, "top1722_linhas_banco", pd.DataFrame())
        linhas_s = getattr(resultado, "top1722_linhas", pd.DataFrame())
        qtd_b = len(linhas_b) if not linhas_b.empty else 0
        qtd_s = len(linhas_s) if not linhas_s.empty else 0
        sub = (
            f"{qtd_b} créditos banco × {qtd_s} vendas Sankhya · diferença R$ 0,00"
            if qtd_b and qtd_s
            else f"{qtd_top1722} agrupamento{'s' if qtd_top1722 != 1 else ''}"
        )
        itens.append({
            "icone": "ti-credit-card",
            "titulo": "Cartão TOP 1722 — agrupamento por soma total",
            "sub": sub,
            "valor": fmt_brl(valor_top),
            "status": "Conciliado",
            "cor": "verde",
        })

    # TOP 1702 (boleto)
    if qtd_top1702 > 0:
        valor_bol = float(kpis.get("valor_top1702_conciliado", 0.0))
        lb = getattr(resultado, "top1702_linhas_banco", pd.DataFrame())
        ls = getattr(resultado, "top1702_linhas", pd.DataFrame())
        dif1702 = getattr(resultado, "top1702_diferencas", pd.DataFrame())
        qb = len(lb) if not lb.empty else 0
        qs = len(ls) if not ls.empty else 0
        dif_val = float(dif1702["diferenca"].sum()) if (not dif1702.empty and "diferenca" in dif1702.columns) else 0.0
        com_dif = abs(dif_val) >= 0.005
        sub = (
            f"{qb} créditos cobrança banco × {qs} boletos Sankhya · "
            + (f"diferença {fmt_brl(dif_val)} (a investigar)" if com_dif else "diferença R$ 0,00")
        )
        itens.append({
            "icone": "ti-barcode",
            "titulo": "Boleto TOP 1702 — agrupamento por soma total",
            "sub": sub,
            "valor": fmt_brl(valor_bol),
            "status": "Conciliado (c/ diferença)" if com_dif else "Conciliado",
            "cor": "amarelo" if com_dif else "verde",
        })

    # Investimentos
    if tem_invest:
        aplic = df_inv[df_inv["tipo_aplicacao"] == "Aplicação"] if "tipo_aplicacao" in df_inv.columns else pd.DataFrame()
        resg = df_inv[df_inv["tipo_aplicacao"] == "Resgate"] if "tipo_aplicacao" in df_inv.columns else pd.DataFrame()
        qtd_a = len(aplic)
        qtd_r = len(resg)
        val_total = float(df_inv["valor"].abs().sum())
        itens.append({
            "icone": "ti-trending-up",
            "titulo": "Investimentos — aplicações e resgates",
            "sub": f"{qtd_a} aplicaç{'ão' if qtd_a == 1 else 'ões'} · {qtd_r} resgate{'s' if qtd_r != 1 else ''}",
            "valor": fmt_brl(val_total),
            "status": "Excluído da divergência",
            "cor": "amarelo",
        })

    # Estornos anulados
    if qtd_est_anu > 0:
        itens.append({
            "icone": "ti-refresh",
            "titulo": "Estornos anulados — débito e crédito se cancelam",
            "sub": f"{qtd_est_anu} ocorrência{'s' if qtd_est_anu != 1 else ''}",
            "valor": fmt_brl(float(kpis.get("valor_estornos_anulados", 0.0))),
            "status": "Removidos do resumo",
            "cor": "verde",
        })

    # Estornos parciais
    if qtd_est_par > 0:
        itens.append({
            "icone": "ti-scale",
            "titulo": "Estornos parciais — débito reduz crédito",
            "sub": f"{qtd_est_par} ocorrência{'s' if qtd_est_par != 1 else ''}",
            "valor": fmt_brl(float(kpis.get("saldo_estornos_parciais", 0.0))),
            "status": "Saldo restante",
            "cor": "amarelo",
        })

    editorial_lista_excecoes(itens)


# ============================================================
# v5.13 — Componentes OPÇÃO A (cards generosos)
# Layout moderno com hierarquia: card âncora grande + secundários
# com borda superior semântica + chips de exceção.
# ============================================================

def _formata_valor_curto(valor: float) -> str:
    """Formata sem o 'R$ ' (usado em sub-textos do card âncora)."""
    return fmt_brl(valor).replace("R$ ", "")


def opa_card_ancora(label: str, valor: float, receitas: float, despesas: float) -> str:
    """Card âncora (maior, com gradiente): Total Movimentado, com receitas/despesas."""
    return f"""
    <div class="opa-card-ancora">
        <div class="opa-card-ancora-label">{label}</div>
        <div class="opa-card-ancora-valor">{fmt_brl(valor)}</div>
        <div class="opa-rec-desp-row">
            <div class="opa-rec-desp-item">
                <div class="marker verde">
                    <span class="dot"></span>
                    Receitas
                </div>
                <div class="valor">{fmt_brl(receitas)}</div>
            </div>
            <div class="opa-rec-desp-item">
                <div class="marker vermelho">
                    <span class="dot"></span>
                    Despesas
                </div>
                <div class="valor">{fmt_brl(despesas)}</div>
            </div>
        </div>
    </div>
    """


def opa_card_conciliado(valor: float, sub_texto: str = "Match Banco × Sankhya") -> str:
    """Card verde translúcido (Total Conciliado)."""
    return f"""
    <div class="opa-card-conciliado">
        <div class="opa-card-conciliado-label">Total conciliado</div>
        <div class="opa-card-conciliado-valor">{fmt_brl(valor)}</div>
        <div class="opa-card-conciliado-sub">
            <i class="ti ti-check"></i>{sub_texto}
        </div>
    </div>
    """


def opa_card_donut(percentual: float) -> str:
    """Card com donut SVG do percentual conciliado.
    Círculo de raio 50, perímetro = 2*pi*50 ≈ 314.16.
    O 'offset' faz a parte amarela = percentual%.
    """
    raio = 50
    perimetro = 2 * 3.14159265 * raio  # ≈ 314.16
    pct_clamped = max(0.0, min(100.0, percentual))
    # dashoffset: 0 = círculo cheio. perimetro = vazio.
    dash_offset = perimetro * (1 - pct_clamped / 100)
    pct_str = f"{percentual:.1f}".replace(".", ",")
    return f"""
    <div class="opa-card-donut">
        <svg width="140" height="140" viewBox="0 0 120 120">
            <circle cx="60" cy="60" r="{raio}" fill="none" stroke="rgba(255,255,255,0.06)" stroke-width="6"/>
            <circle cx="60" cy="60" r="{raio}" fill="none" stroke="{CORES['amarelo']}" stroke-width="6"
                stroke-linecap="round" stroke-dasharray="{perimetro:.2f}"
                stroke-dashoffset="{dash_offset:.2f}"
                transform="rotate(-90 60 60)"/>
        </svg>
        <div class="opa-card-donut-text">
            <div class="opa-card-donut-pct">{pct_str}%</div>
            <div class="opa-card-donut-label">Conciliado</div>
        </div>
    </div>
    """


def opa_card_secundario(
    label: str,
    valor: str,
    sub_html: str = "",
    cor: str = "",
    icone_tabler: str = "",
) -> str:
    """Card secundário com borda superior semântica + ícone em pílula.
    cor: '' | 'vermelho' | 'amarelo' | 'verde'
    """
    icone_html = ""
    if icone_tabler:
        icone_html = f"""
        <div class="opa-card-sec-icon {cor}">
            <i class="ti {icone_tabler}"></i>
        </div>
        """
    sub = f'<div class="opa-card-sec-sub">{sub_html}</div>' if sub_html else ""
    return f"""
    <div class="opa-card-secundario {cor}">
        <div class="opa-card-sec-head">
            <div>
                <div class="opa-card-sec-label">{label}</div>
                <div class="opa-card-sec-valor {cor}">{valor}</div>
            </div>
            {icone_html}
        </div>
        {sub}
    </div>
    """


def opa_render_chip_excecao(icone: str, label: str, valor_principal: str, valor_sub: str = "", cor: str = "") -> str:
    """Chip horizontal pra Exceções (ícone redondo + label + valor inline)."""
    sub_html = f'<span class="light">· {valor_sub}</span>' if valor_sub else ""
    return f"""
    <div class="opa-chip {cor}">
        <div class="opa-chip-icon {cor}">
            <i class="ti {icone}"></i>
        </div>
        <div>
            <div class="opa-chip-label">{label}</div>
            <div class="opa-chip-valor">{valor_principal} {sub_html}</div>
        </div>
    </div>
    """


def render_resumo_executivo_opcao_a(resultado: ResultadoConciliacao, kpis: dict):
    """v5.13: Resumo Executivo no estilo Opção A.
    Layout em camadas:
    1) Section title clássico
    2) Linha hero: Card âncora (1.4fr) + Conciliado (1fr) + Donut % (1fr)
    3) Linha secundária: Falta + Divergência + Contas (3 cards iguais)
    4) Seção Exceções e Regras Aplicadas (chips horizontais)
    """
    section_title("RESUMO EXECUTIVO")

    # Linha 1: hero (3 cards com proporção 1.4 / 1 / 1)
    card_ancora = opa_card_ancora(
        label="Total movimentado · banco",
        valor=float(kpis["total_movimentado_banco"]),
        receitas=float(kpis["receitas_banco"]),
        despesas=float(kpis["despesas_banco"]),
    )
    card_conc = opa_card_conciliado(
        valor=float(kpis["total_conciliado"]),
        sub_texto="Match Banco × Sankhya",
    )
    card_donut = opa_card_donut(float(kpis["percentual_conciliado"]))

    st.html(
        f'<div class="opa-grid-hero">{card_ancora}{card_conc}{card_donut}</div>'
    )

    # Linha 2: secundários (3 cards iguais)
    falta = float(kpis["falta_conciliar"])
    falta_r = float(kpis["falta_conciliar_receitas"])
    falta_d = float(kpis["falta_conciliar_despesas"])
    diverg = float(kpis["divergencia_sankhya_banco"])
    qtd_div = int(kpis.get("qtd_divergencia_sankhya_banco", 0))
    n_contas = len(resultado.contas_processadas)

    sub_falta = (
        f'<span class="item"><span class="label">Receitas:</span> '
        f'<span class="{"vermelho-text" if falta_r > 0 else ""}">{fmt_brl(falta_r)}</span></span>'
        f'<span class="item"><span class="label">Despesas:</span> '
        f'<span class="{"vermelho-text" if falta_d > 0 else ""}">{fmt_brl(falta_d)}</span></span>'
    )

    sub_div = f"Sankhya × Banco · {qtd_div} lançamento{'s' if qtd_div != 1 else ''}"

    # Conta(s)
    nome_contas = ", ".join(resultado.contas_processadas[:2])
    if n_contas > 2:
        nome_contas += f" + {n_contas - 2}"
    sub_contas = nome_contas if n_contas > 0 else "—"

    card_falta = opa_card_secundario(
        label="Falta conciliar",
        valor=fmt_brl(falta),
        sub_html=sub_falta,
        cor="vermelho" if falta > 0 else "",
        icone_tabler="ti-alert-triangle",
    )
    card_diverg = opa_card_secundario(
        label="Divergência",
        valor=fmt_brl(diverg),
        sub_html=sub_div,
        cor="vermelho" if diverg > 0 else "",
        icone_tabler="ti-git-compare",
    )
    card_contas = opa_card_secundario(
        label="Contas processadas",
        valor=str(n_contas),
        sub_html=sub_contas,
        cor="amarelo",
        icone_tabler="ti-building-bank",
    )

    st.html(
        f'<div class="opa-grid-secundario">{card_falta}{card_diverg}{card_contas}</div>'
    )

    # Linha 3: contagens (mantém o estilo antigo dos cards, mas com cards menores e arejados)
    cards_contagem = [
        card_kpi("Registros Banco", fmt_int(kpis["qtd_registros_banco"]),
                 f"{fmt_int(kpis['qtd_movimentacoes_banco'])} movimentações"),
        card_kpi("Registros Sistema", fmt_int(kpis["qtd_registros_sistema"]),
                 f"{fmt_int(kpis['qtd_movimentacoes_sistema'])} movimentações"),
        card_kpi("Conciliados", fmt_int(kpis["qtd_conciliados"]),
                 "pares Banco × Sankhya", classe="destaque-verde"),
        card_kpi("", "", ""),
    ]
    st.write("")  # respiro
    render_cards(cards_contagem)

    # Seção Exceções
    _render_excecoes_opcao_a(resultado, kpis)


def _render_excecoes_opcao_a(resultado: ResultadoConciliacao, kpis: dict):
    """Exceções no estilo Opção A — chips horizontais leves (não cards)."""
    qtd_est_anu = kpis.get("qtd_estornos_anulados", 0)
    qtd_est_par = kpis.get("qtd_estornos_parciais", 0)
    qtd_top1722 = kpis.get("qtd_top1722_grupos", 0)
    qtd_top1702 = kpis.get("qtd_top1702_grupos", 0)
    df_inv = resultado.aplicacoes_resgates
    tem_invest = not df_inv.empty

    # Se nenhuma exceção, esconde a seção (não polui)
    if not (qtd_est_anu > 0 or qtd_est_par > 0 or qtd_top1722 > 0 or qtd_top1702 > 0 or tem_invest):
        return

    # Header da seção
    st.html(
        """
        <div class="opa-excecoes-head">
            <span class="opa-excecoes-head-label">Exceções e regras aplicadas</span>
            <span class="opa-excecoes-head-line"></span>
        </div>
        """
    )

    chips = []

    # TOP 1722
    if qtd_top1722 > 0:
        valor_top = float(kpis.get("valor_top1722_conciliado", 0.0))
        chips.append(opa_render_chip_excecao(
            icone="ti-credit-card",
            label="Cartão TOP 1722",
            valor_principal=str(qtd_top1722),
            valor_sub=fmt_brl(valor_top),
            cor="verde",
        ))

    # TOP 1702 (boleto)
    if qtd_top1702 > 0:
        _dif1702 = getattr(resultado, "top1702_diferencas", pd.DataFrame())
        _difv = float(_dif1702["diferenca"].sum()) if (not _dif1702.empty and "diferenca" in _dif1702.columns) else 0.0
        _com_dif = abs(_difv) >= 0.005
        chips.append(opa_render_chip_excecao(
            icone="ti-barcode",
            label="Boleto TOP 1702",
            valor_principal=str(qtd_top1702),
            valor_sub=fmt_brl(float(kpis.get("valor_top1702_conciliado", 0.0)))
            + (f" · dif {fmt_brl(_difv)}" if _com_dif else ""),
            cor="amarelo" if _com_dif else "verde",
        ))

    # Investimentos
    if tem_invest:
        val_total = float(df_inv["valor"].abs().sum())
        chips.append(opa_render_chip_excecao(
            icone="ti-trending-up",
            label="Investimentos",
            valor_principal=str(len(df_inv)),
            valor_sub=fmt_brl(val_total),
            cor="amarelo",
        ))

    # Anulados por Estorno
    if qtd_est_anu > 0:
        chips.append(opa_render_chip_excecao(
            icone="ti-refresh",
            label="Anulados por estorno",
            valor_principal=str(qtd_est_anu),
            valor_sub=fmt_brl(float(kpis.get("valor_estornos_anulados", 0.0))),
            cor="verde",
        ))

    # Estornos Parciais
    if qtd_est_par > 0:
        chips.append(opa_render_chip_excecao(
            icone="ti-scale",
            label="Estornos parciais",
            valor_principal=str(qtd_est_par),
            valor_sub=fmt_brl(float(kpis.get("saldo_estornos_parciais", 0.0))),
            cor="amarelo",
        ))

    st.html(f'<div class="opa-excecoes-grid">{"".join(chips)}</div>')


# ============================================================
# Validações
# ============================================================
NOMES_PROIBIDOS = {"data", "valor", "histórico", "historico", "conta", "—", ""}


def validar_nome_conta(nome: str) -> str | None:
    """Retorna mensagem de erro ou None se válido."""
    if not nome or not nome.strip():
        return "Identificador da conta não pode ser vazio."
    n = nome.lower().strip()
    if n in NOMES_PROIBIDOS:
        return f"Identificador '{nome}' é genérico demais. Use algo como 'Bradesco-12345'."
    if len(nome.strip()) < 3:
        return f"Identificador '{nome}' tem menos de 3 caracteres."
    return None


# ============================================================
# Página: Dashboard
# ============================================================
def pagina_dashboard():
    resultado: ResultadoConciliacao | None = st.session_state.resultado
    if resultado is None:
        st.info(
            "👋 Nenhuma conciliação foi rodada ainda nesta sessão. "
            "Acesse a página **Conciliação** para começar."
        )
        if st.button("➡️ Ir para Conciliação", type="primary"):
            ir_para("Conciliação")
            st.rerun()
        return

    kpis = resultado.kpis_globais()
    from datetime import datetime as _dt

    def _brl(v):
        try:
            v = float(v)
        except Exception:
            return "—"
        s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return ("-" if v < 0 else "") + s

    # ---- dados reais ----
    div = float(kpis.get("divergencia_sankhya_banco", 0.0))
    # v5.47: o veredito olha os DOIS lados — banco sem explicação (falta
    # conciliar) E Sankhya sem confirmação (divergência). Antes só olhava a
    # divergência: quando ela zerava, o painel dizia "✓ Bate" mesmo com
    # centenas de milhares sem explicação no banco.
    falta_dash = float(kpis.get("falta_conciliar", 0.0))
    pend_dash = round(abs(falta_dash) + abs(div), 2)
    bate = pend_dash < 0.005
    rec_b = float(kpis.get("receitas_banco", 0.0))
    desp_b = float(kpis.get("despesas_banco", 0.0))
    liq = rec_b - desp_b
    pct = float(kpis.get("percentual_conciliado", 0.0))
    n_contas = len(resultado.contas_processadas)

    # saldo inicial/final (do extrato do banco), agregando as contas que têm saldo.
    # v5.67: coleta por CONTA (sem somar entre contas diferentes — saldo bancário
    # não é dado somável entre CNPJs). As variáveis saldo_ini/saldo_fim ficam para
    # o caso N=1 (retrocompat com o card antigo); para N>1 usamos `_saldos_por_conta`.
    saldo_ini = saldo_fim = 0.0
    tem_saldo = False
    _saldos_por_conta: list = []  # v5.67: [{"conta": str, "info": dict}, ...]
    for _c in resultado.contas_processadas:
        try:
            _info = resultado.saldo_final_da_conta(_c, exigir_conciliado=False)
        except TypeError:
            # resultado antigo em sessão (sem o parâmetro novo)
            try:
                _info = resultado.saldo_final_da_conta(_c)
            except Exception:
                _info = None
        except Exception:
            _info = None
        if _info and _info.get("tem_saldo_no_extrato"):
            if _info.get("saldo_inicial") is not None:
                saldo_ini += float(_info["saldo_inicial"])
            if _info.get("saldo_final") is not None:
                saldo_fim += float(_info["saldo_final"])
            tem_saldo = True
        # v5.67: guarda por conta para o card "Saldos por conta" quando N>1
        _saldos_por_conta.append({"conta": _c, "info": _info or {}})

    # investimentos (líquido)
    inv_df = getattr(resultado, "aplicacoes_resgates", None)
    inv_net = float(inv_df["valor"].sum()) if (inv_df is not None and not inv_df.empty and "valor" in inv_df.columns) else 0.0

    # período de referência
    per_txt = "—"
    try:
        _d = pd.to_datetime(resultado.banco_completo["data"], errors="coerce").dropna()
        if not _d.empty:
            per_txt = f"{_d.min().strftime('%d/%m/%Y')} a {_d.max().strftime('%d/%m/%Y')}"
    except Exception:
        pass
    gerado = _dt.now().strftime("%d/%m %H:%M")

    # ---- cabeçalho ----
    st.markdown(
        f"<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:14px'>"
        f"<div style='font-size:16px;font-weight:600;color:#fff'>Visão executiva "
        f"<span style='font-size:12px;font-weight:400;color:#9fb3d6'>· referência {per_txt}</span></div>"
        f"<div style='font-size:12px;color:#6f88b8'>painel gerado em {gerado}</div></div>",
        unsafe_allow_html=True,
    )

    # ---- Conta 70 (do session_state) + histórico p/ tendência ----
    c70 = st.session_state.get("c70_dashboard")
    qtd_div = int(kpis.get("qtd_divergencia_sankhya_banco", 0))

    # métricas do fechamento atual (para o histórico)
    try:
        mes_ref = pd.to_datetime(resultado.banco_completo["data"], errors="coerce").dropna().max().strftime("%m/%Y")
    except Exception:
        mes_ref = per_txt
    atual = {
        "Mês": mes_ref,
        "% Conciliado": round(pct, 1),
        "Qtd Divergências": qtd_div,
        "Conta 70 (R$)": round(float(c70["parado"]), 2) if c70 else None,
    }
    # lê histórico salvo (se a pessoa subiu) para a tendência
    hist_df = None
    _hf = st.session_state.get("dash_hist")
    if _hf is not None:
        try:
            _hf.seek(0)
            hist_df = pd.read_excel(_hf)
        except Exception:
            hist_df = None
    prev = None
    if hist_df is not None and not hist_df.empty:
        ant = hist_df[hist_df["Mês"].astype(str) != str(mes_ref)]
        if not ant.empty:
            prev = ant.iloc[-1]

    def _seta(cur, ant_val, bom_subir=True):
        try:
            cur = float(cur); ant_val = float(ant_val)
        except Exception:
            return ""
        if abs(cur - ant_val) < 1e-9:
            return "<span style='color:#9fb3d6'>→</span>"
        subiu = cur > ant_val
        bom = subiu if bom_subir else (not subiu)
        cor = "#7ee0a6" if bom else "#ff9a9a"
        return f"<span style='color:{cor}'>{'▲' if subiu else '▼'}</span>"

    # ================= PAINEL (Opção 2 — duas colunas) =================
    # veredito
    if bate:
        _hero_grad, _hero_fg, _hero_txt, _hero_sub = "linear-gradient(135deg,#124a34,#0e3a29)", "#b6f5cf", "✓ Bate com o banco", "banco 100% explicado · Sankhya 100% confirmado · R$ 0,00 pendente"
    else:
        _partes_hero = []
        if abs(falta_dash) >= 0.005:
            _partes_hero.append(f"R$ {_brl(abs(falta_dash))} do banco sem explicação")
        if abs(div) >= 0.005:
            _partes_hero.append(f"R$ {_brl(abs(div))} do Sankhya sem confirmação")
        _hero_grad, _hero_fg, _hero_txt, _hero_sub = "linear-gradient(135deg,#4a1a1a,#3a1414)", "#ffc2c2", "✗ Não bate com o banco", " · ".join(_partes_hero) if _partes_hero else f"diferença Sankhya × Banco · R$ {_brl(div)}"
    # v5.67: quantas das contas rodadas FECHARAM (por conta, não sobre "9 do grupo").
    # Antes: '1 de 9 · faltam 8' — confundia com total do grupo, mas o app só valida
    # o que foi rodado. Agora: 'N de N rodadas fecharam'.
    _fechadas_rodadas = 0
    try:
        _kpbn = resultado.kpis_por_banco()
        for _c in resultado.contas_processadas:
            _kc = _kpbn.get(_c, {})
            _fc = abs(float(_kc.get("falta_conciliar", 0.0)))
            _dc = abs(float(_kc.get("divergencia_sankhya_banco", 0.0)))
            if round(_fc + _dc, 2) < 0.005:
                _fechadas_rodadas += 1
    except Exception:
        _fechadas_rodadas = n_contas if bate else 0
    _cob_cor = "#7ee0a6" if _fechadas_rodadas == n_contas else "#FAC318"
    # v5.68: rótulo sempre no plural — "Contas rodadas" lê melhor mesmo com 1
    _rodadas_txt = "rodadas"
    _cob_sub = (f"· todas fecharam" if _fechadas_rodadas == n_contas
                else f"· {n_contas - _fechadas_rodadas} a resolver")

    # saldo (da conta rodada) — v5.47: a conta do card agora FECHA de verdade.
    # Antes faltava a linha de investimentos (o líquido que foi p/ aplicação sai
    # do saldo mas não está em Receitas/Despesas) e o "= extrato ✓" era fixo.
    # Agora: inicial + receitas − despesas ± investimentos = final CALCULADO,
    # comparado com o saldo final LIDO do extrato (✓ só quando bate).
    if tem_saldo:
        # Líquido de investimentos do LADO DO BANCO (aplicações/resgates/rend.
        # do extrato) — é o que sai/entra do saldo. O card "Investimentos" ao
        # lado usa a visão deduplicada banco×Sankhya (outra finalidade).
        _inv_liq_saldo = 0.0
        try:
            _bc = resultado.banco_completo
            if _bc is not None and not _bc.empty and "categoria_mov" in _bc.columns:
                _inv_liq_saldo = float(
                    _bc[_bc["categoria_mov"].isin(
                        ["aplicacao", "resgate", "rendimento", "investimento_outro"]
                    )]["valor"].sum()
                )
        except Exception:
            _inv_liq_saldo = 0.0
        if _inv_liq_saldo < -0.005:
            _linha_inv = (
                f"<div style='display:flex;justify-content:space-between;font-size:14px;color:#cdd9f2;padding:8px 0'>"
                f"<span>− Foi p/ investimento (líq.)</span><b style='color:#FAC318;font-weight:600'>{_brl(abs(_inv_liq_saldo))}</b></div>"
            )
        elif _inv_liq_saldo > 0.005:
            _linha_inv = (
                f"<div style='display:flex;justify-content:space-between;font-size:14px;color:#cdd9f2;padding:8px 0'>"
                f"<span>+ Voltou de investimento (líq.)</span><b style='color:#FAC318;font-weight:600'>{_brl(_inv_liq_saldo)}</b></div>"
            )
        else:
            _linha_inv = ""
        _saldo_calc = round(saldo_ini + rec_b - desp_b + _inv_liq_saldo, 2)
        _fecha_saldo = abs(_saldo_calc - saldo_fim) < 0.01
        _selo_saldo = ("<span style='font-size:12px;color:#7ee0a6;font-weight:600;margin-left:6px'>= extrato ✓</span>" if _fecha_saldo
                       else f"<span style='font-size:12px;color:#ffc94d;font-weight:600;margin-left:6px'>calculado {_brl(_saldo_calc)} · conferir ⚠</span>")
        _saldo_inner = (
            f"<div style='display:flex;justify-content:space-between;font-size:14px;color:#cdd9f2;padding:8px 0'><span>Saldo inicial</span><b style='color:#fff;font-weight:600'>{_brl(saldo_ini)}</b></div>"
            f"<div style='display:flex;justify-content:space-between;font-size:14px;color:#cdd9f2;padding:8px 0'><span>+ Receitas</span><b style='color:#7ee0a6;font-weight:600'>{_brl(rec_b)}</b></div>"
            f"<div style='display:flex;justify-content:space-between;font-size:14px;color:#cdd9f2;padding:8px 0'><span>− Despesas</span><b style='color:#ff9a9a;font-weight:600'>{_brl(desp_b)}</b></div>"
            f"{_linha_inv}"
            f"<div style='border-top:1px solid #1d3a72;margin:6px 0 10px'></div>"
            f"<div style='display:flex;justify-content:space-between;align-items:baseline'><span style='font-size:14px;font-weight:700;color:#fff'>= Saldo final</span><span style='font-size:26px;font-weight:800;color:#fff'>{_brl(saldo_fim)}{_selo_saldo}</span></div>"
        )
    else:
        _saldo_inner = "<div style='font-size:12px;color:#9fb3d6;padding:10px 0'>este extrato não traz linha de saldo — o saldo aparece quando o extrato da conta rodada tiver saldo inicial/final.</div>"

    # conta 70
    if c70:
        _ag = c70.get("aging", {})
        _maior = f" · maior R$ {_brl(abs(c70['maior_val']))} ({c70.get('maior_dias', 0)}d)" if c70.get("maior_val") is not None else ""
        _c70_card = (
            "<div style='background:#0b2560;border:1px solid #163062;border-left:3px solid #ff9a9a;border-radius:14px;padding:16px 18px'>"
            f"<div style='display:flex;justify-content:space-between;align-items:baseline'><span style='font-size:12px;letter-spacing:1px;color:#9fb3d6'>PARADO NA CONTA 70</span><b style='font-size:22px;color:#ff9a9a'>R$ {_brl(c70.get('parado', 0))}</b></div>"
            f"<div style='font-size:12px;color:#6f88b8;margin:4px 0 10px'>{c70.get('itens', 0)} itens{_maior}</div>"
            "<div style='display:flex;justify-content:space-between;border-top:1px solid #163062;padding-top:9px'>"
            f"<div style='text-align:center'><div style='font-size:16px;font-weight:800;color:#7ee0a6'>{_ag.get('ate30',0)}</div><div style='font-size:10px;color:#9fb3d6'>≤30d</div></div>"
            f"<div style='text-align:center'><div style='font-size:16px;font-weight:800;color:#FAC318'>{_ag.get('d31_60',0)}</div><div style='font-size:10px;color:#9fb3d6'>31–60</div></div>"
            f"<div style='text-align:center'><div style='font-size:16px;font-weight:800;color:#cdd9f2'>{_ag.get('d61_90',0)}</div><div style='font-size:10px;color:#9fb3d6'>61–90</div></div>"
            f"<div style='text-align:center'><div style='font-size:16px;font-weight:800;color:#ff9a9a'>{_ag.get('mais90',0)}</div><div style='font-size:10px;color:#9fb3d6'>+90d</div></div>"
            "</div></div>"
        )
    else:
        _c70_card = ("<div style='background:#0b2560;border:1px solid #163062;border-left:3px solid #6f88b8;border-radius:14px;padding:16px 18px;font-size:13.5px;color:#9fb3d6'>"
                     "<b style='color:#cdd9f2'>Parado na Conta 70</b><br>abra a aba <b style='color:#fff'>Atrelamento e Numeração — Conta 70</b> e processe para ver aqui.</div>")

    # tendência
    if prev is not None:
        _t_pct = _seta(atual["% Conciliado"], prev.get("% Conciliado"), bom_subir=True)
        _t_div = _seta(atual["Qtd Divergências"], prev.get("Qtd Divergências"), bom_subir=False)
        _t_c70 = _seta(abs(atual["Conta 70 (R$)"] or 0), abs(pd.to_numeric(prev.get("Conta 70 (R$)"), errors="coerce") or 0), bom_subir=False)
        _trend_label = f"TENDÊNCIA vs {prev.get('Mês', 'mês anterior')}"
        _trend_html = f"<span style='font-size:13.5px;color:#cdd9f2'>% conc. <b>{_t_pct}</b> &nbsp; diverg. <b>{_t_div}</b> &nbsp; Conta 70 <b>{_t_c70}</b></span>"
    else:
        _trend_label = "TENDÊNCIA vs mês anterior"
        _trend_html = "<span style='font-size:12px;color:#6f88b8'>suba o histórico abaixo para ativar ▲/▼</span>"

    _dv_cor = "#7ee0a6" if qtd_div == 0 else "#ff9a9a"

    # v5.52: o grid inteiro era UM st.markdown com CSS grid/flex — o Streamlit
    # mede a altura antes da fonte carregar e, com a tipografia maior (v5.50),
    # o conteúdo real ficava mais alto que a medição e VAZAVA por cima do
    # expander "Tendência". Agora cada card é um st.markdown dentro de
    # st.columns nativas: o Streamlit mede card a card e nada sobrepõe.
    # Visual idêntico ao mockup aprovado (mesmos cards, cores e tipografia).
    _card_hero = (
        f"<div style='background:{_hero_grad};border-radius:16px;padding:24px'>"
        "<div style='font-size:12px;letter-spacing:1.2px;color:#6ee3a0;margin-bottom:8px'>FECHAMENTO DO MÊS</div>"
        f"<div style='font-size:30px;font-weight:800;color:{_hero_fg};line-height:1.15'>{_hero_txt}</div>"
        f"<div style='font-size:14px;color:{_hero_fg};margin:8px 0 20px'>{_hero_sub}</div>"
        "<div style='display:flex;justify-content:space-between;align-items:center;border-top:1px solid #ffffff22;padding-top:12px'>"
        f"<span style='font-size:13px;color:#cdd9f2'>Contas {_rodadas_txt}</span>"
        f"<span style='font-size:22px;font-weight:800;color:{_cob_cor}'>{_fechadas_rodadas} <span style='font-size:14px'>de {n_contas} fecharam</span> <span style='font-size:12px'>{_cob_sub}</span></span></div></div>"
    )
    # v5.67: quando são VÁRIAS contas, o card vira "Saldos por conta" com uma
    # linha por conta (nome + saldo final + selo/dif). Somar saldo bancário entre
    # contas diferentes não faz sentido contábil — antes um saldo Frankenstein
    # (INOV movimentos + APOIO saldo inicial) inflava o "calculado" e enganava.
    if n_contas <= 1:
        _card_saldo = (
            "<div style='background:#0b2560;border:1px solid #1b3a6e;border-radius:16px;padding:22px'>"
            "<div style='font-size:12px;letter-spacing:1.2px;color:#9fb3d6;margin-bottom:12px'>SALDO DA CONTA RODADA</div>"
            f"{_saldo_inner}</div>"
        )
    else:
        # monta uma linha por conta a partir de _saldos_por_conta
        _linhas_sc = []
        _n_com_saldo = _n_fecha = 0
        _total_final = 0.0
        for _sc in _saldos_por_conta:
            _conta_sc = _sc["conta"]
            _info_sc = _sc["info"]
            if not _info_sc or not _info_sc.get("tem_saldo_no_extrato"):
                _linhas_sc.append(
                    "<div style='padding:10px 0;border-top:1px solid #10254e'>"
                    "<div style='display:flex;justify-content:space-between;align-items:baseline'>"
                    f"<span style='font-size:13px;color:#7dd0ff;font-weight:600'>{_conta_sc}</span>"
                    "<span style='font-size:12px;color:#6f88b8'>sem saldo no extrato</span></div></div>"
                )
                continue
            _n_com_saldo += 1
            _si = float(_info_sc.get("saldo_inicial") or 0.0)
            _sf = float(_info_sc.get("saldo_final") or 0.0)
            _total_final += _sf
            # movimentos e investimentos LÍQUIDOS DA CONTA (do banco)
            try:
                _bcc = resultado.banco_completo
                _bcc = _bcc[_bcc["conta"] == _conta_sc] if (_bcc is not None and not _bcc.empty and "conta" in _bcc.columns) else None
                if _bcc is not None and not _bcc.empty:
                    _mask_mov = ~_bcc["categoria_mov"].isin(
                        ["saldo", "aplicacao", "resgate", "rendimento", "investimento_outro"]
                    ) if "categoria_mov" in _bcc.columns else pd.Series([True] * len(_bcc), index=_bcc.index)
                    _rec_sc = float(_bcc[_mask_mov & (_bcc["valor"] > 0)]["valor"].sum())
                    _desp_sc = float(-_bcc[_mask_mov & (_bcc["valor"] < 0)]["valor"].sum())
                    _inv_sc = 0.0
                    if "categoria_mov" in _bcc.columns:
                        _inv_sc = float(_bcc[_bcc["categoria_mov"].isin(
                            ["aplicacao", "resgate", "rendimento", "investimento_outro"]
                        )]["valor"].sum())
                else:
                    _rec_sc = _desp_sc = _inv_sc = 0.0
            except Exception:
                _rec_sc = _desp_sc = _inv_sc = 0.0
            _calc = round(_si + _rec_sc - _desp_sc + _inv_sc, 2)
            _fecha_sc = abs(_calc - _sf) < 0.01
            if _fecha_sc:
                _n_fecha += 1
                _selo_sc = "<span style='font-size:11px;color:#7ee0a6;font-weight:600;margin-left:6px'>= extrato ✓</span>"
            else:
                _selo_sc = f"<span style='font-size:11px;color:#ffc94d;font-weight:600;margin-left:6px'>calc. {_brl(_calc)} · dif {_brl(round(_calc - _sf, 2))} ⚠</span>"
            _inv_txt = ""
            if abs(_inv_sc) >= 0.005:
                _inv_txt = f" · investimento (líq.) {_brl(_inv_sc)}"
            _linhas_sc.append(
                "<div style='padding:12px 0;border-top:1px solid #10254e'>"
                "<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:4px'>"
                f"<span style='font-size:13px;color:#7dd0ff;font-weight:600'>{_conta_sc}</span>"
                f"<span style='font-size:18px;font-weight:800;color:#fff'>{_brl(_sf)}{_selo_sc}</span></div>"
                f"<div style='font-size:11px;color:#6f88b8'>inicial {_brl(_si)} · +receitas {_brl(_rec_sc)} · −despesas {_brl(_desp_sc)}{_inv_txt}</div></div>"
            )
        _cab_txt = f"{n_contas} contas rodadas"
        if _n_com_saldo > 0:
            _cab_cor = "#9fb3d6" if _n_fecha == _n_com_saldo else "#ffc94d"
            _cab_extra = (f" · todas fecham com o extrato ✓" if _n_fecha == _n_com_saldo
                          else f" · {_n_fecha} de {_n_com_saldo} fecham · {_n_com_saldo - _n_fecha} a conferir ⚠")
        else:
            _cab_cor = "#6f88b8"
            _cab_extra = " · nenhum extrato traz saldo"
        _rodape_total = ""
        if _n_com_saldo > 1:
            _rodape_total = (
                "<div style='border-top:1px solid #24427e;margin-top:6px;padding-top:12px;"
                "display:flex;justify-content:space-between;align-items:baseline'>"
                "<span style='font-size:12px;color:#9fb3d6'>Total consolidado das contas rodadas</span>"
                f"<span style='font-size:16px;font-weight:700;color:#FAC318'>{_brl(_total_final)}</span></div>"
                "<div style='font-size:10px;color:#6f88b8;text-align:right;margin-top:2px'>"
                "soma dos saldos finais · use como referência, não como saldo bancário unificado</div>"
            )
        _card_saldo = (
            "<div style='background:#0b2560;border:1px solid #1b3a6e;border-radius:16px;padding:20px 22px'>"
            "<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px'>"
            "<span style='font-size:12px;letter-spacing:1.2px;color:#9fb3d6'>SALDOS POR CONTA</span>"
            f"<span style='font-size:11px;color:{_cab_cor}'>{_cab_txt}{_cab_extra}</span></div>"
            + "".join(_linhas_sc)
            + _rodape_total
            + "</div>"
        )
    # v5.67: sufixo de escopo — deixa explícito que os números vêm das contas
    # rodadas nesta sessão (não do grupo inteiro).
    _sufixo_escopo = ("1 conta rodada" if n_contas == 1 else f"{n_contas} contas rodadas")
    _card_mov = (
        "<div style='background:#0b2560;border:1px solid #163062;border-radius:14px;padding:16px 18px'>"
        "<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px'>"
        "<span style='font-size:12px;letter-spacing:1.2px;color:#9fb3d6'>MOVIMENTAÇÃO DO MÊS</span>"
        f"<span style='font-size:10px;color:#6f88b8'>{_sufixo_escopo}</span></div>"
        f"<div style='display:flex;justify-content:space-between'><span style='font-size:14px;color:#cdd9f2'>Receitas</span><b style='font-size:16px;color:#7ee0a6'>{_brl(rec_b)}</b></div>"
        f"<div style='display:flex;justify-content:space-between;margin-top:4px'><span style='font-size:14px;color:#cdd9f2'>Despesas</span><b style='font-size:16px;color:#ff9a9a'>{_brl(desp_b)}</b></div>"
        f"<div style='display:flex;justify-content:space-between;margin-top:4px;border-top:1px solid #163062;padding-top:6px'><span style='font-size:14px;color:#cdd9f2'>Líquido <span style='color:#FAC318'>(foi p/ invest.)</span></span><b style='font-size:16px;color:#fff'>{_brl(liq)}</b></div></div>"
    )
    # v5.63: mini-cards SEPARADOS (um markdown por card). O bloco único com
    # grid de 2 cards era exatamente o padrão que causa sobreposição (v5.56/
    # v5.60): o Streamlit mede a altura antes da fonte carregar e o conteúdo
    # vaza sobre os cards vizinhos. Agora cada mini vai em sua própria coluna.
    _card_mini_div = (
        f"<div style='background:#0b2560;border:1px solid #163062;border-top:3px solid {_dv_cor};border-radius:14px;padding:15px'>"
        "<div style='font-size:12px;letter-spacing:1px;color:#9fb3d6'>DIVERGÊNCIAS</div>"
        f"<div style='font-size:24px;font-weight:800;color:{_dv_cor};margin-top:6px'>{qtd_div}</div>"
        f"<div style='font-size:12px;color:#6f88b8'>R$ {_brl(kpis.get('divergencia_sankhya_banco',0))}</div>"
        f"<div style='font-size:10px;color:#6f88b8;margin-top:2px'>{_sufixo_escopo}</div></div>"
    )
    _card_mini_inv = (
        f"<div style='background:#0b2560;border:1px solid #163062;border-top:3px solid #FAC318;border-radius:14px;padding:15px'>"
        "<div style='font-size:12px;letter-spacing:1px;color:#9fb3d6'>INVESTIMENTOS</div>"
        f"<div style='font-size:24px;font-weight:800;color:#FAC318;margin-top:6px'>{_brl(inv_net)}</div>"
        "<div style='font-size:12px;color:#6f88b8'>líq. aplic − resgates</div>"
        f"<div style='font-size:10px;color:#6f88b8;margin-top:2px'>{_sufixo_escopo}</div></div>"
    )
    _card_trend = (
        f"<div style='background:#0b2560;border:1px solid #163062;border-radius:14px;padding:14px 18px;display:flex;justify-content:space-between;align-items:center'><span style='font-size:12px;letter-spacing:1px;color:#9fb3d6'>{_trend_label}</span>{_trend_html}</div>"
    )
    # v5.60: REVERTIDO para um st.markdown POR CARD. A tentativa v5.56 de juntar
    # os cards em um bloco por coluna (para alinhar as bases) reativou a
    # sobreposição — o Streamlit mede a altura do bloco antes da fonte carregar
    # e o conteúdo vaza sobre o elemento seguinte. Com um markdown por card
    # (formato da v5.52, validado em produção) a medição é por card e NÃO HÁ
    # como sobrepor. Custo aceito e combinado: as colunas são top-alinhadas e
    # podem terminar em alturas diferentes.
    _col_esq, _col_dir = st.columns([1.15, 1], gap="medium")
    with _col_esq:
        st.markdown(_card_hero, unsafe_allow_html=True)
        st.write("")  # v5.67: respiro vertical entre cards (evita "aglomeração")
        st.markdown(_card_saldo, unsafe_allow_html=True)
    with _col_dir:
        st.markdown(_card_mov, unsafe_allow_html=True)
        st.write("")
        _mc1, _mc2 = st.columns(2, gap="small")
        with _mc1:
            st.markdown(_card_mini_div, unsafe_allow_html=True)
        with _mc2:
            st.markdown(_card_mini_inv, unsafe_allow_html=True)
        st.write("")
        if _c70_card:
            st.markdown(_c70_card, unsafe_allow_html=True)
            st.write("")
        st.markdown(_card_trend, unsafe_allow_html=True)

    # ---- histórico de fechamentos (para a tendência) ----
    with st.expander("Tendência — histórico de fechamentos", expanded=False):
        st.caption("Suba a planilha de histórico (opcional) para ver a tendência mês a mês. "
                   "Depois, baixe o histórico atualizado e guarde para o próximo fechamento.")
        st.file_uploader("Histórico de fechamentos (.xlsx)", type=["xlsx"], key="dash_hist")
        novo_hist = pd.concat([hist_df, pd.DataFrame([atual])], ignore_index=True) if hist_df is not None else pd.DataFrame([atual])
        novo_hist = novo_hist.drop_duplicates(subset=["Mês"], keep="last")
        from io import BytesIO as _BIO
        _b = _BIO()
        with pd.ExcelWriter(_b, engine="openpyxl") as _w:
            novo_hist.to_excel(_w, index=False, sheet_name="Histórico")
        st.download_button("⬇️ Baixar histórico atualizado (.xlsx)", data=_b.getvalue(),
                           file_name="historico_fechamentos.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    # ---- detalhe operacional (recolhido) ----
    with st.expander("Detalhe operacional (conciliados, registros, movimentações, por conta)", expanded=False):
        render_cards([
            card_kpi("Conciliados", fmt_int(kpis["qtd_conciliados"]), classe="destaque-verde"),
            card_kpi("Registros processados", fmt_int(kpis["qtd_registros_banco"] + kpis["qtd_registros_sistema"])),
            card_kpi("Movimentações Banco", fmt_int(kpis["qtd_movimentacoes_banco"])),
            card_kpi("Movimentações Sistema", fmt_int(kpis["qtd_movimentacoes_sistema"])),
        ])
        st.divider()
        render_painel_bancos(resultado, mostrar_botao=False)


# ============================================================
# Painel de botões por banco
# ============================================================
def render_painel_bancos(resultado: ResultadoConciliacao, mostrar_botao: bool = True):
    """v6: painel de contas ordenado por atenção, com SELO DE ALERTA.

    - Cards em tom NEUTRO. O vermelho/âmbar aparece só num selo (badge),
      nunca como cor do card.
    - Contas ORDENADAS POR ATENÇÃO: quem tem mais itens a resolver vem no topo.
    - Mostra movimento da conta e quanto há "a resolver".

    `mostrar_botao=False` no Dashboard (cards só informativos).
    """
    contas = resultado.contas_processadas
    if not contas:
        st.warning("Nenhuma conta foi processada.")
        return

    kpis_pb = resultado.kpis_por_banco()

    def _itens_a_resolver(conta: str, k: dict) -> tuple[int, float]:
        """Itens que precisam de análise nesta conta e quanto em R$.

        'Precisa de você' = banco sem explicação + Sankhya sem confirmação +
        lançamentos que não pertencem à conta (espelha a triagem do detalhamento).
        """
        try:
            qtd_nao_pertence = len(resultado.nao_pertence_da_conta(conta))
        except Exception:
            qtd_nao_pertence = 0
        qtd = (
            int(k.get("qtd_pendentes_banco", 0))
            + int(k.get("qtd_divergencia_sankhya_banco", 0))
            + qtd_nao_pertence
        )
        valor = abs(float(k.get("falta_conciliar", 0.0))) + abs(
            float(k.get("divergencia_sankhya_banco", 0.0))
        )
        return qtd, valor

    # Ordena por atenção: mais itens primeiro; empate desempata por R$ a resolver.
    contas_ordenadas = sorted(
        contas,
        key=lambda c: _itens_a_resolver(c, kpis_pb[c]),
        reverse=True,
    )

    def _metricas(conta: str):
        k = kpis_pb[conta]
        mov = float(k.get("total_movimentado_banco", 0.0))
        falta = float(k.get("falta_conciliar", 0.0))
        explic = 100.0 * (mov - falta) / mov if mov > 0 else 0.0
        tot_sis = float(k.get("total_extrato_sistema", 0.0))
        div = float(k.get("divergencia_sankhya_banco", 0.0))
        confirm = 100.0 * (tot_sis - div) / tot_sis if tot_sis > 0 else 0.0
        pct_pares = float(k.get("percentual_conciliado", 0.0))
        qtd, valor = _itens_a_resolver(conta, k)
        return mov, explic, confirm, pct_pares, qtd, valor

    # ---- UMA conta: mantém o card de hoje (a tabela é só pra múltiplas contas) ----
    if len(contas_ordenadas) == 1:
        conta = contas_ordenadas[0]
        mov_conta, pct, _confirm, pct_pares, qtd_itens, valor_resolver = _metricas(conta)
        if qtd_itens == 0:
            badge_classe, badge_texto = "verde", "ok"
        elif qtd_itens <= 2:
            badge_classe = "amarelo"
            badge_texto = f"{qtd_itens} item" if qtd_itens == 1 else f"{qtd_itens} itens"
        else:
            badge_classe, badge_texto = "vermelho", f"{qtd_itens} itens"
        cor_barra = "#46d18a" if qtd_itens == 0 else "#9fb0d0"
        pct_barra = max(0.0, min(100.0, float(pct)))
        if qtd_itens == 0:
            linha_resolver = '<div class="lle-kpi-suffix">nada pendente</div>'
        else:
            cor_txt = "#ff8a8a" if badge_classe == "vermelho" else "#cdbf7a"
            linha_resolver = (
                f'<div class="lle-kpi-suffix" style="color:{cor_txt};">'
                f"{fmt_brl(valor_resolver)} a resolver</div>"
            )
        st.html(
            f"""
            <div class="lle-kpi">
                <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:8px;">
                    <div class="lle-kpi-label">{conta}</div>
                    <span class="lle-badge {badge_classe}">{badge_texto}</span>
                </div>
                <div class="lle-kpi-value" style="font-size:22px;">{fmt_pct(pct)}</div>
                <div style="height:6px; background:#1a2a52; border-radius:999px; overflow:hidden; margin:8px 0;">
                    <div style="width:{pct_barra}%; height:100%; background:{cor_barra};"></div>
                </div>
                <div class="lle-kpi-suffix">mov. {fmt_brl(mov_conta)}</div>
                <div class="lle-kpi-suffix" style="opacity:.7;">{fmt_pct(pct_pares)} em pares</div>
                {linha_resolver}
            </div>
            """
        )
        if mostrar_botao:
            st.button(
                "Ver detalhamento →", key=f"banco_btn_{conta}",
                on_click=selecionar_banco, args=(conta,), use_container_width=True,
            )
        return

    # ---- VÁRIAS contas: card único (Opção 2) — tabela organizada + selo do banco ----
    # v5.45: tudo dentro de UM card (st.container border), fonte padronizada, selo
    # colorido por banco na 1ª coluna, separadores entre linhas e GRUPO como rodapé
    # no mesmo card. O botão "Ver →" continua NATIVO por linha (botão do Streamlit
    # não entra dentro de HTML), então não voltou a virar fileira embaixo.
    _ratios = [3.4, 1.05, 1.05, 0.8, 1.6, 1.15]

    def _cel(html, align="left", cor="#dfe8fb", peso="400", size="14px"):
        return (f'<div style="text-align:{align}; color:{cor}; font-weight:{peso}; '
                f'font-size:{size}; line-height:1.9;">{html}</div>')

    def _cor_banco(nome: str) -> str:
        u = nome.upper()
        if "SICREDI" in u:
            return "#3FA110"
        if "ITAU" in u or "ITAÚ" in u:
            return "#EC7000"
        if "BRADESCO" in u:
            return "#CC092F"
        if "CAIXA" in u:
            return "#0070AF"
        if "SANTANDER" in u:
            return "#EC0000"
        return "#5B6B8C"

    def _selo(nome: str) -> str:
        cor = _cor_banco(nome)
        return (
            '<span style="display:inline-flex;align-items:center;gap:9px;">'
            f'<span style="width:22px;height:22px;border-radius:6px;background:{cor};'
            'display:inline-flex;align-items:center;justify-content:center;font-size:12px;">'
            '🏦</span>'
            f'<span style="color:#fff;font-weight:600;">{nome}</span></span>'
        )

    _sep = '<div style="border-top:1px solid #16305d; margin:0;"></div>'

    with st.container(border=True):
        st.markdown(
            '<div style="color:#fff; font-size:15px; font-weight:600; margin-bottom:6px;">'
            'Contas processadas</div>', unsafe_allow_html=True)

        # Cabeçalho
        hc = st.columns(_ratios)
        for _col, _txt, _al in zip(
            hc, ["CONTA", "EXPLICADO", "CONFIRMADO", "ITENS", "A RESOLVER", ""],
            ["left", "center", "center", "center", "right", "left"],
        ):
            _col.markdown(_cel(_txt, _al, "#8BA3C7", size="11px"), unsafe_allow_html=True)

        total_itens = 0
        total_resolver = 0.0
        fechadas = 0
        for conta in contas_ordenadas:
            _mov, explic, confirm, _pp, qtd, valor = _metricas(conta)
            total_itens += qtd
            total_resolver += valor
            if qtd == 0:
                fechadas += 1
                itens_cell = '<span style="background:#0F8C3B; color:#fff; font-size:11px; padding:3px 9px; border-radius:20px;">ok</span>'
                resolver_cell = '<span style="color:#46d18a;">—</span>'
            else:
                cor_badge = "#9a7b12" if qtd <= 2 else "#D63031"
                itens_cell = f'<span style="background:{cor_badge}; color:#fff; font-size:11px; padding:3px 9px; border-radius:20px;">{qtd}</span>'
                resolver_cell = f'<span style="color:#ff8a8a; font-weight:500;">{fmt_brl(valor)}</span>'

            st.markdown(_sep, unsafe_allow_html=True)
            rc = st.columns(_ratios)
            rc[0].markdown(_cel(_selo(conta)), unsafe_allow_html=True)
            rc[1].markdown(_cel(fmt_pct(explic), "center"), unsafe_allow_html=True)
            rc[2].markdown(_cel(fmt_pct(confirm), "center"), unsafe_allow_html=True)
            rc[3].markdown(_cel(itens_cell, "center"), unsafe_allow_html=True)
            rc[4].markdown(_cel(resolver_cell, "right"), unsafe_allow_html=True)
            if mostrar_botao:
                rc[5].button("Ver →", key=f"banco_btn_{conta}",
                             on_click=selecionar_banco, args=(conta,),
                             use_container_width=True)

        # Rodapé GRUPO (no mesmo card)
        st.markdown('<div style="border-top:1px solid #24427e; margin:0;"></div>', unsafe_allow_html=True)
        gc = st.columns(_ratios)
        gc[0].markdown(
            _cel(f'GRUPO &middot; {fechadas} de {len(contas_ordenadas)} fechadas', "left", "#f4c430", "500"),
            unsafe_allow_html=True)
        gc[1].markdown(_cel("&mdash;", "center", "#6f86ad"), unsafe_allow_html=True)
        gc[2].markdown(_cel("&mdash;", "center", "#6f86ad"), unsafe_allow_html=True)
        gc[3].markdown(_cel(str(total_itens), "center", "#fff", "500"), unsafe_allow_html=True)
        gc[4].markdown(_cel(fmt_brl(total_resolver), "right", "#ff8a8a", "500"), unsafe_allow_html=True)

    st.markdown(
        '<div style="color:#6f86ad; font-size:11px; margin-top:8px;">As % do grupo ficam vazias '
        'de propósito — média de contas diferentes engana. Cada conta tem seu botão "Ver →" à direita.</div>',
        unsafe_allow_html=True)




# ============================================================
# Resumo com termômetros (% explicado) — v7
# ============================================================
def render_resumo_termometros(resultado: ResultadoConciliacao, kpis: dict):
    """v7: cabeçalho do resumo com dois termômetros.

    % explicado = (movimentado - falta conciliar) / movimentado  → conta o cartão
    (explicado pelo agrupamento TOP 1722), não só os pares 1-a-1. Mostra a
    composição (pares × regra de cartão) pra continuar auditável.
    """
    total_banco = float(kpis.get("total_movimentado_banco", 0.0))
    falta = float(kpis.get("falta_conciliar", 0.0))
    total_sis = float(kpis.get("total_extrato_sistema", 0.0))
    diverg = float(kpis.get("divergencia_sankhya_banco", 0.0))

    explicado = 100.0 * (total_banco - falta) / total_banco if total_banco > 0 else 0.0
    pares = float(kpis.get("percentual_conciliado", 0.0))
    regra = max(0.0, explicado - pares)
    sankhya_conf = 100.0 * (total_sis - diverg) / total_sis if total_sis > 0 else 0.0

    qtd_div = int(kpis.get("qtd_divergencia_sankhya_banco", 0))
    qtd_pend = int(kpis.get("qtd_pendentes_banco", 0))
    precisa = qtd_div + qtd_pend

    cor_banco = "#0F8C3B" if explicado >= 99.95 else ("#FAC318" if explicado >= 80 else "#D63031")
    cor_sis = "#0F8C3B" if sankhya_conf >= 99.95 else ("#FAC318" if sankhya_conf >= 80 else "#D63031")

    section_title("RESUMO EXECUTIVO")

    # Selo de veredito
    if precisa == 0 and explicado >= 99.95 and sankhya_conf >= 99.95:
        st.html(
            '<div style="background:rgba(15,140,59,0.12); border:1px solid rgba(15,140,59,0.5); '
            'border-radius:12px; padding:12px 16px; margin-bottom:16px;">'
            '<span style="background:#0F8C3B; color:#fff; font-size:13px; font-weight:600; '
            'padding:4px 12px; border-radius:999px;">&#10003; Conciliação fechada</span>'
            '<span style="color:#7DD87D; font-size:13px; margin-left:10px;">'
            '100% do banco explicado &middot; R$ 0,00 sem explicação</span></div>'
        )
    else:
        st.html(
            '<div style="background:rgba(214,48,49,0.10); border:1px solid rgba(214,48,49,0.4); '
            'border-radius:12px; padding:12px 16px; margin-bottom:16px;">'
            '<span style="color:#FF8A8A; font-size:13px;">&#9888; ' + str(precisa) + ' '
            + ("item precisa" if precisa == 1 else "itens precisam") + ' da sua análise</span></div>'
        )

    def _termo(label, pct, cor, sub):
        return (
            '<div class="lle-kpi">'
            '<div class="lle-kpi-label">' + label + '</div>'
            '<div class="lle-kpi-value" style="color:' + cor + ';">' + fmt_pct(pct) + '</div>'
            '<div style="height:9px; background:rgba(255,255,255,0.06); border-radius:999px; '
            'overflow:hidden; margin:10px 0 0;">'
            '<div style="width:' + str(max(0.0, min(100.0, pct))) + '%; height:100%; background:' + cor + ';"></div></div>'
            '<div class="lle-kpi-suffix">' + sub + '</div></div>'
        )

    # v5.35: nomeia a regra que DE FATO explicou (boleto/cartão), em vez de cravar
    # "cartão". Se não houve agrupamento, mostra só os pares diretos.
    _regras_nomes = []
    if int(kpis.get("qtd_top1702_grupos", 0)) > 0:
        _regras_nomes.append("boleto")
    if int(kpis.get("qtd_top1722_grupos", 0)) > 0:
        _regras_nomes.append("cartão")
    # v5.49: estornos anulados também explicam volume do banco sem ser "pares"
    if int(kpis.get("qtd_estornos_anulados", 0)) > 0:
        _regras_nomes.append("estornos anulados")
    if regra > 0.05:
        _rot = ("por " + " e ".join(_regras_nomes)) if _regras_nomes else "por agrupamento"
        sub_banco = fmt_pct(pares) + " em pares diretos &middot; " + fmt_pct(regra) + " " + _rot
    else:
        sub_banco = fmt_pct(pares) + " em pares diretos"
    sub_sis = (str(qtd_div) + " lançamento" + ("s" if qtd_div != 1 else "") + " sem confirmação"
               if qtd_div else "0 lançamentos sem confirmação no banco")
    st.html(
        '<div style="display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-bottom:14px;">'
        + _termo("Banco explicado pelo ERP", explicado, cor_banco, sub_banco)
        + _termo("Sankhya confirmado no banco", sankhya_conf, cor_sis, sub_sis)
        + '</div>'
    )

    # Cards: Precisa de você + 2 volumes + Investimentos
    if precisa == 0:
        precisa_card = card_kpi("Precisa de você", "0 itens",
                                "nada pendente de decisão", classe="destaque-verde")
    else:
        precisa_card = card_kpi("Precisa de você",
                                str(precisa) + (" item" if precisa == 1 else " itens"),
                                fmt_brl(falta + diverg) + " a resolver", classe="destaque-vermelho")
    # v5.35: a diferença vai no card do lado que movimentou MAIS, e só cita "cartão"
    # quando a conta tem cartão. Sem cartão, fala a verdade: "a analisar".
    # v5.47: separa a parte que são pares anulados por estorno (PIX recebido e
    # devolvido — já explicados, não precisam de análise) do restante a analisar.
    diff_assinado = round(total_banco - total_sis, 2)
    abs_diff = abs(diff_assinado)
    _ea_g = getattr(resultado, "estornos_anulados", None)
    _vol_anul_g = 0.0
    if _ea_g is not None and not getattr(_ea_g, "empty", True):
        try:
            _vol_anul_g = round(
                float(_ea_g["valor_original"].abs().sum())
                + float(_ea_g["valor_estornado"].abs().sum()), 2
            )
        except Exception:
            _vol_anul_g = 0.0
    nota_diff = ""
    if abs_diff >= 0.01:
        _tem_cartao = int(kpis.get("qtd_top1722_grupos", 0)) > 0
        _lado = ("Banco movimentou " + fmt_brl(abs_diff) + " a mais que o Sankhya"
                 if diff_assinado > 0
                 else "Sankhya movimentou " + fmt_brl(abs_diff) + " a mais que o banco")
        if diff_assinado > 0 and _vol_anul_g >= 0.01:
            _resto = round(abs_diff - _vol_anul_g, 2)
            if _resto >= 0.01:
                nota_diff = (_lado + " — " + fmt_brl(_vol_anul_g)
                             + " são estornos anulados (recebido e devolvido); "
                             + fmt_brl(_resto) + " a analisar.")
            else:
                nota_diff = (_lado + " — 100% são estornos anulados "
                             "(recebido e devolvido); nada a analisar.")
        else:
            nota_diff = _lado + (" — pode incluir taxa de cartão lançada 2x; o restante a analisar."
                                 if _tem_cartao else " — a analisar.")
    if diff_assinado > 0:
        banco_card = card_kpi("Movimentado no Banco", fmt_brl(total_banco), nota_diff)
        sankhya_card = card_kpi("Movimentado no Sankhya", fmt_brl(total_sis))
    elif diff_assinado < 0:
        banco_card = card_kpi("Movimentado no Banco", fmt_brl(total_banco))
        sankhya_card = card_kpi("Movimentado no Sankhya", fmt_brl(total_sis), nota_diff)
    else:
        banco_card = card_kpi("Movimentado no Banco", fmt_brl(total_banco))
        sankhya_card = card_kpi("Movimentado no Sankhya", fmt_brl(total_sis))
    render_cards([precisa_card, banco_card, sankhya_card, _card_investimentos(resultado)])


# ============================================================
# Página: Conciliação
# ============================================================
def pagina_conciliacao():
    if st.session_state.fluxo_etapa == "resultado" and st.session_state.resultado is not None:
        tela_resultado()
    else:
        tela_upload()


def tela_upload():
    section_title("CONFIGURAR EXECUÇÃO")

    col_modo, col_data = st.columns([2, 1])
    with col_modo:
        modo = st.radio(
            "Modo de execução",
            ["1 conta por vez", "Várias contas de uma vez"],
            horizontal=True,
        )
    with col_data:
        # v5.10: período opcional — ignora lançamentos fora da janela (ex.: extrato
        # baixado até 02/07, mas você quer analisar só junho). Banco e Sankhya usam
        # a mesma data efetiva, então o filtro simples resolve.
        usar_periodo = st.checkbox(
            "Filtrar por período",
            value=False,
            help="Ignora lançamentos fora da janela, no extrato e no Sankhya. "
            "Útil quando o extrato baixa acumulando dias.",
        )
        if usar_periodo:
            cini, cfim = st.columns(2)
            periodo_ini = cini.date_input("Início", value=date.today().replace(day=1), format="DD/MM/YYYY")
            periodo_fim = cfim.date_input("Fim", value=date.today(), format="DD/MM/YYYY")
            data_ref = periodo_fim
        else:
            periodo_ini = periodo_fim = None
            data_ref = st.date_input(
                "Data de referência",
                value=date.today(),
                format="DD/MM/YYYY",
            )

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        section_title("EXTRATO BANCÁRIO")
        st.caption("Formato padronizado: Data, Histórico, Documento, Valor (R$). Aceita XLS, XLSX ou PDF do extrato bancário (Itaú, Sicredi, Bradesco, Caixa, Santander).")
        if modo == "1 conta por vez":
            # v5.48: aceita VÁRIOS arquivos da MESMA conta (ex.: Itaú PISA salvo
            # por dia — 01, 02, 03...). Todos entram com o mesmo identificador e
            # o app junta tudo antes de conciliar. Proteções:
            #   - arquivo idêntico subido 2x é ignorado (hash), senão as linhas
            #     contariam em dobro;
            #   - se algum arquivo parecer ser de OUTRA conta (agência/conta do
            #     cabeçalho diferentes), avisa antes de executar.
            _arqs_single = st.file_uploader(
                "Arraste o(s) extrato(s) — pode subir vários da MESMA conta (ex.: um por dia)",
                type=["xlsx", "xls", "pdf"],
                key="banco_single",
                accept_multiple_files=True,
                help="Todos os arquivos aqui devem ser da MESMA conta. O app "
                "junta tudo (ex.: extratos diários) e concilia como um período só. "
                "Para contas diferentes, use o modo 'Várias contas de uma vez'.",
            )
            if _arqs_single and not isinstance(_arqs_single, list):
                _arqs_single = [_arqs_single]
            _arqs_single = _arqs_single or []
            # dedup por conteúdo (mesmo arquivo subido em duplicata)
            if _arqs_single:
                import hashlib as _hl_single
                _vistos_single, _unicos, _dups = set(), [], []
                for _f in _arqs_single:
                    try:
                        _h = _hl_single.md5(_f.getvalue()).hexdigest()
                    except Exception:
                        _unicos.append(_f)
                        continue
                    if _h in _vistos_single:
                        _dups.append(_f.name)
                    else:
                        _vistos_single.add(_h)
                        _unicos.append(_f)
                if _dups:
                    st.info(
                        "ℹ️ Arquivo(s) idêntico(s) ignorado(s) para não contar em "
                        "dobro: " + ", ".join(sorted(set(_dups)))
                    )
                _arqs_single = _unicos
            arquivo_banco = _arqs_single[0] if _arqs_single else None
            # v5.8: auto-preenche o nome da conta a partir do nome do arquivo.
            # v5.9: e, quando dá, LÊ a conta do cabeçalho do próprio extrato
            # (banco + agência + conta), pra não depender do nome do arquivo.
            nome_default = ""
            conta_det = None
            if arquivo_banco is not None:
                nome_default = arquivo_banco.name.rsplit(".", 1)[0].strip()
                try:
                    import io as _io_det
                    from src.parsers.deteccao_conta import detectar_conta_extrato
                    _b = _io_det.BytesIO(arquivo_banco.getvalue())
                    _b.name = arquivo_banco.name
                    conta_det = detectar_conta_extrato(_b)
                    _cores_banco = {
                        "Bradesco": "#CC092F", "Santander": "#EC0000", "Sicredi": "#3FA110",
                        "Caixa": "#005CA9", "Itaú": "#EC7000",
                    }
                    if conta_det.confianca == "alta":
                        _cor = _cores_banco.get(conta_det.banco, "#6f88b8")
                        _emp = f" · {conta_det.empresa}" if conta_det.empresa else ""
                        _qtd_arqs = (f' · {len(_arqs_single)} arquivos'
                                     if len(_arqs_single) > 1 else '')
                        st.markdown(
                            f'<div style="display:flex;align-items:center;gap:12px;background:#0b2560;'
                            f'border-radius:10px;border-left:6px solid {_cor};padding:9px 13px;margin:2px 0 8px;">'
                            f'<span style="background:{_cor};color:#fff;font-size:11px;font-weight:700;'
                            f'letter-spacing:.03em;padding:3px 11px;border-radius:6px;">{conta_det.banco.upper()}</span>'
                            f'<span style="color:#eaf0fb;font-size:12.5px;">agência {conta_det.agencia} · '
                            f'conta {conta_det.conta}{_emp}{_qtd_arqs}</span></div>',
                            unsafe_allow_html=True,
                        )
                        if conta_det.identificador:
                            nome_default = conta_det.identificador
                    else:
                        # v5.54: se agência/conta FORAM lidas (só faltou o nome do
                        # banco — ex.: Santander diário, que não escreve o banco),
                        # o selo diz a verdade em vez de "não consegui ler".
                        if getattr(conta_det, "conta", "") or getattr(conta_det, "agencia", ""):
                            _pedacos = []
                            if conta_det.agencia:
                                _pedacos.append(f"agência {conta_det.agencia}")
                            if conta_det.conta:
                                _pedacos.append(f"conta {conta_det.conta}")
                            st.markdown(
                                '<div style="display:flex;align-items:center;gap:12px;background:#0b2560;'
                                'border-radius:10px;border-left:6px solid #FAC318;padding:9px 13px;margin:2px 0 8px;">'
                                '<span style="background:#FAC318;color:#041747;font-size:11px;font-weight:700;'
                                'letter-spacing:.03em;padding:3px 11px;border-radius:6px;">CONTA LIDA</span>'
                                f'<span style="color:#eaf0fb;font-size:12.5px;">{" · ".join(_pedacos)} '
                                '— o extrato não diz o banco; confirme o nome na lista abaixo</span></div>',
                                unsafe_allow_html=True,
                            )
                        else:
                            st.markdown(
                                '<div style="display:flex;align-items:center;gap:12px;background:#0b2560;'
                                'border-radius:10px;border-left:6px solid #6f88b8;padding:9px 13px;margin:2px 0 8px;">'
                                '<span style="background:#6f88b8;color:#041747;font-size:11px;font-weight:700;'
                                'letter-spacing:.03em;padding:3px 11px;border-radius:6px;">NÃO IDENTIFICADO</span>'
                                '<span style="color:#9fb3d6;font-size:12.5px;">não consegui ler a conta do cabeçalho '
                                '— confirme na lista abaixo</span></div>',
                                unsafe_allow_html=True,
                            )
                except Exception:
                    conta_det = None

            # v5.48: com vários arquivos, confere se TODOS parecem ser da mesma
            # conta (agência+conta do cabeçalho). Divergência = aviso, não trava
            # — o cabeçalho pode falhar num arquivo e estar certo nos outros.
            if len(_arqs_single) > 1 and conta_det is not None and getattr(conta_det, "confianca", "") == "alta":
                try:
                    import io as _io_chk
                    from src.parsers.deteccao_conta import detectar_conta_extrato as _det_chk
                    _divergentes = []
                    for _f in _arqs_single[1:]:
                        try:
                            _bc = _io_chk.BytesIO(_f.getvalue())
                            _bc.name = _f.name
                            _d2 = _det_chk(_bc)
                            if (getattr(_d2, "confianca", "") == "alta"
                                    and getattr(_d2, "conta_digitos", "")
                                    and getattr(conta_det, "conta_digitos", "")
                                    and _d2.conta_digitos != conta_det.conta_digitos):
                                _divergentes.append(f"{_f.name} (conta {_d2.conta})")
                        except Exception:
                            continue
                    if _divergentes:
                        st.warning(
                            "⚠️ Estes arquivos parecem ser de OUTRA conta e vão ser "
                            "somados como se fossem da mesma: "
                            + "; ".join(_divergentes)
                            + ". Se são contas diferentes, use o modo "
                            "'Várias contas de uma vez'."
                        )
                except Exception:
                    pass

            # v5.35: se o Sankhya já foi enviado, oferece os nomes de conta dele
            # numa lista (string idêntica à que o matcher usa), em vez de um campo
            # de texto onde o nome entra com underline/espaço/maiúscula errados.
            contas_sankhya = _ler_contas_sankhya_da_sessao()
            if contas_sankhya:
                opcoes = contas_sankhya + [_OPCAO_DIGITAR]
                idx_match = _melhor_match_conta(nome_default, contas_sankhya)
                # v5.9: se lemos o número da conta no extrato, ele é o critério
                # mais forte — casa com a conta do Sankhya que contém esse número.
                _casou_por_digitos = False
                if conta_det is not None and getattr(conta_det, "conta_digitos", ""):
                    import re as _re_conta
                    alvo = conta_det.conta_digitos
                    for _j, _opt in enumerate(contas_sankhya):
                        if alvo and alvo in _re_conta.sub(r"\D", "", str(_opt)):
                            idx_match = _j
                            _casou_por_digitos = True
                            break
                # v5.51: conta do Sankhya SEM número (ex.: "ITAU E.E. APOIO")
                # — casa pela EMPRESA do cabeçalho do extrato. Só seleciona se
                # exatamente UMA conta bater; ambiguidade = não chuta.
                if (not _casou_por_digitos and idx_match is None
                        and conta_det is not None and getattr(conta_det, "empresa", "")):
                    _emp_n = _norm_conta(conta_det.empresa)
                    _tok_banco = {"itau", "itaú", "bradesco", "sicredi",
                                  "santander", "caixa", "banco", "ag", "cc", "c", "conta"}
                    _cands_emp = []
                    for _j, _opt in enumerate(contas_sankhya):
                        _toks = [t for t in _norm_conta(_opt).split()
                                 if t not in _tok_banco and not t.isdigit() and len(t) > 1]
                        if _toks and all(t in _emp_n.split() for t in _toks):
                            _cands_emp.append(_j)
                    if len(_cands_emp) == 1:
                        idx_match = _cands_emp[0]
                # v5.53: caso ITAU KING — a conta 74839-5 pertence juridicamente
                # à LLE FERRAGENS, e o Sankhya a chama "ITAU KING": nem dígitos
                # nem empresa casam. No modo 1 conta, se o Sankhya enviado tem
                # EXATAMENTE UMA conta, ela é a conta desta conciliação —
                # pré-seleciona (a lista continua na tela para trocar).
                if idx_match is None and len(contas_sankhya) == 1:
                    idx_match = 0
                idx_default = idx_match if idx_match is not None else len(opcoes) - 1
                escolha = st.selectbox(
                    "Extrato Bancário (identificador da conta)",
                    opcoes,
                    index=idx_default,
                    key="conta_single_sel",
                    help="Nomes lidos do Sankhya. Escolha o da conta deste extrato "
                    "— assim casa sem erro de digitação.",
                )
                if escolha == _OPCAO_DIGITAR:
                    nome_conta = st.text_input(
                        "Nome da conta (nova — ainda não está no Sankhya)",
                        value=nome_default,
                        key="conta_single_novo",
                    ).strip()
                else:
                    nome_conta = escolha
            else:
                st.caption(
                    "Suba o relatório do Sankhya ao lado para escolher a conta "
                    "numa lista. Por enquanto, digite o identificador."
                )
                nome_conta = st.text_input(
                    "Extrato Bancário (identificador da conta)",
                    value=nome_default,
                    placeholder="ex: Bradesco-CC-12345",
                    key="conta_single",
                    help="Rótulo único da conta. Mínimo 3 caracteres.",
                ).strip()
            # v5.48: TODOS os arquivos entram com o MESMO identificador — o
            # pipeline concatena os extratos (ex.: diários) antes de conciliar.
            arquivos_banco = (
                [(nome_conta, _f) for _f in _arqs_single]
                if _arqs_single and nome_conta
                else []
            )
        else:
            contas_sankhya = _ler_contas_sankhya_da_sessao()
            arquivos_multi = st.file_uploader(
                "Arraste os extratos (um por conta)",
                type=["xlsx", "xls", "pdf"],
                accept_multiple_files=True,
                key="banco_multi",
            )
            if not (arquivos_multi or []):
                st.caption(
                    "Suba os extratos e o Sankhya — depois escolha a conta de "
                    "cada arquivo numa lista."
                )
            # v5.64: DEDUP por CONTEÚDO. Contas diferentes geram extratos com o
            # MESMO nome de arquivo (ex.: '01.07.2026.xls'), então o nome não
            # serve nem de chave nem de critério de duplicata — só descartamos
            # quando os BYTES são idênticos (o mesmo arquivo subido duas vezes).
            import hashlib as _hl_multi
            _unicos_m: list = []
            _dups_m: list = []
            _vistos_m: set = set()
            for _f in (arquivos_multi or []):
                try:
                    _h = _hl_multi.md5(_f.getvalue()).hexdigest()
                except Exception:
                    _h = None
                if _h is not None and _h in _vistos_m:
                    _dups_m.append(_f.name)
                    continue
                if _h is not None:
                    _vistos_m.add(_h)
                _unicos_m.append((_f, _h))
            if _dups_m:
                st.info(
                    "ℹ️ Arquivo(s) de conteúdo idêntico ignorado(s) para não "
                    "contar em dobro: " + ", ".join(sorted(set(_dups_m)))
                )
            # v5.64: detecção de cabeçalho POR ARQUIVO. Com nomes que são só
            # datas, o cabeçalho (agência/conta/empresa) é a única pista
            # confiável para pré-selecionar a conta certa do Sankhya.
            _det_por_arquivo: dict = {}
            _digitos_vistos: dict = {}
            for _f, _h in _unicos_m:
                try:
                    _det = _detecta_conta_bytes(_f.name, _f.getvalue())
                except Exception:
                    _det = None
                _det_por_arquivo[id(_f)] = _det
                _dig = getattr(_det, "conta_digitos", "") if _det is not None else ""
                if _dig:
                    _digitos_vistos.setdefault(_dig, []).append(_f.name)
            _repetidos = {d: ns for d, ns in _digitos_vistos.items() if len(ns) > 1}
            if _repetidos:
                _txt_rep = " · ".join(
                    "conta " + d + ": " + ", ".join(ns) for d, ns in _repetidos.items()
                )
                st.warning(
                    "⚠️ Arquivos que parecem ser da MESMA conta (mesmo número no "
                    "cabeçalho): " + _txt_rep + ". Se receberem o mesmo nome abaixo, "
                    "o app soma tudo como uma conta só (ex.: extratos diários). "
                    "Se não era a intenção, confira os arquivos."
                )
            arquivos_banco = []
            _widget_key_multi: dict = {}
            for _i_m, (f, _h) in enumerate(_unicos_m):
                # v5.64: chave do widget = impressão digital do CONTEÚDO — dois
                # arquivos com o mesmo nome (contas diferentes) não derrubam
                # mais a tela com StreamlitDuplicateElementKey.
                hkey = _h[:10] if _h else f"pos{_i_m}"
                _widget_key_multi[id(f)] = hkey
                base = f.name.rsplit(".", 1)[0].strip()
                _det = _det_por_arquivo.get(id(f))
                if contas_sankhya:
                    opcoes = contas_sankhya + [_OPCAO_DIGITAR]
                    idx_match = _melhor_match_conta(base, contas_sankhya)
                    # v5.64: pré-seleção pelo CABEÇALHO — mesma régua do modo
                    # '1 conta por vez': 1º número da conta (v5.9); 2º empresa
                    # com match ÚNICO (v5.51); 3º Sankhya com UMA conta só
                    # (v5.53). Ambiguidade = não chuta, cai em 'Digitar…'.
                    if idx_match is None and _det is not None and getattr(_det, "conta_digitos", ""):
                        import re as _re_m
                        _alvo_dig = _det.conta_digitos
                        for _j, _opt in enumerate(contas_sankhya):
                            if _alvo_dig and _alvo_dig in _re_m.sub(r"\D", "", str(_opt)):
                                idx_match = _j
                                break
                    if idx_match is None and _det is not None and getattr(_det, "empresa", ""):
                        _emp_n = _norm_conta(_det.empresa)
                        _tok_banco = {"itau", "itaú", "bradesco", "sicredi",
                                      "santander", "caixa", "banco", "ag", "cc", "c", "conta"}
                        _cands_emp = []
                        for _j, _opt in enumerate(contas_sankhya):
                            _toks = [t for t in _norm_conta(_opt).split()
                                     if t not in _tok_banco and not t.isdigit() and len(t) > 1]
                            if _toks and all(t in _emp_n.split() for t in _toks):
                                _cands_emp.append(_j)
                        if len(_cands_emp) == 1:
                            idx_match = _cands_emp[0]
                    if idx_match is None and len(contas_sankhya) == 1:
                        idx_match = 0
                    idx_default = (
                        idx_match if idx_match is not None else len(opcoes) - 1
                    )
                    # v5.64: transparência — mostra o que foi lido do cabeçalho
                    if _det is not None and (getattr(_det, "conta", "") or getattr(_det, "agencia", "")):
                        _ped = []
                        if getattr(_det, "banco", ""):
                            _ped.append(str(_det.banco))
                        if getattr(_det, "agencia", ""):
                            _ped.append("ag " + str(_det.agencia))
                        if getattr(_det, "conta", ""):
                            _ped.append("cc " + str(_det.conta))
                        if getattr(_det, "empresa", ""):
                            _ped.append(str(_det.empresa))
                        st.caption("“" + f.name + "” — cabeçalho: " + " · ".join(_ped))
                    escolha = st.selectbox(
                        f"Conta de “{f.name}”",
                        opcoes,
                        index=idx_default,
                        key=f"conta_multi_sel_{hkey}",
                        help="Nome da conta como aparece no Sankhya.",
                    )
                    if escolha == _OPCAO_DIGITAR:
                        nome_arq = st.text_input(
                            f"Nome da conta (nova) para “{f.name}”",
                            value=base,
                            key=f"conta_multi_novo_{hkey}",
                        ).strip()
                    else:
                        nome_arq = escolha
                else:
                    nome_arq = base
                if nome_arq:
                    arquivos_banco.append((nome_arq, f))

    with col2:
        section_title("EXTRATO SANKHYA CONCILIAÇÃO")
        st.caption("Relatório de Conciliação Bancária exportado do ERP.")
        arquivo_sistema = st.file_uploader(
            "Arraste o(s) relatório(s) do sistema",
            type=["xlsx", "xls"],
            key="sistema",
            accept_multiple_files=True,
            help="Pode subir mais de um (ex.: um relatório do Sankhya por conta/banco). O app junta todos antes de conciliar.",
        )
        coluna_conta_sistema = st.text_input(
            "Extrato Sankhya Conciliação — coluna da conta",
            value="",
            placeholder="(deixe vazio para auto-detectar)",
            key="coluna_conta_sistema",
            help="Nome exato da coluna do ERP que identifica a conta. Se vazio, tenta detectar.",
        )

    st.divider()
    section_title("EXTRATO DA ADQUIRENTE (OPCIONAL)")
    st.caption(
        "Só pra contas que recebem cartão. Aceita GetNet (Recebíveis) e PagBank/PagSeguro. "
        "Pode subir vários. Conta sem cartão não precisa subir nada — tudo funciona igual."
    )
    arquivos_adquirente = st.file_uploader(
        "Arraste o extrato da adquirente (Cielo / GetNet / PagBank)",
        type=["xlsx", "xls", "csv"],
        key="adquirente",
        accept_multiple_files=True,
        help="Usado pra dar NOME à diferença de cartão (aluguel, tarifa, estorno) na conciliação "
        "e pra alimentar a Auditoria de Cartões (taxa cobrada × taxa de contrato). "
        "Cielo: relatório 'Recebíveis Detalhado' (o líquido por dia bate com os depósitos do banco).",
    )

    st.divider()
    section_title("PENDÊNCIAS DE DIAS ANTERIORES (OPCIONAL)")
    arquivo_pendencias = st.file_uploader(
        "Relatório anterior (lê a aba 'Pendências Consolidadas')",
        type=["xlsx"],
        key="pendencias",
    )

    st.divider()

    # Validações
    erros_validacao: list[str] = []
    for nome, _ in arquivos_banco:
        erro = validar_nome_conta(nome)
        if erro:
            erros_validacao.append(erro)

    pode_executar = (
        bool(arquivos_banco) and bool(arquivo_sistema) and not erros_validacao
    )

    # v5.35: rede de segurança — nome do banco quase-igual a uma conta do Sankhya
    # (só muda espaço/underline/maiúscula). Avisa ANTES de conciliar e corrige
    # num clique, em vez de separar calado em duas contas.
    if contas_sankhya and arquivos_banco:
        norm_sankhya = {_norm_conta(c): c for c in contas_sankhya}
        _nomes_avisados: set[str] = set()  # v5.48: vários arquivos = mesma conta → 1 aviso só
        for _nome, _f in arquivos_banco:
            if _nome in contas_sankhya or _nome in _nomes_avisados:
                continue
            _alvo = norm_sankhya.get(_norm_conta(_nome))
            if _alvo and _alvo != _nome:
                _nomes_avisados.add(_nome)
                _c1, _c2 = st.columns([4, 1])
                with _c1:
                    st.warning(
                        f"⚠️ “{_nome}” parece ser a mesma conta que "
                        f"**{_alvo}** no Sankhya (só muda espaço, underline ou "
                        f"maiúscula). Com nomes diferentes elas viram duas contas "
                        f"e não conciliam."
                    )
                with _c2:
                    # v5.64: no modo múltiplo a chave do widget é por hash do
                    # conteúdo (nomes de arquivo se repetem entre contas).
                    _hk_fix = None
                    if modo != "1 conta por vez":
                        try:
                            _hk_fix = _widget_key_multi.get(id(_f))
                        except NameError:
                            _hk_fix = None
                    if st.button(f"Usar “{_alvo}”",
                                 key=f"fix_conta_{_hk_fix or _f.name}"):
                        if modo == "1 conta por vez":
                            st.session_state["conta_single_sel"] = _alvo
                        elif _hk_fix:
                            st.session_state[f"conta_multi_sel_{_hk_fix}"] = _alvo
                        else:
                            st.session_state[f"conta_multi_sel_{_f.name}"] = _alvo
                        st.rerun()

    # v5.8: mensagem espec\u00edfica em vez do gen\u00e9rico "Aguardando upload"
    if not pode_executar:
        faltando: list[str] = []
        # No modo "1 conta por vez": pode faltar o nome da conta mesmo com arquivo
        if modo == "1 conta por vez":
            tem_arquivo_banco = bool(locals().get("arquivo_banco"))
            tem_nome_conta = bool(locals().get("nome_conta", "").strip())
            if not tem_arquivo_banco:
                faltando.append("**Extrato Bancário** (.xlsx ou .xls)")
            elif not tem_nome_conta:
                faltando.append("**Identificador da conta** (campo de texto abaixo do extrato banc\u00e1rio)")
        else:
            if not arquivos_banco:
                faltando.append("**Extratos Banc\u00e1rios** (.xlsx ou .xls)")
        if not arquivo_sistema:
            faltando.append("**Relat\u00f3rio Sankhya** (.xlsx ou .xls)")

        if faltando and not erros_validacao:
            if len(faltando) == 1:
                st.warning(f"⏳ Falta: {faltando[0]}")
            else:
                lista = "\n".join(f"- {f}" for f in faltando)
                st.warning(f"⏳ Para executar a conciliação ainda falta:\n\n{lista}")
        elif erros_validacao:
            for e in erros_validacao:
                st.error(f"❌ {e}")

    # Para sidebar — recupera config
    rodar_fuzzy = st.session_state.get("rodar_fuzzy", True)
    tolerancia = int(st.session_state.get("tolerancia", 2))

    if st.button(
        "▶️ Executar conciliação",
        type="primary",
        disabled=not pode_executar,
        use_container_width=True,
    ):
        with st.spinner("Processando... isso pode levar alguns segundos."):
            try:
                dfs_banco = []
                _previas_removidas = 0
                for nome_conta, arq in arquivos_banco:
                    df = carregar_extrato_banco(
                        arq, conta=nome_conta, ano_referencia=data_ref.year
                    )
                    # v5.62: extrato consolidado NÃO contém futuro. Arquivos
                    # diários nomeados DD_MM_AAAA (Bradesco/Itaú) podem trazer
                    # PRÉVIA do próximo dia útil (Bradesco: lançamentos futuros
                    # com dcto 2374xx) — o consolidado vem no arquivo do próprio
                    # dia, com outro histórico/ordem/saldo, duplicando tudo.
                    # Descarta lançamentos com data POSTERIOR à data do arquivo
                    # (sanidade: só se o nome condiz com o período do conteúdo).
                    _m_dt = re.search(r"(\d{2})[_\-.](\d{2})[_\-.](\d{4})", str(getattr(arq, "name", "")))
                    if _m_dt is not None and "data" in df.columns:
                        try:
                            _dt_arq = pd.Timestamp(int(_m_dt.group(3)), int(_m_dt.group(2)), int(_m_dt.group(1)))
                            _dts = pd.to_datetime(df["data"], errors="coerce")
                            if _dts.notna().any() and _dts.min() <= _dt_arq:
                                _fut = _dts > _dt_arq
                                if bool(_fut.any()):
                                    _previas_removidas += int(_fut.sum())
                                    df = df[~_fut].reset_index(drop=True)
                        except (ValueError, TypeError):
                            pass
                    # v5.61: guarda a origem pra detectar sobreposição de períodos
                    df["_arq_origem"] = getattr(arq, "name", "")
                    dfs_banco.append(df)
                banco = pd.concat(dfs_banco, ignore_index=True) if dfs_banco else pd.DataFrame()
                if _previas_removidas:
                    st.info(
                        f"ℹ️ {_previas_removidas} lançamento(s) de PRÉVIA descartado(s) — "
                        "data posterior à do arquivo (ex.: Bradesco mostra o próximo dia "
                        "útil como prévia; o consolidado entra no arquivo do próprio dia)."
                    )

                # v5.61/v5.62: DEDUP ENTRE ARQUIVOS, com prova de POSIÇÃO NA
                # CADEIA: mesma conta+data+valor+SALDO CORRENTE = a mesma
                # transação, mesmo que o histórico mude entre arquivos (a cauda
                # do Bradesco reescreve 'APLIC.AUTOM.INVESTFACIL*' como
                # 'APLIC.INVEST FACIL'). Duas transações reais de mesmo valor no
                # mesmo dia têm saldos correntes DIFERENTES — chave é segura.
                # Linha SEM saldo nunca é removida (sem prova, não apaga).
                if not banco.empty and "_saldo_linha" in banco.columns:
                    _antes_x = len(banco)
                    _tem_sx = banco["_saldo_linha"].notna()
                    _dup_x = banco.duplicated(
                        subset=["conta", "data", "valor", "_saldo_linha"],
                        keep="first",
                    ) & _tem_sx
                    if bool(_dup_x.any()):
                        banco = banco[~_dup_x].reset_index(drop=True)
                        st.info(
                            f"ℹ️ {_antes_x - len(banco)} lançamento(s) removido(s) por estarem "
                            "repetidos em mais de um arquivo (períodos sobrepostos) — "
                            "mesma transação comprovada pelo saldo corrente."
                        )
                    # sem saldo na linha não há prova — só avisa a possível duplicidade
                    try:
                        _sem_s = banco[banco["_saldo_linha"].isna()
                                       & ~banco["historico"].astype(str).str.contains("SALDO", na=False)]
                        if not _sem_s.empty and "_arq_origem" in _sem_s.columns:
                            _gx = _sem_s.groupby(["conta", "data", "historico", "valor"])["_arq_origem"].nunique()
                            _susp = _gx[_gx > 1]
                            if len(_susp):
                                st.warning(
                                    f"⚠️ {int(len(_susp))} lançamento(s) aparecem em MAIS DE UM arquivo "
                                    "da mesma conta e podem estar duplicados (esse extrato não traz "
                                    "saldo por linha, então não removi nada). Confira se os arquivos "
                                    "enviados cobrem períodos sobrepostos."
                                )
                    except Exception:
                        pass
                if not banco.empty:
                    banco = banco.drop(columns=["_saldo_linha", "_arq_origem"], errors="ignore")

                # v3.6: força tipos consistentes após concat de múltiplos extratos.
                # Sem isso, se um extrato vier com 'data' como string e outro como datetime,
                # o merge no pipeline quebra com 'merge on object and datetime64 columns'.
                if not banco.empty:
                    if "data" in banco.columns:
                        banco["data"] = pd.to_datetime(banco["data"], errors="coerce")
                    if "valor" in banco.columns:
                        banco["valor"] = pd.to_numeric(banco["valor"], errors="coerce").fillna(0.0)
                    # Remove linhas com data inválida (NaT) — não dá pra conciliar sem data
                    linhas_antes = len(banco)
                    banco = banco.dropna(subset=["data"]).reset_index(drop=True)
                    linhas_descartadas = linhas_antes - len(banco)
                    if linhas_descartadas > 0:
                        st.warning(
                            f"⚠️ {linhas_descartadas} linha(s) do extrato bancário "
                            f"foram descartadas por terem data inválida ou vazia."
                        )

                # Múltiplos relatórios do Sankhya: carrega cada um e junta tudo.
                _lista_sistema = arquivo_sistema if isinstance(arquivo_sistema, list) else [arquivo_sistema]
                _dfs_sistema = []
                for _arq_sis in _lista_sistema:
                    _dfs_sistema.append(
                        carregar_relatorio_sistema(
                            _arq_sis,
                            coluna_conta=coluna_conta_sistema or None,
                        )
                    )
                sistema = (
                    pd.concat(_dfs_sistema, ignore_index=True)
                    if _dfs_sistema else pd.DataFrame()
                )

                # v3.6: mesmo tratamento para o sistema
                if not sistema.empty:
                    if "data" in sistema.columns:
                        sistema["data"] = pd.to_datetime(sistema["data"], errors="coerce")
                    if "valor" in sistema.columns:
                        sistema["valor"] = pd.to_numeric(sistema["valor"], errors="coerce").fillna(0.0)
                    sistema = sistema.dropna(subset=["data"]).reset_index(drop=True)

                # v5.10: aplica o filtro de período (quando ligado) — ignora tudo
                # fora da janela, no extrato E no Sankhya. Como os dois usam a mesma
                # data efetiva, o corte simples fecha certo (feriado empurra p/ o
                # próximo período nos dois lados). O saldo se reajusta às bordas
                # porque saldo_final_da_conta pega a 1ª/última linha de saldo por data.
                if usar_periodo and periodo_ini and periodo_fim:
                    _ini = pd.Timestamp(periodo_ini).normalize()
                    _fim = pd.Timestamp(periodo_fim).normalize()
                    _ab, _as = len(banco), len(sistema)
                    if not banco.empty and "data" in banco.columns:
                        _dn = banco["data"].dt.normalize()
                        banco = banco[(_dn >= _ini) & (_dn <= _fim)].reset_index(drop=True)
                    if not sistema.empty and "data" in sistema.columns:
                        _dn = sistema["data"].dt.normalize()
                        sistema = sistema[(_dn >= _ini) & (_dn <= _fim)].reset_index(drop=True)
                    st.info(
                        f"📅 Período {periodo_ini.strftime('%d/%m/%Y')}–{periodo_fim.strftime('%d/%m/%Y')}: "
                        f"extrato {_ab}→{len(banco)} linha(s) · Sankhya {_as}→{len(sistema)} linha(s). "
                        "Lançamentos fora da janela foram ignorados."
                    )

                if modo == "1 conta por vez" and not sistema.empty and (sistema["conta"] == "—").all():
                    sistema["conta"] = arquivos_banco[0][0]
                    st.info(
                        f"ℹ️ Relatório do sistema sem coluna de conta — atribuído a '{arquivos_banco[0][0]}'."
                    )
                elif modo != "1 conta por vez" and not sistema.empty:
                    # Modo várias contas: avisar se há contas no Sankhya que não existem no banco
                    contas_banco = {nome for nome, _ in arquivos_banco}
                    contas_sis = set(sistema["conta"].unique()) - {"—"}
                    if contas_sis and not (contas_sis & contas_banco):
                        st.warning(
                            f"⚠️ As contas no Sankhya **{sorted(contas_sis)}** não "
                            f"coincidem com nenhum dos identificadores dos extratos "
                            f"bancários enviados **{sorted(contas_banco)}**. Resultado "
                            f"será 0% conciliado. Verifique se o identificador da conta "
                            f"informado para cada extrato bancário corresponde ao valor "
                            f"da coluna de conta no relatório do Sankhya."
                        )
                    elif (sistema["conta"] == "—").any():
                        qtd = int((sistema["conta"] == "—").sum())
                        msg = (
                            f"⚠️ {qtd} linha(s) do Sankhya estão sem identificador de "
                            f"conta."
                        )
                        if coluna_conta_sistema:
                            msg += (
                                f" Você informou que a coluna da conta no Sankhya é "
                                f"**'{coluna_conta_sistema}'** mas ela não foi encontrada "
                                f"no arquivo ou veio vazia. Verifique o nome exato da "
                                f"coluna no Sankhya (atenção a maiúsculas e acentos)."
                            )
                        else:
                            msg += (
                                " Inclua uma coluna 'Conta' na planilha do Sankhya antes "
                                "de subir, ou informe o nome dela no campo "
                                "'Extrato Sankhya — coluna da conta', ou troque para o "
                                "modo '1 conta por vez'."
                            )
                        st.warning(msg)

                pendencias = carregar_pendencias_anteriores(arquivo_pendencias)

                # Detecta reprocessamento (já existe execução anterior nesta sessão)
                exec_anterior = st.session_state.id_execucao_atual
                versao = 1
                id_origem = None
                if exec_anterior:
                    versao = 2  # versionamento simples
                    id_origem = exec_anterior

                # v5.59: tarifas comprovadas pela adquirente permitem fechar o
                # dia de cartão líquido de tarifa (regra da Débora, 14/07).
                _tarifas_adq_exec = None
                try:
                    _adq_exec = _adquirentes_da_sessao()
                    if (_adq_exec is not None and not getattr(_adq_exec, "empty", True)
                            and "categoria" in _adq_exec.columns):
                        _tarifas_adq_exec = _adq_exec[
                            _adq_exec["categoria"].isin(["aluguel", "tarifa"])
                        ][["data", "valor"]].copy()
                        if _tarifas_adq_exec.empty:
                            _tarifas_adq_exec = None
                except Exception:
                    _tarifas_adq_exec = None
                resultado = executar_pipeline(
                    banco, sistema,
                    data_referencia=datetime.combine(data_ref, datetime.min.time()),
                    tolerancia_dias=tolerancia,
                    rodar_fuzzy=rodar_fuzzy,
                    tarifas_adquirente=_tarifas_adq_exec,
                )

                # Fase 1: rede de segurança — avisa se banco e Sankhya cobrem períodos
                # diferentes (descasamento infla a 'Falta Conciliar', como no susto do PDF).
                st.session_state.aviso_periodo = _checar_periodo(banco, sistema)

                id_exec = novo_id_execucao()

                # v5.34: NÃO gera o Excel aqui. Em volume alto (ex.: Santander, 12k+69k)
                # a geração levava ~3 min e segurava a tela em "Processando...". O Excel
                # passa a ser gerado sob demanda, quando o usuário clica em "Gerar Excel".

                # Snapshot append-only
                try:
                    salvar_snapshot(
                        id_exec, banco, sistema,
                        parametros={
                            "data_referencia": resultado.data_referencia,
                            "tolerancia_dias": tolerancia,
                            "rodar_fuzzy": rodar_fuzzy,
                            "modo": modo,
                            "contas": resultado.contas_processadas,
                            "versao": versao,
                            "id_origem": id_origem,
                        },
                        relatorio_xlsx=None,
                    )
                    registrar_execucao(
                        id_exec=id_exec,
                        data_referencia=resultado.data_referencia,
                        contas=resultado.contas_processadas,
                        tolerancia_dias=tolerancia,
                        kpis=resultado.kpis_globais(),
                        arquivos_inputs=[n for n, _ in arquivos_banco] + ["sistema.xlsx"],
                        status="reprocessado" if versao > 1 else "processado",
                        versao=versao,
                        id_origem=id_origem,
                    )
                except Exception as e:
                    st.warning(f"⚠️ Auditoria não pôde ser salva no disco: {e}. "
                               "Os resultados estão disponíveis para download mesmo assim.")

                st.session_state.resultado = resultado
                st.session_state.adquirente_df = _adquirentes_da_sessao()
                # v5.43: guarda os BYTES crus dos extratos de adquirente numa chave
                # própria (não some ao trocar de página) pra a Auditoria de Cartões
                # reaproveitar sem exigir novo upload.
                _adq_files = st.session_state.get("adquirente") or []
                if _adq_files and not isinstance(_adq_files, list):
                    _adq_files = [_adq_files]
                try:
                    st.session_state.adquirente_bytes = [
                        (f.name, f.getvalue()) for f in _adq_files
                    ]
                except Exception:
                    st.session_state.adquirente_bytes = []
                st.session_state.pendencias_anteriores = pendencias
                st.session_state.id_execucao_atual = id_exec
                st.session_state.xlsx_atual = None  # v5.34: gera sob demanda (volume)
                st.session_state.csvs_zip_atual = None  # força regerar quando precisar
                st.session_state.fluxo_etapa = "resultado"
                st.rerun()

            except Exception as e:
                st.error(f"❌ Erro durante o processamento:\n\n```\n{e}\n```")
                import traceback
                with st.expander("Stack trace completo"):
                    st.code(traceback.format_exc())


# ============================================================
# Tela de RESULTADO (única, pós-upload)
# ============================================================
def _checar_periodo(banco: pd.DataFrame, sistema: pd.DataFrame) -> str | None:
    """Compara o intervalo de datas do banco e do Sankhya; retorna um aviso se
    cobrirem períodos diferentes (causa comum de 'Falta Conciliar' inflada)."""
    try:
        if "data" not in banco.columns or "data" not in sistema.columns:
            return None
        bmin, bmax = banco["data"].min(), banco["data"].max()
        smin, smax = sistema["data"].min(), sistema["data"].max()
        if pd.isna(bmin) or pd.isna(bmax) or pd.isna(smin) or pd.isna(smax):
            return None
        if abs((bmax - smax).days) > 3 or abs((bmin - smin).days) > 3:
            return (
                f"⚠️ **Períodos diferentes.** O extrato do banco vai de "
                f"{bmin:%d/%m/%Y} a {bmax:%d/%m/%Y} e o Sankhya de "
                f"{smin:%d/%m/%Y} a {smax:%d/%m/%Y}. Quando os dois não cobrem o "
                f"mesmo período, sobram lançamentos sem par e a 'Falta Conciliar' "
                f"sobe. Confira se os arquivos são do mesmo mês."
            )
    except Exception:
        return None
    return None


def tela_resultado():
    resultado: ResultadoConciliacao = st.session_state.resultado
    kpis = resultado.kpis_globais()

    aviso_periodo = st.session_state.get("aviso_periodo")
    if aviso_periodo:
        st.warning(aviso_periodo)

    # Topo: botão voltar (menor) + ações principais (em destaque, amarelas)
    col_top1, col_top2, col_top3 = st.columns([1, 2, 2])
    with col_top1:
        st.button(
            "← Nova conciliação",
            on_click=voltar_upload,
            use_container_width=True,
            key="btn_voltar_topo",
        )
    with col_top2:
        nome = f"conciliacao_{resultado.data_referencia.strftime('%Y%m%d')}.xlsx"
        if st.session_state.get("xlsx_atual"):
            st.download_button(
                "⬇️ Excel (todas as abas)",
                data=st.session_state.xlsx_atual,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        elif st.button("⬇️ Gerar Excel", type="primary", use_container_width=True,
                       key="btn_gerar_xlsx"):
            with st.spinner("Gerando Excel… em volume grande pode levar alguns minutos."):
                st.session_state.xlsx_atual = gerar_relatorio_excel(
                    resultado,
                    pendencias_anteriores=st.session_state.pendencias_anteriores,
                )
            st.rerun()
    with col_top3:
        nome_zip = f"conciliacao_{resultado.data_referencia.strftime('%Y%m%d')}_csvs.zip"
        if st.session_state.get("csvs_zip_atual"):
            st.download_button(
                "⬇️ CSVs (zip)",
                data=st.session_state.csvs_zip_atual,
                file_name=nome_zip,
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )
        elif st.button("⬇️ Gerar CSVs (zip)", type="primary", use_container_width=True,
                       key="btn_gerar_csvs"):
            with st.spinner("Gerando CSVs…"):
                st.session_state.csvs_zip_atual = gerar_csvs_zip(resultado)
            st.rerun()

    st.divider()

    # v3.9: RESUMO EXECUTIVO global só aparece quando NENHUMA conta está selecionada.
    # Quando o usuário entra no detalhamento de uma conta específica, mostra direto
    # o detalhe da conta — evita confundir KPIs globais com KPIs da conta selecionada.
    if not st.session_state.banco_conta_selecionada:
        render_resumo_termometros(resultado, kpis)

        # v5.14: Removida a seção "Exceções e Regras Aplicadas" (estornos + TOP 1722).
        # A lógica continua rodando — mas o resumo não exibe mais esses cards.
        # TOP 1722 aparece na aba dedicada dentro do detalhamento da conta.

    # Painel de bancos OU detalhe do banco selecionado
    if st.session_state.banco_conta_selecionada:
        tela_detalhamento_banco(resultado, st.session_state.banco_conta_selecionada)
    else:
        section_title("CONTAS PROCESSADAS — CLIQUE PARA DETALHAR")
        render_painel_bancos(resultado)
        # v5.14: Removida "CONCILIAÇÃO POR TIPO DE LANÇAMENTO" do resumo.
        # Os mesmos dados aparecem dentro do detalhamento de cada conta (sem duplicar).


# ============================================================
# Detalhamento por banco — cards + abas internas
# ============================================================
def _volume_estornos_anulados_por_dia(resultado, conta: str) -> dict:
    """v5.47: volume (valor absoluto das DUAS pontas) dos pares anulados por
    estorno, por dia. Um PIX recebido e devolvido conta 2× no volume do banco,
    mas o par se anula e já está explicado na aba ♻️ Estornos — por isso NÃO é
    diferença Banco × Sankhya e precisa ser descontado do banner e do detalhe
    por dia. Sem esse desconto, um dia 100% de estornos aparecia com valor alto
    e nenhuma linha listada (o número do dia não fechava com as linhas)."""
    ea = getattr(resultado, "estornos_anulados", None)
    if ea is None or getattr(ea, "empty", True):
        return {}
    df = ea
    if "conta" in df.columns:
        df = df[df["conta"] == conta]
    if df.empty:
        return {}
    vol: dict = {}
    for _, r in df.iterrows():
        try:
            d1 = pd.Timestamp(r["data_original"]).normalize()
            d2 = pd.Timestamp(r["data_estorno"]).normalize()
            vol[d1] = vol.get(d1, 0.0) + abs(float(r["valor_original"]))
            vol[d2] = vol.get(d2, 0.0) + abs(float(r["valor_estornado"]))
        except Exception:
            continue
    return vol


def _diferenca_bxs_por_dia(resultado: "ResultadoConciliacao", conta: str) -> list[tuple]:
    """Diferença |banco| − |Sankhya| por dia, com o MESMO filtro do KPI de
    'movimentado' (exclui saldo/aplicação/resgate/rendimento) e DESCONTANDO os
    pares anulados por estorno (v5.47 — eles se anulam e não são diferença).
    Retorna lista de (data, dif) só com os dias que divergem (|dif| >= 0,01),
    ordenada do maior para o menor em módulo. Serve pra apontar no alerta EM QUE
    DIA está a diferença — e agora o valor de cada dia fecha exatamente com as
    linhas listadas no detalhe."""
    _excl = ("saldo", "aplicacao", "resgate", "rendimento")

    def _abs_por_dia(df):
        if df is None or getattr(df, "empty", True):
            return pd.Series(dtype=float)
        d = df
        if "conta" in d.columns:
            d = d[d["conta"] == conta]
        if "categoria_mov" in d.columns:
            d = d[~d["categoria_mov"].isin(_excl)]
        if d.empty:
            return pd.Series(dtype=float)
        dia = pd.to_datetime(d["data"], errors="coerce").dt.normalize()
        return d["valor"].abs().groupby(dia).sum()

    gb = _abs_por_dia(resultado.banco_completo)
    gs = _abs_por_dia(resultado.sistema_completo)
    vol_anulados = _volume_estornos_anulados_por_dia(resultado, conta)
    dias = sorted(set(gb.index) | set(gs.index))
    out = []
    for dia in dias:
        dif = round(
            float(gb.get(dia, 0.0))
            - float(vol_anulados.get(dia, 0.0))
            - float(gs.get(dia, 0.0)),
            2,
        )
        if abs(dif) >= 0.01:
            out.append((dia, dif))
    out.sort(key=lambda x: abs(x[1]), reverse=True)
    return out


def _explicar_diferenca_cartao(resultado, conta: str, adq: pd.DataFrame):
    """Tenta explicar a diferença de volume de cada dia pela adquirente.

    Mecanismo confirmado: tarifa/aluguel do GetNet entram 2× no volume (o Sankhya
    lança a tarifa como despesa E dentro do bruto; o banco já vem líquido). Então
    se a diferença do dia == 2 × (aluguel+tarifa GetNet daquele dia), está EXPLICADA.
    Só marca 'explicado' quando bate EXATO — nunca crava.

    Retorna (partes, todos_explicados):
      partes = [(data, dif_assinada, texto_ou_None), ...]  (ordenada por |dif|)
    """
    dias = _diferenca_bxs_por_dia(resultado, conta)
    if not dias:
        return [], True
    tem_adq = adq is not None and not adq.empty
    adq_dt = (
        pd.to_datetime(adq["data"], errors="coerce").dt.normalize() if tem_adq else None
    )
    partes = []
    todos = True
    for d, v in dias:
        alvo = round(abs(v), 2)
        expl = None
        if tem_adq:
            gn = adq[
                (adq_dt == pd.Timestamp(d))
                & (adq["adquirente"] == "GetNet")
                & (adq["categoria"].isin(["aluguel", "tarifa"]))
            ]
            soma = round(float(gn["valor"].abs().sum()), 2)
            if soma > 0 and round(2 * soma, 2) == alvo:
                itens = "; ".join(
                    f"{desc} ({fmt_brl(abs(val))})"
                    for desc, val in zip(gn["descricao"], gn["valor"])
                )
                expl = (
                    itens
                    + " — a adquirente descontou essa taxa <b>dentro de um repasse do dia</b>: "
                    "entrou o valor cheio e saiu a taxa (uma entrada e uma saída do mesmo valor). "
                    "O Sankhya registra as duas pontas; o banco já recebe líquido — por isso conta 2× no volume."
                )
        if expl is None:
            todos = False
        partes.append((d, v, expl))
    return partes, todos


def _esc(s) -> str:
    """Escapa texto para inserção segura em HTML (histórico do lançamento)."""
    import html as _html
    return _html.escape(str(s if s is not None else ""))


def _linhas_pendentes_do_dia(resultado, conta, dia, lado):
    """[(historico, valor), ...] das linhas que NÃO casaram naquele dia.
    lado='banco' -> pendentes_banco (estão no banco, faltam no Sankhya);
    lado='sankhya' -> pendentes_sistema (baixados no Sankhya, sem par no banco).
    Exclui saldo/aplicação/resgate/rendimento (mesmo filtro do KPI de movimentado).
    """
    _excl = ("saldo", "aplicacao", "resgate", "rendimento")
    df = resultado.pendentes_banco if lado == "banco" else resultado.pendentes_sistema
    if df is None or getattr(df, "empty", True):
        return []
    d = df
    if "conta" in d.columns:
        d = d[d["conta"] == conta]
    if "categoria_mov" in d.columns:
        d = d[~d["categoria_mov"].isin(_excl)]
    if d.empty or "data" not in d.columns:
        return []
    dd = pd.to_datetime(d["data"], errors="coerce").dt.normalize()
    d = d[dd == pd.Timestamp(dia)]
    if d.empty:
        return []
    linhas = [(str(h), float(v)) for h, v in zip(d["historico"], d["valor"])]

    # v5.56: COMPACTA o cartão — em dia de cartão que não fechou, listar 113
    # notas uma a uma inunda o alerta e esconde o resto. Cartão vira UMA linha
    # de resumo por lado (contagem + soma); o que não é cartão continua linha a
    # linha. Nada é omitido: a soma exibida é a soma real das linhas.
    try:
        if lado == "banco":
            from src.matching.cartao_top1722 import _eh_cartao_no_banco as _ehc
            _m = [(_ehc(h) and v > 0) for h, v in linhas]
        else:
            if "top_baixa" in d.columns:
                _m = (d["top_baixa"].astype(str).str.strip() == "1722").tolist()
            else:
                _m = [False] * len(linhas)
        _n_cart = sum(_m)
        if _n_cart > 3:
            _soma_cart = round(sum(v for (_, v), f in zip(linhas, _m) if f), 2)
            _resto = [lv for lv, f in zip(linhas, _m) if not f]
            if lado == "banco":
                _rotulo = (f"— {_n_cart} depósitos de CARTÃO (adquirente) somando · "
                           f"suba o extrato da adquirente na conciliação p/ detalhar")
            else:
                _rotulo = (f"— {_n_cart} notas de CARTÃO baixadas (TOP 1722) somando · "
                           f"suba o extrato da adquirente na conciliação p/ detalhar")
            linhas = [(_rotulo, _soma_cart)] + _resto
    except Exception:
        pass
    return linhas


def _explicar_diferenca_por_dia(resultado, conta):
    """Explica a diferença Banco×Sankhya de cada dia pela ORIGEM real das linhas
    que não casaram — SEM supor cartão. Retorna lista de dicts ordenada por |dif|:
      {data, dif, banco:[(h,v)], sankhya:[(h,v)], divergentes:[(h,vb,vs)]}
    banco    -> lançamento no banco sem baixa no Sankhya (ex.: despesa não lançada)
    sankhya  -> baixado no Sankhya, sem par no banco
    divergentes -> mesmo lançamento com valor diferente nos dois lados
    v5.49: dias com PENDÊNCIA entram mesmo quando a diferença líquida do dia é
    zero (lados que se compensam) — antes esses dias sumiam do detalhe.
    """
    dias_dif = dict(_diferenca_bxs_por_dia(resultado, conta))
    # dias com pendências (banco/sankhya), ainda que a dif líquida seja 0
    _excl = ("saldo", "aplicacao", "resgate", "rendimento")
    dias_pend: set = set()
    for _df in (getattr(resultado, "pendentes_banco", None),
                getattr(resultado, "pendentes_sistema", None)):
        if _df is None or getattr(_df, "empty", True) or "data" not in _df.columns:
            continue
        d = _df
        if "conta" in d.columns:
            d = d[d["conta"] == conta]
        if "categoria_mov" in d.columns:
            d = d[~d["categoria_mov"].isin(_excl)]
        if not d.empty:
            dias_pend.update(pd.to_datetime(d["data"], errors="coerce").dropna().dt.normalize().unique())
    todos_dias = set(dias_dif) | {pd.Timestamp(x) for x in dias_pend}
    if not todos_dias:
        return []
    dv = getattr(resultado, "divergencias", None)
    cols_dv = list(getattr(dv, "columns", [])) if dv is not None else []
    tem_dv = dv is not None and not getattr(dv, "empty", True) and "valor_sistema" in cols_dv
    if tem_dv and "conta" in cols_dv:
        dv = dv[dv["conta"] == conta]
    out = []
    for dia in sorted(todos_dias):
        dif = float(dias_dif.get(dia, 0.0))
        divergentes = []
        if tem_dv and "data" in cols_dv:
            dd = pd.to_datetime(dv["data"], errors="coerce").dt.normalize()
            for _, row in dv[dd == pd.Timestamp(dia)].iterrows():
                h = row.get("historico_banco") or row.get("historico_sistema") or ""
                divergentes.append((str(h), row.get("valor_banco"), row.get("valor_sistema")))
        item = {
            "data": dia,
            "dif": dif,
            "banco": _linhas_pendentes_do_dia(resultado, conta, dia, "banco"),
            "sankhya": _linhas_pendentes_do_dia(resultado, conta, dia, "sankhya"),
            "divergentes": divergentes,
        }
        # v5.58: quando os DOIS lados do dia têm cartão, junta num bloco único —
        # o que importa é o RESÍDUO (ex.: 0,45), não as somas cheias, que davam
        # a impressão de o cartão inteiro estar pendente.
        try:
            import re as _re_c
            def _pega_cartao(lista):
                for idx, (h, v) in enumerate(lista):
                    if str(h).startswith("—") and "CARTÃO" in str(h):
                        m = _re_c.search(r"(\d+)", str(h))
                        return idx, int(m.group(1)) if m else 0, float(v)
                return None, 0, 0.0
            _ib, _qb, _vb = _pega_cartao(item["banco"])
            _is, _qs, _vs = _pega_cartao(item["sankhya"])
            if _ib is not None and _is is not None:
                item["banco"] = [x for j, x in enumerate(item["banco"]) if j != _ib]
                item["sankhya"] = [x for j, x in enumerate(item["sankhya"]) if j != _is]
                item["cartao"] = {
                    "qtd_banco": _qb, "soma_banco": round(_vb, 2),
                    "qtd_sankhya": _qs, "soma_sankhya": round(_vs, 2),
                    "residuo": round(_vs - _vb, 2),  # >0 = Sankhya baixou a mais
                }
        except Exception:
            pass
        # v5.57: se a adquirente (GetNet/Cielo) foi enviada, confere se
        # aluguel/tarifa DELA naquele dia explica EXATAMENTE o resíduo de
        # cartão do dia (Sankhya baixou as vendas cheias, a adquirente
        # descontou a tarifa do repasse). Só anota com prova ao centavo.
        try:
            _adq = st.session_state.get("adquirente_df")
            _cart = item.get("cartao")
            if (_adq is not None and not getattr(_adq, "empty", True)
                    and "categoria" in _adq.columns and _cart):
                _residuo_c = float(_cart["residuo"])  # >0 = Sankhya baixou a mais
                if _residuo_c > 0.004:
                    _da = pd.to_datetime(_adq["data"], errors="coerce").dt.normalize()
                    _t = _adq[(_da == pd.Timestamp(dia)) & (_adq["categoria"].isin(["aluguel", "tarifa"]))]
                    _soma_t = round(float(pd.to_numeric(_t["valor"], errors="coerce").abs().sum()), 2)
                    if _soma_t > 0 and abs(_soma_t - _residuo_c) < 0.005:
                        _nomes_t = " + ".join(
                            str(x)[:40] for x in _t.get("descricao", _t.get("tipo", pd.Series(dtype=str))).fillna("").tolist() if str(x).strip()
                        ) or "aluguel/tarifa"
                        item["nota_cartao"] = (
                            "A adquirente descontou <b>" + fmt_brl(_soma_t)
                            + "</b> do repasse deste dia (" + _esc(_nomes_t) + ") — comprovado no "
                            "extrato da adquirente. As baixas de venda estão corretas; "
                            "falta lançar essa <b>despesa de tarifa</b> no Sankhya."
                        )
        except Exception:
            pass
        if item["banco"] or item["sankhya"] or item["divergentes"] or abs(dif) >= 0.01:
            out.append(item)
    # ordena pelo "tamanho" do dia: maior entre |dif| e a soma das linhas listadas
    def _peso(i):
        soma = (sum(abs(v) for _, v in i["banco"])
                + sum(abs(v) for _, v in i["sankhya"]))
        return max(abs(i["dif"]), soma)
    out.sort(key=_peso, reverse=True)
    return out


def _render_alerta_diferenca_por_dia(dif_bs, explicacao, nota_extra: str = "",
                                     falta: float = 0.0, diverg: float = 0.0):
    """Banner enxuto + seta (expander) que explica a diferença por dia e por
    origem. NUNCA rotula como cartão — só descreve o que não casou em cada dia.
    v5.47: `dif_bs` já vem LÍQUIDO de estornos anulados; `nota_extra` (HTML)
    informa quanto foi anulado, pra conta fechar com os cards acima.
    v5.49: quando a diferença líquida é ~zero mas há lançamentos sem par dos
    dois lados (mesmo dinheiro aberto de formas diferentes), o cabeçalho
    mostra os dois lados em vez de um falso 'R$ 0,00'."""
    if abs(dif_bs) >= 0.01:
        _cabeca = ('&#9888;&#65039; <b>Diferença Banco &times; Sankhya: '
                   + fmt_brl(abs(dif_bs)) + '</b> &middot; <b>a analisar por dia</b>')
    else:
        _lados = []
        if abs(falta) >= 0.005:
            _lados.append(fmt_brl(abs(falta)) + ' no banco sem par')
        if abs(diverg) >= 0.005:
            _lados.append(fmt_brl(abs(diverg)) + ' no Sankhya sem par')
        _cabeca = ('&#9888;&#65039; <b>Lançamentos sem par: ' + ' &middot; '.join(_lados)
                   + '</b> &middot; os totais se compensam (diferença líquida R$ 0,00), '
                   'mas cada lançamento precisa do seu par &middot; <b>a analisar por dia</b>')
    st.html(
        '<div style="background:#2a1d10;border-left:4px solid #FAC318;border-radius:8px;'
        'padding:10px 14px;margin:10px 0 2px 0;color:#e9eef7;font-size:14px;line-height:1.5;">'
        + _cabeca + ' &middot; abra o detalhe abaixo para ver o que aconteceu em cada dia.'
        + (nota_extra or '') + '</div>'
    )
    with st.expander("Ver o que aconteceu em cada dia", expanded=False):
        if not explicacao:
            st.caption(
                "A diferença não pôde ser detalhada por dia com os lançamentos atuais. "
                "Confira as abas “Sem baixa no Sankhya” e “Divergências (Sankhya × Banco)”."
            )
            return

        def _tabela(linhas, cor_valor):
            html = ""
            for h, v in linhas:
                cor = "#ff9a9a" if float(v) < 0 else "#7ee0a6"
                sinal = "" if float(v) < 0 else "+"
                html += (
                    '<div style="display:grid;grid-template-columns:2.2fr 1fr;font-size:12px;'
                    'color:#cdd9f2;padding:6px 0;border-bottom:1px solid #10254e;">'
                    '<span>' + _esc(h) + '</span>'
                    '<span style="text-align:right;color:' + cor + ';">'
                    + sinal + fmt_brl(abs(float(v))) + '</span></div>'
                )
            return html

        for item in explicacao:
            d = item["data"].strftime("%d/%m/%Y")
            blocos = ""
            # v5.58: bloco único de CARTÃO do dia — só o resíduo em destaque
            _cart = item.get("cartao")
            if _cart:
                _res = float(_cart["residuo"])
                if _res > 0.004:
                    _res_txt = ("<span style='color:#ffc94d;font-weight:800;'>resta " + fmt_brl(_res)
                                + "</span> <span style='color:#9fb3d6;font-size:10px;'>Sankhya baixou a mais que entrou</span>")
                elif _res < -0.004:
                    _res_txt = ("<span style='color:#ffc94d;font-weight:800;'>resta " + fmt_brl(abs(_res))
                                + "</span> <span style='color:#9fb3d6;font-size:10px;'>entrou no banco a mais que as baixas</span>")
                else:
                    _res_txt = "<span style='color:#7ee0a6;font-weight:800;'>fecha ao centavo</span>"
                blocos += (
                    '<div style="padding:8px 16px 2px 16px;"><span style="background:#3a2e0b;'
                    'color:#FAC318;font-size:10px;padding:3px 9px;border-radius:20px;">'
                    'cartão do dia &middot; TOP 1722 &times; depósitos da adquirente</span></div>'
                    '<div style="display:flex;justify-content:space-between;align-items:baseline;'
                    'padding:6px 16px 10px 16px;font-size:12px;color:#cdd9f2;gap:12px;">'
                    '<span>' + str(_cart["qtd_banco"]) + ' depósitos no banco (' + fmt_brl(_cart["soma_banco"]) + ') '
                    '&times; ' + str(_cart["qtd_sankhya"]) + ' notas baixadas (' + fmt_brl(_cart["soma_sankhya"]) + ')</span>'
                    '<span style="text-align:right;white-space:nowrap;">' + _res_txt + '</span></div>'
                )
            if item["banco"]:
                blocos += (
                    '<div style="padding:8px 16px 2px 16px;"><span style="background:#123f2e;'
                    'color:#a8f0c8;font-size:10px;padding:3px 9px;border-radius:20px;">'
                    'no banco &middot; falta lançar no Sankhya</span></div>'
                    '<div style="padding:2px 16px 8px 16px;">' + _tabela(item["banco"], "#ff9a9a") + '</div>'
                )
            if item["sankhya"]:
                blocos += (
                    '<div style="padding:8px 16px 2px 16px;"><span style="background:#3a2436;'
                    'color:#f2b6d8;font-size:10px;padding:3px 9px;border-radius:20px;">'
                    'baixado no Sankhya &middot; sem par no banco</span></div>'
                    '<div style="padding:2px 16px 8px 16px;">' + _tabela(item["sankhya"], "#7ee0a6") + '</div>'
                )
            for h, vb, vs in item["divergentes"]:
                _vb = fmt_brl(abs(float(vb))) if vb is not None and pd.notna(vb) else "—"
                _vs = fmt_brl(abs(float(vs))) if vs is not None and pd.notna(vs) else "—"
                blocos += (
                    '<div style="padding:8px 16px 2px 16px;"><span style="background:#3a2e0b;'
                    'color:#FAC318;font-size:10px;padding:3px 9px;border-radius:20px;">'
                    'mesmo lançamento &middot; valor divergente</span></div>'
                    '<div style="padding:2px 16px 8px 16px;font-size:12px;color:#cdd9f2;">'
                    + _esc(h) + ' &nbsp; <span style="color:#9fb3d6;">banco</span> ' + _vb
                    + ' &nbsp; <span style="color:#9fb3d6;">Sankhya</span> ' + _vs + '</div>'
                )
            if not blocos:
                blocos = (
                    '<div style="padding:8px 16px 12px 16px;font-size:12px;color:#9fb3d6;">'
                    'Diferença de volume neste dia sem linha isolada — confira as abas de pendências.</div>'
                )
            # v5.57: tarifa da adquirente COMPROVADA para o resíduo do dia
            if item.get("nota_cartao"):
                blocos += (
                    '<div style="margin:0 16px 12px 16px;background:#0c2b1a;border-left:3px solid #0F8C3B;'
                    'border-radius:6px;padding:8px 12px;font-size:12px;color:#c9efd9;">'
                    '&#9989; ' + item["nota_cartao"] + '</div>'
                )
            # v5.57: cabeçalho do dia com DIREÇÃO e SINAL — antes mostrava só o
            # valor absoluto e não dava pra somar os dias e chegar no alerta.
            _dif_d = float(item["dif"])
            if abs(_dif_d) < 0.005:
                _cab_val = "R$ 0,00 &middot; lados se compensam"
            elif _dif_d > 0:
                _cab_val = "+" + fmt_brl(_dif_d) + ' <span style="font-size:10px;color:#9fb3d6;">banco a mais</span>'
            else:
                _cab_val = "&minus;" + fmt_brl(abs(_dif_d)) + ' <span style="font-size:10px;color:#9fb3d6;">Sankhya a mais</span>'
            st.html(
                '<div style="background:#0b2560;border:1px solid #163062;border-radius:10px;'
                'margin:0 0 10px 0;overflow:hidden;">'
                '<div style="display:flex;justify-content:space-between;align-items:center;'
                'background:#071f52;padding:10px 16px;">'
                '<span style="font-size:13px;font-weight:700;color:#fff;">' + d + '</span>'
                '<span style="font-size:13px;font-weight:800;color:#FAC318;">'
                + _cab_val + '</span></div>' + blocos + '</div>'
            )
        # v5.57: rodapé de CONFERÊNCIA — a soma algébrica dos dias tem que bater
        # com o valor do alerta. Se não bater, o próprio app avisa.
        _soma_dias = round(sum(float(i["dif"]) for i in explicacao), 2)
        _bate_soma = abs(_soma_dias - round(float(dif_bs), 2)) < 0.01
        st.html(
            '<div style="display:flex;justify-content:space-between;font-size:12px;'
            'padding:8px 16px;border-top:1px dashed #1d3a72;color:#9fb3d6;">'
            '<span>Soma dos dias (com sinal: + banco a mais, &minus; Sankhya a mais)</span>'
            '<span style="font-weight:700;color:' + ("#7ee0a6" if _bate_soma else "#ffc94d") + ';">'
            + ("+" if _soma_dias >= 0 else "&minus;") + fmt_brl(abs(_soma_dias))
            + (" = valor do alerta &#10003;" if _bate_soma else " &ne; valor do alerta &#9888; (parte da diferença não tem dia identificável)")
            + '</span></div>'
        )


_ROTULOS_CARTAO = {
    "estorno": "Estorno / Cancelamento",
    "aluguel": "Aluguel de maquininha",
    "tarifa": "Tarifa / Excedentes",
    "repasse": "Repasse (cai no banco)",
    "venda": "Vendas",
    "saldo": "Saldo",
    "outros": "Outros",
}


def render_tab_diferenca_cartao(adq: pd.DataFrame, resultado, conta: str):
    """Aba 'Diferença de Cartão': usa o extrato da adquirente pra dar NOME às
    tarifas/aluguéis/estornos e amarrar isso à diferença de volume por dia.
    Só mostra o que está no arquivo — nada é cravado."""
    st.caption(
        "Extrato da adquirente (GetNet/PagBank) — dá nome às tarifas, aluguéis e "
        "estornos de cartão e amarra à diferença de volume. Nada é cravado sem estar no arquivo."
    )
    if adq is None or adq.empty:
        st.info(
            "Nenhum extrato de adquirente foi enviado nesta execução. Suba o GetNet "
            "(Recebíveis) e/ou PagBank na tela de upload — é opcional — pra ver o mapa de cartão aqui."
        )
        return

    # Mapa por ADQUIRENTE + categoria. O PagBank/PagSeguro é extrato de RECEBIMENTO:
    # não detalha vendas de forma confiável, então não exibimos 'venda'/'saldo' dele.
    g = (
        adq.assign(_abs=adq["valor"].abs())
        .groupby(["adquirente", "categoria"])
        .agg(qtd=("valor", "size"), total=("_abs", "sum"))
        .reset_index()
    )
    g = g[~((g["adquirente"] == "PagBank") & (g["categoria"].isin(["venda", "saldo"])))]
    g = g.sort_values(["adquirente", "total"], ascending=[True, False])
    mapa = pd.DataFrame(
        {
            "Adquirente": g["adquirente"].values,
            "Categoria": [_ROTULOS_CARTAO.get(c, c) for c in g["categoria"]],
            "Qtd": g["qtd"].astype(int).values,
            "Total": [fmt_brl(v) for v in g["total"]],
        }
    )
    st.markdown("**Mapa de cartão — por adquirente e categoria (valores do extrato)**")
    st.dataframe(mapa, hide_index=True, use_container_width=True)

    # Amarra a diferença de volume (por dia) ao que a adquirente mostra naquele dia.
    dias = _diferenca_bxs_por_dia(resultado, conta)
    if dias:
        adq_dt = pd.to_datetime(adq["data"], errors="coerce").dt.normalize()
        linhas = []
        for d, v in dias[:12]:
            no_dia = adq[
                (adq_dt == pd.Timestamp(d))
                & (adq["categoria"].isin(["aluguel", "tarifa", "estorno"]))
            ]
            if not no_dia.empty:
                no_dia = no_dia.reindex(
                    no_dia["valor"].abs().sort_values(ascending=False).index
                )
                expl = "; ".join(
                    f"{_ROTULOS_CARTAO.get(c, c)}: {fmt_brl(abs(val))} ({desc})"
                    for c, val, desc in zip(
                        no_dia["categoria"], no_dia["valor"], no_dia["descricao"]
                    )
                )
            else:
                expl = "— (não há tarifa/estorno da adquirente neste dia)"
            linhas.append(
                {"Dia": d.strftime("%d/%m"), "Diferença de volume": fmt_brl(abs(v)), "O que a adquirente mostra": expl}
            )
        st.markdown("**Diferença de volume Banco × Sankhya, dia a dia — explicada pela adquirente**")
        st.caption(
            "Tarifa/aluguel entram 2× no volume (o Sankhya lança a tarifa e o bruto; "
            "o banco já vem líquido), por isso a diferença costuma ser ~2× o valor da tarifa."
        )
        st.dataframe(pd.DataFrame(linhas), hide_index=True, use_container_width=True)

    # Detalhe das tarifas/aluguéis/estornos
    custos = adq[adq["categoria"].isin(["estorno", "aluguel", "tarifa"])].copy()
    if not custos.empty:
        custos = custos.sort_values("data")
        det = pd.DataFrame(
            {
                "Dia": pd.to_datetime(custos["data"]).dt.strftime("%d/%m"),
                "Adquirente": custos["adquirente"].values,
                "Bandeira": custos["bandeira"].values,
                "Lançamento": custos["descricao"].values,
                "Categoria": [_ROTULOS_CARTAO.get(c, c) for c in custos["categoria"]],
                "Valor": [fmt_brl(v) for v in custos["valor"]],
            }
        )
        st.markdown("**Tarifas, aluguéis e estornos (detalhe)**")
        st.dataframe(det, hide_index=True, use_container_width=True)


def _render_regua_conferencia_sankhya(resultado: ResultadoConciliacao, conta: str, k: dict):
    """v5.65: conferência com o rodapé do Sankhya — tabela espelho em expander.

    Layout aprovado (15/07): fica no TOPO do detalhamento; o título do expander
    já traz o veredito (fecha / valor a conferir) e ele vem ABERTO só quando há
    diferença. Dentro, uma tabela espelho do rodapé do Sankhya, coluna a coluna:
    Crédito = receitas + resgates · Débito = despesas + aplicações (o rodapé
    soma investimentos nos totais; o app os separa nos cards). O selo verde só
    aparece quando banco e Sankhya fecham ao centavo — sem prova, a diferença
    fica como 'a conferir' (zero falso positivo).
    """
    rec_b = float(k.get("receitas_banco", 0.0)); desp_b = float(k.get("despesas_banco", 0.0))
    rec_s = float(k.get("receitas_sistema", 0.0)); desp_s = float(k.get("despesas_sistema", 0.0))

    # aplicações/resgates POR ORIGEM (sem dedup — o rodapé do Sankhya soma as
    # linhas do próprio Sankhya; o extrato, as do banco). Exclui SALDO.
    inv = resultado.aplicacoes_resgates_da_conta(conta)
    ap_b = rg_b = ap_s = rg_s = 0.0
    # v5.66: rendimentos entram como CRÉDITO na régua — o rodapé do Sankhya
    # soma pelo campo Receita/Despesa; o app tratava rendimento como
    # "investimento" e tirava do crédito, gerando um falso R$ 0,86 na APOIO.
    rend_b = rend_s = 0.0
    if inv is not None and not inv.empty and "tipo_aplicacao" in inv.columns:
        inv = inv.copy()
        if "historico" in inv.columns:
            inv = inv[~inv["historico"].astype(str).str.upper().str.contains("SALDO", na=False)]
        _orig = inv["origem"].astype(str) if "origem" in inv.columns else pd.Series("", index=inv.index)
        _eh_sk = _orig.str.contains("Sankhya", case=False, na=False)
        _val = pd.to_numeric(inv["valor"], errors="coerce").abs().fillna(0.0)
        _ap = inv["tipo_aplicacao"] == "Aplicação"
        _rg = inv["tipo_aplicacao"] == "Resgate"
        ap_s = float(_val[_ap & _eh_sk].sum()); rg_s = float(_val[_rg & _eh_sk].sum())
        ap_b = float(_val[_ap & ~_eh_sk].sum()); rg_b = float(_val[_rg & ~_eh_sk].sum())
        # rendimentos: nem aplicação nem resgate (podem estar como
        # "Rendimento" em tipo_aplicacao OU marcados em categoria_mov).
        _rend = ~(_ap | _rg)
        if "categoria_mov" in inv.columns:
            _rend = _rend | (inv["categoria_mov"] == "rendimento")
        rend_s = float(_val[_rend & _eh_sk].sum())
        rend_b = float(_val[_rend & ~_eh_sk].sum())

    cred_b = round(rec_b + rg_b + rend_b, 2); cred_s = round(rec_s + rg_s + rend_s, 2)
    deb_b = round(desp_b + ap_b, 2); deb_s = round(desp_s + ap_s, 2)

    # v5.70: desconto par-a-par de tarifas da adquirente já lançadas no Sankhya.
    # Antes, uma tarifa (ex.: GetNet R$ 309,85) descontada no repasse aparecia
    # duplicada como "diferença": Sankhya tinha receita bruta a mais + despesa
    # que o banco não tem — os dois lados iguais em valor, descrevendo o MESMO
    # fato. A régua somava |cred| + |deb| e mostrava R$ 619,71 "a conferir".
    # Agora, para cada tarifa da adquirente COM PAR EXATO no Sankhya (valor +
    # data ±3 dias + histórico contendo palavra da adquirente), descontamos o
    # valor dos dois lados — o fato já está resolvido no ERP.
    #
    # Regra do par (opção B aprovada): histórico do Sankhya deve conter uma
    # dessas palavras: GETNET / CIELO / PAGBANK / PAGSEGURO / REDE, OU as
    # expressões do lançamento manual usado hoje pela Débora (TARIFA ALUGUEL /
    # PLAT DIGITAL / MAQUININHA / MAQUINHA / MAQUINETA). Ambiguidade (2+ pares
    # candidatos) NÃO desconta — mantém como "a conferir" (zero falso positivo).
    tarifas_lancadas_valor = 0.0
    try:
        _adq_df = st.session_state.get("adquirente_df")
        if _adq_df is not None and not _adq_df.empty and "categoria" in _adq_df.columns:
            _tar_adq = _adq_df[_adq_df["categoria"].isin(["aluguel", "tarifa"])].copy()
            if not _tar_adq.empty:
                _tar_adq["_dt"] = pd.to_datetime(_tar_adq["data"], errors="coerce").dt.normalize()
                _tar_adq["_val"] = pd.to_numeric(_tar_adq["valor"], errors="coerce").abs().round(2)
                # despesas do Sankhya desta conta (sem investimentos/saldo)
                _sk_full = resultado.sistema_completo
                if "conta" in _sk_full.columns:
                    _sk_full = _sk_full[_sk_full["conta"] == conta]
                _sk_desp = _sk_full[_sk_full["valor"] < 0].copy() if not _sk_full.empty else pd.DataFrame()
                if "categoria_mov" in _sk_desp.columns:
                    _sk_desp = _sk_desp[
                        ~_sk_desp["categoria_mov"].isin(
                            ["saldo", "aplicacao", "resgate", "rendimento", "investimento_outro"]
                        )
                    ]
                _PALS_ADQ = (
                    "GETNET", "GET NET", "CIELO", "PAGBANK", "PAG BANK",
                    "PAGSEGURO", "PAG SEGURO", "REDE",
                    "TARIFA ALUGUEL", "PLAT DIGITAL", "MAQUININHA",
                    "MAQUINHA", "MAQUINETA",
                )
                if not _sk_desp.empty:
                    _sk_desp = _sk_desp.copy()
                    _sk_desp["_dt"] = pd.to_datetime(_sk_desp["data"], errors="coerce").dt.normalize()
                    _sk_desp["_val"] = pd.to_numeric(_sk_desp["valor"], errors="coerce").abs().round(2)
                    _hup = _sk_desp["historico"].astype(str).str.upper()
                    _mask_pal = _hup.apply(lambda h: any(p in h for p in _PALS_ADQ))
                    _sk_desp = _sk_desp[_mask_pal]
                # match par-a-par
                _idx_usados: set = set()
                for _, tarifa in _tar_adq.iterrows():
                    _v = float(tarifa["_val"])
                    _d = tarifa["_dt"]
                    if pd.isna(_d) or _v <= 0:
                        continue
                    _cand = _sk_desp[
                        (_sk_desp["_val"] == round(_v, 2))
                        & ((_sk_desp["_dt"] - _d).abs() <= pd.Timedelta(days=3))
                        & (~_sk_desp.index.isin(_idx_usados))
                    ] if not _sk_desp.empty else _sk_desp
                    if len(_cand) == 1:
                        tarifas_lancadas_valor += _v
                        _idx_usados.add(_cand.index[0])
                    # 0 candidatos (não lançada) ou >=2 (ambíguo): não desconta
    except Exception:
        tarifas_lancadas_valor = 0.0

    # aplica o desconto — cada tarifa lançada aparecia dobrada: uma vez no
    # crédito (Sankhya recebeu bruto a mais) e uma vez no débito (Sankhya
    # lançou despesa que o banco não tem). Subtrair dos DOIS lados alinha.
    if tarifas_lancadas_valor > 0.005:
        cred_s = round(cred_s - tarifas_lancadas_valor, 2)
        deb_s = round(deb_s - tarifas_lancadas_valor, 2)
    tot_b = round(cred_b + deb_b, 2); tot_s = round(cred_s + deb_s, 2)
    dif_c = round(cred_b - cred_s, 2); dif_d = round(deb_b - deb_s, 2)
    dif_t = round(tot_b - tot_s, 2)
    fecha = abs(dif_c) < 0.005 and abs(dif_d) < 0.005

    if fecha:
        titulo = ("✅ Conferência com o rodapé do Sankhya — "
                  "Crédito, Débito e Total fecham com o banco ao centavo")
    elif abs(dif_t) >= 0.005:
        titulo = ("⚠️ Conferência com o rodapé do Sankhya — "
                  + fmt_brl(abs(dif_t)) + " a conferir")
    else:
        # lados divergem mas se compensam no total — não exibir um falso R$ 0,00
        titulo = ("⚠️ Conferência com o rodapé do Sankhya — Crédito e Débito "
                  "com diferenças que se compensam · a conferir")

    def _selo_dif(dif):
        if abs(dif) < 0.005:
            return '<span style="color:#7ee0a6;font-weight:600;">&#10003; fecha ao centavo</span>'
        _sinal = "&minus;" if dif < 0 else "+"
        return ('<span style="color:#ffc94d;font-weight:600;">' + _sinal
                + fmt_brl(abs(dif)) + ' a conferir</span>')

    _td_lbl = 'style="padding:5px 8px;text-align:right;color:#9fb3d6;border-bottom:1px solid #163062;"'
    _td_sk = 'style="padding:9px 8px 1px;text-align:right;color:#fff;font-weight:700;font-size:15px;"'
    _td_dec = 'style="padding:0 8px 9px;text-align:right;color:#6f88b8;font-size:11px;"'
    _td_bco = 'style="padding:7px 8px;text-align:right;color:#cdd9f2;border-top:1px solid #163062;"'
    _td_dif = 'style="padding:7px 8px;text-align:right;border-top:1px solid #163062;"'

    with st.expander(titulo, expanded=not fecha):
        st.html(
            '<table style="width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed;">'
            '<tr><td style="width:20%;padding:5px 8px;"></td>'
            '<td ' + _td_lbl + '>Crédito</td>'
            '<td ' + _td_lbl + '>Débito</td>'
            '<td ' + _td_lbl + '>Movimentação total</td></tr>'
            '<tr><td style="padding:9px 8px 1px;color:#FAC318;font-weight:600;">Rodapé do Sankhya</td>'
            '<td ' + _td_sk + '>' + fmt_brl(cred_s) + '</td>'
            '<td ' + _td_sk + '>' + fmt_brl(deb_s) + '</td>'
            '<td ' + _td_sk + '>' + fmt_brl(tot_s) + '</td></tr>'
            '<tr><td style="padding:0 8px 9px;color:#6f88b8;font-size:11px;">como o app chegou</td>'
            '<td ' + _td_dec + '>Receitas ' + fmt_brl(rec_s) + ' + Resgates ' + fmt_brl(rg_s) + (' + Rendimentos ' + fmt_brl(rend_s) if rend_s >= 0.005 else '') + '</td>'
            '<td ' + _td_dec + '>Despesas ' + fmt_brl(desp_s) + ' + Aplicações ' + fmt_brl(ap_s) + '</td>'
            '<td ' + _td_dec + '>Operacional ' + fmt_brl(round(rec_s + desp_s, 2)) + ' + Invest. ' + fmt_brl(round(ap_s + rg_s + rend_s, 2)) + '</td></tr>'
            '<tr><td style="padding:7px 8px;color:#cdd9f2;border-top:1px solid #163062;">Extrato do banco</td>'
            '<td ' + _td_bco + '>' + fmt_brl(cred_b) + '</td>'
            '<td ' + _td_bco + '>' + fmt_brl(deb_b) + '</td>'
            '<td ' + _td_bco + '>' + fmt_brl(tot_b) + '</td></tr>'
            '<tr><td style="padding:7px 8px;color:#9fb3d6;border-top:1px solid #163062;">Diferença (banco − Sankhya)</td>'
            '<td ' + _td_dif + '>' + _selo_dif(dif_c) + '</td>'
            '<td ' + _td_dif + '>' + _selo_dif(dif_d) + '</td>'
            '<td ' + _td_dif + '>' + _selo_dif(dif_t) + '</td></tr>'
            '</table>'
        )


def tela_detalhamento_banco(resultado: ResultadoConciliacao, conta: str):
    col_back, _ = st.columns([1, 5])
    with col_back:
        if st.button("← Voltar ao painel"):
            st.session_state.banco_conta_selecionada = None
            st.rerun()

    section_title(f"DETALHAMENTO · {conta}")

    k = resultado.kpis_da_conta(conta)
    # v5.65: conferência com o rodapé do Sankhya no TOPO do detalhamento —
    # expander com veredito no título; abre sozinho só quando há diferença.
    _render_regua_conferencia_sankhya(resultado, conta, k)
    sub_banco_c = _card_total_com_rec_desp(k["receitas_banco"], k["despesas_banco"])
    sub_sankhya_c = _card_total_com_rec_desp(k["receitas_sistema"], k["despesas_sistema"])
    cards = [
        card_kpi_html("Movimentação Operacional · Banco", fmt_brl(k["total_movimentado_banco"]),
                      sub_banco_c),
        card_kpi_html("Movimentação Operacional · Sankhya", fmt_brl(k["total_extrato_sistema"]),
                      sub_sankhya_c),
        _card_investimentos_da_conta(resultado, conta),
        card_kpi("Índice de Conciliação", fmt_pct(k["percentual_conciliado"]),
                 "conferido em pares", classe="destaque-amarelo"),
    ]
    render_cards(cards)

    # Card Falta Conciliar vertical + Falta Lançar + Divergência + Investimentos da conta
    sub_fc = _card_falta_conciliar_vertical(
        k["falta_conciliar_receitas"],
        k["falta_conciliar_despesas"],
    )
    fonte_fl = ("via Sankhya 'Conciliado=Não'"
                if k["fonte_falta_lancar"] == "sankhya_conciliado_nao"
                else "pendência do sistema")
    falta_c = float(k["falta_conciliar"])
    div_c = float(k["divergencia_sankhya_banco"])
    qtd_div_c = int(k["qtd_divergencia_sankhya_banco"])

    # Banco sem explicação (verde quando zero, vermelho quando há)
    if falta_c == 0:
        card_banco_sem_exp = card_kpi(
            "Pendente no Extrato Bancário", "R$ 0,00",
            "movimento do banco sem baixa no Sankhya", classe="destaque-verde")
    else:
        card_banco_sem_exp = card_kpi_html(
            "Pendente no Extrato Bancário", fmt_brl(falta_c), sub_fc, classe="destaque-vermelho")

    # Sankhya sem confirmação: junta valor + contagem num card só
    cor_div = "destaque-verde" if (div_c == 0 and qtd_div_c == 0) else "destaque-vermelho"
    qtd_txt = (fmt_int(qtd_div_c) + " item") if qtd_div_c == 1 else (fmt_int(qtd_div_c) + " itens")
    valor_div_html = (fmt_brl(div_c)
                      + ' <span style="font-size:14px; color:#8BA3C7; font-weight:400;">&middot; '
                      + qtd_txt + '</span>')
    card_sankhya_sem_conf = card_kpi_html(
        "Pendente no Sankhya", valor_div_html,
        '<div class="lle-kpi-suffix">lançamentos do ERP que o banco não confirmou</div>',
        classe=cor_div)

    # v5.49: o card decompõe movimentação × investimentos — antes mostrava a
    # soma total dos pares (com aplicações/resgates) ao lado de um "movimentado"
    # que exclui investimentos, e parecia que conferiu mais do que movimentou.
    _conf_mov = float(k.get("total_conciliado_movimentacao", k["total_conciliado"]))
    _conf_inv = float(k.get("total_conciliado_investimentos", 0.0))
    _suf_conf = ("movimentação em pares" if _conf_inv < 0.01 else
                 "movimentação em pares · + " + fmt_brl(_conf_inv) + " de investimentos casados")
    cards2 = [
        card_banco_sem_exp,
        card_sankhya_sem_conf,
        card_kpi("Total Conciliado", fmt_brl(_conf_mov),
                 _suf_conf, classe="destaque-verde"),
    ]
    render_cards(cards2)

    # v5.38: ALERTA da diferença Banco × Sankhya — antes ela só aparecia no texto do
    # explainer e nas abas, então passava batido. Agora grita num banner sempre que
    # existir, apontando pra onde investigar.
    # v5.47: pares ANULADOS POR ESTORNO (PIX recebido e devolvido) se anulam e já
    # estão explicados na aba ♻️ — eles NÃO contam como diferença. O banner mostra
    # só o que precisa de análise; quando há anulados, a nota fecha a conta com os
    # cards acima (líquido + anulados = banco − sankhya).
    dif_bruta = round(float(k["total_movimentado_banco"]) - float(k["total_extrato_sistema"]), 2)
    _vol_anulados = round(sum(_volume_estornos_anulados_por_dia(resultado, conta).values()), 2)
    dif_bs = round(dif_bruta - _vol_anulados, 2)
    # v5.49: o alerta também dispara quando a diferença LÍQUIDA é zero mas há
    # lançamentos sem par dos dois lados (ex.: 1 PIX de R$ 306 no banco aberto
    # em 4 notas no Sankhya — os totais se compensam e o banner ficava mudo).
    _falta_al = float(k.get("falta_conciliar", 0.0))
    _diverg_al = float(k.get("divergencia_sankhya_banco", 0.0))
    _tem_pend_al = round(abs(_falta_al) + abs(_diverg_al), 2) >= 0.01
    if abs(dif_bs) >= 0.01 or _tem_pend_al:
        _adq_alert = st.session_state.get("adquirente_df")
        _partes, _todos = _explicar_diferenca_cartao(resultado, conta, _adq_alert)
        _val = fmt_brl(abs(dif_bs))
        _tem_adq = _adq_alert is not None and not getattr(_adq_alert, "empty", True)
        _nota_anulados = (
            ' &middot; <span style="color:#9fb3d6;">além de ' + fmt_brl(_vol_anulados)
            + ' em PIX recebidos e devolvidos que se anulam (aba ♻️ Estornos — não precisam de análise)</span>'
            if _vol_anulados >= 0.01 else ''
        )

        if abs(dif_bs) >= 0.01 and _tem_adq and _partes and _todos:
            # Só quando HÁ adquirente E ela explica 100% da diferença (taxa comprovada).
            st.html(
                '<div style="background:#0c2b1a;border-left:4px solid #0F8C3B;border-radius:8px;'
                'padding:10px 14px;margin:10px 0 2px 0;color:#e9eef7;font-size:14px;line-height:1.5;">'
                '&#9989; <b>Diferença Banco &times; Sankhya: ' + _val + '</b> &middot; '
                '<b>identificada 100% pela adquirente</b> &middot; taxa descontada no repasse '
                '(não é erro de conciliação). Detalhe na aba &ldquo;Diferença de Cartão&rdquo;.'
                + _nota_anulados + '</div>'
            )
        else:
            # Caso geral (com ou sem cartão): explica por dia, pela origem real.
            # NUNCA rotula como cartão por padrão.
            _explic = _explicar_diferenca_por_dia(resultado, conta)
            _render_alerta_diferenca_por_dia(
                dif_bs, _explic, nota_extra=_nota_anulados,
                falta=_falta_al, diverg=_diverg_al,
            )
    elif abs(dif_bruta) >= 0.01 and _vol_anulados >= 0.01:
        # v5.47: a diferença entre os cards existe, mas é 100% de pares anulados
        # por estorno — nada a analisar. Banner verde explicando por que os
        # totais dos cards não batem entre si.
        st.html(
            '<div style="background:#0c2b1a;border-left:4px solid #0F8C3B;border-radius:8px;'
            'padding:10px 14px;margin:10px 0 2px 0;color:#e9eef7;font-size:14px;line-height:1.5;">'
            '&#9989; <b>Diferença Banco &times; Sankhya: ' + fmt_brl(abs(dif_bruta)) + '</b> &middot; '
            '<b>100% são PIX recebidos e devolvidos que se anulam</b> (anulados por estorno) '
            '&middot; nada a analisar. Detalhe na aba &ldquo;&#9851;&#65039; Estornos &middot; pares anulados&rdquo;.</div>'
        )
    info_saldo = resultado.saldo_final_da_conta(conta)
    if info_saldo is not None:
        render_card_saldo_final(info_saldo)

    # Notas recolhíveis — explicam os cards acima, sem poluir
    _mov_c = float(k["total_movimentado_banco"])
    _falta_c2 = float(k["falta_conciliar"])
    _explic = 100.0 * (_mov_c - _falta_c2) / _mov_c if _mov_c > 0 else 0.0
    _pares = float(k["percentual_conciliado"])
    _regra = max(0.0, _explic - _pares)
    _diff = abs(float(k["total_extrato_sistema"]) - _mov_c)
    # v5.35: só fala "cartão" se a conta TEM cartão; nomeia a regra real.
    _tem_cartao = int(k.get("qtd_top1722_grupos", 0)) > 0
    _tem_boleto = int(k.get("qtd_top1702_grupos", 0)) > 0
    _nomes = []
    if _tem_boleto:
        _nomes.append("boleto (TOP 1702)")
    if _tem_cartao:
        _nomes.append("cartão (TOP 1722)")
    # v5.49: estornos anulados (PIX recebido e devolvido) também explicam volume
    if int(k.get("qtd_estornos_anulados", 0)) > 0:
        _nomes.append("estornos anulados (recebido e devolvido)")
    _rotulo = " e ".join(_nomes) if _nomes else "agrupamento por soma"
    with st.expander("Entenda os cards acima"):
        if _regra > 0.01:
            st.markdown(
                "**% conferido em pares (" + fmt_pct(_pares) + "):** casou em pares diretos (1 a 1). "
                "Os " + fmt_pct(_regra) + " restantes são **" + _rotulo + "**, que casa pela soma "
                "total — também explicado. Somando, o banco está **" + fmt_pct(_explic)
                + " explicado** (é o número que aparece no resumo)."
            )
        else:
            st.markdown(
                "**% conferido em pares (" + fmt_pct(_pares) + "):** casou em pares diretos (1 a 1). "
                "O banco está **" + fmt_pct(_explic) + " explicado** (é o número do resumo)."
            )
        _inv = resultado.aplicacoes_resgates_da_conta(conta).copy()
        if not _inv.empty:
            if "historico" in _inv.columns:
                _inv = _inv[~_inv["historico"].astype(str).str.upper().str.contains("SALDO", na=False)]
            _cols = [c for c in ["data", "valor", "conta", "tipo_aplicacao"] if c in _inv.columns]
            if _cols and "origem" in _inv.columns:
                _inv["_o"] = _inv["origem"].apply(lambda x: 0 if "Sankhya" in str(x) else 1)
                _inv = _inv.sort_values("_o").drop_duplicates(subset=_cols, keep="first")
            _ap = float(_inv[_inv["tipo_aplicacao"] == "Aplicação"]["valor"].abs().sum()) if "tipo_aplicacao" in _inv.columns else 0.0
            _rg = float(_inv[_inv["tipo_aplicacao"] == "Resgate"]["valor"].abs().sum()) if "tipo_aplicacao" in _inv.columns else 0.0
            _liq = _rg - _ap
            _frase_liq = (
                "Aplicou mais do que resgatou — **não é prejuízo**, o dinheiro ficou guardado "
                "em aplicação e continua sendo da empresa." if _liq < 0
                else "Resgatou mais do que aplicou no mês."
            )
            _msg_inv = (
                "**Investimentos no período:** líquido de **" + fmt_brl(_liq) + "** "
                "(resgatado " + fmt_brl(_rg) + " − aplicado " + fmt_brl(_ap) + "). " + _frase_liq
            )
            # v5.38: escapa o cifrão pra o Streamlit NÃO interpretar R$...R$ como LaTeX.
            st.markdown(_msg_inv.replace("$", "\\$"))
        if _diff > 0.01:
            # v5.47: separa a parte anulada por estorno (recebido e devolvido —
            # já explicada na aba ♻️) do restante que precisa de análise.
            _vol_anul_c = round(
                sum(_volume_estornos_anulados_por_dia(resultado, conta).values()), 2
            )
            _resto_c = round(_diff - _vol_anul_c, 2)
            if _vol_anul_c >= 0.01 and _resto_c < 0.01:
                st.markdown((
                    "**Diferença entre Banco e Sankhya (" + fmt_brl(_diff) + "):** é "
                    "**100% de PIX recebidos e devolvidos que se anulam** (anulados por "
                    "estorno — veja a aba ♻️). Nada a analisar."
                ).replace("$", "\\$"))
            elif _vol_anul_c >= 0.01:
                _txt_causa = (
                    "pode incluir a taxa de cartão lançada **duas vezes** no Sankhya "
                    "(uma no valor bruto, outra como despesa), lançamento não lançado, ou item a conciliar."
                    if _tem_cartao else
                    "o app **não crava** a causa — pode ser lançamento não lançado no "
                    "Sankhya (ex.: devolução), lançamento a mais/duplicado, tarifa, ou item a conciliar."
                )
                st.markdown((
                    "**Diferença entre Banco e Sankhya (" + fmt_brl(_diff) + "):** "
                    + fmt_brl(_vol_anul_c) + " são **PIX recebidos e devolvidos que se "
                    "anulam** (aba ♻️ — não precisam de análise). Restam **"
                    + fmt_brl(_resto_c) + "** a analisar: " + _txt_causa
                ).replace("$", "\\$"))
            elif _tem_cartao:
                st.markdown((
                    "**Diferença entre Banco e Sankhya (" + fmt_brl(_diff) + "):** pode incluir a "
                    "taxa de cartão lançada **duas vezes** no Sankhya (uma no valor bruto, outra como "
                    "despesa). O restante precisa ser analisado."
                ).replace("$", "\\$"))
            else:
                st.markdown((
                    "**Diferença entre Banco e Sankhya (" + fmt_brl(_diff) + "):** precisa ser "
                    "analisada. O app **não crava** a causa — pode ser cartão de adquirente "
                    "(Cielo/Stone/Getnet) que não fechou no agrupamento, lançamento a mais/"
                    "duplicado no Sankhya, tarifa, ou item a conciliar."
                ).replace("$", "\\$"))

    # Download específico desse banco — v5.35: SOB DEMANDA.
    # Antes, o Excel da conta era gerado ao ABRIR o detalhamento; numa conta de
    # alto volume (ex.: Santander, 55k linhas) isso levava ~3,4 min e a tela
    # parecia travada ("Ver detalhamento" não respondia). Agora só gera ao clicar
    # — o detalhamento abre na hora, igual ao Excel global.
    excel_conta_key = (
        f"xlsx_conta_{st.session_state.get('id_execucao_atual', 'novo')}_{conta}"
    )
    nome_xls_conta = (
        f"conciliacao_{conta}_{resultado.data_referencia.strftime('%Y%m%d')}.xlsx"
    )
    if st.session_state.get(excel_conta_key):
        st.download_button(
            f"⬇️ Baixar relatório de {conta}",
            data=st.session_state[excel_conta_key],
            file_name=nome_xls_conta,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    elif st.button(f"⬇️ Gerar Excel de {conta}", type="primary",
                   use_container_width=True, key=f"btn_xls_conta_{conta}"):
        with st.spinner("Gerando Excel desta conta… em volume grande pode levar alguns minutos."):
            try:
                st.session_state[excel_conta_key] = gerar_relatorio_excel_de_conta(
                    resultado, conta
                )
            except Exception as e:
                st.warning(f"Não foi possível gerar o Excel deste banco: {e}")
        st.rerun()

    st.divider()

    # Abas internas — agora com Excesso no Sankhya
    div_conta = resultado.divergencias_da_conta(conta)
    aplic_conta = resultado.aplicacoes_resgates_da_conta(conta)
    poss_dup_conta = resultado.possiveis_duplicidades_da_conta(conta)
    excesso_conta = resultado.excesso_sankhya_da_conta(conta)
    diverg_consolidada = resultado.divergencias_sankhya_banco(conta)

    # v5.0/v5.1: filtra estornos e TOP 1722 pela conta.
    # `getattr` com fallback pra DataFrame vazio garante compatibilidade quando o objeto
    # `resultado` em sessão veio de uma versão anterior do código (sem esses campos).
    def _filtrar_conta_seguro(attr_name: str) -> pd.DataFrame:
        df = getattr(resultado, attr_name, None)
        if df is None or df.empty or "conta" not in df.columns:
            return pd.DataFrame()
        return df[df["conta"] == conta]

    estornos_anu_conta = _filtrar_conta_seguro("estornos_anulados")
    estornos_par_conta = _filtrar_conta_seguro("estornos_parciais")
    top1722_grupos_conta = _filtrar_conta_seguro("top1722_grupos")
    top1722_diff_conta = _filtrar_conta_seguro("top1722_diferencas")

    tabs_nomes = ["✅ Conciliadas", "⏳ Sem baixa no Sankhya",
                  "⚠️ Divergências (Sankhya × Banco)",
                  "🏦 Não Pertence à Conta"]
    if not div_conta.empty:
        tabs_nomes.append("💲 Diferença de Valor")
    if not poss_dup_conta.empty:
        tabs_nomes.append("🔍 Possíveis Duplicidades")
    if not excesso_conta.empty:
        tabs_nomes.append("📥 Excesso no Sankhya")
    if not aplic_conta.empty:
        tabs_nomes.append("💰 Aplicações e Resgates")
    # v5.0: abas novas
    if not estornos_anu_conta.empty:
        tabs_nomes.append(f"♻️ Estornos · pares anulados ({len(estornos_anu_conta)})")
    if not estornos_par_conta.empty:
        tabs_nomes.append(f"⚖️ Estornos Parciais ({len(estornos_par_conta)})")
    if not top1722_grupos_conta.empty:
        tabs_nomes.append(f"🃏 Cartão TOP 1722 ({len(top1722_grupos_conta)})")
    if not top1722_diff_conta.empty:
        tabs_nomes.append(f"⚠️ TOP 1722 Diferença ({len(top1722_diff_conta)})")

    # v5.41: aba "Diferença de Cartão" — só quando há extrato de adquirente enviado.
    _adq_df = st.session_state.get("adquirente_df")
    _tem_adq = _adq_df is not None and not _adq_df.empty
    if _tem_adq:
        tabs_nomes.append("💳 Diferença de Cartão")

    tabs = st.tabs(tabs_nomes)
    idx = 0
    with tabs[idx]:
        render_tab_conciliadas(resultado.conciliados_da_conta(conta), conta)
    idx += 1
    with tabs[idx]:
        render_tab_pendentes(resultado, conta)
    idx += 1
    with tabs[idx]:
        render_tab_divergencia_consolidada(diverg_consolidada, conta)
    idx += 1
    with tabs[idx]:
        render_tab_nao_pertence(resultado.nao_pertence_da_conta(conta), conta)
    if not div_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_divergencias(div_conta, conta)
    if not poss_dup_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_possiveis_duplicidades(poss_dup_conta, conta)
    if not excesso_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_excesso_sankhya(excesso_conta, conta)
    if not aplic_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_aplicacoes(aplic_conta, conta)
    # v5.0: abas novas — render
    if not estornos_anu_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_estornos_anulados(estornos_anu_conta, conta)
    if not estornos_par_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_estornos_parciais(estornos_par_conta, conta)
    if not top1722_grupos_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_top1722_grupos(
                top1722_grupos_conta,
                getattr(resultado, "top1722_linhas", pd.DataFrame()),
                conta,
                getattr(resultado, "top1722_linhas_banco", pd.DataFrame()),
            )
    if not top1722_diff_conta.empty:
        idx += 1
        with tabs[idx]:
            render_tab_top1722_diferenca(top1722_diff_conta, conta)
    if _tem_adq:
        idx += 1
        with tabs[idx]:
            render_tab_diferenca_cartao(_adq_df, resultado, conta)

    # v3.8: nova seção "POR TIPO DE LANÇAMENTO" filtrada por conta
    st.write("")
    st.divider()
    section_title(f"POR TIPO DE LANÇAMENTO · {conta}")
    st.caption(
        "Mesma análise mostrada no resumo geral, mas filtrada apenas para esta conta. "
        "Lançamentos do banco e do Sankhya agrupados por categoria, com status (conciliado / pendente)."
    )
    render_subabas_tipo(resultado, conta=conta)


def render_card_saldo_final(info: dict):
    """Card destacado de saldo final quando a conta está 100% conciliada."""
    saldo_final = info.get("saldo_final")
    saldo_inicial = info.get("saldo_inicial")
    mov_liq = info.get("movimentacao_liquida", 0.0)
    periodo_de = info.get("periodo_de")
    periodo_ate = info.get("periodo_ate")
    conta = info.get("conta", "")
    de_str = periodo_de.strftime("%d/%m/%Y") if periodo_de is not None else "—"
    ate_str = periodo_ate.strftime("%d/%m/%Y") if periodo_ate is not None else "—"

    if info.get("tem_saldo_no_extrato"):
        valor_destaque = fmt_brl(saldo_final)
        legenda = (
            f"Saldo inicial: {fmt_brl(saldo_inicial)} · "
            f"Movimentação líquida: {fmt_brl(mov_liq)}"
        )
    else:
        valor_destaque = fmt_brl(mov_liq)
        legenda = (
            "Movimentação líquida do período "
            "(saldo final não está no extrato — informe manualmente se necessário)"
        )

    st.html(f"""
        <div class="lle-kpi" style="border-left-color:{CORES['verde']}; padding:24px 28px; margin-top:12px;">
            <div style="display:flex; justify-content:space-between; align-items:center; gap:24px;">
                <div>
                    <div class="lle-kpi-label" style="color:{CORES['verde']} !important; font-size:13px;">
                        ✓ CONCILIAÇÃO 100% · SALDO FINAL DA CONTA
                    </div>
                    <div style="font-size:32px; font-weight:800; color:{CORES['verde']} !important; margin-top:6px;">
                        {valor_destaque}
                    </div>
                    <div style="font-size:12px; color:{CORES['texto_muted']} !important; margin-top:8px;">
                        {legenda}
                    </div>
                </div>
                <div style="text-align:right;">
                    <span class="lle-badge verde" style="font-size:12px;">100,0% conciliado</span>
                    <div style="font-size:11px; color:{CORES['texto_muted']} !important; margin-top:10px;">
                        Conta: <strong style="color:{CORES['amarelo']};">{conta}</strong><br>
                        Período: {de_str} a {ate_str}
                    </div>
                </div>
            </div>
        </div>
    """)


def _exibir_df(df: pd.DataFrame, nome_arquivo: str, msg_vazio: str = "Nenhum registro nesta categoria."):
    if df.empty:
        st.success(f"🎉 {msg_vazio}")
        return
    df_show = df.drop(columns=[c for c in df.columns if c.startswith("_")], errors="ignore")
    st.dataframe(df_show, use_container_width=True, height=420)
    n_linhas = len(df_show)
    col_xls, col_csv = st.columns(2)
    with col_xls:
        # v5.35: Excel da tabela SOB DEMANDA. Em abas grandes (dezenas de milhares
        # de linhas) montar o xlsx a cada render deixava o detalhamento lento — e
        # st.tabs roda o código de TODAS as abas a cada clique. O CSV continua na
        # hora (é instantâneo, vetorizado).
        xls_key = f"tblxls_{nome_arquivo}_{n_linhas}"
        if st.session_state.get(xls_key):
            st.download_button(
                f"⬇️ Baixar Excel ({n_linhas} linhas)",
                data=st.session_state[xls_key],
                file_name=f"{nome_arquivo}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_{xls_key}",
            )
        elif st.button(
            f"⬇️ Gerar Excel ({n_linhas} linhas)",
            use_container_width=True,
            key=f"gen_{xls_key}",
        ):
            with st.spinner("Gerando Excel…"):
                st.session_state[xls_key] = _df_to_xlsx_bytes(df_show, nome_arquivo)
            st.rerun()
    with col_csv:
        csv_str = df_show.to_csv(index=False, sep=";", encoding="utf-8-sig", decimal=",")
        st.download_button(
            f"⬇️ Baixar CSV ({n_linhas} linhas)",
            data=csv_str.encode("utf-8-sig"),
            file_name=f"{nome_arquivo}.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"dlcsv_{nome_arquivo}_{n_linhas}",
        )


def _df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Dados") -> bytes:
    import io
    import re
    from openpyxl import Workbook
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    # Excel proíbe : \ / ? * [ ] no nome da aba, e limita a 31 chars
    nome_aba = re.sub(r"[:\\/\?\*\[\]]", "_", sheet_name)[:31] or "Dados"
    ws.title = nome_aba
    # header
    ws.append([str(c) for c in df.columns])
    for _, row in df.iterrows():
        ws.append([
            v.to_pydatetime() if isinstance(v, pd.Timestamp)
            else (None if pd.isna(v) else v)
            for v in row
        ])
    wb.save(buf)
    return buf.getvalue()


def render_tab_conciliadas(df: pd.DataFrame, conta: str):
    if df.empty:
        st.success("🎉 Nenhuma linha conciliada nesta conta.")
        return
    cols_preferidas = [
        "banco_data", "banco_historico", "banco_documento", "banco_valor",
        "sistema_data", "sistema_historico", "sistema_documento", "sistema_valor",
        "dias_diferenca", "status", "motivo",
    ]
    cols_existentes = [c for c in cols_preferidas if c in df.columns]
    out = df[cols_existentes].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"conciliadas_{conta}", "Tudo certo — nenhum item nesta aba.")


def render_tab_pendentes(resultado: ResultadoConciliacao, conta: str):
    pb = (
        resultado.pendentes_banco[resultado.pendentes_banco["conta"] == conta].copy()
        if not resultado.pendentes_banco.empty else pd.DataFrame()
    )
    # v5.32: esta aba ("Sem baixa no Sankhya") mostra SÓ o lado do banco — movimento
    # que entrou no banco e ainda não tem baixa no Sankhya (pendentes_banco). O lado
    # oposto (Sankhya sem confirmação = pendentes_sistema) já aparece na aba
    # "Divergências (Sankhya × Banco)"; mostrá-lo aqui também duplicava a mesma linha.
    if not pb.empty:
        pb["origem"] = "Banco (falta baixar no Sankhya)"
    df = pb
    if df.empty:
        st.success("🎉 Nada sem baixa no Sankhya nesta conta.")
        return

    # v5.35: somatório total da coluna (igual ao "Resumo por origem" do Divergência),
    # pra dar o total e permitir comparar.
    _total_pb = float(df["valor"].abs().sum()) if "valor" in df.columns else 0.0
    _qtd_pb = len(df)
    _un_pb = "item" if _qtd_pb == 1 else "itens"
    st.markdown(f"**Total sem baixa no Sankhya:** {fmt_brl(_total_pb)} · {fmt_int(_qtd_pb)} {_un_pb}")

    # v5.11: editor inline pra trocar o Tipo de uma linha pendente
    st.info(
        "💡 Voc\u00ea pode **editar a coluna Tipo** clicando na c\u00e9lula. "
        "Use isso quando o sistema classificou errado (ex.: 'GETNET' como 'Outros'). "
        "Mudan\u00e7as ficam salvas s\u00f3 nesta sess\u00e3o do navegador."
    )

    from src.classificacao import TIPOS_PRINCIPAIS

    cols = ["origem", "data", "historico", "documento", "valor", "tipo", "natureza"]
    cols = [c for c in cols if c in df.columns]
    df_visu = df[cols].copy()

    # Aplica edi\u00e7\u00f5es pr\u00e9vias salvas em session_state
    chave_edicoes = f"edicoes_tipo_pendentes_{conta}"
    edicoes = st.session_state.get(chave_edicoes, {})
    if edicoes and "tipo" in df_visu.columns:
        for chave, novo_tipo in edicoes.items():
            # chave = (data, historico, valor)
            try:
                data_chave, hist_chave, val_chave = chave
                mask = (
                    (df_visu["data"].astype(str) == str(data_chave))
                    & (df_visu["historico"].astype(str) == str(hist_chave))
                    & (df_visu["valor"].round(2) == round(float(val_chave), 2))
                )
                df_visu.loc[mask, "tipo"] = novo_tipo
            except Exception:
                pass

    # Op\u00e7\u00f5es do dropdown
    opcoes_tipo = sorted(set(TIPOS_PRINCIPAIS + ["Outros", "Salário/Folha", "Imposto", "Transferência"]))

    edited = st.data_editor(
        df_visu,
        column_config={
            "origem": st.column_config.TextColumn("Origem", disabled=True),
            "data": st.column_config.DateColumn("Data", format="DD/MM/YYYY", disabled=True),
            "historico": st.column_config.TextColumn("Hist\u00f3rico", disabled=True),
            "documento": st.column_config.TextColumn("Documento", disabled=True),
            "valor": st.column_config.NumberColumn("Valor (R$)", format="%.2f", disabled=True),
            "tipo": st.column_config.SelectboxColumn(
                "Tipo (edit\u00e1vel)",
                options=opcoes_tipo,
                required=False,
                help="Clique para mudar a classifica\u00e7\u00e3o desta linha.",
            ),
            "natureza": st.column_config.TextColumn("Natureza", disabled=True),
        },
        use_container_width=True,
        hide_index=True,
        key=f"editor_pendentes_{conta}",
    )

    # Salva edi\u00e7\u00f5es no session_state se mudou
    if not edited.equals(df_visu):
        novas_edicoes = dict(edicoes)
        for i, row in edited.iterrows():
            if i < len(df_visu):
                orig = df_visu.iloc[i]
                if row["tipo"] != orig["tipo"]:
                    chave = (str(orig["data"]), str(orig["historico"]), round(float(orig["valor"]), 2))
                    novas_edicoes[chave] = row["tipo"]
        st.session_state[chave_edicoes] = novas_edicoes
        if len(novas_edicoes) > len(edicoes):
            st.toast(f"✅ {len(novas_edicoes) - len(edicoes)} classificações atualizadas", icon="✅")

    # Bot\u00e3o pra resetar edi\u00e7\u00f5es
    if edicoes:
        col_a, col_b = st.columns([3, 1])
        with col_b:
            if st.button(f"🔄 Resetar edições ({len(edicoes)})",
                         key=f"reset_edicoes_{conta}", use_container_width=True):
                st.session_state[chave_edicoes] = {}
                st.rerun()


def render_tab_nao_pertence(df: pd.DataFrame, conta: str):
    if df.empty:
        st.success("🎉 Nenhum lançamento parece estar na conta errada.")
        return
    out = df.copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"nao_pertence_{conta}")


def render_tab_divergencias(df: pd.DataFrame, conta: str):
    if df.empty:
        st.success("🎉 Sem divergências de valor nesta conta.")
        return
    cols = ["data", "historico_banco", "valor_banco", "historico_sistema",
            "valor_sistema", "diferenca", "documento_banco", "documento_sistema",
            "motivo"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"divergencias_{conta}")


def render_tab_aplicacoes(df: pd.DataFrame, conta: str):
    """Aplicações e Resgates da conta."""
    if df.empty:
        st.info("Nenhuma aplicação ou resgate identificado para esta conta.")
        return
    st.info(
        "💡 Estas linhas **não entram** no Total do Extrato Bancário (são movimentações "
        "entre conta corrente e investimento). Use esta lista para identificar "
        "lançamentos faltantes e confirmar manualmente quando necessário."
    )
    cols = ["origem", "tipo_aplicacao", "data", "historico",
            "documento", "valor", "categoria_mov"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"aplicacoes_resgates_{conta}")


# v5.0: renderizadores das abas novas
def render_tab_estornos_anulados(df: pd.DataFrame, conta: str):
    """v5.0: pares pagamento/recebimento + estorno com saldo zero."""
    if df.empty:
        st.info("Nenhum par anulado por estorno nesta conta.")
        return
    st.info(
        "♻️ **Pares anulados.** Estes lançamentos têm um par correspondente (recebimento + "
        "estorno OU pagamento + reversão) com saldo líquido **R$ 0,00**. "
        "Eles **não aparecem** em Pendentes, Falta Conciliar, Conta 70 ou Divergências."
    )
    out = df.copy()
    if "data_original" in out.columns:
        out["data_original"] = pd.to_datetime(out["data_original"]).dt.strftime("%d/%m/%Y")
    if "data_estorno" in out.columns:
        out["data_estorno"] = pd.to_datetime(out["data_estorno"]).dt.strftime("%d/%m/%Y")
    if "valor_original" in out.columns:
        out["valor_original"] = out["valor_original"].apply(fmt_brl)
    if "valor_estornado" in out.columns:
        out["valor_estornado"] = out["valor_estornado"].apply(fmt_brl)
    if "saldo_liquido" in out.columns:
        out["saldo_liquido"] = out["saldo_liquido"].apply(fmt_brl)
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"estornos_anulados_{conta}")


def render_tab_estornos_parciais(df: pd.DataFrame, conta: str):
    """v5.0: estornos parciais (saldo ≠ zero) — a diferença volta pra análise."""
    if df.empty:
        st.info("Nenhum estorno parcial nesta conta.")
        return
    st.warning(
        "⚖️ **Estornos parciais.** Aqui o estorno **não anulou totalmente** o lançamento "
        "original. A diferença foi mantida no fluxo como linha sintética para análise "
        "(marcada com **`[ESTORNO PARCIAL]`** no histórico)."
    )
    out = df.copy()
    if "data_original" in out.columns:
        out["data_original"] = pd.to_datetime(out["data_original"]).dt.strftime("%d/%m/%Y")
    if "data_estorno" in out.columns:
        out["data_estorno"] = pd.to_datetime(out["data_estorno"]).dt.strftime("%d/%m/%Y")
    for c in ("valor_original", "valor_estornado", "saldo_liquido"):
        if c in out.columns:
            out[c] = out[c].apply(fmt_brl)
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"estornos_parciais_{conta}")


def render_tab_top1722_grupos(grupos: pd.DataFrame, todas_linhas: pd.DataFrame, conta: str,
                                linhas_banco: pd.DataFrame = None):
    """v5.2: agrupamento TOP 1722 por SOMA TOTAL por conta.
    Cada linha = 1 conta agrupada (não mais 1 crédito banco)."""
    if grupos.empty:
        st.info("Nenhum agrupamento TOP 1722 nesta conta.")
        return

    grupo = grupos.iloc[0]
    qtd_creditos = int(grupo.get("qtd_creditos_banco", 0))
    valor_banco = float(grupo.get("valor_banco_total", 0))
    qtd_sankhya = int(grupo.get("qtd_linhas_sankhya", 0))
    valor_sankhya = float(grupo.get("valor_sankhya_total", 0))
    diferenca = float(grupo.get("diferenca", 0))

    st.success(
        f"🃏 **Conciliação por Agrupamento — Cartão TOP 1722** · {conta}\n\n"
        f"O sistema somou todos os **créditos de cartão no banco** ({qtd_creditos} lançamentos = "
        f"**{fmt_brl(valor_banco)}**) e todas as **vendas TOP 1722 no Sankhya** ({qtd_sankhya} lançamentos = "
        f"**{fmt_brl(valor_sankhya)}**). Diferença: **{fmt_brl(diferenca)}**. "
        f"Como bateu, todas as linhas foram tiradas de Pendentes/Divergência."
    )

    # Card-resumo do agrupamento
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Créditos no Banco", fmt_int(qtd_creditos), fmt_brl(valor_banco))
    with col2:
        st.metric("Vendas TOP 1722 Sankhya", fmt_int(qtd_sankhya), fmt_brl(valor_sankhya))
    with col3:
        st.metric("Diferença", fmt_brl(diferenca), "R$ 0,00" if diferenca == 0 else None)

    # Detalhamento
    if linhas_banco is not None and not linhas_banco.empty:
        with st.expander(f"🔍 Ver composição — créditos do banco ({qtd_creditos} linhas)"):
            lb = linhas_banco[linhas_banco["conta"] == conta].copy()
            if "data" in lb.columns:
                lb["data"] = pd.to_datetime(lb["data"]).dt.strftime("%d/%m/%Y")
            if "valor" in lb.columns:
                lb["valor"] = lb["valor"].apply(fmt_brl)
            cols = [c for c in ("data", "historico", "documento", "valor") if c in lb.columns]
            lb = lb[cols]
            lb.columns = [c.replace("_", " ").title() for c in lb.columns]
            st.dataframe(lb, use_container_width=True, hide_index=True)

    if not todas_linhas.empty:
        with st.expander(f"🔍 Ver composição — vendas no Sankhya ({qtd_sankhya} linhas)"):
            ls = todas_linhas[todas_linhas["conta"] == conta].copy()
            if "data" in ls.columns:
                ls["data"] = pd.to_datetime(ls["data"]).dt.strftime("%d/%m/%Y")
            if "valor" in ls.columns:
                ls["valor"] = ls["valor"].apply(fmt_brl)
            cols = [c for c in ("data", "historico", "documento", "valor") if c in ls.columns]
            ls = ls[cols]
            ls.columns = [c.replace("_", " ").title() for c in ls.columns]
            _exibir_df(ls, f"top1722_sankhya_{conta}")


def render_tab_top1722_diferenca(df: pd.DataFrame, conta: str):
    """v5.2: agrupamento com diferença (taxa) OU diferença grande não-agrupada."""
    if df.empty:
        st.info("Sem diferenças TOP 1722 nesta conta.")
        return

    linha = df.iloc[0]
    valor_banco = float(linha.get("valor_banco_total", 0))
    valor_sankhya = float(linha.get("valor_sankhya_total", 0))
    diferenca = float(linha.get("diferenca", 0))
    pct = float(linha.get("percentual_diferenca", 0))
    status = str(linha.get("status", ""))
    motivo = str(linha.get("motivo", ""))

    st.warning(
        f"⚠️ **TOP 1722 com Diferença — a analisar** · {conta}\n\n"
        f"Há uma diferença de **{fmt_brl(abs(diferenca))}** ({pct:.2f}%) entre o cartão no "
        f"Banco e no Sankhya. **Não foi confirmada como taxa** — pode ser venda ainda sem "
        f"baixa, taxa, ou estorno. As linhas **continuam em Pendentes/Divergência** (não "
        f"foram escondidas). Confira as transações abaixo antes de fechar."
    )

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Banco", fmt_brl(valor_banco), f"{int(linha.get('qtd_creditos_banco',0))} créditos")
    with col2:
        st.metric("Sankhya TOP 1722", fmt_brl(valor_sankhya), f"{int(linha.get('qtd_linhas_sankhya',0))} linhas")
    with col3:
        st.metric("Diferença", fmt_brl(diferenca), f"{pct:.2f}%")

    if motivo:
        st.caption(f"💡 {motivo}")


def render_tab_possiveis_duplicidades(df: pd.DataFrame, conta: str):
    """Possíveis duplicidades (3 de 4 campos batendo)."""
    if df.empty:
        st.success("🎉 Nenhuma possível duplicidade detectada.")
        return
    st.warning(
        "⚠️ **REVISAR MANUALMENTE.** Estes lançamentos têm 3 de 4 campos iguais "
        "(data, histórico, valor, documento). Não é certeza de duplicidade — "
        "pode ser apenas coincidência (ex: pagamentos similares de clientes diferentes)."
    )
    cols = ["origem", "data", "historico", "documento", "valor", "motivo"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"possiveis_duplicidades_{conta}")


def render_tab_excesso_sankhya(df: pd.DataFrame, conta: str):
    """Excesso no Sankhya: lançamentos do Sankhya sem contrapartida no banco (v3)."""
    if df.empty:
        st.success("🎉 Sankhya não tem lançamentos excedentes em relação ao banco.")
        return
    st.warning(
        "⚠️ **A BASE DA VERDADE É O EXTRATO BANCÁRIO.** "
        "Estes lançamentos aparecem MAIS vezes no Sankhya do que no extrato bancário "
        "para o mesmo perfil (data + valor + histórico + conta). "
        "Possível lançamento duplicado no ERP — revisar."
    )
    cols = ["data", "historico", "documento", "valor",
            "qtd_sankhya", "qtd_banco", "excedente_total", "motivo"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"excesso_sankhya_{conta}")


def render_tab_divergencia_consolidada(df: pd.DataFrame, conta: str):
    """v3.4: Aba 'Divergências (Sankhya × Banco)' — visão consolidada."""
    st.info(
        "📌 **DIVERGÊNCIA = tudo o que o Sankhya tem a mais que o extrato bancário.** "
        "O banco é a base da verdade. Inclui 3 origens:\n\n"
        "- **Sem par no banco**: lançamentos do Sankhya que o app não conseguiu casar com "
        "nenhuma linha do banco (decisão da comparação banco × Sankhya — não depende mais da "
        "flag `Conciliado` do ERP).\n"
        "- **Excesso no Sankhya**: mesma data+valor+conta aparece mais vezes no Sankhya do que no banco.\n"
        "- **Valor diferente**: mesma chave (data+histórico+conta) com valor diferente entre Sankhya e Banco."
    )
    if df.empty:
        st.success("🎉 Sem divergências — Sankhya está alinhado com o banco!")
        return

    # Resumo por origem
    if "origem_divergencia" in df.columns:
        # v5.37: soma COM SINAL (líquido). Antes somava o valor absoluto, então
        # +13,06 e −9,33 davam R$ 22,39 (sem sentido) em vez de R$ 3,73.
        resumo = df.groupby("origem_divergencia").agg(
            quantidade=("valor", "count"),
            total=("valor", "sum"),
        ).reset_index()
        resumo.columns = ["Origem da Divergência", "Quantidade", "Valor Total (líquido)"]
        resumo["Valor Total (líquido)"] = resumo["Valor Total (líquido)"].apply(fmt_brl)
        st.markdown("**Resumo por origem:**")
        st.dataframe(resumo, use_container_width=True, hide_index=True)
        st.write("")

    cols = ["origem_divergencia", "origem", "data", "historico", "documento", "valor", "conta"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"divergencias_sankhya_banco_{conta}")


# Mantido por retrocompat: redireciona para a nova visão consolidada
def render_tab_falta_lancar(df: pd.DataFrame, conta: str, fonte: str):
    """[DEPRECATED v3.4] Substituído por render_tab_divergencia_consolidada."""
    if df.empty:
        st.success("🎉 Nada para lançar — tudo conciliado!")
        return
    cols = ["data", "historico", "documento", "valor", "tipo", "natureza", "conciliado"]
    cols = [c for c in cols if c in df.columns]
    out = df[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in out.columns]
    _exibir_df(out, f"falta_lancar_{conta}")


# ============================================================
# Subabas por TIPO de lançamento (Boleto, Pix, Tarifa...)
# ============================================================
def render_subabas_tipo(resultado: ResultadoConciliacao, conta: str | None = None):
    """v3.3: Subabas por tipo mostram pares conciliados + pendentes (banco e sankhya)
    juntos, todos com coluna Status. O tipo dos pares conciliados vem do lado banco
    (histórico mais padronizado). Aba 'Outros' = sem categoria reconhecida.

    v3.8: Se `conta` for fornecido, filtra apenas registros daquela conta.
    """
    _tipos_base = [t for t in TIPOS_PRINCIPAIS if t not in ("Pagamentos", "Recebimentos", "Outros")]
    tipos_disponiveis = ["Todos"] + _tipos_base + ["Pagamentos", "Recebimentos", "Outros"]
    tabs = st.tabs(tipos_disponiveis)

    # v3.10: visão unificada é cara (concat de 3 DataFrames). Cache no session_state.
    # Chave inclui id_execucao_atual pra invalidar quando o resultado muda.
    # v5.15: adicionada versão na chave pra invalidar cache antigo após deploy do filtro
    # de aplicações/resgates.
    cache_key = f"visao_unificada_v515_{st.session_state.get('id_execucao_atual', 'novo')}"
    df_unif = st.session_state.get(cache_key)
    if df_unif is None:
        df_unif = _montar_visao_unificada(resultado)
        st.session_state[cache_key] = df_unif

    if conta is not None and not df_unif.empty and "conta" in df_unif.columns:
        df_unif = df_unif[df_unif["conta"] == conta].copy()

    for _itab, (tab, tipo) in enumerate(zip(tabs, tipos_disponiveis)):
        with tab:
            if tipo == "Todos":
                df = df_unif.copy()
            elif tipo == "Pagamentos":
                df = df_unif[df_unif["natureza"] == "Pagamento"].copy()
            elif tipo == "Recebimentos":
                df = df_unif[df_unif["natureza"] == "Recebimento"].copy()
            else:
                df = df_unif[df_unif["tipo"] == tipo].copy()

            # v5.12: Aba Cartão — quando regra TOP 1722 esvazia tudo, mostra a visão
            # banco × sankhya do agrupamento (mesmas linhas que sumiram de Pendentes/Diverg).
            mostrar_top1722_aqui = (tipo == "Cartão")

            if df.empty and not mostrar_top1722_aqui:
                if tipo == "Outros":
                    st.info(
                        "🎉 Nenhum lançamento sem categoria. "
                        "Esta aba mostra lançamentos cujo histórico não bateu com "
                        "nenhuma das categorias conhecidas (Pix, Boleto, Tarifa, TED/DOC, "
                        "Débito Automático, Cartão, Salário/Folha, Imposto, Transferência)."
                    )
                else:
                    st.info(f"Nenhum lançamento do tipo **{tipo}**.")
                continue

            # KPIs do tipo (só se houver lançamentos regulares)
            if not df.empty:
                qtd = len(df)
                valor_total = float(df["valor"].abs().sum())
                receitas = float(df[df["valor"] > 0]["valor"].sum())
                despesas = float(df[df["valor"] < 0]["valor"].abs().sum())

                sub_rec_desp = _card_total_com_rec_desp(receitas, despesas)

                # Contagens por status
                qtd_conciliado = int((df["status"] == "Conciliado").sum())
                qtd_pend_b = int((df["status"] == "Pendente Banco").sum())
                qtd_pend_s = int((df["status"] == "Pendente Sankhya").sum())

                label_qtd = f"Lançamentos {tipo}" if tipo not in ("Todos", "Outros") else f"Lançamentos {tipo.lower()}"
                cards = [
                    card_kpi(label_qtd, fmt_int(qtd),
                             f"Conciliados: {qtd_conciliado} · Pend. Banco: {qtd_pend_b} · Pend. Sankhya: {qtd_pend_s}"),
                    card_kpi_html("Valor total (absoluto)", fmt_brl(valor_total),
                                  sub_rec_desp, classe="destaque-amarelo"),
                    card_kpi("Contas envolvidas",
                             fmt_int(df["conta"].nunique())),
                ]
                render_cards(cards)

                # Tabela: status + origem + dados padronizados
                if tipo == "Recebimentos":
                    st.caption(
                        "💡 **Recebimentos** = qualquer entrada na conta (TED/DOC de cliente, "
                        "transferência recebida, depósito, crédito de cartão, devolução, etc.). "
                        "Use os tipos específicos (Pix, Boleto, Cartão, TED/DOC) pra ver "
                        "categorias isoladas."
                    )
                elif tipo == "Outros":
                    st.caption(
                        "📌 Lançamentos cujo histórico não bateu com nenhuma categoria conhecida. "
                        "Use esta lista para identificar termos que valem regras novas."
                    )

                cols_show = ["status", "origem", "data", "conta", "historico",
                             "documento", "valor", "tipo", "natureza"]
                cols_show = [c for c in cols_show if c in df.columns]
                out = df[cols_show].copy()
                out.columns = [c.title() for c in out.columns]
                _exibir_df(out, f"tipo{_itab}_{tipo.lower().replace(' ', '_').replace('/', '_')}")

            # v5.12: Bloco TOP 1722 (banco × sankhya) — aparece na aba Cartão
            if mostrar_top1722_aqui:
                _render_bloco_top1722_banco_sankhya(resultado, conta, mostrou_acima=(not df.empty))


def _render_bloco_top1722_banco_sankhya(
    resultado: ResultadoConciliacao,
    conta: str | None,
    mostrou_acima: bool,
):
    """v5.12: Mostra os lançamentos do agrupamento TOP 1722 em visão banco × sankhya
    (lado a lado). Esses lançamentos foram removidos de Pendentes/Divergências pela
    regra de agrupamento, então a aba Cartão ficaria vazia sem este bloco.
    """
    linhas_banco = getattr(resultado, "top1722_linhas_banco", pd.DataFrame())
    linhas_sank = getattr(resultado, "top1722_linhas", pd.DataFrame())

    # Filtra por conta se informado
    if conta is not None:
        if not linhas_banco.empty and "conta" in linhas_banco.columns:
            linhas_banco = linhas_banco[linhas_banco["conta"] == conta]
        if not linhas_sank.empty and "conta" in linhas_sank.columns:
            linhas_sank = linhas_sank[linhas_sank["conta"] == conta]

    if linhas_banco.empty and linhas_sank.empty:
        if not mostrou_acima:
            st.info(
                "Nenhum lançamento do tipo **Cartão**. "
                "Se houver crédito de cartão no banco, ele aparece aqui "
                "(agrupado com vendas TOP 1722 do Sankhya)."
            )
        return

    if mostrou_acima:
        st.divider()

    qtd_b = len(linhas_banco)
    qtd_s = len(linhas_sank)
    val_b = float(linhas_banco["valor"].abs().sum()) if not linhas_banco.empty and "valor" in linhas_banco.columns else 0.0
    val_s = float(linhas_sank["valor"].abs().sum()) if not linhas_sank.empty and "valor" in linhas_sank.columns else 0.0
    diff = val_b - val_s

    st.success(
        f"🃏 **Cartão TOP 1722 — Agrupamento Banco × Sankhya**\n\n"
        f"Esses lançamentos foram conciliados por **soma agrupada** (não 1-pra-1): "
        f"{qtd_b} créditos no banco ({fmt_brl(val_b)}) × {qtd_s} vendas TOP 1722 no Sankhya "
        f"({fmt_brl(val_s)}). Diferença: **{fmt_brl(diff)}**."
    )

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Créditos no Banco", fmt_int(qtd_b), fmt_brl(val_b))
    with col2:
        st.metric("Vendas TOP 1722 Sankhya", fmt_int(qtd_s), fmt_brl(val_s))
    with col3:
        st.metric("Diferença", fmt_brl(diff), "R$ 0,00" if diff == 0 else None)

    # Tabelas lado a lado
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**🏦 Banco — Créditos de Cartão**")
        if linhas_banco.empty:
            st.info("Sem créditos no banco para o filtro atual.")
        else:
            lb = linhas_banco.copy()
            if "data" in lb.columns:
                lb["data"] = pd.to_datetime(lb["data"], errors="coerce").dt.strftime("%d/%m/%Y")
            if "valor" in lb.columns:
                lb["valor"] = lb["valor"].apply(fmt_brl)
            cols_b = [c for c in ("data", "conta", "historico", "documento", "valor") if c in lb.columns]
            lb = lb[cols_b]
            lb.columns = [c.replace("_", " ").title() for c in lb.columns]
            st.dataframe(lb, use_container_width=True, hide_index=True)

    with col_b:
        st.markdown("**📋 Sankhya — Vendas TOP 1722**")
        if linhas_sank.empty:
            st.info("Sem vendas TOP 1722 no Sankhya para o filtro atual.")
        else:
            ls = linhas_sank.copy()
            if "data" in ls.columns:
                ls["data"] = pd.to_datetime(ls["data"], errors="coerce").dt.strftime("%d/%m/%Y")
            if "valor" in ls.columns:
                ls["valor"] = ls["valor"].apply(fmt_brl)
            cols_s = [c for c in ("data", "conta", "historico", "documento", "valor") if c in ls.columns]
            ls = ls[cols_s]
            ls.columns = [c.replace("_", " ").title() for c in ls.columns]
            st.dataframe(ls, use_container_width=True, hide_index=True)


def _montar_visao_unificada(resultado: ResultadoConciliacao) -> pd.DataFrame:
    """Une conciliados (lado banco) + pendentes banco + pendentes sankhya em um único
    DataFrame com coluna 'status' e 'origem'.

    Para pares conciliados, usa o tipo/natureza do lado BANCO (histórico padronizado).

    v5.15: Exclui aplicações/resgates/SALDO da visão genérica — esses lançamentos
    são tratamento especial e aparecem só no card de Investimentos. Antes apareciam
    duplicados em 'Recebimentos' (Por Tipo de Lançamento).
    """
    frames = []

    # Pares conciliados — pega o lado banco
    if not resultado.conciliados.empty:
        c = resultado.conciliados.copy()
        # Renomeia banco_* → colunas canônicas
        renomeio = {
            "banco_data": "data",
            "banco_historico": "historico",
            "banco_documento": "documento",
            "banco_valor": "valor",
            "banco_conta": "conta",
            "banco_tipo": "tipo",
            "banco_natureza": "natureza",
        }
        cols_pegar = [k for k in renomeio if k in c.columns]
        d = c[cols_pegar].rename(columns=renomeio).copy()
        d["status"] = "Conciliado"
        d["origem"] = "Banco (conciliado)"
        frames.append(d)

    # Pendentes do banco
    if not resultado.pendentes_banco.empty:
        d = resultado.pendentes_banco.copy()
        d["status"] = "Pendente Banco"
        d["origem"] = "Banco"
        cols_keep = [c for c in ["data","historico","documento","valor","conta","tipo","natureza"]
                     if c in d.columns]
        d = d[cols_keep + ["status", "origem"]]
        frames.append(d)

    # Pendentes do sankhya
    if not resultado.pendentes_sistema.empty:
        d = resultado.pendentes_sistema.copy()
        d["status"] = "Pendente Sankhya"
        d["origem"] = "Sankhya"
        cols_keep = [c for c in ["data","historico","documento","valor","conta","tipo","natureza"]
                     if c in d.columns]
        d = d[cols_keep + ["status", "origem"]]
        frames.append(d)

    if not frames:
        return pd.DataFrame(columns=["status","origem","data","conta","historico",
                                      "documento","valor","tipo","natureza"])

    out = pd.concat(frames, ignore_index=True)

    # v5.15: filtra aplicações/resgates/rendimentos/SALDO da visão genérica
    # (esses lançamentos pertencem ao card de Investimentos, não a Recebimentos)
    if "historico" in out.columns:
        h = out["historico"].astype(str).str.upper()
        mask_invest = (
            h.str.contains("APLIC", na=False)
            | h.str.contains("RESG", na=False)
            | h.str.contains("REND PAGO", na=False)
            | h.str.contains("SALDO APLIC", na=False)
            | h.str.contains("AUT MAIS", na=False)
        )
        out = out[~mask_invest].reset_index(drop=True)

    return out


def _filtrar_por(resultado: ResultadoConciliacao, filtro_fn) -> pd.DataFrame:
    """Aplica filtro_fn nos dois lados (banco + sistema) e marca origem."""
    frames = []
    if not resultado.banco_completo.empty:
        d = filtro_fn(resultado.banco_completo).copy()
        if not d.empty:
            d["origem"] = "Banco"
            frames.append(d)
    if not resultado.sistema_completo.empty:
        d = filtro_fn(resultado.sistema_completo).copy()
        if not d.empty:
            d["origem"] = "Sistema"
            frames.append(d)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ============================================================
# Página: Histórico (execuções e reprocessamentos)
# ============================================================
def pagina_historico():
    execucoes = listar_execucoes()
    if not execucoes:
        st.info(
            "📂 Nenhuma execução registrada ainda. "
            "Cada conciliação processada gera uma entrada aqui, com snapshot dos arquivos "
            "e do relatório (auditoria append-only)."
        )
        return

    section_title(f"{len(execucoes)} EXECUÇÃO(ÕES) REGISTRADAS")

    busca = st.text_input(
        "🔍 Buscar por conta, data ou ID",
        placeholder="ex: Bradesco ou 20260515",
        label_visibility="collapsed",
    )

    filtradas = execucoes
    if busca:
        b = busca.lower().strip()
        filtradas = [
            e for e in execucoes
            if b in " ".join(e.get("contas", [])).lower()
            or b in e.get("id", "").lower()
            or b in e.get("data_referencia", "").lower()
        ]

    if not filtradas:
        st.warning("Nenhuma execução encontrada para essa busca.")
        return

    for exec_data in filtradas:
        with st.container():
            kpis = exec_data.get("kpis", {})
            contas = exec_data.get("contas", [])
            contas_str = ", ".join(contas[:3]) + (f" +{len(contas)-3}" if len(contas) > 3 else "")
            data_ref = exec_data.get("data_referencia", "")[:10]
            try:
                data_ref_fmt = pd.to_datetime(data_ref).strftime("%d/%m/%Y")
            except Exception:
                data_ref_fmt = data_ref
            ts_fmt = exec_data.get("timestamp", "")[:19].replace("T", " ")
            pct = kpis.get("percentual_conciliado", 0)
            classe_pct = "verde" if pct >= 95 else ("amarelo" if pct >= 70 else "vermelho")
            status = exec_data.get("status", "processado")
            versao = exec_data.get("versao", 1)

            st.html(
                f"""
                <div class="lle-kpi" style="margin-bottom:8px;">
                    <div style="display:flex; align-items:center; gap:14px; flex-wrap:wrap;">
                        <span class="lle-badge azul">{exec_data.get("id","")}</span>
                        <strong style="color:{CORES['amarelo']}; font-size:14px;">
                            {data_ref_fmt}
                        </strong>
                        <span style="color:{CORES['texto_muted']};">·</span>
                        <span style="font-weight:600;">{contas_str}</span>
                        <span class="lle-badge {classe_pct}">{fmt_pct(pct)} conciliado</span>
                        <span class="lle-badge azul">v{versao} · {status}</span>
                        <span style="color:{CORES['texto_muted']}; margin-left:auto; font-size:12px;">
                            {ts_fmt}
                        </span>
                    </div>
                    <div style="margin-top:10px; font-size:12px; color:{CORES['texto_muted']};">
                        Conciliados: {fmt_int(kpis.get('qtd_conciliados', 0))} ·
                        Pendentes banco: {fmt_int(kpis.get('qtd_pendentes_banco', 0))} ·
                        Pendentes sistema: {fmt_int(kpis.get('qtd_pendentes_sistema', 0))} ·
                        Total processado: {fmt_brl(kpis.get('total_extrato_bancario', 0))}
                    </div>
                </div>
                """
            )

            # Botão de baixar snapshot
            from src.auditoria import carregar_snapshot_relatorio
            xlsx_snap = carregar_snapshot_relatorio(exec_data.get("id", ""))
            if xlsx_snap:
                st.download_button(
                    f"⬇️ Baixar snapshot desta execução",
                    data=xlsx_snap,
                    file_name=f"snapshot_{exec_data.get('id','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"snap_{exec_data.get('id','')}",
                )


# ============================================================
# Página: Sobre — explicação das regras
# ============================================================
def pagina_sobre():
    tab1, tab2, tab3 = st.tabs(["Como funciona", "Regras de negócio", "Identidade visual"])

    with tab1:
        st.markdown(f"""
### 🏦 Conciliação Bancária — Grupo LLE

Sistema que bate diariamente o **extrato bancário** com o **extrato do sistema Sankhya**,
identificando conciliações, pendências, divergências e lançamentos na conta errada.

#### Fluxo
1. **Upload** dos extratos (1 ou várias contas) + relatório do ERP.
2. Processamento automático com as regras definidas pela equipe financeira.
3. **Tela única de resultado** com cards executivos, painel por banco e detalhamento por tipo.
4. **Download** em Excel (multi-aba) e CSV (zip) — globalmente ou por banco.
5. Auditoria de execuções fica em `data/outputs/auditoria.jsonl` (append-only — nada é sobrescrito).

#### Persistência sem banco
- O **Excel gerado é o "estado"**. Suba o do dia anterior em "Pendências de dias anteriores"
  para acompanhar há quantos dias cada pendência está aberta.
- Cada execução também salva um **snapshot completo** em `data/outputs/execucoes/{{id}}/`
  com os inputs, parâmetros e resultado, permitindo **reprocessamento** sem perder histórico.
        """)

    with tab2:
        st.markdown(f"""
### Regras determinantes da conciliação

**Match exato (conciliação automática):**
- Valor: precisa ser **exatamente igual** — sem tolerância de centavos.
- Conta: precisa ser **igual**.
- Data: tolerância de **±N dias corridos** (default 2 — cobre fim de semana e feriado curto).
- Match é **1-pra-1**. Cada lançamento do banco casa com no máximo um do sistema.

**Total Extrato Bancário (v2):**
- Soma absoluta SOMENTE de movimentações reais.
- **Exclui linhas de SALDO, APLICAÇÃO, RESGATE, INVESTIMENTO**, compra/venda de CDB/RDB/LCI/LCA/Tesouro etc.
- Essas linhas aparecem em aba dedicada ("Aplicações e Resgates"), mas não contam no total.

**Falta Conciliar:**
- Total que existe no extrato bancário mas não no Sankhya.
- Separado em **Receitas** e **Despesas** (despesas exibidas positivas).

**Falta Lançar — fonte automática:**
- Se o Sankhya tem a coluna **Conciliado** preenchida, usa as linhas com `Conciliado=Não`.
- Senão usa pendentes pós-match. O card mostra a fonte usada.

**Divergência de valor:**
- Mesma data + histórico exato (após normalização) + conta, mas valores diferentes.

**Duplicidade (estrita):**
- 4 de 4 campos iguais (data, histórico, valor, documento).
- 5 boletos legítimos com docs diferentes NÃO viram duplicidade.

**Possíveis Duplicidades (NOVO):**
- 3 de 4 campos batendo — aba própria, marcadas como "REVISAR MANUALMENTE".

**Não Pertence à Conta:**
- Pendência com candidato perfeito em outra conta (mesmo valor + data próxima).

**Receitas e Despesas Absolutas:**
- Separadas, sem compensação. Despesas em valor positivo no resumo.

**Saldo Final (NOVO):**
- Aparece quando a conta está 100% conciliada.
- Usa SALDO INICIAL/FINAL do extrato; senão usa movimentação líquida com aviso.

**Sugestões Fuzzy:**
- Aba complementar para revisão manual. Não entra na conciliação automática.
        """)

    with tab3:
        st.markdown(f"""
### Identidade Visual — Grupo LLE

Seguindo o Manual da Marca (Fev/2026):

| Cor | Hex | Uso |
|-----|-----|-----|
| Azul institucional escuro | `#041747` | Fundo principal |
| Azul primário | `#0071FE` | Botões e destaques |
| Amarelo | `#FAC318` | Sidebar e elementos de destaque |
| Verde | `#0F8C3B` | Indicadores positivos |
| Branco | `#FFFFFF` | Textos sobre fundo escuro |

**Fonte:** Montserrat (Google Fonts).

**Logo:** versão texto branco sobre fundo escuro (não distorcer, não mudar cor do símbolo).

📩 Dúvidas de identidade visual: marketing@grupolle.com.br
        """)


# ============================================================
# Página: Cadastro de Taxas (v4.0 — módulo CARTÃO)
# ============================================================
def pagina_cadastro_taxas():
    from src.cartao import carregar_cadastro_taxas, MODALIDADES_VALIDAS

    section_title("CADASTRO DE TAXAS — CONTRATOS COM ADQUIRENTES")

    st.markdown(
        "Esta tela mostra as taxas contratadas com as adquirentes (Stone, Cielo, Rede, etc) "
        "que servem de base para a **Auditoria de Taxas**. "
        "As taxas ficam num arquivo `taxas.xlsx` que você gerencia. "
        "Suba ele aqui sempre que houver alteração contratual."
    )

    st.write("")
    arq = st.file_uploader(
        "Suba o arquivo de taxas (.xlsx, .xls ou .csv)",
        type=["xlsx", "xls", "csv"],
        key="upload_taxas",
        help=(
            "Colunas obrigatórias: adquirente, modalidade, parcelas, taxa_mdr. "
            "Colunas opcionais: taxa_antecipacao, prazo_dias, vigencia_inicio, vigencia_fim."
        ),
    )

    if arq is not None:
        try:
            cadastro = carregar_cadastro_taxas(arq)
            st.session_state["cadastro_taxas"] = cadastro
            st.success(f"✅ Cadastro carregado com {len(cadastro)} taxas.")
        except Exception as e:
            st.error(f"❌ Erro ao ler taxas.xlsx: {e}")

    cadastro = st.session_state.get("cadastro_taxas")

    if cadastro is None or cadastro.empty:
        st.info(
            "🙋 Nenhum cadastro carregado. Suba o arquivo `taxas.xlsx` acima.\n\n"
            "**Formato esperado**: uma planilha Excel com as colunas\n"
            "- `adquirente` (Stone, Cielo, Rede, Getnet, …)\n"
            f"- `modalidade` ({', '.join(MODALIDADES_VALIDAS)})\n"
            "- `parcelas` (número inteiro: 1 para débito/à vista, 2-12 para parcelado)\n"
            "- `taxa_mdr` (`1,39%` ou `0.0139`)\n"
            "- `taxa_antecipacao` (opcional)\n"
            "- `prazo_dias` (opcional: D+N)\n"
            "- `vigencia_inicio` / `vigencia_fim` (opcional; deixe vazio se ainda vigente)\n\n"
            "💡 Há um arquivo de exemplo em `data/samples/taxas_exemplo.xlsx`."
        )
        return

    # Resumo
    section_title("RESUMO DO CADASTRO")
    qtd_adq = cadastro["adquirente"].nunique()
    qtd_mod = cadastro["modalidade"].nunique()
    cards = [
        card_kpi("Adquirentes cadastradas", fmt_int(qtd_adq)),
        card_kpi("Modalidades distintas", fmt_int(qtd_mod)),
        card_kpi("Total de regras", fmt_int(len(cadastro))),
        card_kpi("Taxa MDR média (débito)",
                 fmt_pct(cadastro[cadastro["modalidade"] == "Débito"]["taxa_mdr"].mean() * 100)
                 if (cadastro["modalidade"] == "Débito").any() else "—"),
    ]
    render_cards(cards)

    st.write("")
    section_title("CADASTRO DETALHADO")

    visu = cadastro.copy()
    visu["taxa_mdr"] = (visu["taxa_mdr"] * 100).round(4).astype(str) + "%"
    visu["taxa_antecipacao"] = (visu["taxa_antecipacao"] * 100).round(4).astype(str) + "%"
    visu["prazo_dias"] = "D+" + visu["prazo_dias"].astype(str)
    visu["vigencia_inicio"] = visu["vigencia_inicio"].dt.strftime("%d/%m/%Y").fillna("—")
    visu["vigencia_fim"] = visu["vigencia_fim"].dt.strftime("%d/%m/%Y").fillna("(vigente)")
    visu.columns = [c.replace("_", " ").title() for c in visu.columns]
    st.dataframe(visu, use_container_width=True, hide_index=True)

    st.caption(
        "💡 Para editar: abra o `taxas.xlsx` no Excel, faça as alterações e suba novamente. "
        "Histórico de alterações fica preservado no Excel via versionamento manual."
    )


# ============================================================
# Página: Auditoria de Taxas (v4.0 — módulo CARTÃO)
# ============================================================
def pagina_auditoria_taxas():
    from src.cartao import (
        carregar_relatorio_adquirente,
        auditar_taxas,
        consolidar_historico,
        eh_extrato_getnet_cru,
        carregar_extrato_getnet_cru,
        resumir_extrato_getnet,
    )
    from io import BytesIO

    section_title("AUDITORIA DE CARTÕES — ADQUIRENTES")

    cadastro = st.session_state.get("cadastro_taxas")
    if cadastro is None or cadastro.empty:
        st.warning(
            "⚠️ Você precisa carregar o **Cadastro de Taxas** primeiro. "
            "Vá em `💳 CARTÃO → 🏦 Cadastro de Taxas` e suba o `taxas.xlsx`."
        )
        return

    st.markdown(
        "Suba aqui o **relatório padronizado das adquirentes** do período que quer auditar. "
        "O sistema compara a taxa que a adquirente cobrou contra a taxa contratada no cadastro "
        "e marca cada lançamento como **OK**, **Arredondamento**, **Discrepância** ou **Sem contrato**."
    )

    # v4.1: bloco de histórico acumulado
    with st.expander("📁 Subir auditorias anteriores (opcional)", expanded=False):
        st.markdown(
            "Suba aqui Excels de auditorias **anteriores** (gerados pelo próprio sistema) "
            "para que os KPIs e tabelas acumulem o período completo. "
            "Útil pra ver o impacto financeiro de várias auditorias somadas."
        )
        arqs_hist = st.file_uploader(
            "Arquivos de auditorias anteriores (vários permitidos)",
            type=["xlsx"],
            key="upload_historico_auditorias",
            accept_multiple_files=True,
            help=(
                "Sobe os arquivos baixados por 'Baixar auditoria completa' em sessões "
                "anteriores. O sistema agrega tudo num único resultado consolidado."
            ),
        )
        historico_df = pd.DataFrame()
        if arqs_hist:
            historico_df, avisos = consolidar_historico(arqs_hist)
            if historico_df.empty:
                st.warning(
                    "Nenhum dos arquivos enviados pôde ser lido como auditoria. "
                    "Confira se são Excels gerados pelo botão 'Baixar auditoria completa'."
                )
            else:
                st.success(
                    f"✅ {len(historico_df)} lançamentos de {len(arqs_hist)} arquivo(s) "
                    f"de histórico carregados."
                )
                for av in avisos:
                    st.warning(av)

    st.write("")

    col_data, col_arq = st.columns([1, 2])
    with col_data:
        from datetime import date, timedelta
        data_ate = st.date_input(
            "Data a auditar",
            value=date.today() - timedelta(days=1),
            format="DD/MM/YYYY",
            help="Padrão: ontem (modelo D-1). Mude para auditar outro período.",
        )
    with col_arq:
        arq = st.file_uploader(
            "Relatório da adquirente (.xlsx, .xls ou .csv)",
            type=["xlsx", "xls", "csv"],
            key="upload_relatorio_adq",
            help=(
                "Aceita 2 formatos:\n\n"
                "1) **Extrato CRU da GETNET** — o XLSX baixado direto do portal "
                "(Recebíveis > Completos > Detalhado). O sistema detecta e converte sozinho.\n\n"
                "2) **Relatório padronizado** — XLSX com colunas: data_venda, adquirente, "
                "modalidade, parcelas, valor_bruto, taxa_aplicada, valor_liquido (opcional)."
            ),
        )

    # Recupera histórico do session_state (carregado no expander acima)
    historico_df = locals().get("historico_df", pd.DataFrame())

    # v5.43: reaproveita os extratos de adquirente já enviados na conciliação.
    # Só o GetNet cru é auditável por venda; PagBank (recebimento) é ignorado aqui.
    relatorio_sessao, adq_ignorados = _relatorio_auditoria_da_sessao()
    usar_sessao = False
    if arq is None and not relatorio_sessao.empty:
        st.success(
            f"🟩 Encontrei **{len(relatorio_sessao)} venda(s) de GetNet** que você já subiu "
            "na conciliação. Não precisa subir de novo."
        )
        usar_sessao = st.checkbox(
            "Auditar os extratos que já enviei na conciliação",
            value=True,
            help="Usa os mesmos arquivos de adquirente da tela de conciliação. "
            "Se preferir auditar outro período, suba um arquivo ao lado.",
        )
        if adq_ignorados:
            st.caption(
                "Obs.: estes não entram na auditoria por venda (não trazem taxa por "
                "venda, ex.: PagBank é recebimento): " + ", ".join(adq_ignorados)
            )

    if arq is None and not usar_sessao and (historico_df is None or historico_df.empty):
        st.info(
            "🙋 Suba o relatório da adquirente para iniciar a auditoria "
            "(ou suba o extrato da adquirente na conciliação — ele aparece aqui automaticamente).\n\n"
            "**Formato 1 — Extrato CRU da GETNET (recomendado):**\n"
            "- Baixe direto do portal GETNET: *Recebíveis > Completos > Detalhado*\n"
            "- Suba o XLSX como veio, o sistema converte sozinho\n\n"
            "**Formato 2 — Relatório padronizado:**\n"
            "- Planilha Excel/CSV com colunas: `data_venda`, `adquirente`, `modalidade`, "
            "`parcelas`, `valor_bruto`, `taxa_aplicada`, `valor_liquido` (opcional)\n"
            "- Exemplo em `data/samples/relatorio_adquirente_exemplo.xlsx`"
        )
        return

    # Carrega relatório atual (se subiu) — detecta formato automaticamente
    if arq is not None:
        try:
            # v5.16: detecta se é extrato CRU da GETNET e converte automaticamente
            if eh_extrato_getnet_cru(arq):
                relatorio = carregar_extrato_getnet_cru(arq)
                # Resumo informativo do que foi lido do extrato GETNET
                resumo_getnet = resumir_extrato_getnet(relatorio)
                periodo_str = ""
                if resumo_getnet["data_min"] and resumo_getnet["data_max"]:
                    periodo_str = (
                        f" · Período: {resumo_getnet['data_min'].strftime('%d/%m/%Y')} → "
                        f"{resumo_getnet['data_max'].strftime('%d/%m/%Y')}"
                    )
                st.success(
                    f"🟧 **Extrato GETNET detectado e convertido automaticamente.** "
                    f"{resumo_getnet['qtd']} vendas · "
                    f"Bruto: {fmt_brl(resumo_getnet['bruto_total'])} · "
                    f"Líquido: {fmt_brl(resumo_getnet['liquido_total'])} · "
                    f"Taxa média real: {fmt_pct(resumo_getnet['taxa_media']*100)}"
                    f"{periodo_str}"
                )
                # Mostra taxa real por bandeira/modalidade num expander (pra você conferir
                # antes de cadastrar o contrato)
                if not resumo_getnet["por_modalidade"].empty:
                    with st.expander("📊 Taxa real cobrada por bandeira/modalidade (informativo)", expanded=False):
                        st.caption(
                            "Esta é a taxa que a GETNET **efetivamente cobrou** em cada bandeira/modalidade neste período. "
                            "Use como referência para cadastrar a taxa contratada no `taxas.xlsx`."
                        )
                        visu_mod = resumo_getnet["por_modalidade"].copy()
                        visu_mod["bruto"] = visu_mod["bruto"].apply(fmt_brl)
                        visu_mod["liquido"] = visu_mod["liquido"].apply(fmt_brl)
                        visu_mod["taxa_real"] = (visu_mod["taxa_real"] * 100).round(4).astype(str) + "%"
                        visu_mod.columns = [c.replace("_", " ").title() for c in visu_mod.columns]
                        st.dataframe(visu_mod, use_container_width=True, hide_index=True)
            else:
                # Formato padronizado clássico
                relatorio = carregar_relatorio_adquirente(arq)
        except Exception as e:
            st.error(f"❌ Erro ao ler o relatório: {e}")
            return
        if relatorio.empty:
            st.warning("⚠️ O relatório está vazio ou não tem linhas válidas.")
            relatorio = pd.DataFrame()
    elif usar_sessao and not relatorio_sessao.empty:
        relatorio = relatorio_sessao
        resumo_getnet = resumir_extrato_getnet(relatorio)
        periodo_str = ""
        if resumo_getnet["data_min"] and resumo_getnet["data_max"]:
            periodo_str = (
                f" · Período: {resumo_getnet['data_min'].strftime('%d/%m/%Y')} → "
                f"{resumo_getnet['data_max'].strftime('%d/%m/%Y')}"
            )
        st.success(
            f"🟧 **Auditando os extratos da conciliação.** "
            f"{resumo_getnet['qtd']} vendas · "
            f"Bruto: {fmt_brl(resumo_getnet['bruto_total'])} · "
            f"Líquido: {fmt_brl(resumo_getnet['liquido_total'])} · "
            f"Taxa média real: {fmt_pct(resumo_getnet['taxa_media']*100)}"
            f"{periodo_str}"
        )
    else:
        relatorio = pd.DataFrame()

    # Filtra pelo período escolhido (padrão: data exata = data_ate)
    if not relatorio.empty:
        col_p1, _ = st.columns(2)
        with col_p1:
            usar_filtro = st.checkbox("Filtrar por data específica", value=False,
                                      help="Marque para auditar apenas a data selecionada acima.")
        if usar_filtro:
            relatorio = relatorio[
                relatorio["data_venda"].dt.date == data_ate
            ]
            if relatorio.empty:
                st.warning(f"⚠️ Nenhuma transação encontrada para {data_ate.strftime('%d/%m/%Y')}.")
                return

    # v4.1: roda auditoria com histórico opcional
    res = auditar_taxas(
        relatorio,
        cadastro,
        historico=historico_df if not historico_df.empty else None,
    )
    k = res.kpis

    if not historico_df.empty:
        st.info(
            f"📈 Auditoria **consolidada**: {len(relatorio)} lançamento(s) atual(is) + "
            f"{len(historico_df)} do histórico = **{len(res.detalhado)} total**."
        )

    st.write("")
    section_title("INDICADORES")

    cards1 = [
        card_kpi("Volume Bruto de Vendas", fmt_brl(k["volume_bruto"]),
                 f"{fmt_int(k['qtd_total'])} transações"),
        card_kpi("Valor Líquido Recebido", fmt_brl(k["valor_liquido"]),
                 classe="destaque-verde"),
        card_kpi("Total de Taxas Pagas", fmt_brl(k["taxas_pagas"])),
        card_kpi("Taxa Média Efetiva", fmt_pct(k["taxa_media_efetiva"] * 100),
                 classe="destaque-amarelo"),
    ]
    render_cards(cards1)

    cards2 = [
        card_kpi("OK", fmt_int(k["qtd_ok"]),
                 "conforme contrato", classe="destaque-verde"),
        card_kpi("Arredondamento", fmt_int(k.get("qtd_arredondamento", 0)),
                 "centavo da adquirente"),
        card_kpi("Discrepâncias", fmt_int(k["qtd_divergencias"]),
                 "diferença real de taxa",
                 classe="destaque-vermelho" if k["qtd_divergencias"] > 0 else ""),
        card_kpi("Sem Contrato", fmt_int(k["qtd_sem_contrato"]),
                 "modalidade não cadastrada",
                 classe="destaque-amarelo" if k["qtd_sem_contrato"] > 0 else ""),
    ]
    render_cards(cards2)

    cards3 = [
        card_kpi("Você pagou a mais", fmt_brl(k.get("impacto_pagou_mais", 0.0)),
                 "só discrepâncias",
                 classe="destaque-vermelho" if k.get("impacto_pagou_mais", 0.0) > 0 else ""),
        card_kpi("Você pagou a menos", fmt_brl(abs(k.get("impacto_pagou_menos", 0.0))),
                 "só discrepâncias (a seu favor)"),
        card_kpi("Arredondamento (total)", fmt_brl(k.get("arredondamento_total", 0.0)),
                 "não é cobrança indevida"),
    ]
    render_cards(cards3)

    st.divider()

    # Tabela com abas: OK / arredondamento / discrepância / sem contrato / tudo
    tabs = st.tabs([
        f"✅ OK ({k['qtd_ok']})",
        f"🔧 Arredondamento ({k.get('qtd_arredondamento', 0)})",
        f"⚠️ Discrepância ({k['qtd_divergencias']})",
        f"❓ Sem contrato ({k['qtd_sem_contrato']})",
        "📋 Tudo",
    ])

    # v5.18: 'bandeira' incluída pra facilitar identificar a origem da divergência
    # (Visa, Mastercard, Elo, etc — vem do parser GETNET).
    cols_show = [
        "data_venda", "adquirente", "bandeira", "modalidade", "parcelas",
        "valor_bruto", "taxa_aplicada", "taxa_esperada",
        "diferenca_pp", "diferenca_rs", "status", "motivo",
    ]
    cols_show = [c for c in cols_show if c in res.detalhado.columns]

    def _formatar_visualizacao(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        out = df[cols_show].copy()
        if "taxa_aplicada" in out.columns:
            out["taxa_aplicada"] = (out["taxa_aplicada"] * 100).round(4).astype(str) + "%"
        if "taxa_esperada" in out.columns:
            out["taxa_esperada"] = out["taxa_esperada"].apply(
                lambda x: f"{float(x) * 100:.4f}%" if pd.notna(x) else "—"
            )
        if "diferenca_pp" in out.columns:
            out["diferenca_pp"] = out["diferenca_pp"].apply(
                lambda x: f"{x:+.4f} p.p." if pd.notna(x) else "—"
            )
        if "diferenca_rs" in out.columns:
            out["diferenca_rs"] = out["diferenca_rs"].apply(
                lambda x: fmt_brl(x) if pd.notna(x) else "—"
            )
        if "valor_bruto" in out.columns:
            out["valor_bruto"] = out["valor_bruto"].apply(fmt_brl)
        if "data_venda" in out.columns:
            out["data_venda"] = out["data_venda"].dt.strftime("%d/%m/%Y")
        out.columns = [c.replace("_", " ").title() for c in out.columns]
        return out

    with tabs[0]:
        ok = res.detalhado[res.detalhado["status"] == "OK"]
        if ok.empty:
            st.info("Nenhuma transação OK neste período.")
        else:
            st.success(f"✅ **{len(ok)} transações conforme contrato.**")
            st.dataframe(_formatar_visualizacao(ok), use_container_width=True, hide_index=True)

    with tabs[1]:
        arr = res.detalhado[res.detalhado["status"] == "Arredondamento"]
        if arr.empty:
            st.success("✅ Nenhum arredondamento a destacar.")
        else:
            st.info(
                f"🔧 **{len(arr)} lançamentos com arredondamento da adquirente** "
                "(diferença de taxa até 0,02 pp). Total "
                f"**{fmt_brl(k.get('arredondamento_total', 0.0))}** — não é cobrança indevida, "
                "é o centavo que a adquirente arredonda por transação."
            )
            st.dataframe(_formatar_visualizacao(arr), use_container_width=True, hide_index=True)

    with tabs[2]:
        if res.divergentes.empty:
            st.success("🎉 Nenhuma discrepância de taxa. Tudo dentro do contrato (fora arredondamento).")
        else:
            _mais = k.get("impacto_pagou_mais", 0.0)
            _menos = abs(k.get("impacto_pagou_menos", 0.0))
            st.warning(
                f"⚠️ **{len(res.divergentes)} discrepâncias de taxa** (acima de 0,02 pp). "
                f"Você pagou a mais **{fmt_brl(_mais)}** e a menos **{fmt_brl(_menos)}**."
            )
            visu = _formatar_visualizacao(res.divergentes)
            st.dataframe(visu, use_container_width=True, hide_index=True)

            # Download das discrepâncias
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                _formatar_visualizacao(res.divergentes).to_excel(
                    writer, sheet_name="Discrepancias", index=False
                )
                # Aba "Bruto" com dados originais (sem formatação) pra reanálise
                res.divergentes[cols_show].to_excel(
                    writer, sheet_name="Bruto", index=False
                )
            buf.seek(0)
            st.download_button(
                "⬇️ Baixar Excel com discrepâncias",
                data=buf.getvalue(),
                file_name=f"discrepancias_taxas_{data_ate.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

    with tabs[3]:
        sc = res.detalhado[res.detalhado["status"] == "Sem contrato"]
        if sc.empty:
            st.success("✅ Todas as transações têm taxa cadastrada.")
        else:
            st.warning(
                f"❓ **{len(sc)} transações sem contrato cadastrado.** "
                f"Cadastre a taxa correspondente em `🏦 Cadastro de Taxas` para que entrem na auditoria."
            )
            st.dataframe(_formatar_visualizacao(sc), use_container_width=True, hide_index=True)

    with tabs[4]:
        st.dataframe(_formatar_visualizacao(res.detalhado),
                     use_container_width=True, hide_index=True)

        # Resumo por adquirente (só discrepâncias)
        if not res.divergentes.empty:
            st.write("")
            st.markdown("**Discrepâncias por adquirente:**")
            por_adq = res.divergentes_por_adquirente()
            por_adq["impacto"] = por_adq["impacto"].apply(fmt_brl)
            por_adq.columns = ["Adquirente", "Quantidade", "Impacto (R$)"]
            st.dataframe(por_adq, use_container_width=True, hide_index=True)

    # v4.1: download da auditoria COMPLETA (pra usar como histórico no próximo dia)
    st.write("")
    st.divider()
    section_title("EXPORTAR AUDITORIA COMPLETA")
    st.caption(
        "💡 Baixe esse arquivo e guarde. Na próxima auditoria, suba-o em "
        "**'📁 Subir auditorias anteriores'** acima pra ver os números acumulados."
    )
    buf_completo = BytesIO()
    with pd.ExcelWriter(buf_completo, engine="openpyxl") as writer:
        # Dados brutos (com tipos preservados) — esse é o arquivo que o sistema relê
        res.detalhado.to_excel(writer, sheet_name="Auditoria", index=False)
        # Aba formatada pra leitura humana
        _formatar_visualizacao(res.detalhado).to_excel(
            writer, sheet_name="Visualização", index=False
        )
    buf_completo.seek(0)
    from datetime import date as _date
    nome_arquivo = f"auditoria_taxas_{_date.today().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        "⬇️ Baixar auditoria completa (.xlsx)",
        data=buf_completo.getvalue(),
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


# ============================================================
# Página: Conta 70 (v5.0 — módulo de controle provisório)
# ============================================================
def _c70_payload(uploaded):
    """file_uploader (1 ou vários) -> tupla de (bytes, nome). Empilha sem dedup.
    Vazio () quando nada foi subido. Mantém compatível com 1 arquivo."""
    if not uploaded:
        return ()
    if isinstance(uploaded, list):
        return tuple((f.getvalue(), f.name) for f in uploaded)
    return ((uploaded.getvalue(), uploaded.name),)


@st.cache_data(show_spinner="Processando a Conta 70 (uma vez)…")
def _c70_processar(capa_payload, sk_payload):
    """Lê Capa + Sankhya (um ou VÁRIOS arquivos, empilhados linha-a-linha sem
    dedup) e atrela. Cache por conjunto de arquivos. Com 1 arquivo, idêntico a antes."""
    import io as _io
    from src.conta70.casamento import carregar_movimento, atrelar

    def _n(b, nome):
        x = _io.BytesIO(b)
        x.name = nome
        return x

    capa = pd.concat([carregar_movimento(_n(b, nm)) for b, nm in capa_payload], ignore_index=True)
    sk = pd.concat([carregar_movimento(_n(b, nm)) for b, nm in sk_payload], ignore_index=True)
    _m = pd.to_numeric(capa["numero"], errors="coerce").max()
    ult = int(_m) if pd.notna(_m) else 0
    res = atrelar(sk, capa, ultimo_numero=ult)
    # acumulado da Capa inteira (conta 70): receita, despesa e diferença
    _rd = capa["receita_despesa"].astype(str).str.upper()
    acum_rec = float(capa[_rd.str.contains("RECEITA", na=False)]["valor"].abs().sum())
    acum_desp = float(capa[_rd.str.contains("DESPESA", na=False)]["valor"].abs().sum())
    return res.detalhado, dict(res.kpis), int(res.proximo_numero), ult, acum_rec, acum_desp


@st.cache_data(show_spinner=False)
def _c70_faturamento(fat_payload):
    import io as _io
    from src.conta70.casamento import carregar_faturamento

    def _n(b, nome):
        x = _io.BytesIO(b)
        x.name = nome
        return x

    return pd.concat([carregar_faturamento(_n(b, nm)) for b, nm in fat_payload], ignore_index=True)


@st.cache_data(show_spinner="Montando o Mapa de recebimentos…")
def _mapa_c70_construir(capa_payload, fat_payload):
    """Constrói o Mapa (um ou VÁRIOS arquivos, empilhados sem dedup).
    fat_payload pode ser vazio (sem notas)."""
    import io as _io
    from src.conta70.casamento import carregar_faturamento
    from src.conta70.mapa_recebimentos import carregar_capa_bruta, construir_mapa

    def _n(b, nome):
        x = _io.BytesIO(b); x.name = nome; return x

    capa = pd.concat([carregar_capa_bruta(_n(b, nm)) for b, nm in capa_payload], ignore_index=True)
    fat = None
    if fat_payload:
        try:
            fat = pd.concat([carregar_faturamento(_n(b, nm)) for b, nm in fat_payload], ignore_index=True)
        except Exception:
            fat = None
    m, resumo = construir_mapa(capa, fat)
    return m, resumo


def _render_conta70_mapa_recebimentos():
    """v5.13 — Mapa de Recebimentos (Conta 70). Visão aprovada na prévia:
    três blocos com origem, status em cores da identidade LLE, alerta de aging,
    reconciliação e download do Excel. Reaproveita os arquivos já subidos na aba
    'Atrelamento e Numeração' — não pede upload de novo."""
    from src.conta70.mapa_recebimentos import CORES_STATUS, exportar_mapa_excel

    section_title("MAPA DE RECEBIMENTOS — CONTA 70")
    st.markdown(
        "Cruza a **Capa da Conta 70** com as **notas emitidas não baixadas**. A NF só é sugerida "
        "quando o **valor** bate (exato) ou por **somatório** (a conferir) — nunca só pelo CNPJ."
    )

    # v5.19: com o submenu em pílulas, os uploaders (na aba Atrelamento) não
    # renderizam junto do Mapa — então lemos os arquivos de uma chave persistente
    # que a aba de Atrelamento preenche a cada render.
    _saved = st.session_state.get("_c70_saved", {})
    _capa_payload = _saved.get("capa", ())
    _fat_payload = _saved.get("fat", ())
    if not _capa_payload:
        st.info(
            "Suba a **Capa da Conta 70** na aba **Atrelamento e Numeração** — os mesmos arquivos "
            "valem aqui. As **notas emitidas** (opcional) habilitam a sugestão de NF por valor."
        )
        return

    try:
        m, R = _mapa_c70_construir(_capa_payload, _fat_payload)
    except Exception as e:
        st.error(f"Não consegui montar o Mapa: {e}")
        return

    # ---- cards de resumo ----
    _ate = ""
    if R.get("capa_ate"):
        try:
            _ate = pd.to_datetime(R["capa_ate"]).strftime("%d/%m/%Y")
        except Exception:
            _ate = ""
    # v5.15: cards enxutos (4) — remove os já resolvidos (Baixadas/Saídas) e a
    # duplicidade "Em aberto (não atrelado)"; mantém a terminologia técnica.
    render_cards([
        card_kpi("Saldo em aberto na Conta 70", fmt_brl(R["saldo_desp_aberto"]),
                 f"{fmt_int(R['qtd_desp_aberto'])} recebimentos sem baixa" + (f" · Capa até {_ate}" if _ate else ""),
                 classe="destaque-vermelho"),
        card_kpi("NF sugerida (valor exato)", fmt_int(R["exato"]),
                 "valor bate — é só dar baixa" + (f" · soma a conferir: {fmt_int(R['soma_conferir'])}" if R["soma_conferir"] else ""),
                 classe="destaque-verde" if R["exato"] > 0 else ""),
        card_kpi("Pendentes de baixa", fmt_int(R["pendente"]), "identificados, faltam baixar",
                 classe="destaque-amarelo" if R["pendente"] > 0 else ""),
        card_kpi("Sem identificação", fmt_int(R["sem_id"]), "nem CPF/CNPJ nem nome",
                 classe="destaque-vermelho" if R["sem_id"] > 0 else ""),
    ])

    if not _fat_payload:
        st.caption("💡 Suba as **notas emitidas (com CNPJ)** na aba ao lado para habilitar a sugestão de NF por valor.")

    # ---- detalhe expansível da(s) NF sugerida(s): data, valor e qual NF baixar ----
    _sug = m[m["nf_baixar"].astype(str).str.strip() != ""].copy()
    if len(_sug) > 0:
        _tm = _sug["tipo_match"].astype(str) if "tipo_match" in _sug.columns else pd.Series([""] * len(_sug), index=_sug.index)
        _n_exato = int((_tm.str.lower() == "exato").sum())
        with st.expander(f"🔎 Ver as NF sugeridas — {fmt_int(len(_sug))} (valor exato: {fmt_int(_n_exato)})"):
            _det = pd.DataFrame({
                "Data": pd.to_datetime(_sug["data"], errors="coerce").dt.strftime("%d/%m/%Y"),
                "Banco": _sug["banco"].astype(str),
                "Valor": _sug["valor"].map(lambda v: fmt_brl(abs(float(v)))),
                "CPF / CNPJ": _sug["identificacao"].astype(str),
                "Parceiro (nota)": (_sug["parceiro_nf"].astype(str) if "parceiro_nf" in _sug.columns else ""),
                "Tipo": _tm.str.lower().map({"exato": "Valor exato", "soma": "Soma (conferir)"}).fillna(_tm),
                "NF a baixar": _sug["nf_baixar"].astype(str),
            })
            st.dataframe(_det, hide_index=True, width="stretch")
            st.caption("As de **valor exato** podem ser baixadas contra o recebimento; as de **soma (conferir)** são só indicação — a decisão é sua.")

    # ---- alerta de aging (só em aberto) ----
    ab = m[m["aberto"]].copy()
    ab["dias"] = pd.to_numeric(ab["dias"], errors="coerce")
    # v5.15: 4 faixas; 90+ em destaque (prioridade — o valor crítico da conta).
    faixas = [
        ("15–30 dias", 15, 30, "destaque-amarelo"),
        ("31–60 dias", 30, 60, "destaque-amarelo"),
        ("61–90 dias", 60, 90, "destaque-amarelo"),
        ("90+ dias (prioridade)", 90, 10**9, "destaque-vermelho"),
    ]
    cards_al = []
    for rot, lo, hi, classe in faixas:
        sub = ab[(ab["dias"] > lo) & (ab["dias"] <= hi)]
        cards_al.append(card_kpi(rot, fmt_int(len(sub)), fmt_brl(float(sub["valor"].sum())), classe=classe))
    section_title("ALERTA — PARADOS EM ABERTO (aging)")
    render_cards(cards_al)

    # ---- tabela ----
    st.markdown("##### Mapa detalhado")
    # v5.17: seletor "Mostrar" amarelo — mesmo padrão das abas (marcador + :has() + irmão ~), alvo amplo
    _S = '[data-testid="stElementContainer"]:has(.c70selmark)'
    st.markdown(
        "<style>"
        f'{_S} ~ div [data-baseweb="select"]>div{{background-color:#FAC318!important;border-color:#d9a800!important}}'
        f'{_S} ~ div [data-baseweb="select"]>div>div{{background:transparent!important}}'
        f'{_S} ~ div [data-baseweb="select"] input{{background:transparent!important}}'
        f'{_S} ~ div [data-baseweb="select"] *{{color:#041747!important}}'
        f'{_S} ~ div [data-baseweb="select"] svg{{fill:#041747!important}}'
        "</style>",
        unsafe_allow_html=True,
    )
    st.markdown('<span class="c70selmark"></span>', unsafe_allow_html=True)
    filtro = st.selectbox(
        "Mostrar", ["Em aberto (pendências)", "Só com sugestão de NF", "Alerta (15+ dias)", "Tudo"],
        key="c70_mapa_filtro",
    )
    if filtro == "Em aberto (pendências)":
        mf = m[m["aberto"]].copy()
    elif filtro == "Só com sugestão de NF":
        mf = m[m["nf_baixar"].astype(str) != ""].copy()
    elif filtro == "Alerta (15+ dias)":
        mf = m[m["aberto"] & (pd.to_numeric(m["dias"], errors="coerce") > 15)].copy()
    else:
        mf = m.copy()

    st.caption(f"{fmt_int(len(mf))} linhas · valor negativo = despesa (saída); positivo = receita (entrada).")

    def _dstr(d):
        try:
            return pd.to_datetime(d).strftime("%d/%m/%Y") if pd.notna(d) else ""
        except Exception:
            return ""

    disp = pd.DataFrame({
        "Nº Capa 70": mf["num_txt"].values,
        "Data": mf["data"].map(_dstr).values,
        "Banco": mf["banco"].values,
        "Identificação (CPF/CNPJ / origem)": mf["identificacao"].values,
        "Valor": [fmt_brl(v) for v in pd.to_numeric(mf["valor"], errors="coerce").fillna(0)],
        "R/D": mf["rd"].values,
        "NF a baixar (sugestão)": mf["nf_baixar"].values,
        "Parceiro da NF": mf["parceiro_nf"].values,
        "Tipo de match": mf["tipo_match"].values,
        "Status": mf["status"].values,
        "Parceiro efetivo": mf["parceiro_efetivo"].values,
        "Justificativa do vínculo": mf["justificativa"].values,
    })
    _keys = mf["status_key"].tolist()

    def _style_status(col):
        out = []
        for k in _keys:
            if k in CORES_STATUS:
                bg, fg = CORES_STATUS[k]
                out.append(f"background-color:#{bg};color:#{fg};font-weight:600")
            else:
                out.append("")
        return out

    try:
        styler = disp.style.apply(_style_status, subset=["Status"])
        st.dataframe(styler, use_container_width=True, hide_index=True, height=520)
    except Exception:
        st.dataframe(disp, use_container_width=True, hide_index=True, height=520)

    # ---- download do Excel (mesmo layout aprovado) ----
    st.divider()
    cA, cB = st.columns([1, 3])
    if cA.button("📥 Gerar Mapa em Excel", key="c70_mapa_gerar"):
        try:
            xls = exportar_mapa_excel(m, R)
            st.session_state["c70_mapa_xls"] = xls
        except Exception as e:
            st.error(f"Não consegui gerar o Excel: {e}")
    if st.session_state.get("c70_mapa_xls"):
        from datetime import date as _date
        cB.download_button(
            "⬇️ Baixar Mapa_Recebimentos_Conta70.xlsx",
            data=st.session_state["c70_mapa_xls"],
            file_name=f"Mapa_Recebimentos_Conta70_{_date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="c70_mapa_download",
        )


def _render_conta70_casamento_numeracao():
    """v5.7 — Atrelamento, numeração e esteira da Conta 70.

    Três uploads: Capa consolidada (só leitura), movimentação do Sankhya e notas
    emitidas não baixadas (com CNPJ). A Capa original nunca é alterada — a "capa
    atualizada" sai como arquivo novo, completo/acumulado.
    """
    import io as _io
    from datetime import date as _date
    from src.conta70.casamento import diagnosticar, sugerir_atrelamentos_cnpj, gerar_capa_acumulada

    def _money(v):
        try:
            return ("R$ " + f"{float(v):,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "R$ 0,00"

    def _brl(v):
        """Formato brasileiro com sinal: 1.750,00 / -1.028,76 (sem R$)."""
        try:
            v = float(v)
        except Exception:
            return ""
        s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return ("-" if v < 0 else "") + s

    section_title("ATRELAMENTO E NUMERAÇÃO — CONTA 70")
    st.markdown(
        "Suba a **Capa**, a **movimentação da Conta 70 do Sankhya** e (opcional) as "
        "**notas com CNPJ**. O app atrela, numera em sequência e organiza o que está aberto na esteira."
    )
    with st.expander("Como funciona / como conferir com o Sankhya"):
        st.markdown(
            "O app lê a **Capa** (consolidada, somente leitura) e a **movimentação da Conta 70 do "
            "Sankhya**. Ele atrela cada lançamento pela **identidade do histórico** (CNPJ/CPF/dados "
            "bancários), numera em sequência e organiza o que está aberto numa **esteira** com "
            "diagnóstico e ação. Quando você sobe as **notas emitidas não baixadas** (com CNPJ), o app "
            "sugere atrelamentos onde o CNPJ da nota bate com o recebimento.\n\n"
            "**Para conferir com o Sankhya:** cada número (atrelamento) cobre os **dois lados** da "
            "operação — a receita e a despesa do mesmo item recebem o **mesmo número**. A capa "
            "atualizada sai completa/acumulada, com o sinal original (despesa negativa) e o número "
            "único bancário, para bater linha a linha com o Sankhya. **Sua capa original nunca é alterada.**"
        )

    c1, c2, c3 = st.columns(3)
    up_capa = c1.file_uploader("📄 Capa da Conta 70 (consolidada · só leitura)", type=["xlsx", "xls", "xlsm", "csv"], accept_multiple_files=True, key="c70_capa")
    up_sk = c2.file_uploader("📄 Movimentação Conta 70 (Sankhya)", type=["xlsx", "xls", "xlsm", "csv"], accept_multiple_files=True, key="c70_sk")
    up_fat = c3.file_uploader("📄 Notas emitidas não baixadas (com CNPJ)", type=["xlsx", "xls", "xlsm", "csv"], accept_multiple_files=True, key="c70_fat")
    # v5.22: guarda os bytes numa chave persistente SÓ quando o uploader tem
    # arquivo — nunca sobrescreve com vazio. Assim trocar de aba NÃO perde os
    # arquivos. Todo o processamento (e o Mapa) lê destas chaves, não do widget.
    if up_capa:
        st.session_state["_c70_bytes_capa"] = _c70_payload(up_capa)
    if up_sk:
        st.session_state["_c70_bytes_sk"] = _c70_payload(up_sk)
    if up_fat:
        st.session_state["_c70_bytes_fat"] = _c70_payload(up_fat)
    capa_payload = st.session_state.get("_c70_bytes_capa", ())
    sk_payload = st.session_state.get("_c70_bytes_sk", ())
    fat_payload = st.session_state.get("_c70_bytes_fat", ())
    # compat com o Mapa (que lê _c70_saved)
    st.session_state["_c70_saved"] = {"capa": capa_payload, "sk": sk_payload, "fat": fat_payload}
    st.caption("Pode subir **mais de um arquivo** por campo — as linhas são **empilhadas** (sem remover repetições). Só junte relatórios da **mesma tela** e de **períodos que não se sobreponham**, senão os valores contam em dobro.")

    if not capa_payload or not sk_payload:
        st.caption("Suba pelo menos a Capa e a movimentação do Sankhya. As notas emitidas são opcionais (habilitam os atrelamentos sugeridos por CNPJ).")
        return

    # ---- parte pesada em cache: instantâneo nos cliques seguintes ----
    try:
        d, k, prox, ultimo, acum_rec, acum_desp = _c70_processar(capa_payload, sk_payload)
        d = d.copy()
    except Exception as e:
        st.error(f"Não consegui ler um dos arquivos: {e}")
        return

    pend = d[d["situacao"].isin(["Aguardando baixa", "A conferir"])]
    esteira = diagnosticar(pend)
    n_ident = k["ja_identificado"] + k["herdado"]
    acum_dif = acum_rec - acum_desp

    # publica um resumo pro Dashboard (só leitura; nunca chuta)
    try:
        _dias = pd.to_numeric(esteira["dias"], errors="coerce") if "dias" in esteira.columns else pd.Series(dtype=float)
        _aging = {
            "ate30": int((_dias <= 30).sum()),
            "d31_60": int(((_dias > 30) & (_dias <= 60)).sum()),
            "d61_90": int(((_dias > 60) & (_dias <= 90)).sum()),
            "mais90": int((_dias > 90).sum()),
        }
        _maior_val = _maior_dias = None
        if not esteira.empty:
            _im = esteira["valor"].abs().idxmax()
            _maior_val = float(esteira.loc[_im, "valor"])
            _maior_dias = int(pd.to_numeric(pd.Series([esteira.loc[_im, "dias"]]), errors="coerce").fillna(0).iloc[0])
        st.session_state["c70_dashboard"] = {
            "parado": float(acum_dif),
            "itens": int(len(pend)),
            "aging": _aging,
            "maior_val": _maior_val,
            "maior_dias": _maior_dias,
        }
    except Exception:
        pass

    # cards do acumulado da Capa inteira (conta 70) — em cima
    render_cards([
        card_kpi("Receita acumulada", _money(acum_rec), "toda a Capa · Conta 70", classe="destaque-verde"),
        card_kpi("Despesa acumulada", _money(acum_desp), "toda a Capa · Conta 70", classe="destaque-vermelho"),
        card_kpi("Diferença acumulada", _money(acum_dif), "receita − despesa"),
    ])
    # cards de progresso do período — embaixo
    render_cards([
        card_kpi("Identificado", fmt_int(n_ident), "já com número na capa", classe="destaque-verde" if n_ident > 0 else ""),
        card_kpi("Atrelado agora", fmt_int(k["numerado_agora"]), "números novos", classe="destaque-amarelo" if k["numerado_agora"] > 0 else ""),
        card_kpi("Na esteira", fmt_int(len(pend)), "abertos a resolver"),
        card_kpi("Último número usado", fmt_int(ultimo), "na capa"),
    ])

    # ---- SELEÇÃO: marca à vontade e confirma no fim (dentro de um form,
    #      então marcar NÃO recarrega a tela — só o Confirmar processa) ----
    st.markdown("##### 🔗 Selecionar atrelamentos")
    st.caption("Marque à vontade nos **sugeridos** e/ou na **esteira** — nada é processado até você clicar "
               "em **Confirmar selecionados** no fim. Confirmar = “esse recebimento é o pagamento dessa nota”.")

    # prepara SUGERIDOS (fora do form)
    vis = None
    sug = None
    if not fat_payload:
        st.caption("💡 Suba as notas emitidas (com CNPJ) para o app sugerir atrelamentos pelo CNPJ do histórico.")
    else:
        fat = None
        try:
            fat = _c70_faturamento(fat_payload)
        except Exception as e:
            st.warning(f"Não consegui ler o faturamento: {e}")
        if fat is not None:
            if int((fat["cnpj"] != "").sum()) == 0:
                st.info("O faturamento veio **sem CNPJ preenchido**. Exporte com a coluna CNPJ/CPF para habilitar as sugestões.")
            else:
                _sug = sugerir_atrelamentos_cnpj(esteira, fat)
                if _sug.empty:
                    st.caption("Nenhum CNPJ do faturamento bateu com as entradas abertas.")
                else:
                    sug = _sug.drop_duplicates(subset=["idx", "nota", "valor_recebido"]).reset_index(drop=True)
                    _rds = sug["receita_despesa"].astype(str).str.upper()
                    _sinal = _rds.map(lambda x: -1 if "DESPESA" in x else 1)
                    vis = pd.DataFrame({
                        "Confirmar": False,
                        "R/D": _rds.map(lambda x: "Despesa" if "DESPESA" in x else "Receita").values,
                        "CNPJ": sug["cnpj"].values,
                        "Cliente": sug["nome"].astype(str).str.slice(0, 28).values,
                        "Nota": sug["nota"].astype(str).values,
                        "Recebido": [_brl(x * sg) for x, sg in zip(sug["valor_recebido"].astype(float), _sinal)],
                        "Valor da nota": [_brl(x) for x in pd.to_numeric(sug["valor_nota"], errors="coerce").fillna(0)],
                        "Confere": sug["valor_fecha"].map(lambda b: "✅ bate" if b else "⚠️ conferir valor").values,
                    })

    # prepara ESTEIRA + filtros (fora do form, pra os filtros reagirem na hora)
    est = esteira.copy()
    vis2 = None
    if est.empty:
        st.caption("Esteira: nenhuma pendência aberta no momento. 🎉")
    else:
        contagem = est["diagnostico"].value_counts()
        chips = " &nbsp;·&nbsp; ".join(f"<b>{nome}:</b> {qtd}" for nome, qtd in contagem.items())
        st.markdown("**Esteira — pendências abertas.** Valor negativo = despesa (saída); positivo = receita (entrada).")
        st.markdown(f"<div style='color:#9fb3d6;font-size:13px;margin:2px 0 8px'>{chips}</div>", unsafe_allow_html=True)
        with st.container(key="c70filtros"):
            st.markdown(
                "<style>"
                '.st-key-c70filtros [data-baseweb="select"]>div{border-left:4px solid #FAC318!important}'
                '.st-key-c70filtros [data-baseweb="input"]{border-left:4px solid #FAC318!important}'
                '.st-key-c70filtros [data-baseweb="base-input"]{border-left:4px solid #FAC318!important}'
                "</style>",
                unsafe_allow_html=True,
            )
            fc1, fc2, fc3, fc4, fc5 = st.columns([1, 1.2, 1, 1.3, 1.6])
            f_prio = fc1.selectbox("Prioridade", ["todas", "Alta", "Média", "Baixa"], key="c70_fprio")
            f_diag = fc2.selectbox("Tipo de pendência", ["todos"] + sorted(est["diagnostico"].dropna().unique().tolist()), key="c70_fdiag")
            f_banco = fc3.selectbox("Banco", ["todos"] + sorted(est["banco"].dropna().unique().tolist()), key="c70_fbanco")
            f_col = fc4.selectbox("Filtrar por coluna", ["Todas as colunas", "CNPJ / Histórico", "Data", "Banco", "R/D", "Valor", "Diagnóstico"], key="c70_fcol")
            busca = fc5.text_input("Buscar", key="c70_busca", placeholder="digite e tecle Enter")
        view = est
        if f_prio != "todas":
            view = view[view["prioridade"] == f_prio]
        if f_diag != "todos":
            view = view[view["diagnostico"] == f_diag]
        if f_banco != "todos":
            view = view[view["banco"] == f_banco]
        if busca.strip():
            termo = busca.strip().lower()
            dig = re.sub(r"\D", "", busca)  # só dígitos (para CNPJ com/sem pontuação)
            hist_l = view["historico"].astype(str).str.lower()
            hist_dig = view["historico"].astype(str).str.replace(r"\D", "", regex=True)
            ident_dig = view["identidade"].astype(str).str.replace(r"\D", "", regex=True)
            if f_col == "CNPJ / Histórico":
                m = hist_l.str.contains(termo, na=False)
                if dig:
                    m = m | hist_dig.str.contains(dig, na=False) | ident_dig.str.contains(dig, na=False)
                view = view[m]
            elif f_col == "Data":
                dstr = pd.to_datetime(view["data"], errors="coerce").dt.strftime("%d/%m/%Y")
                view = view[dstr.str.contains(termo, na=False)]
            elif f_col == "Banco":
                view = view[view["banco"].astype(str).str.lower().str.contains(termo, na=False)]
            elif f_col == "R/D":
                view = view[view["receita_despesa"].astype(str).str.lower().str.contains(termo, na=False)]
            elif f_col == "Valor":
                view = view[view["valor"].abs().round(2).astype(str).str.contains(termo, na=False)]
            elif f_col == "Diagnóstico":
                view = view[view["diagnostico"].astype(str).str.lower().str.contains(termo, na=False)]
            else:  # Todas as colunas
                m = (hist_l.str.contains(termo, na=False)
                     | view["valor"].abs().round(2).astype(str).str.contains(termo, na=False)
                     | view["banco"].astype(str).str.lower().str.contains(termo, na=False)
                     | view["diagnostico"].astype(str).str.lower().str.contains(termo, na=False))
                if dig:
                    m = m | hist_dig.str.contains(dig, na=False) | ident_dig.str.contains(dig, na=False)
                view = view[m]
        st.caption(f"{len(view)} de {len(est)} pendentes")
        _rdv = view["receita_despesa"].astype(str).str.upper()
        valor_sinal = view["valor"].abs() * _rdv.map(lambda x: -1 if "DESPESA" in x else 1)
        vis2 = pd.DataFrame({
            "Atrelar": False,
            "Data": pd.to_datetime(view["data"], errors="coerce"),
            "Banco": view["banco"].values,
            "R/D": _rdv.map(lambda x: "Despesa" if "DESPESA" in x else "Receita").values,
            "Histórico": view["historico"].astype(str).str.slice(0, 44).values,
            "Valor": [_brl(v) for v in valor_sinal],
            "Dias": view["dias"].values,
            "Diagnóstico": view["diagnostico"].values,
            "Ação": view["acao"].values,
        }, index=view.index)

    # FORM: marcar aqui NÃO recarrega; só o Confirmar processa
    ed = None
    ed2 = None
    with st.form("c70_selecao", border=True):
        if vis is not None:
            st.markdown("**Atrelamentos sugeridos (CNPJ bateu com o histórico):**")
            ed = st.data_editor(
                vis, hide_index=True, use_container_width=True, key="c70_sug_ed",
                column_config={
                    "Confirmar": st.column_config.CheckboxColumn("Confirmar"),
                },
                disabled=["R/D", "CNPJ", "Cliente", "Nota", "Recebido", "Valor da nota", "Confere"],
            )
        if vis2 is not None:
            st.markdown("**Esteira — marque para atrelar manualmente:**")
            ed2 = st.data_editor(
                vis2, hide_index=True, use_container_width=True, key="c70_est_ed",
                column_config={
                    "Atrelar": st.column_config.CheckboxColumn("Atrelar"),
                    "Data": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
                    "Dias": st.column_config.NumberColumn("Dias", format="%d"),
                },
                disabled=[c for c in vis2.columns if c != "Atrelar"],
            )
        submitted = st.form_submit_button("✅ Confirmar selecionados", type="primary")

    if submitted:
        # coleta os índices marcados (sugeridos + esteira)
        selecionados = []
        if ed is not None and sug is not None:
            for pos, marc in enumerate(ed["Confirmar"].tolist()):
                if marc:
                    selecionados.append(int(sug.iloc[pos]["idx"]))
        if ed2 is not None and vis2 is not None:
            for idx, marc in zip(vis2.index.tolist(), ed2["Atrelar"].tolist()):
                if marc:
                    selecionados.append(idx)
        selecionados = list(dict.fromkeys(selecionados))  # únicos, mantém ordem

        # UM número por operação: mesma identidade + mesmo valor absoluto
        # (a receita e a despesa do mesmo item recebem o MESMO número)
        conf = {}
        conf_nota = {}
        grupos = {}
        s = prox
        sug_notas = {}
        if sug is not None:
            for _p in range(len(sug)):
                sug_notas[int(sug.iloc[_p]["idx"])] = str(sug.iloc[_p].get("nota", "")).strip()
        for idx in selecionados:
            if idx not in d.index:
                continue
            ident = str(d.at[idx, "identidade"])
            valor_abs = round(abs(float(d.at[idx, "valor"])), 2)
            chave = (ident, valor_abs)
            if chave not in grupos:
                grupos[chave] = s
                s += 1
            conf[idx] = grupos[chave]
            conf_nota[idx] = sug_notas.get(idx)  # nota se veio dos sugeridos; None se veio da esteira

        st.session_state["c70_confirmados_num"] = conf
        st.session_state["c70_confirmados_nota"] = conf_nota
        st.session_state.pop("c70_capa_bytes", None)
        n_ops = len(grupos)
        if conf:
            faixa = f"número {prox}" if n_ops == 1 else f"números {prox} a {s - 1}"
            st.success(f"{len(conf)} linha(s) confirmada(s) em {n_ops} operação(ões) — {faixa}. "
                       "Receita e despesa da mesma operação recebem o mesmo número. Agora gere a capa abaixo.")
        else:
            st.info("Nada marcado — nenhum atrelamento confirmado.")

    confirmados_persist = st.session_state.get("c70_confirmados_num", {})

    # ---- Gerar capa atualizada — INDEPENDENTE da seleção ----
    st.markdown("##### ⬇️ Gerar capa atualizada")
    if confirmados_persist:
        gc1, gc2 = st.columns([3, 1])
        gc1.caption(
            f"Gera a **capa completa e acumulada** com os números **automáticos** + os "
            f"**{len(confirmados_persist)} que você confirmou**."
        )
        if gc2.button("Limpar confirmados", key="c70_limpar"):
            st.session_state.pop("c70_confirmados_num", None)
            st.session_state.pop("c70_capa_bytes", None)
            st.rerun()
    else:
        st.caption(
            "Gera a **capa completa e acumulada** com os números **automáticos** já aplicados. "
            "Funciona mesmo sem confirmar nada (sai só com os automáticos)."
        )
    if st.button("Gerar capa atualizada", key="c70_gerar", type="primary"):
        conf_nota = st.session_state.get("c70_confirmados_nota", {})
        for idx, num in confirmados_persist.items():
            if idx in d.index:
                d.at[idx, "numero_final"] = num
                d.at[idx, "situacao"] = "Atrelado (confirmado)"
        # coluna I — o que fazer no Sankhya, por número
        acoes = {}
        _novos = pd.to_numeric(d["numero_final"], errors="coerce") > ultimo
        for idx in d.index[_novos]:
            try:
                num = int(d.at[idx, "numero_final"])
            except Exception:
                continue
            sit = str(d.at[idx, "situacao"])
            if sit == "Atrelado (confirmado)":
                nota = (conf_nota or {}).get(idx)
                if nota and str(nota).lower() not in ("", "nan", "none"):
                    acoes[num] = f"Baixar NF {nota} no Sankhya"
                else:
                    acoes[num] = "Atrelado manualmente — conferir baixa"
            elif sit == "Numerado agora":
                acoes.setdefault(num, "Baixa já lançada no Sankhya")
        try:
            with st.spinner("Gerando a capa acumulada… isso leva alguns segundos (arquivo grande)."):
                capa_out, preenchidos, n_novos = gerar_capa_acumulada(
                    _io.BytesIO(capa_payload[0][0]), d, ultimo, acoes=acoes,
                )

                # normaliza a coluna de data (mistura datetime + série do Excel)
                _cdt = next((c for c in capa_out.columns if str(c).strip().lower() in ("dt. lançamento", "dt. lancamento", "data")), None)
                if _cdt is not None:
                    def _fix_data(v):
                        try:
                            if hasattr(v, "year"):
                                return pd.Timestamp(v)
                            n = pd.to_numeric(v, errors="coerce")
                            if pd.notna(n) and 10000 < float(n) < 90000:  # série do Excel
                                return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(n))
                            return pd.to_datetime(v, errors="coerce", dayfirst=True)
                        except Exception:
                            return pd.NaT
                    capa_out[_cdt] = capa_out[_cdt].map(_fix_data)

                _cval = next((c for c in capa_out.columns if str(c).strip().lower() in ("vlr. lançamento", "vlr. lancamento", "valor", "vlr lancamento")), None)

                buf = _io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as w:
                    capa_out.to_excel(w, sheet_name="Capa Conta 70", index=False)
                    from openpyxl.utils import get_column_letter as _gcl
                    ws = w.sheets["Capa Conta 70"]
                    cols = list(capa_out.columns)
                    nrows = len(capa_out)
                    if _cval is not None:
                        _lv = _gcl(cols.index(_cval) + 1)
                        for row in range(2, nrows + 2):
                            ws[f"{_lv}{row}"].number_format = "#,##0.00"
                    if _cdt is not None:
                        _ld = _gcl(cols.index(_cdt) + 1)
                        for row in range(2, nrows + 2):
                            ws[f"{_ld}{row}"].number_format = "DD/MM/YYYY"
            st.session_state["c70_capa_bytes"] = buf.getvalue()
            st.session_state["c70_capa_info"] = (len(capa_out), preenchidos, n_novos - preenchidos)
        except Exception as e:
            st.error(f"Não consegui montar a capa: {e}")

    if st.session_state.get("c70_capa_bytes"):
        linhas, preench, nao_aloc = st.session_state.get("c70_capa_info", (0, 0, 0))
        st.success(
            f"Capa pronta: **{linhas:,} linhas** (acumulada, sinal original), **{preench}** número(s) novo(s) preenchido(s)."
            .replace(",", ".")
            + (f" {nao_aloc} sem linha única na Capa (ficam na esteira, sem chute)." if nao_aloc > 0 else "")
        )
        st.download_button(
            "⬇️ Baixar capa atualizada (.xlsx)",
            data=st.session_state["c70_capa_bytes"],
            file_name=f"capa_conta70_atualizada_{_date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def pagina_conta70():
    # v5.13: submenu — "Atrelamento e Numeração" (fluxo original, intacto) +
    # "Mapa de recebimentos" (visão aprovada). Aditivo: nada foi removido; o
    # fluxo existente passou a viver dentro da primeira aba.
    section_title("CONTA 70 — CONTROLE PROVISÓRIO")

    st.markdown(
        "A **Conta 70** é uma conta contábil provisória onde ficam os recebimentos que ainda **não "
        "foram identificados** (não se sabe de qual NF / cliente / pedido são). Aqui você **atrela** "
        "cada um à sua origem e dá um **número sequencial**, atualizando a capa da conta."
    )

    # v5.19: submenu como PÍLULAS próprias, ISOLADAS num container para o CSS não
    # vazar pros outros botões (ex.: "Gerar capa atualizada"). Ativa amarela.
    if "c70_sub" not in st.session_state:
        st.session_state["c70_sub"] = "atrel"
    _ativo = 1 if st.session_state["c70_sub"] == "atrel" else 2
    with st.container():
        st.markdown('<span class="c70pillmark"></span>', unsafe_allow_html=True)
        _P = '[data-testid="stElementContainer"]:has(.c70pillmark)'
        _hb = f'{_P} ~ div [data-testid="stHorizontalBlock"]'
        st.markdown(
            "<style>"
            f'{_hb} .stButton button{{border-radius:22px!important;border:1px solid #24406f!important;'
            f'background:transparent!important;padding:6px 20px!important;box-shadow:none!important}}'
            f'{_hb} .stButton button, {_hb} .stButton button *{{font-weight:700!important;color:#cdd9f2!important}}'
            f'{_hb} .stButton button:hover{{border-color:#FAC318!important}}'
            f'{_hb}>div:nth-child({_ativo}) .stButton button{{background:#FAC318!important;border-color:#FAC318!important}}'
            f'{_hb}>div:nth-child({_ativo}) .stButton button,'
            f'{_hb}>div:nth-child({_ativo}) .stButton button *{{color:#041747!important;font-weight:800!important}}'
            "</style>",
            unsafe_allow_html=True,
        )
        _pc1, _pc2, _pc3 = st.columns([1.4, 1.4, 3])
        with _pc1:
            if st.button("Atrelamento e Numeração", key="c70pill_atrel",
                         type=("primary" if _ativo == 1 else "secondary"), use_container_width=True):
                st.session_state["c70_sub"] = "atrel"
                st.rerun()
        with _pc2:
            if st.button("Mapa de recebimentos", key="c70pill_mapa",
                         type=("primary" if _ativo == 2 else "secondary"), use_container_width=True):
                st.session_state["c70_sub"] = "mapa"
                st.rerun()
    st.caption("Conta 70 · **v5.22** — se aqui não aparecer v5.22, o deploy ainda não pegou (faça Reboot do app).")
    # v5.21: renderiza AS DUAS seções sempre e esconde a inativa por CSS. Usa a
    # CHAVE do container (st-key-*) — jeito estável — com o marcador como reforço.
    # Assim circular entre as pílulas NÃO perde o estado (uploads/rodada).
    if st.session_state["c70_sub"] == "atrel":
        _hide_key, _hide_mk = "c70sec_mapa", "c70sec-mapamk"
    else:
        _hide_key, _hide_mk = "c70sec_atrel", "c70sec-atrelmk"
    st.markdown(
        "<style>"
        f'.st-key-{_hide_key}{{display:none!important}}'
        f'[data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .{_hide_mk}){{display:none!important}}'
        "</style>",
        unsafe_allow_html=True,
    )
    with st.container(key="c70sec_atrel"):
        st.markdown('<span class="c70sec-atrelmk"></span>', unsafe_allow_html=True)
        _render_conta70_atrelamento_full()
    with st.container(key="c70sec_mapa"):
        st.markdown('<span class="c70sec-mapamk"></span>', unsafe_allow_html=True)
        _render_conta70_mapa_recebimentos()


def _render_conta70_atrelamento_full():
    from src.conta70 import gerar_conta_70, carregar_historico_conta_70, STATUS_VALIDOS
    from io import BytesIO

    # v5.5: Casamento e numeração (Capa × Sankhya) — aditivo e independente da conciliação
    _render_conta70_casamento_numeracao()
    st.divider()

    resultado = st.session_state.get("resultado")
    if resultado is None:
        with st.expander("Visão complementar (vem da conciliação)", expanded=False):
            st.caption(
                "Esta seção é **opcional** e mostra os créditos não identificados a partir do "
                "**fechamento de uma conciliação**. Ela aparece depois que você roda uma conciliação — "
                "o atrelamento e a numeração acima não dependem dela."
            )
        return

    # Upload do histórico (opcional)
    with st.expander("📁 Subir histórico anterior (opcional)", expanded=False):
        st.markdown(
            "Suba aqui o arquivo `conta_70_historico.xlsx` que você mantém com os "
            "status das execuções passadas (ex.: linhas marcadas como **Regularizado**, "
            "com observações, etc). O sistema vai mesclar essas informações com os "
            "créditos da execução atual."
        )
        arq_hist = st.file_uploader(
            "Histórico Conta 70 (.xlsx)",
            type=["xlsx"],
            key="upload_historico_conta70",
            help="Arquivo gerado pelo botão 'Baixar Conta 70' em execuções anteriores, possivelmente editado por você.",
        )
        historico_df = pd.DataFrame()
        if arq_hist is not None:
            historico_df = carregar_historico_conta_70(arq_hist)
            if historico_df.empty:
                st.warning("Arquivo não pôde ser lido ou está vazio. Verifique o formato.")
            else:
                st.success(f"✅ {len(historico_df)} linhas de histórico carregadas.")

    # Gera Conta 70
    historico_df = locals().get("historico_df", pd.DataFrame())
    id_exec = st.session_state.get("id_execucao_atual", "")
    res_c70 = gerar_conta_70(
        resultado.pendentes_banco,
        historico_anterior=historico_df if not historico_df.empty else None,
        id_execucao=id_exec,
    )
    kpis_c70 = res_c70.kpis

    if res_c70.detalhado.empty:
        st.info(
            "A **conciliação bancária** desta sessão não deixou créditos não identificados. "
            "As pendências da Conta 70 (atrelamento Capa × Sankhya) aparecem na **esteira acima** — "
            "esta seção abaixo é só a visão que vem da conciliação."
        )
        return

    st.write("")
    section_title("INDICADORES — CONTA 70")

    cards1 = [
        card_kpi("Total a Lançar", fmt_brl(kpis_c70["total_a_lancar"]),
                 "valores não identificados",
                 classe="destaque-vermelho" if kpis_c70["total_a_lancar"] > 0 else ""),
        card_kpi("Quantidade", fmt_int(kpis_c70["qtd_total"]),
                 "lançamentos no controle"),
        card_kpi("Contas Envolvidas", fmt_int(kpis_c70["qtd_contas"])),
        card_kpi("Regularizados", fmt_int(kpis_c70["qtd_regularizado"]),
                 "identificados em revisões anteriores",
                 classe="destaque-verde" if kpis_c70["qtd_regularizado"] > 0 else ""),
    ]
    render_cards(cards1)

    cards2 = [
        card_kpi("Não Identificado", fmt_int(kpis_c70["qtd_nao_identificado"]),
                 classe="destaque-amarelo" if kpis_c70["qtd_nao_identificado"] > 0 else ""),
        card_kpi("Pendente de NF", fmt_int(kpis_c70["qtd_pendente_nf"])),
        card_kpi("Pendente de Baixa", fmt_int(kpis_c70["qtd_pendente_baixa"])),
        card_kpi("Em Análise", fmt_int(kpis_c70["qtd_em_analise"])),
    ]
    render_cards(cards2)

    st.divider()

    # Agrupamentos
    section_title("RESUMOS")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Por banco/conta:**")
        por_banco = res_c70.por_banco()
        if not por_banco.empty:
            por_banco["Total a Lançar"] = por_banco["Total a Lançar"].apply(fmt_brl)
            st.dataframe(por_banco, use_container_width=True, hide_index=True)
        else:
            st.caption("Sem dados.")
    with col2:
        st.markdown("**Por data:**")
        por_data = res_c70.por_data()
        if not por_data.empty:
            por_data["Total a Lançar"] = por_data["Total a Lançar"].apply(fmt_brl)
            st.dataframe(por_data, use_container_width=True, hide_index=True)
        else:
            st.caption("Sem dados.")

    st.divider()

    section_title("DETALHAMENTO — EDITÁVEL")
    st.caption(
        "Você pode editar o **status** e a **observação** de cada linha. "
        "Depois, clique em **Baixar Conta 70** pra salvar o arquivo. Na próxima execução, "
        "suba esse arquivo em '📁 Subir histórico anterior' pra manter os status."
    )

    # Filtros
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filtro_status = st.multiselect(
            "Filtrar por status",
            options=STATUS_VALIDOS,
            default=[],
            placeholder="Todos",
        )
    with col_f2:
        contas_unicas = sorted(res_c70.detalhado["conta"].unique().tolist())
        filtro_conta = st.multiselect(
            "Filtrar por conta",
            options=contas_unicas,
            default=[],
            placeholder="Todas",
        )
    with col_f3:
        filtro_tipo = st.multiselect(
            "Filtrar por tipo",
            options=sorted(res_c70.detalhado["tipo_recebimento"].unique().tolist()),
            default=[],
            placeholder="Todos",
        )

    df_visu = res_c70.detalhado.copy()
    if filtro_status:
        df_visu = df_visu[df_visu["status"].isin(filtro_status)]
    if filtro_conta:
        df_visu = df_visu[df_visu["conta"].isin(filtro_conta)]
    if filtro_tipo:
        df_visu = df_visu[df_visu["tipo_recebimento"].isin(filtro_tipo)]

    if df_visu.empty:
        st.info("Nenhuma linha corresponde aos filtros.")
        return

    # Tabela editável
    edited = st.data_editor(
        df_visu,
        column_config={
            "data": st.column_config.DateColumn("Data", format="DD/MM/YYYY", disabled=True),
            "conta": st.column_config.TextColumn("Conta", disabled=True),
            "historico": st.column_config.TextColumn("Histórico", disabled=True),
            "documento": st.column_config.TextColumn("Documento", disabled=True),
            "valor": st.column_config.NumberColumn("Valor (R$)", format="%.2f", disabled=True),
            "tipo_recebimento": st.column_config.TextColumn("Tipo", disabled=True),
            "status": st.column_config.SelectboxColumn(
                "Status",
                options=STATUS_VALIDOS,
                required=True,
            ),
            "conta_contabil": st.column_config.TextColumn("Conta Contábil", disabled=True),
            "observacao": st.column_config.TextColumn("Observação", help="Edite aqui pra anotar"),
            "data_analise": st.column_config.DatetimeColumn("Análise", disabled=True),
            "id_execucao": st.column_config.TextColumn("ID Execução", disabled=True),
        },
        use_container_width=True,
        hide_index=True,
        key="editor_conta70",
    )

    st.write("")
    section_title("EXPORTAR")

    # Download Excel + CSV
    col_dl1, col_dl2 = st.columns(2)

    buf_xlsx = BytesIO()
    with pd.ExcelWriter(buf_xlsx, engine="openpyxl") as writer:
        edited.to_excel(writer, sheet_name="Conta 70", index=False)
        res_c70.por_banco().to_excel(writer, sheet_name="Por Banco", index=False)
        res_c70.por_data().to_excel(writer, sheet_name="Por Data", index=False)
        res_c70.por_status().to_excel(writer, sheet_name="Por Status", index=False)
    buf_xlsx.seek(0)
    from datetime import date as _date
    nome_xlsx = f"conta_70_{_date.today().strftime('%Y%m%d')}.xlsx"
    with col_dl1:
        st.download_button(
            "⬇️ Baixar Conta 70 (.xlsx)",
            data=buf_xlsx.getvalue(),
            file_name=nome_xlsx,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_dl2:
        csv_bytes = edited.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Baixar Conta 70 (.csv)",
            data=csv_bytes,
            file_name=f"conta_70_{_date.today().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )


# ============================================================
# Roteamento
# ============================================================
pagina = st.session_state.pagina
_PAGINAS = {
    "Dashboard": pagina_dashboard,
    "Conciliação": pagina_conciliacao,
    "Cadastro de Taxas": pagina_cadastro_taxas,
    "Auditoria de Taxas": pagina_auditoria_taxas,
    "Conta 70": pagina_conta70,
    "Histórico": pagina_historico,
    "Sobre": pagina_sobre,
}
try:
    _fn = _PAGINAS.get(pagina)
    if _fn is not None:
        _fn()
except Exception as _err_pagina:
    st.error(
        "Ocorreu um erro ao montar esta tela. O detalhe técnico está abaixo — "
        "tire um print e mande para o suporte para correção."
    )
    st.exception(_err_pagina)


# ============================================================
# Footer
# ============================================================
st.html(
    f"""
    <div class="lle-footer">
        <strong style="color:{CORES['amarelo']};">Grupo LLE</strong> · Conciliação Bancária ·
        Aplicação interna seguindo o Manual da Marca (Fev/2026).<br>
        Dúvidas sobre identidade visual:
        <a href="mailto:marketing@grupolle.com.br">marketing@grupolle.com.br</a>
    </div>
    """
)
